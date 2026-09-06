from __future__ import annotations

from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .engine import severity_label


def build_reason(
    rule_label: str, rule_cost: float, current_pci: int, suggested_pci: int, earfcn: Any, infeasible_rules: list[dict[str, Any]]
) -> tuple[bool, str]:
    """Per-RULE outcome, not per-sector. A sector touched by two rules at
    once (e.g. Collision AND Mod3 on the same physical sector) can clear
    one while the other stays capacity-limited -- judging by the combined
    cost would wrongly call the resolved side unresolved too, just
    because a different, co-occurring rule couldn't be satisfied. Always
    a genuine reason, backed by the same pigeonhole proof the closed-loop
    optimizer itself used wherever it applies, never a silent blank."""
    resolved = rule_cost == 0

    if resolved:
        if current_pci != suggested_pci:
            return True, f"Resolved: PCI changed {current_pci} -> {suggested_pci} to clear {rule_label}."
        return True, f"Already optimal for {rule_label} -- no PCI change needed."

    matches = [p for p in infeasible_rules if p["rule"] == rule_label and p["earfcn"] == earfcn]
    if matches:
        p = matches[0]
        return False, (
            f"Cannot be resolved: {p['clique_size']} sectors are mutual neighbors on EARFCN {p['earfcn']}, "
            f"but {p['rule']} only offers {p['capacity']} distinct value(s) -- mathematically proven "
            "(pigeonhole principle), not a search limitation."
        )
    return False, f"Not resolved: no better PCI found for {rule_label} given the current neighbor state."


# Plain-language explanation of what each check/sheet actually means --
# written for a client reading the report, not for someone who already
# knows this codebase. Printed as a banner note at the top of that check's
# own sheet. "OrderN" (N=2..5) is internal hop-distance vocabulary that
# means nothing on its own to a reader -- this is what makes it legible.
RULE_DESCRIPTIONS: dict[str, str] = {
    "Collision": (
        "Two neighbor sectors reusing the EXACT same PCI on the same EARFCN -- direct interference "
        "risk, the most severe check. This is also what \"PCI reuse 1 hop apart\" means structurally."
    ),
    "Confusion": (
        "Two sectors sharing a common serving neighbor, confirmed by REAL observed handover data -- a "
        "UE can genuinely get confused between them. Structurally this is what \"PCI reuse 2 hops "
        "apart\" means, but this specific check is based on actual driven handover events, not just "
        "graph distance."
    ),
    "Grouped": "Two neighbor sectors sharing the same PCI Group (PCI // 3) -- a lower-severity SSS-level overlap.",
    "Co-centric": (
        "Informational only -- sectors at the same site pointing in similar directions. Not a fault; "
        "nothing here needs fixing."
    ),
    "Order2": (
        "Two sectors reusing the same PCI exactly 2 hops apart in the neighbor graph -- structurally "
        "similar to Confusion (see that sheet), but this is a graph-distance check rather than "
        "real-handover-verified."
    ),
    "Order3": (
        "Two sectors reusing the same PCI exactly 3 hops apart. This is a PCI reuse-DISTANCE audit "
        "metric -- it has NO standard 3GPP/RF interference meaning at this distance (a UE realistically "
        "never hears cells this far apart at once). Resolved where possible anyway, for good PCI-reuse "
        "hygiene across the network, not because it's a fault."
    ),
    "Order4": (
        "Two sectors reusing the same PCI exactly 4 hops apart. Same reuse-distance audit as the Order3 "
        "sheet -- no real interference meaning at this distance, resolved where possible for good "
        "network hygiene."
    ),
    "Order5": (
        "Two sectors reusing the same PCI exactly 5 hops apart. Same reuse-distance audit as the "
        "Order3/Order4 sheets -- no real interference meaning at this distance, resolved where possible "
        "for good network hygiene."
    ),
}


def _rule_description(rule_label: str) -> str:
    if rule_label in RULE_DESCRIPTIONS:
        return RULE_DESCRIPTIONS[rule_label]
    if rule_label.startswith("Mod"):
        n = rule_label[3:]
        meaning = " (PSS group -- real 3GPP planning meaning)" if n == "3" else " (checked because it was explicitly requested, no standard 3GPP meaning)"
        return f"Two neighbor sectors whose PCI values share the same remainder when divided by {n} (PCI mod {n}){meaning}."
    return "PCI conflict check."


def build_export_rows(result: dict[str, Any]) -> pd.DataFrame:
    """One row per (sector, rule) actually touched by a checked rule --
    never the whole project's sector list, and never one row combining
    multiple rules' outcomes into a single misleading total. Every row
    states whether THAT rule was resolved on THAT sector, and what
    changed or exactly why it could not be -- same shape as the DB save
    (services.py._build_result_rows), so Excel and DB never disagree.
    Feeds the per-rule DETAIL sheets in export_pci_optimization_excel --
    the client-facing final answer is build_final_summary_rows instead."""
    recs = result.get("recommendations")
    if recs is None or recs.empty:
        return pd.DataFrame()

    verification = result.get("verification") or {}
    infeasible_rules = verification.get("infeasible_rules", [])

    rows = []
    for _, r in recs.iterrows():
        row = r.to_dict()
        rule_costs = row.get("rule_costs") or {}
        for rule_label in sorted(row.get("conflict_types") or []):
            rule_cost = rule_costs.get(rule_label, row["after_cost"])
            resolved, reason = build_reason(rule_label, rule_cost, row["current_pci"], row["suggested_pci"], row["earfcn"], infeasible_rules)
            rows.append(
                {
                    "Site": row["site"],
                    "EARFCN": row["earfcn"],
                    "Rule": rule_label,
                    "Severity": severity_label({rule_label}),
                    "Current PCI": row["current_pci"],
                    "Suggested PCI": row["suggested_pci"],
                    "Changed": bool(row["current_pci"] != row["suggested_pci"]),
                    "Resolved": resolved,
                    "Reason": reason,
                    "Same-EARFCN Sectors Considered": row["num_same_earfcn_sectors"],
                }
            )
    df = pd.DataFrame(rows)
    return df.sort_values(by=["Resolved", "Severity"], ascending=[True, False]).reset_index(drop=True)


def build_final_summary_rows(result: dict[str, Any]) -> pd.DataFrame:
    """ONE row per sector actually touched by the optimizer -- deduplicated
    across every rule/order that flagged it, so the same PCI change is
    never explained 2-3 times under different rule names. Confirmed real
    example this replaces: site 901380 EARFCN 1750 (PCI 0 -> 363) used to
    appear as three separate rows -- under Collision, Order2, AND Order4 --
    each describing the identical change, which reads as three different
    fixes to a client instead of one. This is the sheet a client reads
    first: what to actually change, and whether it's fully done. Per-rule
    reasoning (including why an unresolved one can't be) lives on the
    per-rule detail sheets -- see 'Checks NOT Cleared' below for which one
    to check."""
    recs = result.get("recommendations")
    if recs is None or recs.empty:
        return pd.DataFrame()

    verification = result.get("verification") or {}
    infeasible_rules = verification.get("infeasible_rules", [])

    rows = []
    for _, r in recs.iterrows():
        row = r.to_dict()
        rule_costs = row.get("rule_costs") or {}
        conflict_types = sorted(row.get("conflict_types") or [])
        cleared: list[str] = []
        not_cleared: list[str] = []
        for rule_label in conflict_types:
            rule_cost = rule_costs.get(rule_label, row["after_cost"])
            resolved, _ = build_reason(rule_label, rule_cost, row["current_pci"], row["suggested_pci"], row["earfcn"], infeasible_rules)
            (cleared if resolved else not_cleared).append(rule_label)

        overall_resolved = not not_cleared
        changed = row["current_pci"] != row["suggested_pci"]
        change_text = f"PCI changed {row['current_pci']} -> {row['suggested_pci']}." if changed else "No PCI change needed."
        if overall_resolved:
            final_reason = f"{change_text} Clears: {', '.join(cleared) or 'nothing checked'}."
        else:
            final_reason = (
                f"{change_text} Cleared: {', '.join(cleared) or 'none'}. STILL NOT resolved: "
                f"{', '.join(not_cleared)} -- see that check's own sheet (tab below) for why."
            )

        rows.append(
            {
                "Site": row["site"],
                "EARFCN": row["earfcn"],
                "Current PCI": row["current_pci"],
                "Suggested PCI": row["suggested_pci"],
                "Changed": changed,
                "Checks Cleared": ", ".join(cleared),
                "Checks NOT Cleared": ", ".join(not_cleared),
                "Overall Resolved": overall_resolved,
                "Final Reason": final_reason,
            }
        )
    df = pd.DataFrame(rows)
    return df.sort_values(by=["Overall Resolved", "Site"], ascending=[True, True]).reset_index(drop=True)


def export_pci_optimization_excel(result: dict[str, Any], output_dir: Path) -> Path | None:
    """Local Excel report, same tools/*-established pattern as tilt
    recommendation's RF_Optimization_Report.xlsx (ML/outputs/<feature>/...).
    Returns None (no file written) if nothing was touched -- an empty
    report isn't useful and shouldn't be produced.

    Layout: Summary (project metadata) -> Final Optimization (ONE row per
    sector, deduplicated across rules -- the client-facing final answer,
    what to actually change) -> one DETAIL sheet per check that fired
    (Collision, Confusion, Order2, Order3, ...), each carrying its own
    plain-language description banner so "Order3" etc. is never left
    unexplained. This replaces the old single flat "Recommendations" sheet,
    which repeated the same PCI change under every rule that touched it --
    confusing once Order 2-5 reuse-distance checks started contributing
    rows alongside Collision/Confusion/Mod/Grouped."""
    export_df = build_export_rows(result)
    final_df = build_final_summary_rows(result)
    if export_df.empty or final_df.empty:
        return None

    output_dir.mkdir(parents=True, exist_ok=True)
    output_path = output_dir / "PCI_Optimization_Report.xlsx"

    verification = result.get("verification") or {}
    summary_df = pd.DataFrame(
        [
            {"Metric": "Project ID", "Value": result.get("project_id")},
            {"Metric": "Region", "Value": result.get("region")},
            {"Metric": "Operator", "Value": result.get("operator")},
            {"Metric": "Neighbor distance (m)", "Value": result.get("neighbor_distance_m")},
            {"Metric": "Sites considered", "Value": result.get("site_count")},
            {"Metric": "Sectors considered", "Value": result.get("sector_count")},
            {"Metric": "Sectors touched (see 'Final Optimization' sheet)", "Value": len(final_df)},
            {"Metric": "Sectors fully resolved", "Value": int(final_df["Overall Resolved"].sum())},
            {"Metric": "Sectors NOT fully resolved", "Value": int((~final_df["Overall Resolved"]).sum())},
            {"Metric": "Closed-loop iterations", "Value": verification.get("iterations")},
            {"Metric": "Verified clean (0 remaining)", "Value": verification.get("verified_clean")},
            {"Metric": "Stopped reason", "Value": verification.get("stopped_reason")},
        ]
    )

    wb = Workbook()
    wb.remove(wb.active)

    header_fill = PatternFill(fill_type="solid", fgColor="D9EAF7")
    resolved_fill = PatternFill(fill_type="solid", fgColor="C6EFCE")
    not_resolved_fill = PatternFill(fill_type="solid", fgColor="FFC7CE")
    bold = Font(bold=True)
    note_font = Font(italic=True, color="555555")

    def _write_sheet(name: str, df: pd.DataFrame, highlight_col: str | None, note: str | None = None) -> None:
        ws = wb.create_sheet(name)
        header_row = 1
        if note:
            ws.append([note])
            last_col = max(len(df.columns), 1)
            ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=last_col)
            ws["A1"].font = note_font
            ws["A1"].alignment = Alignment(wrap_text=True, vertical="top")
            ws.row_dimensions[1].height = 30
            header_row = 2
        ws.append(list(df.columns))
        for cell in ws[header_row]:
            cell.font = bold
            cell.fill = header_fill
        for _, row in df.iterrows():
            ws.append(list(row))
            if highlight_col:
                fill = resolved_fill if row[highlight_col] else not_resolved_fill
                for cell in ws[ws.max_row]:
                    cell.fill = fill
        ws.freeze_panes = f"A{header_row + 1}"
        for idx, col in enumerate(df.columns, start=1):
            longest = df[col].astype(str).str.len().max() if len(df) else len(col)
            ws.column_dimensions[get_column_letter(idx)].width = max(12, min(60, int(longest) + 2))

    _write_sheet("Summary", summary_df, highlight_col=None)
    _write_sheet(
        "Final Optimization",
        final_df,
        highlight_col="Overall Resolved",
        note=(
            "FINAL recommended PCI per sector -- apply these changes. A sector can be flagged by "
            "several checks at once (Collision, Confusion, PCI reuse-distance Order2-5, etc.); each "
            "check's own sheet (tabs below) explains what that check means. 'Overall Resolved' is True "
            "only when EVERY check that flagged this sector is now clear."
        ),
    )

    for rule_label in sorted(export_df["Rule"].unique()):
        sheet_df = export_df[export_df["Rule"] == rule_label].drop(columns=["Rule"]).reset_index(drop=True)
        _write_sheet(rule_label, sheet_df, highlight_col="Resolved", note=_rule_description(rule_label))

    wb.save(output_path)
    return output_path
