import uuid
import threading
import pandas as pd
import os
import datetime
import json
import traceback
from sqlalchemy import create_engine, text
from openpyxl import load_workbook
from tools.lte_tilt_recommandation.recommendation_engine import (
    TiltEngineConfig,
    export_engine_report,
    run_recommendation_engine,
)
from tools.lte_tilt_recommandation.candidate_validation import (
    build_combined_kpi_grid_impact,
    prepare_grid_metrics_export,
    prepare_scope_export,
)
from tools.lte_tilt_recommandation.cell_identity import canonical_cell_id
from utils.python_bridge import PythonBridgeError, _filter_complete_site_prediction_identity, get_bridge_client

# Global dictionary to track job status
JOBS = {}
DEFAULT_THRESHOLD_FILE = "lte_tilt_recommendation_transformed.csv"

BASELINE_IDENTITY_COLUMNS = [
    "project_id",
    "job_id",
    "grid_id",
    "lat",
    "lon",
    "nodeb_id_cell_id",
    "frontend_site_sector_key",
    "sector",
    "site_id",
    "node_b_id",
    "nodeb_id",
    "cell_id",
    "operator",
    "Technology",
    "technology",
    "pred_rsrp",
    "pred_rsrq",
    "pred_sinr",
]

BASELINE_TOPOLOGY_COLUMNS = [
    "serving_pci",
    "serving_earfcn",
    "serving_frequency_mhz",
    "best_interferer_cell_id",
    "best_interferer_pci",
    "best_interferer_earfcn",
    "best_interferer_distance_m",
    "best_interferer_azimuth_delta_deg",
    "best_interferer_proxy_phys_dbm",
    "neighbor_1_cell_id",
    "neighbor_1_pci",
    "neighbor_1_earfcn",
    "neighbor_1_proxy_rsrp_dbm",
    "neighbor_1_distance_m",
    "neighbor_1_azimuth_delta_deg",
    "neighbor_2_cell_id",
    "neighbor_2_pci",
    "neighbor_2_earfcn",
    "neighbor_2_proxy_rsrp_dbm",
    "neighbor_2_distance_m",
    "neighbor_2_azimuth_delta_deg",
    "interference_gap_db",
    "interference_ratio_linear",
    "interference_sum_proxy_dbm",
    "sinr_proxy_db",
    "rsrq_proxy_db",
    "same_earfcn_interferer_count",
    "dominant_interferer_count",
    "interference_selection_mode",
]

GRID_ANALYTICS_COLUMNS = [
    "project_id",
    "grid_id",
    "center_lat",
    "center_lon",
    "min_lat",
    "max_lat",
    "min_lon",
    "max_lon",
    "baseline_point_count",
    "baseline_avg_rsrp",
    "baseline_avg_rsrq",
    "baseline_avg_sinr",
    "operator",
    "created_at",
    "scenario_id",
]


def _clean_id_series(series: pd.Series) -> pd.Series:
    out = series.astype(str).str.strip()
    return out.str[:-2].where(out.str.endswith(".0"), out)


def _rf_identity_series(series: pd.Series) -> pd.Series:
    out = series.astype(str).str.strip()
    out = out.replace({"": pd.NA, "None": pd.NA, "none": pd.NA, "nan": pd.NA, "NaN": pd.NA, "<NA>": pd.NA})
    out = out.str.replace("|", "_", regex=False)
    out = out.str.replace(r"\.0(?=_|$)", "", regex=True)
    return out


def _build_node_cell_id(nodeb_series: pd.Series, cell_series: pd.Series) -> pd.Series:
    nodeb = _clean_id_series(nodeb_series)
    cell = _clean_id_series(cell_series)
    cell_has_node = cell.str.contains("_", regex=False).fillna(False)
    missing_nodeb = nodeb.isin(["", "None", "none", "nan", "NaN", "<NA>"])
    direct_cell = cell.where(cell_has_node & missing_nodeb, "")
    built = (nodeb + "_" + cell).str.strip("_")
    return direct_cell.where(direct_cell.ne(""), built)


def _build_site_sector_node_cell_id(site_series: pd.Series, cell_series: pd.Series) -> pd.Series:
    site = site_series.astype(str).str.strip()
    site = site.str[:-2].where(site.str.endswith(".0"), site)
    site = site.mask(site.isin(["", "None", "none", "nan", "NaN", "<NA>"]), "")
    cell = _clean_id_series(cell_series)
    return (site + "_" + cell).str.strip("_")


def _get_table_columns(current_engine, table_name: str) -> set[str]:
    query = text(
        """
        SELECT column_name
        FROM information_schema.columns
        WHERE table_schema = DATABASE()
          AND table_name = :table_name
        """
    )
    with current_engine.connect() as conn:
        schema_df = pd.read_sql(query, conn, params={"table_name": table_name})
    if schema_df.empty:
        return set()
    schema_col = "column_name" if "column_name" in schema_df.columns else schema_df.columns[0]
    return set(schema_df[schema_col].astype(str).tolist())


def _select_existing_columns(current_engine, table_name: str, requested_cols: list[str]) -> list[str]:
    available = _get_table_columns(current_engine, table_name)
    return [col for col in requested_cols if col in available]


def _fetch_latest_baseline_job_id(current_engine, project_id: int) -> str:
    query = text(
        """
        SELECT job_id
        FROM lte_prediction_baseline_results
        WHERE project_id = :project_id
        ORDER BY created_at DESC
        LIMIT 1
        """
    )
    with current_engine.connect() as conn:
        row = conn.execute(query, {"project_id": int(project_id)}).fetchone()
    if not row or row[0] is None:
        raise ValueError(f"No baseline results found for project {project_id}")
    return str(row[0])


def _prepare_tilt_antenna_df(antenna_df: pd.DataFrame) -> pd.DataFrame:
    out = antenna_df.copy()
    out = out.loc[:, ~out.columns.duplicated()].copy()
    if "cell_id" in out.columns and "local_cell_id" not in out.columns:
        out["local_cell_id"] = _clean_id_series(out["cell_id"])
    strict_identity_col = next(
        (col for col in ["site_prediction_key", "site_cell_sector_band_operator_key"] if col in out.columns),
        None,
    )
    if strict_identity_col:
        out["Node_Cell_ID"] = _rf_identity_series(out[strict_identity_col])
    elif "Node_Cell_ID" not in out.columns and {"site", "cell_id"}.issubset(out.columns):
        if "nodeb_id" in out.columns:
            nodeb_available = ~_clean_id_series(out["nodeb_id"]).isin(["", "None", "none", "nan", "NaN", "<NA>"])
            out["Node_Cell_ID"] = _build_node_cell_id(out["nodeb_id"], out["cell_id"])
        else:
            nodeb_available = pd.Series(False, index=out.index)
            out["Node_Cell_ID"] = _clean_id_series(out["cell_id"])
        out.loc[~nodeb_available, "Node_Cell_ID"] = _build_site_sector_node_cell_id(out.loc[~nodeb_available, "site"], out.loc[~nodeb_available, "cell_id"])
    elif "Node_Cell_ID" not in out.columns and {"nodeb_id", "cell_id"}.issubset(out.columns):
        out["Node_Cell_ID"] = _build_node_cell_id(out["nodeb_id"], out["cell_id"])
    if "lat" not in out.columns and "latitude" in out.columns:
        out["lat"] = out["latitude"]
    if "lon" not in out.columns and "longitude" in out.columns:
        out["lon"] = out["longitude"]
    if "electrical_tilt" not in out.columns and "e_tilt" in out.columns:
        out["electrical_tilt"] = out["e_tilt"]
    if "mechanical_tilt" not in out.columns and "m_tilt" in out.columns:
        out["mechanical_tilt"] = out["m_tilt"]
    if "antenna_height" not in out.columns and "height" in out.columns:
        out["antenna_height"] = out["height"]
    if "dashboard_site_id" not in out.columns and "nodeb_id" in out.columns:
        out["dashboard_site_id"] = _clean_id_series(out["nodeb_id"])
    if "Node_Cell_ID" in out.columns:
        out["Node_Cell_ID"] = _rf_identity_series(out["Node_Cell_ID"])
    alias_cols = {
        "latitude": "lat",
        "longitude": "lon",
        "e_tilt": "electrical_tilt",
        "m_tilt": "mechanical_tilt",
        "height": "antenna_height",
    }
    drop_aliases = [alias for alias, canonical in alias_cols.items() if alias in out.columns and canonical in out.columns]
    if drop_aliases:
        out = out.drop(columns=drop_aliases)
        print(f"[TILT][ANTENNA_PREP] dropped_optimizer_alias_columns={drop_aliases}")
    return out


def _prepare_tilt_log_df(log_df: pd.DataFrame, antenna_df: pd.DataFrame) -> pd.DataFrame:
    out = log_df.copy()
    canonical_node_cell = pd.Series(pd.NA, index=out.index, dtype="object")
    if "nodeb_id_cell_id" in out.columns and "Node_Cell_ID" not in out.columns:
        out["Node_Cell_ID"] = out["nodeb_id_cell_id"].astype(str).str.strip()
    if "nodeb_id_cell_id" in out.columns:
        nodeb_cell_key = _rf_identity_series(out["nodeb_id_cell_id"])
        canonical_node_cell = canonical_node_cell.where(canonical_node_cell.notna(), nodeb_cell_key)
    if "frontend_site_sector_key" in out.columns:
        frontend_key = _rf_identity_series(out["frontend_site_sector_key"])
        canonical_node_cell = canonical_node_cell.where(canonical_node_cell.notna(), frontend_key)
        out["Node_Cell_ID"] = frontend_key.where(frontend_key.ne(""), out.get("Node_Cell_ID", frontend_key))
    if "Node_Cell_ID" in out.columns:
        node_cell_key = _rf_identity_series(out["Node_Cell_ID"])
        canonical_node_cell = canonical_node_cell.where(canonical_node_cell.notna(), node_cell_key)
    if "Node_Cell_ID" in out.columns:
        node_cell = out["Node_Cell_ID"].astype(str).str.strip()
        split_source = node_cell.str.split("_", n=1, expand=True)
        if "node_b_id" not in out.columns and split_source.shape[1] >= 1:
            out["node_b_id"] = split_source[0]
        if "nodeb_id" not in out.columns and split_source.shape[1] >= 1:
            out["nodeb_id"] = split_source[0]
        if "cell_id" not in out.columns and split_source.shape[1] >= 2:
            out["cell_id"] = split_source[1]
    if "node_b_id" in out.columns and "nodeb_id" not in out.columns:
        out["nodeb_id"] = _clean_id_series(out["node_b_id"])
    elif "nodeb_id" in out.columns:
        out["nodeb_id"] = _clean_id_series(out["nodeb_id"])
    if "cell_id" in out.columns:
        out["cell_id"] = _clean_id_series(out["cell_id"])
        out["local_cell_id"] = out["cell_id"]
    if canonical_node_cell.notna().any():
        out["Node_Cell_ID"] = canonical_node_cell
    elif {"nodeb_id", "cell_id"}.issubset(out.columns):
        out["Node_Cell_ID"] = _build_node_cell_id(out["nodeb_id"], out["cell_id"])

    if "Node_Cell_ID" in antenna_df.columns:
        ant = antenna_df.copy()
        ant["Node_Cell_ID"] = ant["Node_Cell_ID"].astype(str).str.strip()
        merge_cols = [
            col for col in [
                "Node_Cell_ID",
                "Technology",
                "lat",
                "lon",
                "azimuth",
                "electrical_tilt",
                "mechanical_tilt",
                "tx_power",
                "antenna_height",
                "dashboard_site_id",
            ] if col in ant.columns
        ]
        if merge_cols:
            ant = ant[merge_cols].drop_duplicates(subset=["Node_Cell_ID"], keep="last")
            out = out.merge(ant, on="Node_Cell_ID", how="left", suffixes=("", "_site"))
    if "Technology" not in out.columns:
        out["Technology"] = "4G"
    else:
        out["Technology"] = out["Technology"].fillna("4G")
    if "operator" not in out.columns:
        out["operator"] = "Unknown"
    return out


def _fetch_baseline_log_df(current_engine, project_id: int, operator_input, is_all_operators: bool) -> tuple[pd.DataFrame, str]:
    baseline_job_id = _fetch_latest_baseline_job_id(current_engine, int(project_id))
    requested_cols = list(dict.fromkeys(BASELINE_IDENTITY_COLUMNS + BASELINE_TOPOLOGY_COLUMNS + ["created_at"]))
    select_cols = _select_existing_columns(current_engine, "lte_prediction_baseline_results", requested_cols)
    required_cols = {"lat", "lon", "pred_rsrp", "pred_rsrq", "pred_sinr"}
    missing_required = sorted(required_cols.difference(select_cols))
    if missing_required:
        raise ValueError(f"Baseline table is missing required columns: {missing_required}")
    if "job_id" not in select_cols:
        raise ValueError("Baseline table is missing required job_id column for latest-run isolation.")

    filters = ["project_id = :pid", "job_id = :job_id"]
    params = {"pid": int(project_id), "job_id": baseline_job_id}
    if not is_all_operators and "operator" in select_cols:
        filters.append("LOWER(TRIM(operator)) = :op")
        params["op"] = str(operator_input).strip().lower()

    query = text(
        f"""
        SELECT {", ".join(f"`{col}`" for col in select_cols)}
        FROM lte_prediction_baseline_results
        WHERE {" AND ".join(filters)}
        """
    )
    log_dfs = []
    with current_engine.connect() as conn:
        for chunk in pd.read_sql(query, conn, params=params, chunksize=50000):
            print(f"[TILT][BASELINE_FETCH] chunk_rows={len(chunk)} baseline_job_id={baseline_job_id}")
            log_dfs.append(chunk)
    if not log_dfs:
        raise ValueError(f"No baseline rows found for project {project_id} job_id={baseline_job_id}")
    log_df = pd.concat(log_dfs, ignore_index=True)
    print(
        f"[TILT][BASELINE_FETCH] rows={len(log_df)} baseline_job_id={baseline_job_id} "
        f"columns_loaded={len(select_cols)} topology_columns={len([c for c in BASELINE_TOPOLOGY_COLUMNS if c in select_cols])}"
    )
    return log_df, baseline_job_id


def _fetch_grid_analytics_df(current_engine, project_id: int, operator_input, is_all_operators: bool) -> pd.DataFrame:
    select_cols = _select_existing_columns(current_engine, "grid_analytics_results", GRID_ANALYTICS_COLUMNS)
    if not {"project_id", "grid_id"}.issubset(select_cols):
        print("[TILT][GRID_ANALYTICS_FETCH] skipped reason=missing_table_or_required_columns")
        return pd.DataFrame()

    filters = ["project_id = :pid"]
    params = {"pid": int(project_id)}
    if not is_all_operators and "operator" in select_cols:
        filters.append("LOWER(TRIM(operator)) = :op")
        params["op"] = str(operator_input).strip().lower()

    selected_scenario_id = None
    scenario_filter = ""
    if "scenario_id" in select_cols:
        scenario_query = text(
            f"""
            SELECT scenario_id, MAX(created_at) AS max_created, COUNT(*) AS row_count
            FROM grid_analytics_results
            WHERE {" AND ".join(filters)}
            GROUP BY scenario_id
            ORDER BY row_count DESC, max_created DESC
            LIMIT 1
            """
        )
        with current_engine.connect() as conn:
            scenario_row = conn.execute(scenario_query, params).fetchone()
        if scenario_row is not None:
            selected_scenario_id = scenario_row[0]
            if selected_scenario_id is None:
                scenario_filter = " AND scenario_id IS NULL"
            else:
                scenario_filter = " AND scenario_id = :scenario_id"
                params["scenario_id"] = selected_scenario_id

    order_cols = [col for col in ["created_at", "grid_id"] if col in select_cols]
    order_sql = ", ".join(f"{col} DESC" if col == "created_at" else f"{col} ASC" for col in order_cols)
    query = text(
        f"""
        SELECT {", ".join(f"`{col}`" for col in select_cols)}
        FROM grid_analytics_results
        WHERE {" AND ".join(filters)}
        {scenario_filter}
        {f"ORDER BY {order_sql}" if order_sql else ""}
        """
    )
    with current_engine.connect() as conn:
        grid_df = pd.read_sql(query, conn, params=params)
    print(
        f"[TILT][GRID_ANALYTICS_FETCH] rows={len(grid_df)} selected_scenario_id={selected_scenario_id} "
        f"operator_filter_applied={bool(not is_all_operators and 'operator' in select_cols)}"
    )
    return grid_df


def _bridge_latest_baseline_job_id(bridge, project_id: int, region: str, operator_input, is_all_operators: bool) -> str:
    params = {
        "projectId": int(project_id),
        "region": region,
    }
    if not is_all_operators:
        params["operator"] = operator_input
    payload = bridge._request("GET", "GetLatestLteBaselineJobId", params=params)
    job_id = payload.get("JobId") or payload.get("jobId")
    if not job_id:
        raise ValueError(f"No baseline results found for project {project_id}")
    return str(job_id)


def _bridge_fetch_baseline_log_df(bridge, project_id: int, region: str, operator_input, is_all_operators: bool) -> tuple[pd.DataFrame, str]:
    baseline_job_id = _bridge_latest_baseline_job_id(bridge, int(project_id), region, operator_input, is_all_operators)
    params = {
        "projectId": int(project_id),
        "region": region,
        "jobId": baseline_job_id,
    }
    if not is_all_operators:
        params["operator"] = operator_input
    page_limit = int(os.getenv("PYTHON_BRIDGE_BASELINE_PAGE_SIZE", "50000"))
    log_df = bridge.get_rows(
        "GetLteBaselineRows",
        params,
        limit=page_limit,
        progress_label="lte_baseline",
    )
    if log_df.empty:
        raise ValueError(f"No baseline rows found for project {project_id} job_id={baseline_job_id}")
    if "job_id" in log_df.columns:
        log_df = log_df.loc[log_df["job_id"].astype(str) == baseline_job_id].copy()
    if not is_all_operators and "operator" in log_df.columns:
        requested_operator = str(operator_input).strip().lower()
        log_df = log_df.loc[log_df["operator"].astype(str).str.strip().str.lower() == requested_operator].copy()
    if log_df.empty:
        raise ValueError(f"No baseline rows found for project {project_id} job_id={baseline_job_id}")
    print(f"[TILT][BASELINE_FETCH] source=python_bridge rows={len(log_df)} baseline_job_id={baseline_job_id}")
    return log_df, baseline_job_id


def _bridge_fetch_grid_analytics_df(bridge, project_id: int, operator_input, is_all_operators: bool) -> pd.DataFrame:
    try:
        grid_df, _grid_size = bridge.get_grid_analytics(int(project_id))
    except PythonBridgeError as exc:
        print(f"[TILT][GRID_ANALYTICS_FETCH] source=python_bridge skipped=True reason={exc}")
        return pd.DataFrame()
    if not is_all_operators and "operator" in grid_df.columns:
        grid_df = grid_df.loc[grid_df["operator"].astype(str).str.strip().str.lower() == str(operator_input).strip().lower()].copy()
    print(f"[TILT][GRID_ANALYTICS_FETCH] source=python_bridge rows={len(grid_df)}")
    return grid_df


def _bridge_next_scenario_id(bridge, project_id: int) -> int:
    payload = bridge._request("GET", "GetNextRfOptimizationScenarioId", params={"projectId": int(project_id)})
    return int(payload.get("ScenarioId") or payload.get("scenarioId") or 1)


def _bridge_save_rf_optimization_results(bridge, project_id: int, scenario_id: int, df: pd.DataFrame) -> int:
    rows = []
    safe_df = df.replace({pd.NA: None}).where(pd.notna(df), None)
    for row in safe_df.to_dict(orient="records"):
        rows.append(
            {
                "operator": row.get("operator"),
                "cell_id": row.get("cell_id"),
                "technology": row.get("technology"),
                "parameter": row.get("parameter"),
                "current_value": None if row.get("current_value") is None else str(row.get("current_value")),
                "recommended_value": None if row.get("recommended_value") is None else str(row.get("recommended_value")),
                "reason": row.get("reason"),
                "swap_sector_detected": row.get("swap_sector_detected"),
                "rsrp_threshold": row.get("rsrp_threshold"),
                "rsrq_threshold": row.get("rsrq_threshold"),
                "sinr_threshold": row.get("sinr_threshold"),
                "created_at": row.get("created_at").isoformat() if hasattr(row.get("created_at"), "isoformat") else row.get("created_at"),
            }
        )
    payload = bridge._request(
        "POST",
        "SaveRfOptimizationResults",
        json={"ProjectId": int(project_id), "ScenarioId": int(scenario_id), "Rows": rows},
    )
    return int(payload.get("Inserted") or payload.get("inserted") or 0)


def _safe_nunique(df, col):
    return int(df[col].nunique(dropna=True)) if col in df.columns else "n/a"


def _safe_minmax(df, col):
    if col not in df.columns:
        return "n/a"
    series = pd.to_numeric(df[col], errors="coerce").dropna()
    if series.empty:
        return "n/a"
    return f"{series.min():.4f}..{series.max():.4f}"


def _log_df(stage, df):
    print(f"[TILT][{stage}] shape={df.shape}")
    print(f"[TILT][{stage}] columns={list(df.columns)}")
    for col in ["cell_id", "nodeb_id", "node_b_id", "operator", "final_operator", "Technology"]:
        if col in df.columns:
            print(f"[TILT][{stage}] distinct_{col}={_safe_nunique(df, col)}")
    for col in ["rsrp", "rsrq", "sinr", "pred_rsrp", "pred_rsrq", "pred_sinr"]:
        if col in df.columns:
            print(f"[TILT][{stage}] range_{col}={_safe_minmax(df, col)}")


def _normalize_constraint_bool(value) -> bool:
    if isinstance(value, bool):
        return value
    if pd.isna(value):
        return False
    return str(value).strip().lower() in {"true", "1", "yes", "y"}


def _to_clean_cell_id(value) -> str:
    return canonical_cell_id(value)


def _to_threshold_cell_id(value) -> str:
    return canonical_cell_id(value)


def _resolve_threshold_file_path(cfg: dict, project_id: int, project_root: str) -> str:
    supplied = str(cfg.get("threshold_file_path", "") or cfg.get("threshold_file", "")).strip()
    if supplied:
        return supplied
    default_path = os.path.join(project_root, DEFAULT_THRESHOLD_FILE)
    if os.path.exists(default_path):
        return default_path
    return ""


def _load_constraint_df(file_path: str) -> pd.DataFrame:
    if not file_path:
        return pd.DataFrame()
    ext = os.path.splitext(file_path)[1].lower()
    if ext in {".xlsx", ".xls"}:
        df = pd.read_excel(file_path)
    else:
        df = pd.read_csv(file_path)

    df = df.copy()
    df.columns = [str(col).strip() for col in df.columns]
    lower_map = {str(col).strip().lower(): col for col in df.columns}
    if "cell_id" not in lower_map:
        raise ValueError("Constraint file must contain a cell_id column")

    rename_map = {}
    for required in [
        "cell_id",
        "min_m_tilt", "max_m_tilt",
        "min_e_tilt", "max_e_tilt",
        "min_height", "max_height",
        "min_azimuth", "max_azimuth",
        "min_tx_power", "max_tx_power",
        "optimised",
    ]:
        if required in lower_map:
            rename_map[lower_map[required]] = required
    df = df.rename(columns=rename_map)
    df["cell_id"] = df["cell_id"].map(_to_threshold_cell_id)
    if "optimised" not in df.columns:
        df["optimised"] = False
    df["optimised"] = df["optimised"].map(_normalize_constraint_bool)

    numeric_cols = [
        "min_m_tilt", "max_m_tilt",
        "min_e_tilt", "max_e_tilt",
        "min_height", "max_height",
        "min_azimuth", "max_azimuth",
        "min_tx_power", "max_tx_power",
    ]
    for col in numeric_cols:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors="coerce")
    df = df.drop_duplicates(subset=["cell_id"], keep="last")
    return df


def _angular_distance_deg(a: float, b: float) -> float:
    return abs((float(a) - float(b) + 180.0) % 360.0 - 180.0)


def _azimuth_in_range(value: float, min_az: float, max_az: float) -> bool:
    value = float(value) % 360.0
    min_az = float(min_az) % 360.0
    max_az = float(max_az) % 360.0
    if min_az <= max_az:
        return min_az <= value <= max_az
    return value >= min_az or value <= max_az


def _clamp_azimuth(value: float, min_az: float, max_az: float) -> float:
    value = float(value) % 360.0
    min_norm = float(min_az) % 360.0
    max_norm = float(max_az) % 360.0
    if _azimuth_in_range(value, min_norm, max_norm):
        return value
    return min_norm if _angular_distance_deg(value, min_norm) <= _angular_distance_deg(value, max_norm) else max_norm


def _apply_constraint_ranges(reco_df: pd.DataFrame, constraint_df: pd.DataFrame) -> pd.DataFrame:
    if reco_df.empty or constraint_df.empty:
        return reco_df

    out = reco_df.copy()
    out["Cell ID"] = out["Cell ID"].map(_to_clean_cell_id)
    constraint_map = constraint_df.set_index("cell_id").to_dict("index")

    out["Constraint Applied"] = "No"
    out["Constraint Source"] = ""
    out["Allowed Min"] = pd.NA
    out["Allowed Max"] = pd.NA

    for idx, row in out.iterrows():
        cell_id = _to_threshold_cell_id(row.get("Cell ID"))
        param = str(row.get("Parameter", "")).strip().lower()
        cfg = constraint_map.get(cell_id)
        if not cfg or not bool(cfg.get("optimised")):
            continue

        if param == "etilt":
            min_col, max_col = "min_e_tilt", "max_e_tilt"
            is_azimuth = False
        elif param == "azimuth":
            min_col, max_col = "min_azimuth", "max_azimuth"
            is_azimuth = True
        elif param in {"tx power", "power"}:
            min_col, max_col = "min_tx_power", "max_tx_power"
            is_azimuth = False
        elif param in {"mechanical tilt", "mtilt"}:
            min_col, max_col = "min_m_tilt", "max_m_tilt"
            is_azimuth = False
        elif param in {"height", "antenna height"}:
            min_col, max_col = "min_height", "max_height"
            is_azimuth = False
        else:
            continue

        min_allowed = cfg.get(min_col)
        max_allowed = cfg.get(max_col)
        if pd.isna(min_allowed) or pd.isna(max_allowed):
            continue

        rec_value = pd.to_numeric(pd.Series([row.get("Recommended Value")]), errors="coerce").iloc[0]
        if pd.isna(rec_value):
            continue

        constrained_value = rec_value
        if is_azimuth:
            constrained_value = _clamp_azimuth(rec_value, float(min_allowed), float(max_allowed))
            changed = not _azimuth_in_range(rec_value, float(min_allowed), float(max_allowed))
        else:
            constrained_value = float(min(max(rec_value, float(min_allowed)), float(max_allowed)))
            changed = not pd.isna(rec_value) and float(constrained_value) != float(rec_value)

        out.at[idx, "Allowed Min"] = min_allowed
        out.at[idx, "Allowed Max"] = max_allowed
        out.at[idx, "Constraint Source"] = "cell_threshold_file"
        if changed:
            out.at[idx, "Recommended Value"] = constrained_value
            out.at[idx, "Constraint Applied"] = "Yes"
            existing_reason = str(row.get("Reason", "") or "").strip()
            suffix = f" Recommended value constrained within configured range [{min_allowed}, {max_allowed}] for this cell."
            out.at[idx, "Reason"] = (existing_reason + suffix).strip()
    return out


def _log_constraint_summary(raw_reco_df: pd.DataFrame, constrained_reco_df: pd.DataFrame, constraint_df: pd.DataFrame, threshold_file_path: str):
    if constraint_df.empty:
        print("[TILT][CONSTRAINT_APPLY] used=False reason=constraint_df_empty")
        return

    eligible_cells = int(constraint_df.loc[constraint_df["optimised"] == True, "cell_id"].nunique()) if "optimised" in constraint_df.columns else 0
    eligible_rows = 0
    if not raw_reco_df.empty and not constraint_df.empty:
        optimised_cells = set(constraint_df.loc[constraint_df["optimised"] == True, "cell_id"].astype(str).tolist()) if "optimised" in constraint_df.columns else set()
        eligible_rows = int(raw_reco_df["Cell ID"].map(_to_threshold_cell_id).isin(optimised_cells).sum())

    applied_rows = 0
    if "Constraint Applied" in constrained_reco_df.columns:
        applied_rows = int((constrained_reco_df["Constraint Applied"].astype(str).str.strip().str.lower() == "yes").sum())

    print(
        f"[TILT][CONSTRAINT_APPLY] used=True path={threshold_file_path} "
        f"constraint_rows={len(constraint_df)} eligible_cells={eligible_cells} "
        f"eligible_recommendation_rows={eligible_rows} applied_rows={applied_rows}"
    )


def _write_recommendations_back_to_workbook(output_file: str, reco_df: pd.DataFrame):
    if reco_df.empty or not os.path.exists(output_file):
        return
    wb = load_workbook(output_file)
    if "Recommendations" not in wb.sheetnames:
        return
    ws = wb["Recommendations"]
    headers = [ws.cell(row=1, column=col_idx).value for col_idx in range(1, ws.max_column + 1)]
    header_map = {str(header).strip(): idx for idx, header in enumerate(headers, start=1) if header is not None}
    required_headers = [h for h in ["Recommended Value", "Reason"] if h not in header_map]
    if required_headers:
        return

    optional_cols = ["Constraint Applied", "Allowed Min", "Allowed Max", "Constraint Source"]
    next_col = ws.max_column + 1
    for col in optional_cols:
        if col not in header_map and col in reco_df.columns:
            ws.cell(row=1, column=next_col, value=col)
            header_map[col] = next_col
            next_col += 1

    for row_idx, (_, row) in enumerate(reco_df.iterrows(), start=2):
        if row_idx > ws.max_row:
            break
        for col_name, value in row.items():
            if col_name not in header_map:
                continue
            ws.cell(row=row_idx, column=header_map[col_name], value=None if pd.isna(value) else value)
    wb.save(output_file)

# ==========================================================
# MULTI-REGION DATABASE ENGINES
# Aggressive Connection Recycling prevents SQLAlchemy from using "dead" connections
# ==========================================================
engine = {
    # India Database (uses your original DATABASE_URL)
    "india": create_engine(
        os.getenv("DATABASE_URL"),
        pool_size=10, max_overflow=20, pool_recycle=300, pool_pre_ping=True
    ) if os.getenv("DATABASE_URL") else None,
    
    # Taiwan Database (Matches your exact .env variable name)
    "taiwan": create_engine(
        os.getenv("DATABASE_URL_Taiwan"), 
        pool_size=10, max_overflow=20, pool_recycle=300, pool_pre_ping=True
    ) if os.getenv("DATABASE_URL_Taiwan") else None
}

class RFOptimizationService:

    def submit(self, cfg):
        job_id = str(uuid.uuid4())
        JOBS[job_id] = {"status": "queued"}
        threading.Thread(target=self._run, args=(job_id, cfg), daemon=True).start()
        return {"job_id": job_id}

    def get(self, job_id):
        return JOBS.get(job_id)

    def _get_next_scenario_id(self, project_id, current_engine):
        """Finds the next scenario number for a given project using the correct regional DB."""
        query = text("SELECT COALESCE(MAX(scenario_id), 0) + 1 FROM rf_optimization_results WHERE project_id = :pid")
        try:
            with current_engine.connect() as conn:
                result = conn.execute(query, {"pid": project_id}).scalar()
                return int(result) if result else 1
        except Exception as e:
            print(f"Error fetching scenario ID: {e}")
            return 1

    def _run(self, job_id, cfg):
        try:
            # ==========================================
            # 1. REGION & ENGINE SELECTION
            # ==========================================
            region = str(cfg.get("region", "india")).lower()
            current_engine = engine.get(region)
            bridge = get_bridge_client()
            use_bridge = bridge is not None

            if current_engine is None and not use_bridge:
                raise ValueError(f"Database config for region '{region}' not found or .env variable is missing.")

            self._update(job_id, "running", f"Initializing {region.upper()} optimization job...")
            project_id = cfg["project_id"]
            print(
                f"[TILT][JOB_START] job_id={job_id} project_id={project_id} region={region} "
                f"operator={cfg.get('operator')} rsrp={cfg.get('rsrp', -105)} "
                f"rsrq={cfg.get('rsrq', -15)} sinr={cfg.get('sinr', 0)}"
            )
            
            operator_input = cfg.get("operator")
            is_all_operators = not operator_input or str(operator_input).lower() in ["all", "none", ""]

            r_thresh = str(cfg.get("rsrp", -105))
            q_thresh = str(cfg.get("rsrq", -15))
            s_thresh = str(cfg.get("sinr", 0))

            # ==========================================
            # 2. FETCH ANTENNA DATA FIRST (Prevents Timeout)
            # ==========================================
            self._update(job_id, "running", "Fetching antenna records...")
            if use_bridge:
                antenna_df = bridge.get_rows("GetLteTiltAntennaRows", {"projectId": int(project_id)}, limit=50000)
                print(f"[TILT][ANTENNA_FETCH] source=python_bridge rows={len(antenna_df)}")
            else:
                ant_query = text("""
                    SELECT
                        site_prediction.*,
                        cluster AS provider,
                        cluster AS operator_name
                    FROM site_prediction
                    WHERE tbl_project_id = :pid
                """)
                with current_engine.connect() as conn:
                    antenna_df = pd.read_sql(ant_query, conn, params={"pid": project_id})
                antenna_df = _filter_complete_site_prediction_identity(antenna_df, endpoint="direct:site_prediction_tilt")
            antenna_df = _prepare_tilt_antenna_df(antenna_df)
            _log_df("ANTENNA_FETCH", antenna_df)

            # ==========================================
            # 3. FETCH LARGE LOG DATA IN CHUNKS
            # ==========================================
            self._update(job_id, "running", "Downloading calibrated baseline rows...")
            if use_bridge:
                log_df, baseline_job_id = _bridge_fetch_baseline_log_df(
                    bridge,
                    project_id,
                    region,
                    operator_input,
                    is_all_operators,
                )
            else:
                log_df, baseline_job_id = _fetch_baseline_log_df(
                    current_engine,
                    project_id,
                    operator_input,
                    is_all_operators,
                )
            log_df = _prepare_tilt_log_df(log_df, antenna_df)
            _log_df("BASELINE_FETCH_COMBINED", log_df)

            # ==========================================
            # 4. ROBUST OPERATOR MAPPING
            # ==========================================
            self._update(job_id, "running", "Processing optimization script...")

            if "Node_Cell_ID" in log_df.columns:
                log_df["clean_key"] = _clean_id_series(log_df["Node_Cell_ID"])
            else:
                log_df["clean_key"] = (
                    _clean_id_series(log_df["node_b_id"]) + "_" +
            _clean_id_series(log_df["cell_id"])
                )
            log_df["clean_key"] = _rf_identity_series(log_df["clean_key"])
            operator_map = log_df.drop_duplicates("clean_key").set_index("clean_key")["operator"].to_dict()
            print(f"[TILT][OPERATOR_MAP] mapped_keys={len(operator_map)}")

            # ==========================================
            # 5. PREPARE PATHS & RUN PRODUCTION ENGINE
            # ==========================================
            current_dir = os.path.dirname(os.path.abspath(__file__))
            root_dir = os.path.normpath(os.path.join(current_dir, "..", ".."))
            temp_dir = os.path.normpath(os.path.join(root_dir, "outputs", f"temp_{job_id}"))
            os.makedirs(temp_dir, exist_ok=True)

            self._update(job_id, "running", "Fetching geo-feature rows...")
            if use_bridge:
                geo_df = bridge.get_rows(
                    "GetLtePredictionGeoFeatures",
                    {"projectId": int(project_id), "region": region},
                    limit=50000,
                    progress_label="lte_geo_features",
                )
                print(f"[TILT][GEO_FETCH] source=python_bridge rows={len(geo_df)}")
            else:
                geo_query = text(
                    """
                    SELECT
                        lat,
                        lon,
                        nodeb_id_cell_id,
                        clutter_class,
                        morphology_cluster,
                        building_count,
                        building_area_ratio,
                        avg_building_area_m2,
                        road_length_m,
                        green_ratio,
                        water_ratio,
                        los_blocker_count,
                        los_blocked_ratio,
                        max_blocker_height_m,
                        diffraction_proxy_db,
                        nlos_flag,
                        terrain_elevation_m,
                        terrain_slope_deg,
                        proxy_site_elevation_m,
                        terrain_relief_to_site_m,
                        site_count_250m,
                        site_count_500m,
                        serving_distance_m,
                        nearest_site_distance_m,
                        mean_nearest3_site_distance_m,
                        azimuth_delta_deg
                    FROM lte_prediction_geo_features
                    WHERE project_id = :pid
                      AND region = :region
                    """
                )
                with current_engine.connect() as conn:
                    geo_df = pd.read_sql(geo_query, conn, params={"pid": project_id, "region": region})
            if "nodeb_id_cell_id" in geo_df.columns:
                geo_df["Node_Cell_ID"] = geo_df["nodeb_id_cell_id"].astype(str).str.strip()
            _log_df("GEO_FETCH", geo_df)

            self._update(job_id, "running", "Fetching frontend grid analytics...")
            if use_bridge:
                grid_analytics_df = _bridge_fetch_grid_analytics_df(
                    bridge,
                    project_id,
                    operator_input,
                    is_all_operators,
                )
            else:
                grid_analytics_df = _fetch_grid_analytics_df(
                    current_engine,
                    project_id,
                    operator_input,
                    is_all_operators,
                )
            _log_df("GRID_ANALYTICS_FETCH", grid_analytics_df)

            scenario_id = _bridge_next_scenario_id(bridge, project_id) if use_bridge else self._get_next_scenario_id(project_id, current_engine)
            threshold_file_path = _resolve_threshold_file_path(cfg, project_id, root_dir)
            if threshold_file_path:
                print(f"[TILT][CONSTRAINT_FILE] path={threshold_file_path}")
            else:
                print(f"[TILT][CONSTRAINT_FILE] path=n/a project_id={project_id}")
            constraint_df = _load_constraint_df(threshold_file_path) if threshold_file_path else pd.DataFrame()
            if not constraint_df.empty:
                print(
                    f"[TILT][CONSTRAINT_FETCH] rows={len(constraint_df)} "
                    f"optimised_rows={int(constraint_df['optimised'].sum()) if 'optimised' in constraint_df.columns else 0}"
                )

            self._update(job_id, "running", "Running production tilt recommendation engine...")
            engine_config = TiltEngineConfig(
                project_id=int(project_id),
                region=region,
                operator=None if is_all_operators else operator_input,
                mode=cfg.get("mode") or cfg.get("kpi_mode") or cfg.get("recommendation_mode") or "combined_weighted",
                rsrp_threshold=float(r_thresh),
                rsrq_threshold=float(q_thresh),
                sinr_threshold=float(s_thresh),
                rsrp_weight=float(cfg.get("rsrp_weight", 34.0)),
                rsrq_weight=float(cfg.get("rsrq_weight", 33.0)),
                sinr_weight=float(cfg.get("sinr_weight", 33.0)),
                validate_candidates=bool(cfg.get("validate_candidates", True)),
                max_validation_candidates=int(cfg.get("max_validation_candidates", cfg.get("max_candidates", 25))),
                radius_m=float(cfg.get("radius_m", cfg.get("radius", 500.0))),
                grid_resolution_m=float(cfg.get("grid_resolution_m", cfg.get("grid_resolution", 30.0))),
                workers=int(cfg.get("n_workers", cfg.get("workers", 1))),
                impact_radius_m=float(cfg.get("impact_radius_m", cfg.get("radius_m", cfg.get("radius", 500.0)))),
                neighbor_site_count=int(cfg.get("neighbor_site_count", 3)),
                max_interference_sites=int(cfg.get("max_interference_sites", 10)),
                baseline_job_id=baseline_job_id,
                coordinate_passes=int(cfg.get("coordinate_passes", 2)),
                candidate_workers=int(cfg.get("candidate_workers", 1)),
                bad_grid_coverage_pct=float(cfg.get("bad_grid_coverage_pct", 80.0)),
                max_group_cells=int(cfg.get("max_group_cells", 0)),
                max_neighbors_per_update_cell=int(cfg.get("max_neighbors_per_update_cell", 2)),
                rf_debug_log_path=os.path.join(temp_dir, "tilt_rf_debug.log"),
                constraint_map=constraint_df.set_index("cell_id").to_dict("index") if not constraint_df.empty else {},
            )
            print(
                "[TILT][ENGINE_CONFIG] "
                f"project_id={engine_config.project_id} region={engine_config.region} "
                f"operator={engine_config.operator or 'all'} "
                f"thresholds=({engine_config.rsrp_threshold},{engine_config.rsrq_threshold},{engine_config.sinr_threshold}) "
                f"weights=({engine_config.rsrp_weight},{engine_config.rsrq_weight},{engine_config.sinr_weight}) "
                f"validate_candidates={engine_config.validate_candidates} "
                f"radius_m={engine_config.radius_m} grid_resolution_m={engine_config.grid_resolution_m} "
                f"workers={engine_config.workers} impact_radius_m={engine_config.impact_radius_m} "
                f"neighbor_site_count={engine_config.neighbor_site_count} "
                f"max_interference_sites={engine_config.max_interference_sites} "
                f"candidate_workers={engine_config.candidate_workers} "
                f"coordinate_passes={engine_config.coordinate_passes} "
                f"bad_grid_coverage_pct={engine_config.bad_grid_coverage_pct} "
                f"max_group_cells={engine_config.max_group_cells} "
                f"max_neighbors_per_update_cell={engine_config.max_neighbors_per_update_cell}"
            )
            engine_outputs = run_recommendation_engine(
                log_df=log_df,
                antenna_df=antenna_df,
                geo_df=geo_df,
                grid_analytics_df=grid_analytics_df,
                config=engine_config,
            )
            output_file = os.path.join(temp_dir, "RF_Optimization_Report.xlsx")
            export_engine_report(engine_outputs, output_file)
            candidate_eval_df = engine_outputs.get("candidate_evaluations", pd.DataFrame())
            if isinstance(candidate_eval_df, pd.DataFrame) and not candidate_eval_df.empty:
                candidate_eval_df.to_csv(os.path.join(temp_dir, "candidate_validation_results.csv"), index=False)
            grid_scores_df = engine_outputs.get("grid_scores", pd.DataFrame())
            if isinstance(grid_scores_df, pd.DataFrame) and not grid_scores_df.empty:
                grid_scores_df.to_csv(os.path.join(temp_dir, "frontend_grid_scores.csv"), index=False)
            best_after_df = engine_outputs.get("best_candidate_after", pd.DataFrame())
            if isinstance(best_after_df, pd.DataFrame) and not best_after_df.empty:
                weights = {
                    "rsrp": float(engine_config.rsrp_weight) / max(float(engine_config.rsrp_weight + engine_config.rsrq_weight + engine_config.sinr_weight), 1e-9),
                    "rsrq": float(engine_config.rsrq_weight) / max(float(engine_config.rsrp_weight + engine_config.rsrq_weight + engine_config.sinr_weight), 1e-9),
                    "sinr": float(engine_config.sinr_weight) / max(float(engine_config.rsrp_weight + engine_config.rsrq_weight + engine_config.sinr_weight), 1e-9),
                }
                thresholds = {"rsrp": float(r_thresh), "rsrq": float(q_thresh), "sinr": float(s_thresh)}
                before_scope_df = prepare_scope_export(log_df, thresholds, weights, "before")
                after_scope_df = prepare_scope_export(best_after_df, thresholds, weights, "after")
                before_scope_df.to_csv(os.path.join(temp_dir, "best_candidate_before_scope.csv.gz"), index=False, compression="gzip")
                after_scope_df.to_csv(os.path.join(temp_dir, "best_candidate_after_scope.csv.gz"), index=False, compression="gzip")
                before_bad_combined = before_scope_df.loc[before_scope_df["is_bad_combined"].fillna(False)].copy()
                after_bad_combined = after_scope_df.loc[after_scope_df["is_bad_combined"].fillna(False)].copy()
                before_bad_combined.to_csv(os.path.join(temp_dir, "best_candidate_before_bad_combined.csv.gz"), index=False, compression="gzip")
                after_bad_combined.to_csv(os.path.join(temp_dir, "best_candidate_after_bad_combined.csv.gz"), index=False, compression="gzip")
                before_grid_df = prepare_grid_metrics_export(before_scope_df, thresholds, weights)
                after_grid_df = prepare_grid_metrics_export(after_scope_df, thresholds, weights)
                before_grid_df.to_csv(os.path.join(temp_dir, "best_candidate_before_grid_metrics.csv"), index=False)
                after_grid_df.to_csv(os.path.join(temp_dir, "best_candidate_after_grid_metrics.csv"), index=False)
                combined_impact_df = build_combined_kpi_grid_impact(before_grid_df, after_grid_df, thresholds, weights)
                if not combined_impact_df.empty:
                    combined_impact_df.to_csv(os.path.join(temp_dir, "combined_kpi_grid_impact.csv"), index=False)
                best_summary = {}
                best_metrics_df = engine_outputs.get("best_candidate_metrics", pd.DataFrame())
                if isinstance(best_metrics_df, pd.DataFrame) and not best_metrics_df.empty:
                    best_summary = best_metrics_df.iloc[0].to_dict()
                elif isinstance(candidate_eval_df, pd.DataFrame) and not candidate_eval_df.empty:
                    ranked_eval = candidate_eval_df.copy()
                    for col in ["constraints_passed", "score", "net_bad_reduction", "combined_weighted_tie_break", "new_bad_samples"]:
                        if col in ranked_eval.columns:
                            ranked_eval[col] = pd.to_numeric(ranked_eval[col], errors="coerce")
                    sort_cols = [col for col in ["constraints_passed", "score", "net_bad_reduction", "combined_weighted_tie_break"] if col in ranked_eval.columns]
                    ascending = [False] * len(sort_cols)
                    if "new_bad_samples" in ranked_eval.columns:
                        sort_cols.append("new_bad_samples")
                        ascending.append(True)
                    if sort_cols:
                        ranked_eval = ranked_eval.sort_values(sort_cols, ascending=ascending, na_position="last")
                    best_summary = ranked_eval.iloc[0].to_dict()
                best_summary.update({
                    "kpi_mode": str(engine_config.mode),
                    "thresholds": thresholds,
                    "weights": weights,
                    "artifact_source": "production_coordinate_search_rf_validated_combined_weighted",
                    "candidate_summary_source": "best_candidate_metrics" if isinstance(best_metrics_df, pd.DataFrame) and not best_metrics_df.empty else "ranked_candidate_evaluations",
                })
                with open(os.path.join(temp_dir, "best_candidate_summary.json"), "w", encoding="utf-8") as fh:
                    json.dump(best_summary, fh, indent=2, default=str)
            reco_df = engine_outputs["recommendations"].copy()
            if reco_df.empty:
                print(
                    f"[TILT][RECOMMENDATIONS_EMPTY] project_id={project_id} "
                    f"baseline_job_id={baseline_job_id} reason=no_actionable_rows"
                )

            # ==========================================
            # 6. SAVE RESULTS MAPPING CORRECT OPERATOR
            # ==========================================
            self._update(job_id, "running", "Saving recommendations to database...")
            if not constraint_df.empty:
                raw_reco_df = reco_df.copy()
                reco_df = _apply_constraint_ranges(reco_df, constraint_df)
                _log_constraint_summary(raw_reco_df, reco_df, constraint_df, threshold_file_path)
                _write_recommendations_back_to_workbook(output_file, reco_df)
            else:
                print("[TILT][CONSTRAINT_APPLY] used=False reason=no_constraint_file_or_no_rows")
            _log_df("RECOMMENDATIONS_FETCH", reco_df)
            if reco_df.empty:
                print(
                    f"[TILT][DB_WRITE] skipped=True reason=no_rf_validated_recommendations "
                    f"project_id={project_id} scenario_id={scenario_id}"
                )
                JOBS[job_id].update({
                    "status": "done",
                    "output": output_file,
                    "scenario": scenario_id,
                    "recommendation_rows": 0,
                    "inserted": 0,
                })
                return
            
            # Ensure "ALL" is replaced by the actual cell's operator
            reco_df["final_operator"] = reco_df["Cell ID"].astype(str).map(operator_map).fillna(
                operator_input if (operator_input and not is_all_operators) else "Unknown"
            )

            db_save_df = pd.DataFrame({
                "project_id": project_id,
                "scenario_id": scenario_id,
                "operator": reco_df["final_operator"],
                "cell_id": reco_df["Cell ID"],
                "technology": reco_df["Technology"],
                "parameter": reco_df["Parameter"],
                "current_value": reco_df["Current Value"],
                "recommended_value": reco_df["Recommended Value"],
                "reason": reco_df["Reason"],
                "swap_sector_detected": reco_df["Swap Sector Detected"],
                "rsrp_threshold": float(r_thresh),
                "rsrq_threshold": float(q_thresh),
                "sinr_threshold": float(s_thresh),
                "created_at": datetime.datetime.now()
            })
            _log_df("DB_PAYLOAD", db_save_df)
            print(
                f"[TILT][DB_WRITE] table=rf_optimization_results mode=append "
                f"rows={len(db_save_df)} project_id={project_id} scenario_id={scenario_id}"
            )

            if use_bridge:
                inserted = _bridge_save_rf_optimization_results(bridge, project_id, scenario_id, db_save_df)
                print(f"[TILT][DB_WRITE_DONE] source=python_bridge rows={inserted}")
            else:
                # Fast DB saving with chunks using the correct regional engine
                with current_engine.begin() as conn:
                    db_save_df.to_sql("rf_optimization_results", conn, if_exists="append", index=False, method="multi", chunksize=1000)
                inserted = len(db_save_df)

            JOBS[job_id].update({
                "status": "done",
                "output": output_file,
                "scenario": scenario_id,
                "recommendation_rows": len(db_save_df),
                "inserted": inserted,
            })

        except Exception as e:
            print(f"Error in RF Service: {str(e)}")
            traceback.print_exc()
            JOBS[job_id].update({"status": "failed", "error": str(e)})

    def _update(self, job_id, status, msg):
        JOBS[job_id].update({"status": status, "progress": msg})
