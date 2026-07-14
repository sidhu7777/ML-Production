from __future__ import annotations

import argparse
import json
import os
import shutil
import sys
import logging
import time
import io
import subprocess
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Iterable

import numpy as np
import pandas as pd
import joblib
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


ML_ROOT = Path(__file__).resolve().parents[2]
os.chdir(ML_ROOT)
if str(ML_ROOT) not in sys.path:
    sys.path.insert(0, str(ML_ROOT))

from tests.coverage_prediction.coverage_artifact_locator import resolve_coverage_artifact_path
from tests.coverage_prediction import lte_coverage_test as coverage_test
from tests.coverage_prediction import train_model1_coverage_xgboost as model1_base
from tests.coverage_prediction import train_model1_hybrid_target_experiment as model1_hybrid
from tests.coverage_prediction import train_model2_hybrid_capacity_xgboost as model2_hybrid
from tests.coverage_prediction import build_model3_hybrid_load_balancing_dataset as model3_builder
from tools.lte_prediction import ml_engine as base_ml


DEFAULT_MODEL3_DATASET = (
    ML_ROOT / "models" / "model3_hybrid_load_balancing_experiment" / "model3_load_balancing_dataset.csv"
)
DEFAULT_MODEL3_SUMMARY = (
    ML_ROOT / "models" / "model3_hybrid_load_balancing_experiment" / "model3_load_balancing_summary.json"
)
DEFAULT_OUTPUT_ROOT = ML_ROOT / "tests" / "output" / "model3_business_rule_recommendation"
DEFAULT_STABLE_OUTPUT_DIR = ML_ROOT / "models" / "model3_hybrid_load_balancing_experiment"
DEFAULT_RRC_SECTOR_CAPACITY = 400.0
DEFAULT_CONGESTION_THRESHOLD = 80.0

BAND_PRIORITY = {
    2300.0: 6,
    2100.0: 5,
    1800.0: 4,
    900.0: 3,
    850.0: 2,
    700.0: 1,
}

SYNTHETIC_BAND_TO_EARFCN = {
    700: 700.0,
    850: 850.0,
    900: 900.0,
    1800: 1750.0,
    2100: 2100.0,
    2300: 2300.0,
}


@dataclass
class Model3RecommendationConfig:
    dataset_path: Path = DEFAULT_MODEL3_DATASET
    summary_path: Path = DEFAULT_MODEL3_SUMMARY
    output_root: Path = DEFAULT_OUTPUT_ROOT
    stable_output_dir: Path = DEFAULT_STABLE_OUTPUT_DIR
    congestion_threshold: float = DEFAULT_CONGESTION_THRESHOLD
    rrc_sector_capacity: float = DEFAULT_RRC_SECTOR_CAPACITY
    sector_split_local_radius_m: float = 900.0
    model1_variant: str = "physical_no_teacher_summary"


def _timestamp() -> str:
    return time.strftime("%Y%m%d_%H%M%S")


def _ensure_dir(path: Path) -> Path:
    path.mkdir(parents=True, exist_ok=True)
    return path


def _save_json(path: Path, payload: dict[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(payload, indent=2, default=str), encoding="utf-8")


def _setup_logger(log_path: Path) -> logging.Logger:
    logger = logging.getLogger(f"model3_business_rule_{log_path.stem}")
    logger.setLevel(logging.INFO)
    logger.handlers.clear()
    logger.propagate = False

    formatter = logging.Formatter("%(asctime)s | %(levelname)s | %(message)s", "%Y-%m-%d %H:%M:%S")

    file_handler = logging.FileHandler(log_path, mode="w", encoding="utf-8")
    file_handler.setFormatter(formatter)
    file_handler.setLevel(logging.INFO)
    logger.addHandler(file_handler)

    console_handler = logging.StreamHandler(sys.stdout)
    console_handler.setFormatter(formatter)
    console_handler.setLevel(logging.INFO)
    logger.addHandler(console_handler)
    return logger


def _pick_col(df: pd.DataFrame, candidates: Iterable[str], required: bool = True) -> str:
    lower_map = {str(c).strip().lower(): c for c in df.columns}
    for candidate in candidates:
        key = str(candidate).strip().lower()
        if key in lower_map:
            return lower_map[key]
    if required:
        raise KeyError(f"Could not find any of: {list(candidates)}")
    return ""


def _first_non_empty(series: pd.Series) -> Any:
    work = series.dropna()
    if work.empty:
        return ""
    for value in work.tolist():
        text = str(value).strip()
        if text and text.lower() != "nan":
            return value
    return ""


def _mode_text(series: pd.Series) -> str:
    work = series.dropna().astype(str).str.strip()
    work = work[work != ""]
    if work.empty:
        return ""
    modes = work.mode(dropna=True)
    return str(modes.iloc[0]) if not modes.empty else str(work.iloc[0])


def _to_num(series: pd.Series) -> pd.Series:
    return pd.to_numeric(series, errors="coerce")


def _fmt_band(value: Any) -> str:
    try:
        if pd.isna(value):
            return ""
    except Exception:
        pass
    try:
        num = float(value)
    except Exception:
        return str(value).strip()
    if float(num).is_integer():
        return str(int(num))
    return f"{num:g}"


def _normalize_identity_text(value: Any) -> str:
    text = str(value).strip()
    if not text or text.lower() == "nan":
        return ""
    try:
        num = float(text)
    except Exception:
        return text
    if float(num).is_integer():
        return str(int(num))
    return f"{num:g}"


def _band_priority(value: Any) -> int:
    try:
        band = float(value)
    except Exception:
        return 0
    return BAND_PRIORITY.get(band, 0)


def _threshold_counts(series: pd.Series, thresholds: Iterable[float]) -> dict[str, int]:
    values = _to_num(series)
    return {f"gt_{int(threshold)}": int((values > threshold).sum()) for threshold in thresholds}


def _scalar_to_int(value: Any, default: int = 0) -> int:
    try:
        if pd.isna(value):
            return int(default)
    except Exception:
        pass
    try:
        return int(float(value))
    except Exception:
        return int(default)


def _scalar_to_float(value: Any, default: float = np.nan) -> float:
    try:
        if pd.isna(value):
            return float(default)
    except Exception:
        pass
    try:
        return float(value)
    except Exception:
        return float(default)


def _scalar_to_bool(value: Any, default: bool = False) -> bool:
    try:
        if pd.isna(value):
            return bool(default)
    except Exception:
        pass
    if isinstance(value, str):
        text = value.strip().lower()
        if text in {"true", "1", "yes", "y"}:
            return True
        if text in {"false", "0", "no", "n", ""}:
            return False
    return bool(value)


def _excel_safe_value(value: Any) -> Any:
    if value is None:
        return None
    try:
        if pd.isna(value):
            return None
    except Exception:
        pass
    if isinstance(value, np.generic):
        return value.item()
    if isinstance(value, (list, tuple, set)):
        return ", ".join(_fmt_band(v) for v in value if _fmt_band(v))
    if isinstance(value, dict):
        return json.dumps(value, ensure_ascii=False, default=str)
    return value


def _archive_root(archive_path: Path) -> str:
    listed = subprocess.check_output(["tar", "-tf", str(archive_path)], text=True)
    first_file = next((line for line in listed.splitlines() if "/" in line and not line.endswith("/")), "")
    if not first_file:
        raise RuntimeError(f"No files found in archive: {archive_path}")
    return first_file.split("/", 1)[0]


def _read_csv_from_archive(archive_path: Path, member_name: str) -> pd.DataFrame:
    root = _archive_root(archive_path)
    raw = subprocess.check_output(["tar", "-xOf", str(archive_path), f"{root}/{member_name}"])
    return pd.read_csv(io.BytesIO(raw))


def _extract_site_id(series: pd.Series) -> pd.Series:
    return series.astype(str).str.rsplit("_", n=1).str[0]


def _haversine_distance_m_series(lat1: pd.Series, lon1: pd.Series, lat2: float, lon2: float) -> pd.Series:
    lat1_rad = np.radians(pd.to_numeric(lat1, errors="coerce"))
    lon1_rad = np.radians(pd.to_numeric(lon1, errors="coerce"))
    lat2_rad = np.radians(float(lat2))
    lon2_rad = np.radians(float(lon2))
    dlat = lat2_rad - lat1_rad
    dlon = lon2_rad - lon1_rad
    a = np.sin(dlat / 2.0) ** 2 + np.cos(lat1_rad) * np.cos(lat2_rad) * np.sin(dlon / 2.0) ** 2
    return 2.0 * 6371000.0 * np.arctan2(np.sqrt(a), np.sqrt(1.0 - a))


def _safe_merge_first(left: pd.Series, right: pd.Series) -> pd.Series:
    return left.combine_first(right) if len(left) == len(right) else left


def _make_local_point_map(pred_df: pd.DataFrame) -> pd.DataFrame:
    keep = [
        col
        for col in [
            "lat",
            "lon",
            "grid_id",
            "grid_row",
            "grid_col",
            "grid_centroid_lat",
            "grid_centroid_lon",
            "cell_area_m2",
        ]
        if col in pred_df.columns
    ]
    if not {"lat", "lon"}.issubset(keep):
        return pd.DataFrame()
    return pred_df[keep].dropna(subset=["lat", "lon"]).drop_duplicates(subset=["lat", "lon"]).copy()


def _merge_point_map_by_coordinates(pred_df: pd.DataFrame, point_map: pd.DataFrame) -> pd.DataFrame:
    if pred_df.empty or point_map.empty or not {"lat", "lon"}.issubset(pred_df.columns) or not {"lat", "lon"}.issubset(point_map.columns):
        return pred_df
    work = pred_df.copy()
    lookup = point_map.copy()
    work["_lat_key"] = pd.to_numeric(work["lat"], errors="coerce").round(6)
    work["_lon_key"] = pd.to_numeric(work["lon"], errors="coerce").round(6)
    lookup["_lat_key"] = pd.to_numeric(lookup["lat"], errors="coerce").round(6)
    lookup["_lon_key"] = pd.to_numeric(lookup["lon"], errors="coerce").round(6)
    add_cols = [col for col in lookup.columns if col not in {"lat", "lon", "_lat_key", "_lon_key"}]
    merged = work.merge(
        lookup[["_lat_key", "_lon_key", *add_cols]].drop_duplicates(subset=["_lat_key", "_lon_key"]),
        on=["_lat_key", "_lon_key"],
        how="left",
        suffixes=("", "_point"),
    )
    merged = merged.drop(columns=["_lat_key", "_lon_key"], errors="ignore")
    return merged


def _ensure_grid_group_columns(df: pd.DataFrame, default_time_bucket: str = "PART_3") -> pd.DataFrame:
    work = df.copy()
    if work.empty:
        return work
    lat_series = pd.to_numeric(work.get("lat"), errors="coerce") if "lat" in work.columns else pd.Series(np.nan, index=work.index)
    lon_series = pd.to_numeric(work.get("lon"), errors="coerce") if "lon" in work.columns else pd.Series(np.nan, index=work.index)
    lat_key = lat_series.round(6)
    lon_key = lon_series.round(6)
    pair_labels = lat_key.astype(str).fillna("nan") + "|" + lon_key.astype(str).fillna("nan")
    pair_codes = pd.Series(pd.factorize(pair_labels)[0] + 1, index=work.index)

    if "grid_id" not in work.columns or pd.to_numeric(work["grid_id"], errors="coerce").notna().sum() == 0:
        work["grid_id"] = pair_codes.astype("Int64")
    else:
        work["grid_id"] = pd.to_numeric(work["grid_id"], errors="coerce").astype("Int64")
        work.loc[work["grid_id"].isna(), "grid_id"] = pair_codes.loc[work["grid_id"].isna()].astype("Int64")

    if "grid_row" not in work.columns or pd.to_numeric(work["grid_row"], errors="coerce").notna().sum() == 0:
        work["grid_row"] = lat_key.rank(method="dense").astype("Int64")
    else:
        work["grid_row"] = pd.to_numeric(work["grid_row"], errors="coerce").astype("Int64")
        missing = work["grid_row"].isna()
        if missing.any():
            work.loc[missing, "grid_row"] = lat_key.loc[missing].rank(method="dense").astype("Int64")

    if "grid_col" not in work.columns or pd.to_numeric(work["grid_col"], errors="coerce").notna().sum() == 0:
        work["grid_col"] = lon_key.rank(method="dense").astype("Int64")
    else:
        work["grid_col"] = pd.to_numeric(work["grid_col"], errors="coerce").astype("Int64")
        missing = work["grid_col"].isna()
        if missing.any():
            work.loc[missing, "grid_col"] = lon_key.loc[missing].rank(method="dense").astype("Int64")

    if "grid_centroid_lat" not in work.columns or pd.to_numeric(work["grid_centroid_lat"], errors="coerce").notna().sum() == 0:
        work["grid_centroid_lat"] = lat_series
    else:
        work["grid_centroid_lat"] = pd.to_numeric(work["grid_centroid_lat"], errors="coerce").fillna(lat_series)

    if "grid_centroid_lon" not in work.columns or pd.to_numeric(work["grid_centroid_lon"], errors="coerce").notna().sum() == 0:
        work["grid_centroid_lon"] = lon_series
    else:
        work["grid_centroid_lon"] = pd.to_numeric(work["grid_centroid_lon"], errors="coerce").fillna(lon_series)

    if "time_bucket" not in work.columns:
        work["time_bucket"] = default_time_bucket
    else:
        work["time_bucket"] = work["time_bucket"].astype(str).replace({"": default_time_bucket, "nan": default_time_bucket}).fillna(default_time_bucket)
    return work


def _match_site_coordinates(pred_df: pd.DataFrame, site_df: pd.DataFrame) -> pd.DataFrame:
    work = pred_df.copy()
    site_work = site_df.copy()
    for col in ["Node_Cell_ID", "cell_id"]:
        if col in work.columns:
            work[col] = work[col].astype(str)
        if col in site_work.columns:
            site_work[col] = site_work[col].astype(str)

    lookup = site_work.rename(
        columns={
            "lat": "serving_site_lat",
            "lon": "serving_site_lon",
            "azimuth": "serving_site_azimuth",
        }
    )
    join_attempts = []
    if "original_cell_id" in work.columns and "cell_id" in lookup.columns:
        join_attempts.append(("original_cell_id", "cell_id"))
    if "original_node_cell_id" in work.columns and "Node_Cell_ID" in lookup.columns:
        join_attempts.append(("original_node_cell_id", "Node_Cell_ID"))
    if "Node_Cell_ID" in work.columns and "Node_Cell_ID" in lookup.columns:
        join_attempts.append(("Node_Cell_ID", "Node_Cell_ID"))

    for left_key, right_key in join_attempts:
        merged = work.merge(
            lookup[[right_key, "serving_site_lat", "serving_site_lon", "serving_site_azimuth"]].drop_duplicates(subset=[right_key]),
            left_on=left_key,
            right_on=right_key,
            how="left",
        )
        if merged["serving_site_lat"].notna().any():
            return merged.drop(columns=[right_key], errors="ignore")
    work["serving_site_lat"] = np.nan
    work["serving_site_lon"] = np.nan
    work["serving_site_azimuth"] = np.nan
    return work


def _build_dense_part3_features(
    *,
    baseline_pred_df: pd.DataFrame,
    corrected_pred_df: pd.DataFrame,
    kpi_df: pd.DataFrame,
    geo_df: pd.DataFrame,
    site_df: pd.DataFrame,
) -> pd.DataFrame:
    pred_df = baseline_pred_df.copy()
    pred_df = _ensure_grid_group_columns(pred_df)
    corrected_pred_df = _ensure_grid_group_columns(corrected_pred_df)
    group_cols = ["grid_id", "grid_row", "grid_col", "grid_centroid_lat", "grid_centroid_lon", "time_bucket"]
    point_map_cols = [col for col in ["lat", "lon", *group_cols] if col in pred_df.columns]
    if {"lat", "lon"}.issubset(point_map_cols):
        point_map = pred_df[point_map_cols].drop_duplicates(subset=["lat", "lon"])
        missing_pred_cols = [col for col in group_cols if col not in pred_df.columns]
        if missing_pred_cols and {"lat", "lon"}.issubset(pred_df.columns):
            pred_df = pred_df.merge(point_map[["lat", "lon", *missing_pred_cols]], on=["lat", "lon"], how="left")
        missing_corrected_cols = [col for col in group_cols if col not in corrected_pred_df.columns]
        if missing_corrected_cols and {"lat", "lon"}.issubset(corrected_pred_df.columns):
            corrected_pred_df = corrected_pred_df.merge(
                point_map[["lat", "lon", *missing_corrected_cols]],
                on=["lat", "lon"],
                how="left",
            )
    if "original_cell_id" in pred_df.columns:
        pred_df["Node_Cell_ID"] = (
            pred_df["original_cell_id"]
            .astype(str)
            .replace({"nan": pd.NA, "": pd.NA})
            .combine_first(pred_df.get("original_node_cell_id", pd.Series(pd.NA, index=pred_df.index)).astype(str))
            .combine_first(pred_df["Node_Cell_ID"].astype(str))
        )
    pred_df = model1_base._add_rf_topology_features(pred_df, site_df)
    dense_df = (
        pred_df.groupby(group_cols, as_index=False)
        .agg(
            baseline_rsrp=("pred_rsrp", "mean"),
            baseline_rsrq=("pred_rsrq", "mean"),
            baseline_sinr=("pred_sinr", "mean"),
            pred_rsrp_min=("pred_rsrp", "min"),
            pred_rsrp_max=("pred_rsrp", "max"),
            pred_rsrp_std=("pred_rsrp", "std"),
            pred_sinr_std=("pred_sinr", "std"),
            measurement_count=("pred_rsrp", "count"),
            unique_cells=("Node_Cell_ID", "nunique"),
            unique_sites=("site_id", "nunique"),
            serving_distance_m=("serving_distance_m", "mean"),
            nearest_site_distance_m=("nearest_site_distance_m", "mean"),
            site_count_250m=("site_count_250m", "mean"),
            site_count_500m=("site_count_500m", "mean"),
            azimuth_delta_deg=("azimuth_delta_deg", "mean"),
        )
    )
    dense_df["bucket_seq"] = 3.0
    dense_df["bucket_min_timestamp"] = pd.to_datetime("2026-02-11 00:00:00")
    dense_df["bucket_max_timestamp"] = pd.to_datetime("2026-05-16 23:59:59")
    dense_df["bucket_mid_timestamp"] = dense_df["bucket_min_timestamp"]

    corrected_required = set(group_cols + ["morphology_cluster", "interference_gap_db", "interference_ratio_linear", "terrain_elevation_m", "terrain_slope_deg", "los_blocked_ratio", "nlos_flag"])
    corrected_features = (
        model1_base._derive_corrected_surface_features(corrected_pred_df)
        if corrected_required.issubset(set(corrected_pred_df.columns))
        else pd.DataFrame()
    )
    if not corrected_features.empty:
        corrected_features = corrected_features.drop_duplicates(subset=group_cols, keep="first")
        dense_df = dense_df.merge(corrected_features, on=group_cols, how="left", validate="one_to_one")
    band_features = model1_base._derive_band_features_from_kpi(kpi_df)
    if not band_features.empty:
        band_features = band_features.drop_duplicates(subset=group_cols, keep="first")
        dense_df = dense_df.merge(band_features, on=group_cols, how="left", validate="one_to_one")
    geo_clean = model1_hybrid.clean_geo_columns(geo_df)
    geo_clean = geo_clean.drop_duplicates(subset=["grid_id", "grid_row", "grid_col", "time_bucket"], keep="first")
    dense_df = dense_df.merge(
        geo_clean,
        on=["grid_id", "grid_row", "grid_col", "time_bucket"],
        how="left",
        validate="many_to_one",
    )
    return dense_df


def _load_resimulation_context(config: Model3RecommendationConfig, logger: logging.Logger) -> dict[str, Any]:
    archive_path = resolve_coverage_artifact_path()
    # Read directly from the archive because the run directory is not always preserved.
    project_sites_df = _read_csv_from_archive(archive_path, "project_sites.csv")
    coverage_rows_df = _read_csv_from_archive(archive_path, "coverage_rows.csv")
    baseline_pred_df = _read_csv_from_archive(archive_path, "baseline_prediction_grid.csv")
    corrected_pred_df = _read_csv_from_archive(archive_path, "bucket_corrected_prediction_grid.csv")
    geo_df = _read_csv_from_archive(archive_path, "bucket_grid_geo_features.csv")
    kpi_df = _read_csv_from_archive(archive_path, "grid_kpi_timeseries.csv")
    with_summary = json.loads(subprocess.check_output(["tar", "-xOf", str(archive_path), f"{_archive_root(archive_path)}/summary.json"], text=True))

    for frame in [coverage_rows_df, baseline_pred_df, corrected_pred_df, geo_df, kpi_df]:
        if "time_bucket" in frame.columns:
            frame["time_bucket"] = frame["time_bucket"].astype(str)

    part3_site_map, _ = coverage_test._build_bucket_site_topologies(
        site_df=project_sites_df,
        buckets=coverage_test.DEFAULT_BUCKETS,
        operator_name=str(with_summary.get("topology_operator") or "Airtel"),
    )
    part3_site_df = part3_site_map.get("PART_3", pd.DataFrame()).copy()
    building_df = pd.DataFrame(columns=["geometry_wkt"])
    logger.info("resim_context local_only_inputs building_rows=%d", len(building_df))
    hybrid_history = model1_hybrid.load_or_build_hybrid_dataset()
    hybrid_history["time_bucket"] = hybrid_history["time_bucket"].astype(str)
    model2_base_df = pd.read_csv(model3_builder.MODEL2_BASE_CSV)
    model2_base_df["time_bucket"] = model2_base_df["time_bucket"].astype(str)
    model2_base_df["grid_id"] = pd.to_numeric(model2_base_df["grid_id"], errors="coerce").astype("Int64")

    model1_features = {}
    numeric_features, categorical_features = model1_hybrid.variant_features(config.model1_variant)
    for target in model1_base.TARGETS:
        model1_features[target] = {
            "model": model1_hybrid.load_model(config.model1_variant, target),
            "features": numeric_features + categorical_features,
        }
    model2_models = {}
    for target_key, target_cfg in model2_hybrid.TARGET_CONFIG.items():
        model2_models[target_cfg["target"]] = joblib.load(model2_hybrid.MODEL_ROOT / target_cfg["model_file"])

    logger.info(
        "resim_context archive=%s part3_sites=%d part3_baseline_rows=%d part3_corrected_rows=%d",
        archive_path,
        len(part3_site_df),
        int((baseline_pred_df["time_bucket"] == "PART_3").sum()) if "time_bucket" in baseline_pred_df.columns else len(baseline_pred_df),
        int((corrected_pred_df["time_bucket"] == "PART_3").sum()) if "time_bucket" in corrected_pred_df.columns else len(corrected_pred_df),
    )
    return {
        "archive_path": archive_path,
        "summary": with_summary,
        "project_sites_df": project_sites_df,
        "coverage_rows_df": coverage_rows_df,
        "baseline_pred_df": baseline_pred_df,
        "corrected_pred_df": corrected_pred_df,
        "geo_df": geo_df,
        "kpi_df": kpi_df,
        "part3_site_df": part3_site_df,
        "building_df": building_df,
        "hybrid_history": hybrid_history,
        "model2_base_df": model2_base_df,
        "model1_models": model1_features,
        "model2_models": model2_models,
    }


def _pick_sector_key(df: pd.DataFrame) -> str:
    return _pick_col(
        df,
        [
            "topology_frontend_site_sector_key",
            "topology_node_cell_sector_key",
            "topology_sector_identity_key",
            "topology_canonical_sector_id",
            "topology_site_sector_band_key",
            "sector_id",
        ],
    )


def _pick_site_key(df: pd.DataFrame) -> str:
    return _pick_col(
        df,
        [
            "topology_site_id",
            "site_id",
            "Site ID",
            "topology_site_identity_key",
        ],
        required=False,
    )


def _build_cell_inventory(df: pd.DataFrame, config: Model3RecommendationConfig) -> tuple[pd.DataFrame, dict[str, Any]]:
    sector_key = _pick_sector_key(df)
    site_key = _pick_site_key(df)

    rows: list[dict[str, Any]] = []
    for node_cell_id, group in df.groupby("Node_Cell_ID", dropna=False):
        sector_id = _first_non_empty(group[sector_key]) if sector_key in group.columns else ""
        site_id = _first_non_empty(group[site_key]) if site_key and site_key in group.columns else ""
        prb = _to_num(group["estimated_prb_utilization_pct"])
        cell_rrc = _to_num(group["estimated_cell_rrc_utilization_pct"])
        rrc_users = _to_num(group["estimated_cell_rrc_connected_users"])

        row = {
            "Node_Cell_ID": node_cell_id,
            "site_id": site_id,
            "sector_id": sector_id,
            "band": _fmt_band(_first_non_empty(group.get("topology_band", pd.Series(dtype=object)))),
            "earfcn": _fmt_band(_first_non_empty(group.get("topology_earfcn", pd.Series(dtype=object)))),
            "topology_original_cell_id": _first_non_empty(group.get("topology_original_cell_id", pd.Series(dtype=object))),
            "topology_original_node_cell_id": _first_non_empty(group.get("topology_original_node_cell_id", pd.Series(dtype=object))),
            "topology_frontend_site_sector_key": _first_non_empty(group.get("topology_frontend_site_sector_key", pd.Series(dtype=object))),
            "grid_count": int(group["grid_id"].nunique(dropna=True)) if "grid_id" in group.columns else int(len(group)),
            "congested_grid_count": int(((prb > config.congestion_threshold) | (cell_rrc > config.congestion_threshold)).sum()),
            "prb_before_pct": float(prb.max()) if prb.notna().any() else np.nan,
            "prb_p90_pct": float(prb.quantile(0.90)) if prb.notna().any() else np.nan,
            "rrc_before_pct": float(cell_rrc.max()) if cell_rrc.notna().any() else np.nan,
            "rrc_users_before": float(rrc_users.max()) if rrc_users.notna().any() else np.nan,
            "estimated_offered_traffic_mbps": float(_to_num(group.get("estimated_offered_traffic_mbps", pd.Series(dtype=float))).max())
            if "estimated_offered_traffic_mbps" in group.columns and _to_num(group["estimated_offered_traffic_mbps"]).notna().any()
            else np.nan,
            "existing_carriers": _first_non_empty(group.get("existing_carriers", pd.Series(dtype=object))),
            "existing_carrier_count": _scalar_to_int(group.get("existing_carrier_count", pd.Series(dtype=float)).iloc[0] if "existing_carrier_count" in group.columns and not group["existing_carrier_count"].empty else 0),
            "available_bands_to_add": _first_non_empty(group.get("available_bands_to_add", pd.Series(dtype=object))),
            "carrier_addition_options": _first_non_empty(group.get("carrier_addition_options", pd.Series(dtype=object))),
            "available_earfcns_to_add": _first_non_empty(group.get("available_earfcns_to_add", pd.Series(dtype=object))),
            "available_earfcn_options": _first_non_empty(group.get("available_earfcn_options", pd.Series(dtype=object))),
            "recommended_band_to_add": _first_non_empty(group.get("recommended_band_to_add", pd.Series(dtype=object))),
            "available_band_options_count": _scalar_to_int(group.get("available_band_options_count", pd.Series(dtype=float)).iloc[0] if "available_band_options_count" in group.columns and not group["available_band_options_count"].empty else 0),
            "max_supported_carriers": _scalar_to_int(group.get("max_supported_carriers", pd.Series(dtype=float)).iloc[0] if "max_supported_carriers" in group.columns and not group["max_supported_carriers"].empty else 0),
            "carrier_addition_possible": _scalar_to_bool(group.get("carrier_addition_possible", pd.Series(dtype=bool)).iloc[0] if "carrier_addition_possible" in group.columns and not group["carrier_addition_possible"].empty else False),
            "carrier_addition_blocked": _scalar_to_bool(group.get("carrier_addition_blocked", pd.Series(dtype=bool)).iloc[0] if "carrier_addition_blocked" in group.columns and not group["carrier_addition_blocked"].empty else False),
            "carrier_addition_reason": _first_non_empty(group.get("carrier_addition_reason", pd.Series(dtype=object))),
            "sector_has_alternate_carrier": _scalar_to_bool(group.get("sector_has_alternate_carrier", pd.Series(dtype=bool)).iloc[0] if "sector_has_alternate_carrier" in group.columns and not group["sector_has_alternate_carrier"].empty else False),
            "sector_capacity_limit": _scalar_to_int(group.get("sector_capacity_limit", pd.Series(dtype=float)).iloc[0] if "sector_capacity_limit" in group.columns and not group["sector_capacity_limit"].empty else 0),
            "hotspot_score": float(_to_num(group["model3_hotspot_score"]).max()) if "model3_hotspot_score" in group.columns and _to_num(group["model3_hotspot_score"]).notna().any() else np.nan,
            "hotspot_rank": float(_to_num(group["model3_hotspot_rank"]).max()) if "model3_hotspot_rank" in group.columns and _to_num(group["model3_hotspot_rank"]).notna().any() else np.nan,
            "sector_key": sector_key,
        }
        rows.append(row)

    cell_df = pd.DataFrame(rows)
    if cell_df.empty:
        return cell_df, {"sector_key": sector_key, "site_key": site_key, "cell_count": 0, "sector_count": 0}

    cell_df["prb_before_pct"] = _to_num(cell_df["prb_before_pct"])
    cell_df["prb_p90_pct"] = _to_num(cell_df["prb_p90_pct"])
    cell_df["rrc_before_pct"] = _to_num(cell_df["rrc_before_pct"])
    cell_df["rrc_users_before"] = _to_num(cell_df["rrc_users_before"])
    cell_df["existing_carrier_count"] = pd.to_numeric(cell_df["existing_carrier_count"], errors="coerce").fillna(0).astype(int)
    cell_df["available_band_options_count"] = pd.to_numeric(cell_df["available_band_options_count"], errors="coerce").fillna(0).astype(int)
    cell_df["max_supported_carriers"] = pd.to_numeric(cell_df["max_supported_carriers"], errors="coerce").fillna(0).astype(int)
    cell_df["sector_capacity_limit"] = pd.to_numeric(cell_df["sector_capacity_limit"], errors="coerce").fillna(0).astype(int)
    cell_df["carrier_addition_possible"] = cell_df["carrier_addition_possible"].astype(bool)
    cell_df["carrier_addition_blocked"] = cell_df["carrier_addition_blocked"].astype(bool)
    cell_df["sector_has_alternate_carrier"] = cell_df["sector_has_alternate_carrier"].astype(bool)
    cell_df["prb_rrc_pressure"] = cell_df[["prb_before_pct", "rrc_before_pct"]].max(axis=1)
    cell_df["congested"] = (cell_df["prb_before_pct"] > config.congestion_threshold) | (
        cell_df["rrc_before_pct"] > config.congestion_threshold
    )
    cell_df["sector_cell_count"] = cell_df.groupby("sector_id")["Node_Cell_ID"].transform("count")
    cell_df["sector_congested_count"] = cell_df.groupby("sector_id")["congested"].transform("sum")

    sector_count = int(cell_df["sector_id"].nunique(dropna=True)) if "sector_id" in cell_df.columns else 0
    summary = {
        "sector_key": sector_key,
        "site_key": site_key,
        "cell_count": int(len(cell_df)),
        "sector_count": sector_count,
        "congested_cell_count": int(cell_df["congested"].sum()),
        "load_balance_candidate_count": 0,
        "carrier_addition_candidate_count": int(cell_df["carrier_addition_possible"].sum()),
        "sector_split_candidate_count": 0,
        "new_site_candidate_count": 0,
    }
    return cell_df, summary


def _band_list_from_text(value: str) -> list[float]:
    if not value:
        return []
    parts = [p.strip() for p in str(value).split(",") if p.strip()]
    values: list[float] = []
    for part in parts:
        try:
            values.append(float(part))
        except Exception:
            continue
    return values


def _choose_load_balance_candidate(sector_cells: pd.DataFrame) -> dict[str, Any] | None:
    congested_sources = sector_cells.loc[sector_cells["congested"]].copy()
    if congested_sources.empty:
        return None

    candidates: list[dict[str, Any]] = []
    for _, source_row in congested_sources.iterrows():
        others = sector_cells.loc[sector_cells["Node_Cell_ID"] != source_row["Node_Cell_ID"]].copy()
        if others.empty:
            continue
        others = others.loc[
            (others["prb_before_pct"] < DEFAULT_CONGESTION_THRESHOLD)
            & (others["rrc_before_pct"] < DEFAULT_CONGESTION_THRESHOLD)
        ].copy()
        if others.empty:
            continue
        others["target_priority"] = others["band"].map(_band_priority)
        others["util_score"] = others[["prb_before_pct", "rrc_before_pct"]].max(axis=1)
        others = others.sort_values(
            ["util_score", "target_priority", "grid_count"],
            ascending=[True, False, False],
            na_position="last",
        )
        target_row = others.iloc[0]
        source_prb = float(source_row["prb_before_pct"]) if pd.notna(source_row["prb_before_pct"]) else np.nan
        source_rrc = float(source_row["rrc_before_pct"]) if pd.notna(source_row["rrc_before_pct"]) else np.nan
        pressure = max(source_prb if pd.notna(source_prb) else 0.0, source_rrc if pd.notna(source_rrc) else 0.0)
        severity = max(0.0, pressure - DEFAULT_CONGESTION_THRESHOLD)
        shift_pct = min(25.0, max(10.0, severity * 1.25))
        projected_prb = max(0.0, source_prb - shift_pct)
        projected_rrc = max(0.0, source_rrc - (shift_pct * 0.90))
        candidates.append(
            {
                "source_node_cell_id": source_row["Node_Cell_ID"],
                "source_band": source_row["band"],
                "target_node_cell_id": target_row["Node_Cell_ID"],
                "target_band": target_row["band"],
                "projected_prb_after_pct": projected_prb,
                "projected_rrc_after_pct": projected_rrc,
                "projected_rrc_users_after": max(0.0, float(source_row["rrc_users_before"]) - (shift_pct * 3.0))
                if pd.notna(source_row["rrc_users_before"])
                else np.nan,
                "resolved": projected_prb <= DEFAULT_CONGESTION_THRESHOLD and projected_rrc <= DEFAULT_CONGESTION_THRESHOLD,
                "source_pressure": pressure,
            }
        )

    if not candidates:
        return None

    resolved = [c for c in candidates if c["resolved"]]
    ranked_pool = resolved if resolved else candidates
    ranked_pool = sorted(
        ranked_pool,
        key=lambda c: (
            0 if c["resolved"] else 1,
            c["projected_prb_after_pct"],
            c["projected_rrc_after_pct"],
            -_band_priority(c["target_band"]),
            -c["source_pressure"],
        ),
    )
    return ranked_pool[0]


def _build_sector_split_topology(
    sector_cells: pd.DataFrame,
    part3_site_df: pd.DataFrame,
    logger: logging.Logger,
) -> tuple[pd.DataFrame, pd.DataFrame]:
    work = part3_site_df.copy()
    original_ids = set()
    for col in ["topology_original_cell_id", "topology_original_node_cell_id"]:
        if col in sector_cells.columns:
            original_ids.update(str(v).strip() for v in sector_cells[col].dropna().astype(str).tolist() if str(v).strip())
    if not original_ids:
        original_ids.update(str(v).strip() for v in sector_cells["Node_Cell_ID"].dropna().astype(str).tolist() if str(v).strip())

    mask = work["Node_Cell_ID"].astype(str).isin(original_ids) | work["cell_id"].astype(str).isin(original_ids)
    source_rows = work.loc[mask].copy()
    if source_rows.empty:
        return work, source_rows

    keep_rows = work.loc[~mask].copy()
    split_rows: list[pd.DataFrame] = []
    for _, row in source_rows.iterrows():
        base_azimuth = _scalar_to_float(row.get("azimuth"), 0.0)
        base_pci = _scalar_to_int(row.get("PCI"), 0)
        base_cell = str(row.get("cell_id") or row.get("Node_Cell_ID") or "").strip()
        for suffix, delta_deg, pci_offset in [("A", -20.0, 17), ("B", 20.0, 34)]:
            child = row.copy()
            child["cell_id"] = f"{base_cell}__SS{suffix}"
            child["Node_Cell_ID"] = str(child["cell_id"])
            child["PCI"] = int((base_pci + pci_offset) % 504)
            child["azimuth"] = float((base_azimuth + delta_deg) % 360.0)
            child["carrier_variant"] = f"sector_split_{suffix.lower()}"
            split_rows.append(pd.DataFrame([child]))
    split_df = pd.concat(split_rows, ignore_index=True) if split_rows else pd.DataFrame(columns=work.columns)
    combined = pd.concat([keep_rows, split_df], ignore_index=True)
    logger.info(
        "sector_split_topology sector=%s source_rows=%d split_rows=%d",
        _first_non_empty(sector_cells["sector_id"]),
        len(source_rows),
        len(split_df),
    )
    return combined, source_rows


def _extract_source_site_rows(sector_cells: pd.DataFrame, part3_site_df: pd.DataFrame) -> pd.DataFrame:
    work = part3_site_df.copy()
    original_ids = set()
    for col in ["topology_original_cell_id", "topology_original_node_cell_id"]:
        if col in sector_cells.columns:
            original_ids.update(str(v).strip() for v in sector_cells[col].dropna().astype(str).tolist() if str(v).strip())
    if not original_ids:
        original_ids.update(str(v).strip() for v in sector_cells["Node_Cell_ID"].dropna().astype(str).tolist() if str(v).strip())
    matched = work.loc[
        work["Node_Cell_ID"].astype(str).isin(original_ids) | work["cell_id"].astype(str).isin(original_ids)
    ].copy()
    dedupe_cols = [col for col in ["Node_Cell_ID", "cell_id", "Site ID", "band", "earfcn", "azimuth", "PCI"] if col in matched.columns]
    if dedupe_cols:
        matched = matched.drop_duplicates(subset=dedupe_cols, keep="first").copy()
    return matched


def _prepare_local_action_context(
    *,
    sector_cells: pd.DataFrame,
    site_df: pd.DataFrame,
    context: dict[str, Any],
    config: Model3RecommendationConfig,
    logger: logging.Logger,
    action_label: str,
    source_rows: pd.DataFrame | None = None,
) -> dict[str, Any] | None:
    baseline_all = context["baseline_pred_df"]
    corrected_all = context["corrected_pred_df"]
    detail_all = context["coverage_rows_df"]
    kpi_all = context["kpi_df"]
    geo_all = context["geo_df"]

    part3_baseline = baseline_all.loc[baseline_all["time_bucket"].astype(str) == "PART_3"].copy()
    part3_corrected = corrected_all.loc[corrected_all["time_bucket"].astype(str) == "PART_3"].copy()
    part3_detail = detail_all.loc[detail_all["time_bucket"].astype(str) == "PART_3"].copy()
    part3_kpi = kpi_all.loc[kpi_all["time_bucket"].astype(str) == "PART_3"].copy()
    part3_geo = geo_all.loc[geo_all["time_bucket"].astype(str) == "PART_3"].copy()

    if source_rows is None:
        source_rows = _extract_source_site_rows(sector_cells, site_df)
    if source_rows.empty:
        logger.info("%s_prepare_failed reason=no_source_rows", action_label)
        return None

    site_lat = pd.to_numeric(source_rows["lat"], errors="coerce").mean()
    site_lon = pd.to_numeric(source_rows["lon"], errors="coerce").mean()
    site_work = site_df.copy()
    site_work["_site_distance_m"] = _haversine_distance_m_series(site_work["lat"], site_work["lon"], site_lat, site_lon)
    site_distance_df = (
        site_work.groupby("Site ID", as_index=False)
        .agg(site_distance_m=("_site_distance_m", "min"))
        .sort_values("site_distance_m", ascending=True)
    )
    nearest_site_ids = site_distance_df["Site ID"].head(25).astype(str).tolist()
    local_site_df = site_work.loc[site_work["Site ID"].astype(str).isin(nearest_site_ids)].copy()
    if local_site_df.empty:
        local_site_df = site_work.copy()

    local_point_map = _make_local_point_map(part3_baseline)
    if local_point_map.empty:
        logger.info("%s_prepare_failed reason=no_point_map", action_label)
        return None
    local_point_map["_site_distance_m"] = _haversine_distance_m_series(local_point_map["lat"], local_point_map["lon"], site_lat, site_lon)
    local_point_map = local_point_map.loc[local_point_map["_site_distance_m"] <= float(config.sector_split_local_radius_m)].copy()
    if local_point_map.empty:
        local_point_map = _make_local_point_map(
            part3_baseline.loc[part3_baseline["Node_Cell_ID"].astype(str).isin(sector_cells["Node_Cell_ID"].astype(str))]
        )
    if local_point_map.empty:
        logger.info("%s_prepare_failed reason=no_local_points", action_label)
        return None

    affected_grid_ids = pd.to_numeric(local_point_map["grid_id"], errors="coerce").dropna().astype("Int64").unique().tolist()
    local_detail = pd.DataFrame()
    if affected_grid_ids and "grid_id" in part3_detail.columns:
        detail_grid_ids = pd.to_numeric(part3_detail["grid_id"], errors="coerce").astype("Int64")
        local_detail = part3_detail.loc[detail_grid_ids.isin(affected_grid_ids)].copy()
    if local_detail.empty:
        local_detail = part3_detail.merge(local_point_map[["lat", "lon"]].drop_duplicates(), on=["lat", "lon"], how="inner")
    if local_detail.empty:
        local_detail = local_point_map[["lat", "lon", "grid_id", "grid_row", "grid_col", "grid_centroid_lat", "grid_centroid_lon"]].drop_duplicates().copy()
        local_detail["time_bucket"] = "PART_3"

    local_kpi = part3_kpi.loc[pd.to_numeric(part3_kpi["grid_id"], errors="coerce").astype("Int64").isin(affected_grid_ids)].copy()
    local_geo = part3_geo.loc[pd.to_numeric(part3_geo["grid_id"], errors="coerce").astype("Int64").isin(affected_grid_ids)].copy()
    logger.info(
        "%s_prepare local_points=%d local_detail_rows=%d local_sites=%d",
        action_label,
        len(local_point_map),
        len(local_detail),
        len(local_site_df),
    )
    return {
        "part3_baseline": part3_baseline,
        "part3_corrected": part3_corrected,
        "local_site_df": local_site_df,
        "local_point_map": local_point_map,
        "local_detail": local_detail,
        "local_kpi": local_kpi,
        "local_geo": local_geo,
    }


def _build_modeled_inventory_from_surface(
    *,
    baseline_local: pd.DataFrame,
    corrected_local: pd.DataFrame,
    local_site_df: pd.DataFrame,
    local_kpi: pd.DataFrame,
    local_geo: pd.DataFrame,
    context: dict[str, Any],
    config: Model3RecommendationConfig,
    forced_assignments: pd.DataFrame | None = None,
) -> tuple[pd.DataFrame, pd.DataFrame]:
    hybrid_history = context["hybrid_history"]
    model2_base_df = context["model2_base_df"]

    baseline_local = coverage_test._attach_carrier_load_share(baseline_local.copy())
    corrected_local = coverage_test._attach_carrier_load_share(corrected_local.copy())
    baseline_local = _ensure_grid_group_columns(baseline_local)
    corrected_local = _ensure_grid_group_columns(corrected_local)

    dense_part3 = _build_dense_part3_features(
        baseline_pred_df=baseline_local,
        corrected_pred_df=corrected_local,
        kpi_df=local_kpi,
        geo_df=local_geo,
        site_df=local_site_df.drop(columns=["_site_distance_m"], errors="ignore"),
    )
    dense_part3["time_bucket"] = "PART_3"
    dense_part3["grid_id"] = pd.to_numeric(dense_part3["grid_id"], errors="coerce").astype("Int64")

    history = hybrid_history.loc[hybrid_history["time_bucket"].astype(str).isin(["PART_1", "PART_2"])].copy()
    history["grid_id"] = pd.to_numeric(history["grid_id"], errors="coerce").astype("Int64")
    infer_df = pd.concat([history, dense_part3], ignore_index=True, sort=False)
    infer_df = model1_base._add_temporal_history_features(infer_df)
    infer_df = model1_base._add_clutter_evolution_features(infer_df)
    infer_part3 = infer_df.loc[infer_df["time_bucket"].astype(str) == "PART_3"].copy()

    for target, spec in context["model1_models"].items():
        infer_part3[target] = spec["model"].predict(infer_part3[spec["features"]])

    base_keep = [
        col
        for col in model2_base_df.columns
        if col
        in [
            "grid_id",
            "time_bucket",
            "sample_count",
            "dl_tpt_mean",
            "ul_tpt_mean",
            "estimated_prb_mean",
            "cqi_mean",
            "dominant_pci",
            "green_ratio",
            "water_ratio",
            "grid_size_m",
            "grid_area_m2",
            "cell_area_m2",
            "road_length_m",
            "building_count",
            "building_area_ratio",
            "park_open_area",
            "open_area_ratio",
            "mall_presence",
            "metro_presence",
            "road_density",
            "geo_snapshot_mode",
            "geo_snapshot_source_ts",
        ]
    ]
    base_context = model2_base_df[base_keep].drop_duplicates(subset=["grid_id", "time_bucket"]).copy()
    work = infer_part3.copy()
    for src, dst in [("pred_rsrp", "rsrp_mean"), ("pred_rsrq", "rsrq_mean"), ("pred_sinr", "sinr_mean")]:
        work[dst] = pd.to_numeric(work[src], errors="coerce")
    merged = work.merge(base_context, on=["grid_id", "time_bucket"], how="left", suffixes=("", "_base"))
    for col in list(merged.columns):
        if col.endswith("_base"):
            original = col[:-5]
            if original in merged.columns:
                merged[original] = merged[original].combine_first(merged[col])
                merged = merged.drop(columns=[col])
            else:
                merged = merged.rename(columns={col: original})
    merged = merged.sort_values(["grid_id", "time_bucket"]).drop_duplicates(subset=["grid_id", "time_bucket"], keep="first").copy()

    join_keys = ["grid_id", "time_bucket"]
    topo_work = corrected_local.copy()
    topo_work["grid_id"] = pd.to_numeric(topo_work["grid_id"], errors="coerce").astype("Int64")
    topo_work["_rank_rsrp"] = pd.to_numeric(topo_work.get("pred_rsrp"), errors="coerce")
    topo_best = topo_work.sort_values(join_keys + ["_rank_rsrp"], ascending=[True, True, False]).drop_duplicates(subset=join_keys, keep="first")

    if forced_assignments is not None and not forced_assignments.empty:
        forced = forced_assignments.copy()
        forced["grid_id"] = pd.to_numeric(forced["grid_id"], errors="coerce").astype("Int64")
        forced["time_bucket"] = forced.get("time_bucket", "PART_3").astype(str)
        topo_full = topo_work.merge(
            forced[["grid_id", "time_bucket", "Node_Cell_ID"]].rename(columns={"Node_Cell_ID": "_forced_node_cell_id"}),
            on=["grid_id", "time_bucket"],
            how="left",
        )
        forced_rows = topo_full.loc[
            topo_full["_forced_node_cell_id"].notna()
            & topo_full["Node_Cell_ID"].astype(str).eq(topo_full["_forced_node_cell_id"].astype(str))
        ].copy()
        forced_rows = forced_rows.sort_values(join_keys + ["_rank_rsrp"], ascending=[True, True, False]).drop_duplicates(subset=join_keys, keep="first")
        forced_keys = forced_rows[join_keys].drop_duplicates()
        topo_best = topo_best.merge(forced_keys.assign(_drop=True), on=join_keys, how="left")
        topo_best = topo_best.loc[topo_best["_drop"] != True].drop(columns=["_drop"], errors="ignore")
        topo_best = pd.concat([topo_best, forced_rows[topo_best.columns.intersection(forced_rows.columns)]], ignore_index=True, sort=False)
        topo_best = topo_best.sort_values(join_keys).drop_duplicates(subset=join_keys, keep="first")

    topo_keep = [
        col
        for col in topo_best.columns
        if col
        in [
            "grid_id",
            "time_bucket",
            "Node_Cell_ID",
            "Site ID",
            "site_identity_key",
            "sector_identity",
            "sector_identity_key",
            "frontend_site_sector_key",
            "node_cell_sector_key",
            "sector",
            "band",
            "earfcn",
            "nodeb_id",
            "PCI",
            "azimuth",
            "canonical_sector_id",
            "site_sector_band_key",
            "rf_identity_key",
            "original_node_cell_id",
            "original_cell_id",
            "carrier_load_share",
        ]
    ]
    topo_best = topo_best[topo_keep].copy()
    topo_best["site_id"] = _extract_site_id(topo_best["Node_Cell_ID"])
    rename_map = {
        "Site ID": "topology_site_id",
        "site_identity_key": "topology_site_identity_key",
        "sector_identity": "topology_sector_identity",
        "sector_identity_key": "topology_sector_identity_key",
        "frontend_site_sector_key": "topology_frontend_site_sector_key",
        "node_cell_sector_key": "topology_node_cell_sector_key",
        "sector": "topology_sector",
        "band": "topology_band",
        "earfcn": "topology_earfcn",
        "nodeb_id": "topology_nodeb_id",
        "PCI": "topology_pci",
        "azimuth": "topology_azimuth",
        "canonical_sector_id": "topology_canonical_sector_id",
        "site_sector_band_key": "topology_site_sector_band_key",
        "rf_identity_key": "topology_rf_identity_key",
        "original_node_cell_id": "topology_original_node_cell_id",
        "original_cell_id": "topology_original_cell_id",
        "carrier_load_share": "topology_carrier_load_share",
    }
    topo_best = topo_best.rename(columns=rename_map)
    merged = merged.merge(topo_best, on=["grid_id", "time_bucket"], how="left", validate="many_to_one")

    enriched = model3_builder.model2_builder._add_model2_features(merged)
    feature_fallbacks = {
        "corrected_rsrp_mean": ["baseline_rsrp", "rsrp_mean"],
        "corrected_rsrq_mean": ["baseline_rsrq", "rsrq_mean"],
        "corrected_sinr_mean": ["baseline_sinr", "sinr_mean"],
    }
    for feature_name in model2_hybrid.ALL_FEATURES:
        if feature_name in enriched.columns:
            continue
        filled = None
        for source_name in feature_fallbacks.get(feature_name, []):
            if source_name in enriched.columns:
                source_series = pd.to_numeric(enriched[source_name], errors="coerce")
                filled = source_series if filled is None else filled.combine_first(source_series)
        enriched[feature_name] = filled if filled is not None else np.nan
    for target, model in context["model2_models"].items():
        enriched[f"{target}_pred"] = model.predict(enriched[model2_hybrid.ALL_FEATURES])

    modeled = enriched.copy()
    modeled, _ = model3_builder._add_sector_carrier_capability_fields(modeled)
    bandwidth = pd.to_numeric(modeled.get("bandwidth_mhz_est"), errors="coerce").replace(0, np.nan).fillna(10.0)
    spectral = model3_builder._estimate_spectral_efficiency_bpshz(
        modeled,
        bandwidth_mhz=bandwidth,
        mimo_layers=model3_builder.DEFAULT_MIMO_LAYERS,
        control_overhead=model3_builder.DEFAULT_CONTROL_OVERHEAD,
    )
    modeled["estimated_dl_capacity_mbps"] = (
        bandwidth * max(1.0, float(model3_builder.DEFAULT_MIMO_LAYERS)) * max(0.1, 1.0 - float(model3_builder.DEFAULT_CONTROL_OVERHEAD)) * spectral
    ).round(3)
    modeled, _ = model3_builder._apply_model3_load_profile(modeled)
    local_inventory, _ = _build_cell_inventory(modeled, config)
    return modeled, local_inventory


def _evaluate_action_inventory(
    *,
    sector_cells: pd.DataFrame,
    after_inventory: pd.DataFrame,
    config: Model3RecommendationConfig,
    logger: logging.Logger,
    action_label: str,
) -> dict[str, Any]:
    before_cells = sector_cells["Node_Cell_ID"].astype(str).tolist()
    source_sector_id = str(_first_non_empty(sector_cells["sector_id"])).strip()
    source_site_id = _normalize_identity_text(_first_non_empty(sector_cells["site_id"]))
    source_frontend_sector_keys = {source_sector_id} if source_sector_id else set()
    source_topology_sector_keys: set[str] = set()
    for col in [
        "topology_frontend_site_sector_key",
        "topology_node_cell_sector_key",
        "topology_sector_identity_key",
        "topology_canonical_sector_id",
        "topology_site_sector_band_key",
        "sector_id",
    ]:
        if col in sector_cells.columns:
            source_topology_sector_keys.update(
                str(value).strip()
                for value in sector_cells[col].dropna().astype(str).tolist()
                if str(value).strip()
            )
    source_frontend_sector_keys.update(source_topology_sector_keys)
    original_lineage_ids = set(value for value in before_cells if str(value).strip())
    for col in ["topology_original_cell_id", "topology_original_node_cell_id"]:
        if col in sector_cells.columns:
            original_lineage_ids.update(
                str(value).strip()
                for value in sector_cells[col].dropna().astype(str).tolist()
                if str(value).strip()
            )

    def _scope_mask(
        df: pd.DataFrame,
        *,
        sector_keys: set[str],
        lineage_ids: set[str],
        site_id: str,
    ) -> pd.Series:
        mask = pd.Series(False, index=df.index)
        if source_sector_id and "sector_id" in df.columns:
            mask = mask | df["sector_id"].astype(str).eq(source_sector_id)
        for col in [
            "sector_id",
            "topology_frontend_site_sector_key",
            "topology_node_cell_sector_key",
            "topology_sector_identity_key",
            "topology_canonical_sector_id",
            "topology_site_sector_band_key",
        ]:
            if col in df.columns and sector_keys:
                mask = mask | df[col].astype(str).isin(sector_keys)
        for col in ["Node_Cell_ID", "topology_original_cell_id", "topology_original_node_cell_id"]:
            if col in df.columns and lineage_ids:
                mask = mask | df[col].astype(str).isin(lineage_ids)
        if site_id and "site_id" in df.columns:
            mask = mask | df["site_id"].map(_normalize_identity_text).eq(site_id)
        return mask

    primary_mask = _scope_mask(
        after_inventory,
        sector_keys=source_frontend_sector_keys,
        lineage_ids=original_lineage_ids,
        site_id="",
    )
    primary_matches = after_inventory.loc[primary_mask].copy()

    affected_sector_keys = set(source_frontend_sector_keys)
    if not primary_matches.empty:
        for col in [
            "sector_id",
            "topology_frontend_site_sector_key",
            "topology_node_cell_sector_key",
            "topology_sector_identity_key",
            "topology_canonical_sector_id",
            "topology_site_sector_band_key",
        ]:
            if col in primary_matches.columns:
                affected_sector_keys.update(
                    str(value).strip()
                    for value in primary_matches[col].dropna().astype(str).tolist()
                    if str(value).strip()
                )
        for col in ["Node_Cell_ID", "topology_original_cell_id", "topology_original_node_cell_id"]:
            if col in primary_matches.columns:
                original_lineage_ids.update(
                    str(value).strip()
                    for value in primary_matches[col].dropna().astype(str).tolist()
                    if str(value).strip()
                )

    final_mask = _scope_mask(
        after_inventory,
        sector_keys=affected_sector_keys,
        lineage_ids=original_lineage_ids,
        site_id="",
    )
    after_candidates = after_inventory.loc[final_mask].copy()
    if after_candidates.empty and source_site_id:
        after_site_ids = after_inventory["site_id"].map(_normalize_identity_text) if "site_id" in after_inventory.columns else pd.Series("", index=after_inventory.index)
        after_candidates = after_inventory.loc[after_site_ids == source_site_id].copy()

    scope_mode = "matched_scope"

    def _summarize_scope(df: pd.DataFrame) -> tuple[float, float, float]:
        prb = _to_num(df["prb_before_pct"])
        rrc = _to_num(df["rrc_before_pct"])
        users = _to_num(df["rrc_users_before"]) if "rrc_users_before" in df.columns else pd.Series(dtype=float)
        return (
            float(prb.max()) if prb.notna().any() else np.nan,
            float(rrc.max()) if rrc.notna().any() else np.nan,
            float(users.max()) if users.notna().any() else np.nan,
        )

    projected_prb, projected_rrc, projected_users = _summarize_scope(after_candidates)
    before_prb = float(_to_num(sector_cells["prb_before_pct"]).max()) if _to_num(sector_cells["prb_before_pct"]).notna().any() else np.nan
    before_rrc = float(_to_num(sector_cells["rrc_before_pct"]).max()) if _to_num(sector_cells["rrc_before_pct"]).notna().any() else np.nan
    before_pressure = max(before_prb if pd.notna(before_prb) else 0.0, before_rrc if pd.notna(before_rrc) else 0.0)

    # If the matched scope is emptied out after topology changes, widen to the same-site
    # post-action scope so we do not falsely mark zero-load source lineages as resolved.
    if (
        source_site_id
        and not after_candidates.empty
        and pd.notna(before_prb)
        and before_prb > 0
        and (not pd.notna(projected_prb) or projected_prb <= 0.0)
        and (not pd.notna(projected_rrc) or projected_rrc <= 0.0)
        and "site_id" in after_inventory.columns
    ):
        site_scope = after_inventory.loc[after_inventory["site_id"].map(_normalize_identity_text).eq(source_site_id)].copy()
        site_scope = site_scope.loc[
            (_to_num(site_scope.get("prb_before_pct", pd.Series(dtype=float))) > 0.0)
            | (_to_num(site_scope.get("rrc_before_pct", pd.Series(dtype=float))) > 0.0)
            | (_to_num(site_scope.get("rrc_users_before", pd.Series(dtype=float))) > 0.0)
        ].copy()
        if not site_scope.empty:
            after_candidates = site_scope
            projected_prb, projected_rrc, projected_users = _summarize_scope(after_candidates)
            scope_mode = "same_site_nonzero_fallback"

    after_pressure = max(projected_prb if pd.notna(projected_prb) else 0.0, projected_rrc if pd.notna(projected_rrc) else 0.0)
    resolved = pd.notna(projected_prb) and pd.notna(projected_rrc) and projected_prb <= config.congestion_threshold and projected_rrc <= config.congestion_threshold
    improved = after_pressure < before_pressure - 0.5
    worsened = after_pressure > before_pressure + 0.5
    if resolved:
        status = "Resolved"
    elif worsened:
        status = "Rejected"
    elif improved:
        status = "Partially Resolved"
    else:
        status = "No Material Change"
    after_cells = after_candidates["Node_Cell_ID"].astype(str).dropna().tolist() if "Node_Cell_ID" in after_candidates.columns else []
    logger.info(
        "%s_done sector=%s before_cells=%s after_cells=%s local_after_rows=%d scope_mode=%s scope_keys=%s lineage_ids=%s before_prb=%.3f before_rrc=%.3f actual_prb=%.3f actual_rrc=%.3f status=%s",
        action_label,
        source_sector_id,
        before_cells,
        after_cells,
        len(after_candidates),
        scope_mode,
        sorted(affected_sector_keys)[:8],
        sorted(original_lineage_ids)[:8],
        before_prb if pd.notna(before_prb) else -1.0,
        before_rrc if pd.notna(before_rrc) else -1.0,
        projected_prb if pd.notna(projected_prb) else -1.0,
        projected_rrc if pd.notna(projected_rrc) else -1.0,
        status,
    )
    return {
        "status": status,
        "projected_prb_after_pct": round(projected_prb, 3) if pd.notna(projected_prb) else np.nan,
        "projected_rrc_after_pct": round(projected_rrc, 3) if pd.notna(projected_rrc) else np.nan,
        "projected_rrc_users_after": round(projected_users, 3) if pd.notna(projected_users) else np.nan,
        "next_step": "" if resolved else "New Site",
        "after_cells": after_cells,
    }


def _run_sector_split_resimulation(
    sector_cells: pd.DataFrame,
    config: Model3RecommendationConfig,
    context: dict[str, Any],
    logger: logging.Logger,
) -> dict[str, Any]:
    part3_site_df = context["part3_site_df"]
    building_df = context["building_df"]

    split_site_df, source_rows = _build_sector_split_topology(sector_cells, part3_site_df, logger)
    if source_rows.empty:
        return {
            "status": "Recommended",
            "action_reason": "Sector split topology could not be mapped back to PART_3 site rows.",
            "projected_prb_after_pct": np.nan,
            "projected_rrc_after_pct": np.nan,
            "projected_rrc_users_after": np.nan,
            "next_step": "New Site",
            "resimulation_required": True,
            "resimulation_flow": "PART_3 local rerun failed at topology mapping",
        }
    local_ctx = _prepare_local_action_context(
        sector_cells=sector_cells,
        site_df=split_site_df,
        context=context,
        config=config,
        logger=logger,
        action_label="sector_split_resim",
        source_rows=source_rows,
    )
    if local_ctx is None:
        return {
            "status": "Recommended",
            "action_reason": "Sector split local context could not be prepared from PART_3 artifacts.",
            "projected_prb_after_pct": np.nan,
            "projected_rrc_after_pct": np.nan,
            "projected_rrc_users_after": np.nan,
            "next_step": "New Site",
            "resimulation_required": True,
            "resimulation_flow": "PART_3 local rerun failed during local context preparation",
        }

    baseline_local = coverage_test._run_project_baseline_prediction(
        project_id=int(context["summary"].get("project_id", 196)),
        region=str(context["summary"].get("region", "india")).lower(),
        site_df=local_ctx["local_site_df"].drop(columns=["_site_distance_m"], errors="ignore"),
        drive_df=local_ctx["local_detail"],
        building_df=building_df,
        baseline_radius_m=float(context["summary"].get("baseline_radius_m", 500.0)),
        grid_size_m=float(context["summary"].get("grid_size_m", 50.0)),
        workers=1,
        max_interference_sites=int(context["summary"].get("max_interference_sites", 50)),
        polygon_wkt=str(context["summary"].get("polygon_wkt", "")).strip() or None,
        use_frontend_grid_sampling=False,
        grid_analytics_scenario_id=context["summary"].get("grid_analytics_scenario_id"),
    )
    if baseline_local.empty:
        return {
            "status": "Recommended",
            "action_reason": "Sector split rerun produced no PART_3 baseline rows.",
            "projected_prb_after_pct": np.nan,
            "projected_rrc_after_pct": np.nan,
            "projected_rrc_users_after": np.nan,
            "next_step": "New Site",
            "resimulation_required": True,
            "resimulation_flow": "PART_3 local rerun produced no baseline rows",
        }
    baseline_local["time_bucket"] = "PART_3"
    baseline_local = _merge_point_map_by_coordinates(
        baseline_local,
        local_ctx["local_point_map"].drop(columns=["_site_distance_m"], errors="ignore"),
    )
    baseline_local = _ensure_grid_group_columns(baseline_local)

    if not local_ctx["local_detail"].empty and {"rsrp", "rsrq", "sinr"}.issubset(local_ctx["local_detail"].columns):
        corrected_local = coverage_test._run_bucket_corrected_predictions(
            baseline_pred_df=baseline_local,
            detail_df=local_ctx["local_detail"].assign(time_bucket="PART_3"),
            site_df_by_bucket={"PART_3": local_ctx["local_site_df"].drop(columns=["_site_distance_m"], errors="ignore")},
            building_df=building_df,
            project_id=int(context["summary"].get("project_id", 196)),
            region=str(context["summary"].get("region", "india")).lower(),
            grid_size_m=float(context["summary"].get("grid_size_m", 50.0)),
            buckets=[("PART_3", "2026-02-11 00:00:00", "2026-05-16 23:59:59")],
            polygon_wkt=str(context["summary"].get("polygon_wkt", "")).strip() or None,
        )
    else:
        corrected_local = pd.DataFrame()
    if corrected_local.empty:
        corrected_local = baseline_local.copy()
    corrected_local["time_bucket"] = "PART_3"
    corrected_local = _merge_point_map_by_coordinates(
        corrected_local,
        local_ctx["local_point_map"].drop(columns=["_site_distance_m"], errors="ignore"),
    )
    corrected_local = _ensure_grid_group_columns(corrected_local)
    _, local_inventory = _build_modeled_inventory_from_surface(
        baseline_local=baseline_local,
        corrected_local=corrected_local,
        local_site_df=local_ctx["local_site_df"],
        local_kpi=local_ctx["local_kpi"],
        local_geo=local_ctx["local_geo"],
        context=context,
        config=config,
    )
    outcome = _evaluate_action_inventory(
        sector_cells=sector_cells,
        after_inventory=local_inventory,
        config=config,
        logger=logger,
        action_label="sector_split_resim",
    )
    status = outcome["status"]
    if status == "Rejected":
        action_reason = "Sector split was rerun through baseline, Model 1, and Model 2, but the affected sector lineage became more congested, so the action should be rejected."
    elif status == "No Material Change":
        action_reason = "Sector split was rerun through baseline, Model 1, and Model 2, but the affected sector lineage did not improve enough to count as resolved."
    else:
        action_reason = "Sector split was rerun on the PART_3 local planning surface and then pushed through Model 1 and Model 2 inference."
    return {
        "status": status,
        "action_reason": action_reason,
        "projected_prb_after_pct": outcome["projected_prb_after_pct"],
        "projected_rrc_after_pct": outcome["projected_rrc_after_pct"],
        "projected_rrc_users_after": outcome["projected_rrc_users_after"],
        "next_step": outcome["next_step"],
        "resimulation_required": True,
        "resimulation_flow": "PART_3 topology split -> baseline rerun -> Model 1 inference -> Model 2 inference -> Model 3 reevaluation",
    }


def _build_carrier_addition_topology(
    sector_cells: pd.DataFrame,
    part3_site_df: pd.DataFrame,
    band_to_add: str,
    logger: logging.Logger,
) -> tuple[pd.DataFrame, pd.DataFrame]:
    work = part3_site_df.copy()
    source_rows = _extract_source_site_rows(sector_cells, part3_site_df)
    if source_rows.empty:
        return work, source_rows
    try:
        band_num = int(float(band_to_add))
    except Exception:
        band_num = 1800
    add_rows = source_rows.copy()
    add_rows["band"] = float(band_num)
    add_rows["earfcn"] = float(SYNTHETIC_BAND_TO_EARFCN.get(int(band_num), band_num))
    add_rows["carrier_variant"] = f"carrier_add_{band_num}"
    add_rows["cell_id"] = add_rows["cell_id"].astype(str).map(lambda value: f"{value}__ADD{band_num}")
    add_rows["Node_Cell_ID"] = add_rows["cell_id"].astype(str)
    if "PCI" in add_rows.columns:
        add_rows["PCI"] = (pd.to_numeric(add_rows["PCI"], errors="coerce").fillna(0).astype(int) + (band_num % 97) + 11) % 504
    combined = pd.concat([work, add_rows], ignore_index=True)
    logger.info(
        "carrier_add_topology sector=%s source_rows=%d add_rows=%d band=%s",
        _first_non_empty(sector_cells["sector_id"]),
        len(source_rows),
        len(add_rows),
        band_to_add,
    )
    return combined, source_rows


def _build_new_site_topology(
    sector_cells: pd.DataFrame,
    part3_site_df: pd.DataFrame,
    logger: logging.Logger,
) -> tuple[pd.DataFrame, pd.DataFrame]:
    work = part3_site_df.copy()
    source_rows = _extract_source_site_rows(sector_cells, part3_site_df)
    if source_rows.empty:
        return work, source_rows
    template = source_rows.copy()
    site_id_seed = pd.to_numeric(work["Site ID"], errors="coerce").dropna().astype(int)
    new_site_id = int(site_id_seed.max()) + 1001 if not site_id_seed.empty else 990001
    lat = pd.to_numeric(template["lat"], errors="coerce").mean()
    lon = pd.to_numeric(template["lon"], errors="coerce").mean()
    lat_offset = 180.0 / 111320.0
    lon_offset = 180.0 / (111320.0 * max(0.2, np.cos(np.radians(lat))))
    new_rows: list[pd.DataFrame] = []
    azimuths = [30.0, 150.0, 270.0]
    bands = template["band"].astype(str).tolist()
    for idx, azimuth in enumerate(azimuths, start=1):
        base = template.iloc[[0]].copy()
        band_value = float(pd.to_numeric(template["band"], errors="coerce").iloc[(idx - 1) % max(1, len(template))])
        base["Site ID"] = new_site_id
        base["lat"] = lat + lat_offset
        base["lon"] = lon + lon_offset
        base["band"] = band_value
        base["earfcn"] = float(SYNTHETIC_BAND_TO_EARFCN.get(int(round(band_value)), int(round(band_value))))
        base["azimuth"] = azimuth
        base["cell_id"] = f"{new_site_id}_{idx}__NS"
        base["Node_Cell_ID"] = base["cell_id"].astype(str)
        if "PCI" in base.columns:
            base["PCI"] = int((100 + idx * 29 + (int(round(band_value)) % 41)) % 504)
        base["carrier_variant"] = "new_site"
        new_rows.append(base)
    new_site_df = pd.concat(new_rows, ignore_index=True) if new_rows else pd.DataFrame(columns=work.columns)
    combined = pd.concat([work, new_site_df], ignore_index=True)
    logger.info(
        "new_site_topology sector=%s source_rows=%d new_site_rows=%d new_site_id=%s",
        _first_non_empty(sector_cells["sector_id"]),
        len(source_rows),
        len(new_site_df),
        new_site_id,
    )
    return combined, source_rows


def _run_carrier_addition_resimulation(
    sector_cells: pd.DataFrame,
    config: Model3RecommendationConfig,
    context: dict[str, Any],
    logger: logging.Logger,
    band_to_add: str,
) -> dict[str, Any]:
    part3_site_df = context["part3_site_df"]
    building_df = context["building_df"]
    added_site_df, source_rows = _build_carrier_addition_topology(sector_cells, part3_site_df, band_to_add, logger)
    if source_rows.empty:
        return {
            "status": "Recommended",
            "action_reason": "Carrier addition topology could not be mapped back to PART_3 site rows.",
            "projected_prb_after_pct": np.nan,
            "projected_rrc_after_pct": np.nan,
            "projected_rrc_users_after": np.nan,
            "next_step": "Sector Split",
            "resimulation_required": True,
            "resimulation_flow": "PART_3 carrier addition rerun failed at topology mapping",
        }
    local_ctx = _prepare_local_action_context(
        sector_cells=sector_cells,
        site_df=added_site_df,
        context=context,
        config=config,
        logger=logger,
        action_label="carrier_add_resim",
        source_rows=source_rows,
    )
    if local_ctx is None:
        return {
            "status": "Recommended",
            "action_reason": "Carrier addition local context could not be prepared from PART_3 artifacts.",
            "projected_prb_after_pct": np.nan,
            "projected_rrc_after_pct": np.nan,
            "projected_rrc_users_after": np.nan,
            "next_step": "Sector Split",
            "resimulation_required": True,
            "resimulation_flow": "PART_3 carrier addition rerun failed during local context preparation",
        }
    baseline_local = coverage_test._run_project_baseline_prediction(
        project_id=int(context["summary"].get("project_id", 196)),
        region=str(context["summary"].get("region", "india")).lower(),
        site_df=local_ctx["local_site_df"].drop(columns=["_site_distance_m"], errors="ignore"),
        drive_df=local_ctx["local_detail"],
        building_df=building_df,
        baseline_radius_m=float(context["summary"].get("baseline_radius_m", 500.0)),
        grid_size_m=float(context["summary"].get("grid_size_m", 50.0)),
        workers=1,
        max_interference_sites=int(context["summary"].get("max_interference_sites", 50)),
        polygon_wkt=str(context["summary"].get("polygon_wkt", "")).strip() or None,
        use_frontend_grid_sampling=False,
        grid_analytics_scenario_id=context["summary"].get("grid_analytics_scenario_id"),
    )
    if baseline_local.empty:
        return {
            "status": "Recommended",
            "action_reason": "Carrier addition rerun produced no PART_3 baseline rows.",
            "projected_prb_after_pct": np.nan,
            "projected_rrc_after_pct": np.nan,
            "projected_rrc_users_after": np.nan,
            "next_step": "Sector Split",
            "resimulation_required": True,
            "resimulation_flow": "PART_3 carrier addition rerun produced no baseline rows",
        }
    baseline_local["time_bucket"] = "PART_3"
    baseline_local = _merge_point_map_by_coordinates(
        baseline_local,
        local_ctx["local_point_map"].drop(columns=["_site_distance_m"], errors="ignore"),
    )
    baseline_local = _ensure_grid_group_columns(baseline_local)
    corrected_local = coverage_test._run_bucket_corrected_predictions(
        baseline_pred_df=baseline_local,
        detail_df=local_ctx["local_detail"].assign(time_bucket="PART_3"),
        site_df_by_bucket={"PART_3": local_ctx["local_site_df"].drop(columns=["_site_distance_m"], errors="ignore")},
        building_df=building_df,
        project_id=int(context["summary"].get("project_id", 196)),
        region=str(context["summary"].get("region", "india")).lower(),
        grid_size_m=float(context["summary"].get("grid_size_m", 50.0)),
        buckets=[("PART_3", "2026-02-11 00:00:00", "2026-05-16 23:59:59")],
        polygon_wkt=str(context["summary"].get("polygon_wkt", "")).strip() or None,
    )
    if corrected_local.empty:
        corrected_local = baseline_local.copy()
    corrected_local["time_bucket"] = "PART_3"
    corrected_local = _merge_point_map_by_coordinates(
        corrected_local,
        local_ctx["local_point_map"].drop(columns=["_site_distance_m"], errors="ignore"),
    )
    corrected_local = _ensure_grid_group_columns(corrected_local)
    _, local_inventory = _build_modeled_inventory_from_surface(
        baseline_local=baseline_local,
        corrected_local=corrected_local,
        local_site_df=local_ctx["local_site_df"],
        local_kpi=local_ctx["local_kpi"],
        local_geo=local_ctx["local_geo"],
        context=context,
        config=config,
    )
    outcome = _evaluate_action_inventory(
        sector_cells=sector_cells,
        after_inventory=local_inventory,
        config=config,
        logger=logger,
        action_label="carrier_add_resim",
    )
    status = outcome["status"]
    action_reason = (
        "Carrier addition was rerun through baseline, Model 1, and Model 2 on the updated PART_3 local topology."
        if status in {"Resolved", "Partially Resolved"}
        else "Carrier addition was rerun through baseline, Model 1, and Model 2, but the affected sector lineage did not improve enough."
    )
    return {
        "status": status,
        "action_reason": action_reason,
        "projected_prb_after_pct": outcome["projected_prb_after_pct"],
        "projected_rrc_after_pct": outcome["projected_rrc_after_pct"],
        "projected_rrc_users_after": outcome["projected_rrc_users_after"],
        "next_step": outcome["next_step"] or "Sector Split",
        "resimulation_required": True,
        "resimulation_flow": "PART_3 carrier addition topology -> baseline rerun -> Model 1 inference -> Model 2 inference -> Model 3 reevaluation",
    }


def _run_new_site_resimulation(
    sector_cells: pd.DataFrame,
    config: Model3RecommendationConfig,
    context: dict[str, Any],
    logger: logging.Logger,
) -> dict[str, Any]:
    part3_site_df = context["part3_site_df"]
    building_df = context["building_df"]
    new_site_df, source_rows = _build_new_site_topology(sector_cells, part3_site_df, logger)
    if source_rows.empty:
        return {
            "status": "Recommended",
            "action_reason": "New-site topology could not be mapped back to PART_3 site rows.",
            "projected_prb_after_pct": np.nan,
            "projected_rrc_after_pct": np.nan,
            "projected_rrc_users_after": np.nan,
            "next_step": "",
            "resimulation_required": True,
            "resimulation_flow": "PART_3 new-site rerun failed at topology mapping",
        }
    local_ctx = _prepare_local_action_context(
        sector_cells=sector_cells,
        site_df=new_site_df,
        context=context,
        config=config,
        logger=logger,
        action_label="new_site_resim",
        source_rows=source_rows,
    )
    if local_ctx is None:
        return {
            "status": "Recommended",
            "action_reason": "New-site local context could not be prepared from PART_3 artifacts.",
            "projected_prb_after_pct": np.nan,
            "projected_rrc_after_pct": np.nan,
            "projected_rrc_users_after": np.nan,
            "next_step": "",
            "resimulation_required": True,
            "resimulation_flow": "PART_3 new-site rerun failed during local context preparation",
        }
    baseline_local = coverage_test._run_project_baseline_prediction(
        project_id=int(context["summary"].get("project_id", 196)),
        region=str(context["summary"].get("region", "india")).lower(),
        site_df=local_ctx["local_site_df"].drop(columns=["_site_distance_m"], errors="ignore"),
        drive_df=local_ctx["local_detail"],
        building_df=building_df,
        baseline_radius_m=float(context["summary"].get("baseline_radius_m", 500.0)),
        grid_size_m=float(context["summary"].get("grid_size_m", 50.0)),
        workers=1,
        max_interference_sites=int(context["summary"].get("max_interference_sites", 50)),
        polygon_wkt=str(context["summary"].get("polygon_wkt", "")).strip() or None,
        use_frontend_grid_sampling=False,
        grid_analytics_scenario_id=context["summary"].get("grid_analytics_scenario_id"),
    )
    if baseline_local.empty:
        return {
            "status": "Recommended",
            "action_reason": "New-site rerun produced no PART_3 baseline rows.",
            "projected_prb_after_pct": np.nan,
            "projected_rrc_after_pct": np.nan,
            "projected_rrc_users_after": np.nan,
            "next_step": "",
            "resimulation_required": True,
            "resimulation_flow": "PART_3 new-site rerun produced no baseline rows",
        }
    baseline_local["time_bucket"] = "PART_3"
    baseline_local = _merge_point_map_by_coordinates(
        baseline_local,
        local_ctx["local_point_map"].drop(columns=["_site_distance_m"], errors="ignore"),
    )
    baseline_local = _ensure_grid_group_columns(baseline_local)
    corrected_local = coverage_test._run_bucket_corrected_predictions(
        baseline_pred_df=baseline_local,
        detail_df=local_ctx["local_detail"].assign(time_bucket="PART_3"),
        site_df_by_bucket={"PART_3": local_ctx["local_site_df"].drop(columns=["_site_distance_m"], errors="ignore")},
        building_df=building_df,
        project_id=int(context["summary"].get("project_id", 196)),
        region=str(context["summary"].get("region", "india")).lower(),
        grid_size_m=float(context["summary"].get("grid_size_m", 50.0)),
        buckets=[("PART_3", "2026-02-11 00:00:00", "2026-05-16 23:59:59")],
        polygon_wkt=str(context["summary"].get("polygon_wkt", "")).strip() or None,
    )
    if corrected_local.empty:
        corrected_local = baseline_local.copy()
    corrected_local["time_bucket"] = "PART_3"
    corrected_local = _merge_point_map_by_coordinates(
        corrected_local,
        local_ctx["local_point_map"].drop(columns=["_site_distance_m"], errors="ignore"),
    )
    corrected_local = _ensure_grid_group_columns(corrected_local)
    _, local_inventory = _build_modeled_inventory_from_surface(
        baseline_local=baseline_local,
        corrected_local=corrected_local,
        local_site_df=local_ctx["local_site_df"],
        local_kpi=local_ctx["local_kpi"],
        local_geo=local_ctx["local_geo"],
        context=context,
        config=config,
    )
    outcome = _evaluate_action_inventory(
        sector_cells=sector_cells,
        after_inventory=local_inventory,
        config=config,
        logger=logger,
        action_label="new_site_resim",
    )
    status = outcome["status"]
    action_reason = (
        "New site was rerun through baseline, Model 1, and Model 2 on the updated PART_3 local topology."
        if status in {"Resolved", "Partially Resolved"}
        else "New site was rerun through baseline, Model 1, and Model 2, but the affected sector lineage did not improve enough."
    )
    return {
        "status": status,
        "action_reason": action_reason,
        "projected_prb_after_pct": outcome["projected_prb_after_pct"],
        "projected_rrc_after_pct": outcome["projected_rrc_after_pct"],
        "projected_rrc_users_after": outcome["projected_rrc_users_after"],
        "next_step": "",
        "resimulation_required": True,
        "resimulation_flow": "PART_3 new-site topology -> baseline rerun -> Model 1 inference -> Model 2 inference -> Model 3 reevaluation",
    }


def _run_load_balance_resimulation(
    sector_cells: pd.DataFrame,
    config: Model3RecommendationConfig,
    context: dict[str, Any],
    logger: logging.Logger,
    load_candidate: dict[str, Any],
) -> dict[str, Any]:
    local_ctx = _prepare_local_action_context(
        sector_cells=sector_cells,
        site_df=context["part3_site_df"],
        context=context,
        config=config,
        logger=logger,
        action_label="load_balance_resim",
        source_rows=_extract_source_site_rows(sector_cells, context["part3_site_df"]),
    )
    if local_ctx is None:
        return {
            "status": "Recommended",
            "action_reason": "Load-balance local context could not be prepared from PART_3 artifacts.",
            "projected_prb_after_pct": np.nan,
            "projected_rrc_after_pct": np.nan,
            "projected_rrc_users_after": np.nan,
            "next_step": "Add Carrier",
            "resimulation_required": True,
            "resimulation_flow": "Model 2 load-balance rerun failed during local context preparation",
        }
    source_id = str(load_candidate["source_node_cell_id"])
    target_id = str(load_candidate["target_node_cell_id"])
    corrected_local = local_ctx["part3_corrected"].merge(
        local_ctx["local_point_map"][["lat", "lon"]].drop_duplicates(), on=["lat", "lon"], how="inner"
    )
    corrected_local["grid_id"] = pd.to_numeric(corrected_local["grid_id"], errors="coerce").astype("Int64")
    corrected_local["time_bucket"] = corrected_local["time_bucket"].astype(str)
    corrected_local["_rsrp"] = pd.to_numeric(corrected_local.get("pred_rsrp"), errors="coerce")
    pair = corrected_local.loc[corrected_local["Node_Cell_ID"].astype(str).isin([source_id, target_id])].copy()
    pair = pair.pivot_table(index=["grid_id", "time_bucket"], columns="Node_Cell_ID", values="_rsrp", aggfunc="max").reset_index()
    if source_id not in pair.columns or target_id not in pair.columns:
        return {
            "status": "No Material Change",
            "action_reason": "Load-balance rerun could not find both source and target carriers on the same local grid set.",
            "projected_prb_after_pct": np.nan,
            "projected_rrc_after_pct": np.nan,
            "projected_rrc_users_after": np.nan,
            "next_step": "Add Carrier",
            "resimulation_required": True,
            "resimulation_flow": "Model 2 grid reassignment could not find overlapping source/target candidates",
        }
    pair["_delta_db"] = pd.to_numeric(pair[target_id], errors="coerce") - pd.to_numeric(pair[source_id], errors="coerce")
    movable = pair.loc[pair["_delta_db"] >= -6.0].sort_values("_delta_db", ascending=False).copy()
    if movable.empty:
        return {
            "status": "No Material Change",
            "action_reason": "Load-balance rerun found no grids where the target carrier had acceptable RF overlap with the source carrier.",
            "projected_prb_after_pct": np.nan,
            "projected_rrc_after_pct": np.nan,
            "projected_rrc_users_after": np.nan,
            "next_step": "Add Carrier",
            "resimulation_required": True,
            "resimulation_flow": "Model 2 grid reassignment found no movable overlapping grids",
        }
    move_count = max(1, int(round(len(movable) * 0.30)))
    forced_assignments = movable.head(move_count)[["grid_id", "time_bucket"]].copy()
    forced_assignments["Node_Cell_ID"] = target_id
    baseline_local = local_ctx["part3_baseline"].merge(
        local_ctx["local_point_map"][["lat", "lon"]].drop_duplicates(), on=["lat", "lon"], how="inner"
    )
    baseline_local = _merge_point_map_by_coordinates(
        baseline_local,
        local_ctx["local_point_map"].drop(columns=["_site_distance_m"], errors="ignore"),
    )
    baseline_local = _ensure_grid_group_columns(baseline_local)
    corrected_local = _merge_point_map_by_coordinates(
        corrected_local,
        local_ctx["local_point_map"].drop(columns=["_site_distance_m"], errors="ignore"),
    )
    corrected_local = _ensure_grid_group_columns(corrected_local)
    _, local_inventory = _build_modeled_inventory_from_surface(
        baseline_local=baseline_local,
        corrected_local=corrected_local,
        local_site_df=local_ctx["local_site_df"],
        local_kpi=local_ctx["local_kpi"],
        local_geo=local_ctx["local_geo"],
        context=context,
        config=config,
        forced_assignments=forced_assignments,
    )
    outcome = _evaluate_action_inventory(
        sector_cells=sector_cells,
        after_inventory=local_inventory,
        config=config,
        logger=logger,
        action_label="load_balance_resim",
    )
    return {
        "status": outcome["status"],
        "action_reason": f"Load balancing reassigned {move_count} overlapping local grids from {source_id} to {target_id} and reran Model 2 aggregation on the updated state.",
        "projected_prb_after_pct": outcome["projected_prb_after_pct"],
        "projected_rrc_after_pct": outcome["projected_rrc_after_pct"],
        "projected_rrc_users_after": outcome["projected_rrc_users_after"],
        "next_step": outcome["next_step"] or "Add Carrier",
        "resimulation_required": True,
        "resimulation_flow": "Grid reassignment -> Model 2 inference -> Model 3 reevaluation",
    }


def _simulate_recommendation(
    sector_cells: pd.DataFrame,
    config: Model3RecommendationConfig,
    logger: logging.Logger,
    context: dict[str, Any] | None = None,
) -> dict[str, Any]:
    congested_sources = sector_cells.loc[sector_cells["congested"]].copy()
    source_row = congested_sources.sort_values(
        ["prb_rrc_pressure", "grid_count"],
        ascending=[False, False],
        na_position="last",
    ).iloc[0]
    source_prb = float(source_row["prb_before_pct"]) if pd.notna(source_row["prb_before_pct"]) else np.nan
    source_rrc = float(source_row["rrc_before_pct"]) if pd.notna(source_row["rrc_before_pct"]) else np.nan
    pressure = max(source_prb if pd.notna(source_prb) else 0.0, source_rrc if pd.notna(source_rrc) else 0.0)
    congested = pressure > config.congestion_threshold

    load_candidate = _choose_load_balance_candidate(sector_cells)
    load_balance_possible = load_candidate is not None
    candidate_band = str(load_candidate["target_band"]) if load_candidate is not None else ""
    candidate_cell = str(load_candidate["target_node_cell_id"]) if load_candidate is not None else ""
    source_cell_id = str(load_candidate["source_node_cell_id"]) if load_candidate is not None else str(source_row["Node_Cell_ID"])

    decision_path: list[str] = []
    action = "Healthy"
    status = "Healthy"
    projected_prb = source_prb
    projected_rrc = source_rrc
    projected_users = float(source_row["rrc_users_before"]) if pd.notna(source_row["rrc_users_before"]) else np.nan
    action_reason = "Healthy."
    resimulation_flow = ""
    logger.info(
        "cell=%s sector=%s band=%s prb=%.3f rrc=%.3f congested=%s",
        source_cell_id,
        source_row["sector_id"],
        source_row["band"],
        source_prb if pd.notna(source_prb) else -1.0,
        source_rrc if pd.notna(source_rrc) else -1.0,
        congested,
    )

    if not congested:
        return {
            "action": action,
            "status": status,
            "decision_path": "No action",
            "load_balance_possible": load_balance_possible,
            "selected_peer_node_cell_id": candidate_cell,
            "selected_peer_band": candidate_band,
            "projected_prb_after_pct": projected_prb,
            "projected_rrc_after_pct": projected_rrc,
            "projected_rrc_users_after": projected_users,
            "action_reason": action_reason,
            "next_step": "",
            "resimulation_required": False,
            "resimulation_flow": "",
        }

    decision_path.append("Congested")

    if load_balance_possible:
        decision_path.append(f"Load balance candidate found: {candidate_band}")
        if context is not None:
            resim = _run_load_balance_resimulation(sector_cells, config, context, logger, load_candidate)
            action = f"Load Balance -> {candidate_band} MHz"
            return {
                "action": action,
                "status": resim["status"],
                "decision_path": " | ".join(decision_path),
                "load_balance_possible": True,
                "selected_peer_node_cell_id": candidate_cell,
                "selected_peer_band": candidate_band,
                "projected_prb_after_pct": resim["projected_prb_after_pct"],
                "projected_rrc_after_pct": resim["projected_rrc_after_pct"],
                "projected_rrc_users_after": resim["projected_rrc_users_after"],
                "action_reason": resim["action_reason"],
                "next_step": resim["next_step"],
                "resimulation_required": True,
                "resimulation_flow": resim["resimulation_flow"],
            }
        projected_prb = float(load_candidate["projected_prb_after_pct"])
        projected_rrc = float(load_candidate["projected_rrc_after_pct"])
        projected_users = float(load_candidate["projected_rrc_users_after"]) if pd.notna(load_candidate["projected_rrc_users_after"]) else np.nan
        if pd.notna(projected_prb) and pd.notna(projected_rrc) and projected_prb <= config.congestion_threshold and projected_rrc <= config.congestion_threshold:
            action = f"Load Balance -> {candidate_band} MHz"
            status = "Resolved"
            action_reason = "Same-sector carrier can absorb load."
            logger.info(
                "decision=load_balance sector=%s source=%s target=%s projected_prb=%.3f projected_rrc=%.3f",
                source_row["sector_id"],
                source_cell_id,
                candidate_cell,
                projected_prb,
                projected_rrc,
            )
            return {
                "action": action,
                "status": status,
                "decision_path": " | ".join(decision_path),
                "load_balance_possible": True,
                "selected_peer_node_cell_id": candidate_cell,
                "selected_peer_band": candidate_band,
                "projected_prb_after_pct": round(projected_prb, 3),
                "projected_rrc_after_pct": round(projected_rrc, 3),
                "projected_rrc_users_after": round(projected_users, 3) if pd.notna(projected_users) else np.nan,
                "action_reason": action_reason,
                "next_step": "",
                "resimulation_required": False,
                "resimulation_flow": "",
            }
        decision_path.append("Load balance insufficient")
        logger.info("decision=load_balance_failed sector=%s source=%s target=%s", source_row["sector_id"], source_cell_id, candidate_cell)

    can_add = bool(sector_cells["carrier_addition_possible"].any()) and not bool(sector_cells["carrier_addition_blocked"].all())
    add_row = sector_cells.loc[sector_cells["carrier_addition_possible"] & ~sector_cells["carrier_addition_blocked"]].copy()
    add_row = add_row.sort_values(["prb_rrc_pressure", "grid_count"], ascending=[False, False], na_position="last")
    add_row = add_row.iloc[0] if not add_row.empty else None
    if can_add and add_row is not None and str(add_row["recommended_band_to_add"]).strip():
        band = str(add_row["recommended_band_to_add"]).strip()
        decision_path.append(f"Carrier addition possible: {band}")
        if context is not None:
            resim = _run_carrier_addition_resimulation(sector_cells, config, context, logger, band)
            return {
                "action": f"Add Carrier -> {band} MHz",
                "status": resim["status"],
                "decision_path": " | ".join(decision_path),
                "load_balance_possible": load_balance_possible,
                "selected_peer_node_cell_id": source_cell_id,
                "selected_peer_band": candidate_band,
                "projected_prb_after_pct": resim["projected_prb_after_pct"],
                "projected_rrc_after_pct": resim["projected_rrc_after_pct"],
                "projected_rrc_users_after": resim["projected_rrc_users_after"],
                "action_reason": resim["action_reason"],
                "next_step": resim["next_step"],
                "resimulation_required": True,
                "resimulation_flow": resim["resimulation_flow"],
            }
        add_factor = 0.65 if load_balance_possible else 0.70
        projected_prb = max(0.0, source_prb * add_factor)
        projected_rrc = max(0.0, source_rrc * add_factor)
        projected_users = max(0.0, projected_users * add_factor) if pd.notna(projected_users) else np.nan
        action = f"Add Carrier -> {band} MHz"
        status = "Resolved" if projected_prb <= config.congestion_threshold and projected_rrc <= config.congestion_threshold else "Partially Resolved"
        action_reason = "Add the next recommended band."
        next_step = "" if status == "Resolved" else "Sector Split"
        logger.info(
            "decision=add_carrier cell=%s band=%s projected_prb=%.3f projected_rrc=%.3f status=%s",
            source_cell_id,
            band,
            projected_prb,
            projected_rrc,
            status,
        )
        return {
            "action": action,
            "status": status,
            "decision_path": " | ".join(decision_path),
            "load_balance_possible": load_balance_possible,
            "selected_peer_node_cell_id": source_cell_id,
            "selected_peer_band": candidate_band,
            "projected_prb_after_pct": round(projected_prb, 3),
            "projected_rrc_after_pct": round(projected_rrc, 3),
            "projected_rrc_users_after": round(projected_users, 3) if pd.notna(projected_users) else np.nan,
            "action_reason": action_reason,
            "next_step": next_step,
            "resimulation_required": False,
            "resimulation_flow": "",
        }

    sector_split_possible = bool(
        int(source_row["sector_cell_count"]) >= 1
        and not can_add
        and (
            int(source_row["sector_congested_count"]) >= 2
            or (
                pressure >= 95.0
                and int(source_row["existing_carrier_count"]) >= int(source_row["max_supported_carriers"])
            )
            or bool(sector_cells["carrier_addition_blocked"].all())
        )
    )
    if sector_split_possible:
        decision_path.append("Sector split branch")
        action = "Sector Split"
        status = "Proposed"
        action_reason = "Sector split is the remaining RF action, so rerun the PART_3 topology before marking it solved."
        resimulation_flow = "PART_3 topology split -> baseline rerun -> Model 1 inference -> Model 2 inference -> Model 3 reevaluation"
        logger.info("decision=sector_split_prepare cell=%s sector=%s", source_cell_id, source_row["sector_id"])
        if context is not None:
            resim = _run_sector_split_resimulation(sector_cells, config, context, logger)
            projected_prb = resim["projected_prb_after_pct"]
            projected_rrc = resim["projected_rrc_after_pct"]
            projected_users = resim["projected_rrc_users_after"]
            status = resim["status"]
            action_reason = resim["action_reason"]
            resimulation_flow = resim["resimulation_flow"]
            next_step = resim["next_step"]
        else:
            projected_prb = np.nan
            projected_rrc = np.nan
            projected_users = np.nan
            next_step = "Run PART_3 rerun"
        return {
            "action": action,
            "status": status,
            "decision_path": " | ".join(decision_path),
            "load_balance_possible": load_balance_possible,
            "selected_peer_node_cell_id": source_cell_id,
            "selected_peer_band": candidate_band,
            "projected_prb_after_pct": round(projected_prb, 3),
            "projected_rrc_after_pct": round(projected_rrc, 3),
            "projected_rrc_users_after": round(projected_users, 3) if pd.notna(projected_users) else np.nan,
            "action_reason": action_reason,
            "next_step": next_step,
            "resimulation_required": True,
            "resimulation_flow": resimulation_flow,
        }

    decision_path.append("New site branch")
    action = "New Site"
    if context is not None:
        resim = _run_new_site_resimulation(sector_cells, config, context, logger)
        return {
            "action": action,
            "status": resim["status"],
            "decision_path": " | ".join(decision_path),
            "load_balance_possible": load_balance_possible,
            "selected_peer_node_cell_id": source_cell_id,
            "selected_peer_band": candidate_band,
            "projected_prb_after_pct": resim["projected_prb_after_pct"],
            "projected_rrc_after_pct": resim["projected_rrc_after_pct"],
            "projected_rrc_users_after": resim["projected_rrc_users_after"],
            "action_reason": resim["action_reason"],
            "next_step": resim["next_step"],
            "resimulation_required": True,
            "resimulation_flow": resim["resimulation_flow"],
        }
    status = "Recommended"
    action_reason = "No viable in-sector or same-site relief remains."
    logger.info(
        "decision=new_site cell=%s sector=%s source_prb=%.3f source_rrc=%.3f",
        source_cell_id,
        source_row["sector_id"],
        source_prb if pd.notna(source_prb) else -1.0,
        source_rrc if pd.notna(source_rrc) else -1.0,
    )
    return {
        "action": action,
        "status": status,
        "decision_path": " | ".join(decision_path),
        "load_balance_possible": load_balance_possible,
        "selected_peer_node_cell_id": source_cell_id,
        "selected_peer_band": candidate_band,
        "projected_prb_after_pct": np.nan,
        "projected_rrc_after_pct": np.nan,
        "projected_rrc_users_after": np.nan,
        "action_reason": action_reason,
        "next_step": "",
        "resimulation_required": False,
        "resimulation_flow": "",
    }


def _build_recommendations(
    cell_df: pd.DataFrame,
    config: Model3RecommendationConfig,
    logger: logging.Logger,
    context: dict[str, Any] | None = None,
) -> pd.DataFrame:
    rows: list[dict[str, Any]] = []
    for sector_id, sector_cells in cell_df.groupby("sector_id", dropna=False):
        sector_cells = sector_cells.sort_values(
            ["congested", "prb_rrc_pressure", "grid_count"],
            ascending=[False, False, False],
            na_position="last",
        ).copy()
        if not bool(sector_cells["congested"].any()):
            continue
        rec = _simulate_recommendation(sector_cells, config, logger, context=context)
        sector_congested_ids = sector_cells.loc[sector_cells["congested"], "Node_Cell_ID"].astype(str).tolist()
        lead_row = sector_cells.iloc[0]
        if rec["action"] == "Sector Split":
            logger.info(
                "sector_split_plan sector=%s congested_cells=%s resimulation_flow=%s",
                sector_id,
                sector_congested_ids,
                rec["resimulation_flow"],
            )
        rows.append(
            {
                "site_id": lead_row["site_id"],
                "sector_id": sector_id,
                "Node_Cell_ID": lead_row["Node_Cell_ID"],
                "sector_congested_node_cell_ids": ", ".join(sector_congested_ids),
                "band": lead_row["band"],
                "earfcn": lead_row["earfcn"],
                "grid_count": int(sector_cells["grid_count"].sum()),
                "congested_grid_count": int(sector_cells["congested_grid_count"].sum()),
                "prb_before_pct": round(float(lead_row["prb_before_pct"]), 3) if pd.notna(lead_row["prb_before_pct"]) else np.nan,
                "rrc_before_pct": round(float(lead_row["rrc_before_pct"]), 3) if pd.notna(lead_row["rrc_before_pct"]) else np.nan,
                "rrc_users_before": round(float(lead_row["rrc_users_before"]), 3) if pd.notna(lead_row["rrc_users_before"]) else np.nan,
                "action": rec["action"],
                "status": rec["status"],
                "decision_path": rec["decision_path"],
                "load_balance_possible": bool(rec["load_balance_possible"]),
                "selected_peer_node_cell_id": rec["selected_peer_node_cell_id"],
                "selected_peer_band": rec["selected_peer_band"],
                "recommended_band_to_add": lead_row["recommended_band_to_add"],
                "available_bands_to_add": lead_row["available_bands_to_add"],
                "carrier_addition_possible": bool(sector_cells["carrier_addition_possible"].any()),
                "carrier_addition_blocked": bool(sector_cells["carrier_addition_blocked"].all()),
                "max_supported_carriers": int(sector_cells["max_supported_carriers"].max()),
                "existing_carrier_count": int(sector_cells["existing_carrier_count"].max()),
                "existing_carriers": lead_row["existing_carriers"],
                "projected_prb_after_pct": rec["projected_prb_after_pct"],
                "projected_rrc_after_pct": rec["projected_rrc_after_pct"],
                "projected_rrc_users_after": rec["projected_rrc_users_after"],
                "action_reason": rec["action_reason"],
                "next_step": rec["next_step"],
                "resimulation_required": bool(rec["resimulation_required"]),
                "resimulation_flow": rec["resimulation_flow"],
            }
        )

    reco_df = pd.DataFrame(rows)
    if reco_df.empty:
        return reco_df
    reco_df["load_balance_possible"] = reco_df["load_balance_possible"].astype(bool)
    reco_df["carrier_addition_possible"] = reco_df["carrier_addition_possible"].astype(bool)
    reco_df["carrier_addition_blocked"] = reco_df["carrier_addition_blocked"].astype(bool)
    reco_df["prb_before_pct"] = _to_num(reco_df["prb_before_pct"])
    reco_df["rrc_before_pct"] = _to_num(reco_df["rrc_before_pct"])
    reco_df["projected_prb_after_pct"] = _to_num(reco_df["projected_prb_after_pct"])
    reco_df["projected_rrc_after_pct"] = _to_num(reco_df["projected_rrc_after_pct"])
    reco_df["projected_rrc_users_after"] = _to_num(reco_df["projected_rrc_users_after"])
    reco_df["priority_score"] = reco_df[["prb_before_pct", "rrc_before_pct"]].max(axis=1)
    reco_df = reco_df.sort_values(
        ["status", "priority_score", "congested_grid_count", "grid_count"],
        ascending=[True, False, False, False],
        na_position="last",
    ).reset_index(drop=True)
    return reco_df


def _build_sector_inventory(cell_df: pd.DataFrame) -> pd.DataFrame:
    rows: list[dict[str, Any]] = []
    for sector_id, group in cell_df.groupby("sector_id", dropna=False):
        rows.append(
            {
                "sector_id": sector_id,
                "site_id": _mode_text(group["site_id"]),
                "cell_count": int(group["Node_Cell_ID"].nunique(dropna=True)),
                "congested_cell_count": int(group["congested"].sum()),
                "existing_carriers": _mode_text(group["existing_carriers"]),
                "existing_carrier_count": int(group["existing_carrier_count"].max()),
                "available_bands_to_add": _mode_text(group["available_bands_to_add"]),
                "recommended_band_to_add": _mode_text(group["recommended_band_to_add"]),
                "carrier_addition_possible": bool(group["carrier_addition_possible"].any()),
                "carrier_addition_blocked": bool(group["carrier_addition_blocked"].any()),
                "max_supported_carriers": int(group["max_supported_carriers"].max()),
                "sector_capacity_limit": int(group["sector_capacity_limit"].max()),
                "sector_has_alternate_carrier": bool(group["sector_has_alternate_carrier"].any()),
                "sector_pressure_max_pct": float(group["prb_rrc_pressure"].max()) if group["prb_rrc_pressure"].notna().any() else np.nan,
            }
        )
    sector_df = pd.DataFrame(rows)
    if sector_df.empty:
        return sector_df
    sector_df["sector_pressure_max_pct"] = _to_num(sector_df["sector_pressure_max_pct"])
    sector_df = sector_df.sort_values(
        ["congested_cell_count", "sector_pressure_max_pct", "cell_count"],
        ascending=[False, False, False],
        na_position="last",
    ).reset_index(drop=True)
    return sector_df


def _build_summary_payload(
    *,
    config: Model3RecommendationConfig,
    cell_df: pd.DataFrame,
    reco_df: pd.DataFrame,
    sector_df: pd.DataFrame,
    run_dir: Path,
    workbook_path: Path,
    source_summary: dict[str, Any] | None,
    runtime_sec: float,
) -> dict[str, Any]:
    action_counts = (
        {str(k): int(v) for k, v in reco_df["action"].value_counts(dropna=False).items()}
        if not reco_df.empty
        else {}
    )
    status_counts = (
        {str(k): int(v) for k, v in reco_df["status"].value_counts(dropna=False).items()}
        if not reco_df.empty
        else {}
    )
    threshold_stats = {}
    for col in ["prb_before_pct", "rrc_before_pct"]:
        if col not in cell_df.columns or cell_df.empty:
            continue
        values = _to_num(cell_df[col])
        threshold_stats[col] = {
            "non_null": int(values.notna().sum()),
            "min": float(values.min()) if values.notna().any() else None,
            "p50": float(values.quantile(0.50)) if values.notna().any() else None,
            "p80": float(values.quantile(0.80)) if values.notna().any() else None,
            "p90": float(values.quantile(0.90)) if values.notna().any() else None,
            "max": float(values.max()) if values.notna().any() else None,
            "threshold_counts": _threshold_counts(values, [50.0, 60.0, 70.0, 80.0, 90.0]),
        }

    summary = {
        "dataset_path": str(config.dataset_path),
        "summary_path": str(config.summary_path),
        "workbook_path": str(workbook_path),
        "run_dir": str(run_dir),
        "rows_in_model3_dataset": int(len(cell_df)),
        "congested_cell_rows": int(len(reco_df)),
        "sector_rows": int(len(sector_df)),
        "action_counts": action_counts,
        "status_counts": status_counts,
        "threshold": float(config.congestion_threshold),
        "rrc_sector_capacity_assumption": float(config.rrc_sector_capacity),
        "threshold_stats": threshold_stats,
        "source_model3_summary": source_summary or {},
        "top_recommendations": (
            reco_df[
                [
                    "sector_id",
                    "Node_Cell_ID",
                    "band",
                    "prb_before_pct",
                    "rrc_before_pct",
                    "action",
                    "status",
                    "projected_prb_after_pct",
                    "projected_rrc_after_pct",
                    "resimulation_required",
                    "resimulation_flow",
                ]
            ]
            .head(10)
            .to_dict(orient="records")
            if not reco_df.empty
            else []
        ),
        "artifacts": {
            "recommendations_csv": str(run_dir / "model3_recommendations.csv"),
            "sector_inventory_csv": str(run_dir / "model3_sector_inventory.csv"),
            "cell_inventory_csv": str(run_dir / "model3_cell_inventory.csv"),
            "summary_json": str(run_dir / "summary.json"),
            "workbook": str(workbook_path),
        },
        "runtime_sec": round(float(runtime_sec), 4),
    }
    return summary


def _write_df_sheet(ws, df: pd.DataFrame) -> None:
    header_fill = PatternFill(fill_type="solid", fgColor="D9EAF7")
    alt_fill = PatternFill(fill_type="solid", fgColor="F7F7F7")
    white_fill = PatternFill(fill_type="solid", fgColor="FFFFFF")
    header_font = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True, shrink_to_fit=True)

    for col_idx, col_name in enumerate(df.columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font
        cell.alignment = center
        cell.fill = header_fill

    for row_idx, row in enumerate(df.itertuples(index=False), start=2):
        fill = alt_fill if row_idx % 2 == 0 else white_fill
        for col_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=_excel_safe_value(value))
            cell.fill = fill
            cell.alignment = center

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for col_idx, col_name in enumerate(df.columns, start=1):
        max_len = len(str(col_name))
        for row_idx in range(2, ws.max_row + 1):
            val = ws.cell(row=row_idx, column=col_idx).value
            if val is None:
                continue
            max_len = max(max_len, len(str(val)))
        if str(df.columns[col_idx - 1]).lower() in {"action_reason", "next_step", "decision_path", "carrier_addition_reason", "sector_congested_node_cell_ids"}:
            max_len = max(max_len, 28)
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 70)


def _style_recommendation_sheet(ws, df: pd.DataFrame) -> None:
    cols = {name: idx + 1 for idx, name in enumerate(df.columns)}
    center = Alignment(horizontal="center", vertical="center", wrap_text=True, shrink_to_fit=True)
    action_fill = {
        "Load Balance": PatternFill(fill_type="solid", fgColor="D9EAD3"),
        "Add Carrier": PatternFill(fill_type="solid", fgColor="D9E8FF"),
        "Sector Split": PatternFill(fill_type="solid", fgColor="FCE5CD"),
        "New Site": PatternFill(fill_type="solid", fgColor="F4CCCC"),
    }
    status_fill = {
        "Resolved": PatternFill(fill_type="solid", fgColor="C6EFCE"),
        "Partially Resolved": PatternFill(fill_type="solid", fgColor="FFEB9C"),
        "Recommended": PatternFill(fill_type="solid", fgColor="D9D9D9"),
        "Proposed": PatternFill(fill_type="solid", fgColor="FFF2CC"),
    }
    action_col = cols.get("action")
    status_col = cols.get("status")
    prb_after_col = cols.get("projected_prb_after_pct")
    rrc_after_col = cols.get("projected_rrc_after_pct")
    add_band_col = cols.get("recommended_band_to_add")
    header_row = 1

    for row_idx in range(2, ws.max_row + 1):
        action = str(ws.cell(row=row_idx, column=action_col).value or "") if action_col else ""
        status = str(ws.cell(row=row_idx, column=status_col).value or "") if status_col else ""
        for col_idx in range(1, ws.max_column + 1):
            ws.cell(row=row_idx, column=col_idx).alignment = center
        fill = None
        for key, palette_fill in action_fill.items():
            if key.lower() in action.lower():
                fill = palette_fill
                break
        if fill is not None:
            for col_idx in range(1, ws.max_column + 1):
                ws.cell(row=row_idx, column=col_idx).fill = fill

        if status_col and status in status_fill:
            ws.cell(row=row_idx, column=status_col).fill = status_fill[status]
        if prb_after_col:
            prb_val = ws.cell(row=row_idx, column=prb_after_col).value
            rrc_val = ws.cell(row=row_idx, column=rrc_after_col).value if rrc_after_col else None
            try:
                if prb_val is not None and rrc_val is not None and float(prb_val) <= DEFAULT_CONGESTION_THRESHOLD and float(rrc_val) <= DEFAULT_CONGESTION_THRESHOLD:
                    ws.cell(row=row_idx, column=prb_after_col).fill = PatternFill(fill_type="solid", fgColor="C6EFCE")
                    ws.cell(row=row_idx, column=rrc_after_col).fill = PatternFill(fill_type="solid", fgColor="C6EFCE")
            except Exception:
                pass
        if add_band_col and str(ws.cell(row=row_idx, column=add_band_col).value or "").strip():
            ws.cell(row=row_idx, column=add_band_col).fill = PatternFill(fill_type="solid", fgColor="DDEEFF")

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for col_idx in range(1, ws.max_column + 1):
        max_len = len(str(ws.cell(row=header_row, column=col_idx).value or ""))
        for row_idx in range(2, ws.max_row + 1):
            val = ws.cell(row=row_idx, column=col_idx).value
            if val is None:
                continue
            max_len = max(max_len, len(str(val)))
        col_name = str(ws.cell(row=header_row, column=col_idx).value or "").lower()
        if col_name in {"action_reason", "next_step", "decision_path", "carrier_addition_reason", "sector_congested_node_cell_ids"}:
            max_len = max(max_len, 32)
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 75)

    for row_idx in range(2, ws.max_row + 1):
        ws.row_dimensions[row_idx].height = 22


def _write_workbook(
    *,
    run_dir: Path,
    cell_df: pd.DataFrame,
    reco_df: pd.DataFrame,
    sector_df: pd.DataFrame,
    summary: dict[str, Any],
) -> Path:
    wb = Workbook()
    wb.remove(wb.active)

    ws_summary = wb.create_sheet("Summary")
    summary_rows = []
    for key, value in [
        ("rows_in_model3_dataset", summary["rows_in_model3_dataset"]),
        ("congested_cell_rows", summary["congested_cell_rows"]),
        ("sector_rows", summary["sector_rows"]),
        ("threshold", summary["threshold"]),
        ("rrc_sector_capacity_assumption", summary["rrc_sector_capacity_assumption"]),
        ("runtime_sec", summary["runtime_sec"]),
    ]:
        summary_rows.append({"metric": key, "value": value})
    for action, count in summary["action_counts"].items():
        summary_rows.append({"metric": f"action::{action}", "value": count})
    for status, count in summary["status_counts"].items():
        summary_rows.append({"metric": f"status::{status}", "value": count})
    for metric_name, stats in summary["threshold_stats"].items():
        summary_rows.append({"metric": f"{metric_name}::p50", "value": stats.get("p50")})
        summary_rows.append({"metric": f"{metric_name}::p80", "value": stats.get("p80")})
        summary_rows.append({"metric": f"{metric_name}::p90", "value": stats.get("p90")})
        summary_rows.append({"metric": f"{metric_name}::max", "value": stats.get("max")})
    _write_df_sheet(ws_summary, pd.DataFrame(summary_rows))

    ws_reco = wb.create_sheet("Recommendations")
    _write_df_sheet(ws_reco, reco_df)
    _style_recommendation_sheet(ws_reco, reco_df)

    ws_cells = wb.create_sheet("Congested Cells")
    _write_df_sheet(ws_cells, cell_df.loc[cell_df["congested"]].copy())

    ws_sector = wb.create_sheet("Sector Inventory")
    _write_df_sheet(ws_sector, sector_df)

    ws_top = wb.create_sheet("Top Hotspots")
    top_cols = [
        "site_id",
        "sector_id",
        "Node_Cell_ID",
        "band",
        "grid_count",
        "prb_before_pct",
        "rrc_before_pct",
        "action",
        "status",
        "projected_prb_after_pct",
        "projected_rrc_after_pct",
        "recommended_band_to_add",
        "decision_path",
        "resimulation_required",
        "resimulation_flow",
    ]
    ws_top_df = reco_df[top_cols].head(25).copy() if not reco_df.empty else pd.DataFrame(columns=top_cols)
    _write_df_sheet(ws_top, ws_top_df)

    split_rows = reco_df.loc[reco_df["action"].astype(str) == "Sector Split"].copy() if not reco_df.empty else pd.DataFrame()
    ws_split = wb.create_sheet("Sector Split Plan")
    split_cols = [
        "site_id",
        "sector_id",
        "Node_Cell_ID",
        "sector_congested_node_cell_ids",
        "band",
        "prb_before_pct",
        "rrc_before_pct",
        "action_reason",
        "next_step",
        "resimulation_flow",
        "projected_prb_after_pct",
        "projected_rrc_after_pct",
    ]
    ws_split_df = split_rows[split_cols].copy() if not split_rows.empty else pd.DataFrame(columns=split_cols)
    _write_df_sheet(ws_split, ws_split_df)

    workbook_path = run_dir / "model3_business_rule_recommendations.xlsx"
    wb.save(workbook_path)

    stable_copy = DEFAULT_STABLE_OUTPUT_DIR / "model3_business_rule_recommendations.xlsx"
    stable_copy.parent.mkdir(parents=True, exist_ok=True)
    try:
        shutil.copy2(workbook_path, stable_copy)
    except PermissionError:
        # Keep the run usable even if a previous workbook copy is briefly locked.
        pass
    return workbook_path


def run_model3_business_rule_recommendation_test(config: Model3RecommendationConfig) -> Path:
    start = time.perf_counter()
    if not config.dataset_path.exists():
        raise FileNotFoundError(f"Model 3 dataset not found: {config.dataset_path}")

    run_dir = _ensure_dir(config.output_root / f"model3_business_rule_{_timestamp()}")
    log_path = run_dir / "log.txt"
    logger = _setup_logger(log_path)
    logger.info("start dataset=%s summary=%s", config.dataset_path, config.summary_path)
    cell_df = pd.read_csv(config.dataset_path)
    source_summary = {}
    if config.summary_path.exists():
        try:
            source_summary = json.loads(config.summary_path.read_text(encoding="utf-8"))
        except Exception:
            source_summary = {}
    resimulation_context = _load_resimulation_context(config, logger)

    cell_inventory, inventory_summary = _build_cell_inventory(cell_df, config)
    logger.info(
        "inventory cell_count=%d sector_count=%d congested_cells=%d carrier_addition_candidates=%d",
        inventory_summary["cell_count"],
        inventory_summary["sector_count"],
        inventory_summary["congested_cell_count"],
        inventory_summary["carrier_addition_candidate_count"],
    )
    recommendations = _build_recommendations(cell_inventory, config, logger, context=resimulation_context)
    sector_inventory = _build_sector_inventory(cell_inventory)

    workbook_path = run_dir / "model3_business_rule_recommendations.xlsx"
    runtime_sec = time.perf_counter() - start
    summary = _build_summary_payload(
        config=config,
        cell_df=cell_inventory,
        reco_df=recommendations,
        sector_df=sector_inventory,
        run_dir=run_dir,
        workbook_path=workbook_path,
        source_summary=source_summary,
        runtime_sec=runtime_sec,
    )

    # Write the workbook once the summary payload is complete.
    workbook_path = _write_workbook(
        run_dir=run_dir,
        cell_df=cell_inventory,
        reco_df=recommendations,
        sector_df=sector_inventory,
        summary=summary,
    )
    runtime_sec = time.perf_counter() - start
    summary["runtime_sec"] = round(float(runtime_sec), 4)
    summary["workbook_path"] = str(workbook_path)
    summary["artifacts"]["workbook"] = str(workbook_path)

    workbook_path = _write_workbook(
        run_dir=run_dir,
        cell_df=cell_inventory,
        reco_df=recommendations,
        sector_df=sector_inventory,
        summary=summary,
    )
    summary["workbook_path"] = str(workbook_path)
    summary["artifacts"]["workbook"] = str(workbook_path)
    summary["artifacts"]["log"] = str(log_path)

    _save_json(run_dir / "summary.json", summary)
    recommendations.to_csv(run_dir / "model3_recommendations.csv", index=False)
    cell_inventory.to_csv(run_dir / "model3_cell_inventory.csv", index=False)
    sector_inventory.to_csv(run_dir / "model3_sector_inventory.csv", index=False)
    shutil.copy2(run_dir / "summary.json", DEFAULT_STABLE_OUTPUT_DIR / "model3_business_rule_recommendation_summary.json")
    shutil.copy2(run_dir / "model3_recommendations.csv", DEFAULT_STABLE_OUTPUT_DIR / "model3_recommendations.csv")
    try:
        shutil.copy2(log_path, DEFAULT_STABLE_OUTPUT_DIR / "model3_business_rule_recommendation.log")
    except PermissionError:
        pass
    print(json.dumps(summary, indent=2, default=str))
    return run_dir


def parse_args() -> Model3RecommendationConfig:
    parser = argparse.ArgumentParser(description="Build a Model 3 business-rule recommendation workbook from PART_3 only.")
    parser.add_argument("--dataset-path", type=Path, default=DEFAULT_MODEL3_DATASET)
    parser.add_argument("--summary-path", type=Path, default=DEFAULT_MODEL3_SUMMARY)
    parser.add_argument("--output-root", type=Path, default=DEFAULT_OUTPUT_ROOT)
    parser.add_argument("--stable-output-dir", type=Path, default=DEFAULT_STABLE_OUTPUT_DIR)
    parser.add_argument("--congestion-threshold", type=float, default=DEFAULT_CONGESTION_THRESHOLD)
    parser.add_argument("--rrc-sector-capacity", type=float, default=DEFAULT_RRC_SECTOR_CAPACITY)
    args = parser.parse_args()
    return Model3RecommendationConfig(
        dataset_path=args.dataset_path,
        summary_path=args.summary_path,
        output_root=args.output_root,
        stable_output_dir=args.stable_output_dir,
        congestion_threshold=args.congestion_threshold,
        rrc_sector_capacity=args.rrc_sector_capacity,
    )


if __name__ == "__main__":
    run_model3_business_rule_recommendation_test(parse_args())
