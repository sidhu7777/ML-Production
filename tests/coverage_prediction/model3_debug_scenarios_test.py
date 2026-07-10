from __future__ import annotations

import argparse
import json
import os
import sys
import time
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


ML_ROOT = Path(__file__).resolve().parents[2]
os.chdir(ML_ROOT)
if str(ML_ROOT) not in sys.path:
    sys.path.insert(0, str(ML_ROOT))

from tests.coverage_prediction import model3_business_rule_recommendation_test as model3_rules


DEFAULT_OUTPUT_ROOT = ML_ROOT / "tests" / "output" / "model3_debug_scenarios"


def _timestamp() -> str:
    return time.strftime("%Y%m%d_%H%M%S")


def _ensure_dir(path: Path) -> Path:
    path.mkdir(parents=True, exist_ok=True)
    return path


def _save_json(path: Path, payload: dict[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(payload, indent=2, default=str), encoding="utf-8")


def _center_sheet(ws) -> None:
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    header_fill = PatternFill(fill_type="solid", fgColor="D9EAF7")
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = center
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
    for column_cells in ws.columns:
        width = max(len(str(cell.value or "")) for cell in column_cells) + 2
        ws.column_dimensions[get_column_letter(column_cells[0].column)].width = min(max(width, 14), 42)


def _write_workbook(path: Path, summary_df: pd.DataFrame, feature_df: pd.DataFrame) -> None:
    wb = Workbook()
    ws_summary = wb.active
    ws_summary.title = "scenario_summary"
    ws_summary.append(summary_df.columns.tolist())
    for row in summary_df.itertuples(index=False, name=None):
        ws_summary.append(list(row))
    _center_sheet(ws_summary)

    ws_features = wb.create_sheet("feature_debug")
    ws_features.append(feature_df.columns.tolist())
    for row in feature_df.itertuples(index=False, name=None):
        ws_features.append(list(row))
    _center_sheet(ws_features)
    wb.save(path)


def _make_logger(output_dir: Path):
    return model3_rules._setup_logger(output_dir / "log.txt")


def _lead_row(sector_cells: pd.DataFrame) -> pd.Series:
    return (
        sector_cells.loc[sector_cells["congested"]]
        .sort_values(["prb_rrc_pressure", "grid_count"], ascending=[False, False], na_position="last")
        .iloc[0]
    )


def _scenario_feature_rows(
    *,
    scenario_name: str,
    scenario_kind: str,
    before_row: pd.Series,
    after_rec: dict[str, Any],
    actual_state_change: str,
    notes: str,
) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    comparisons = [
        ("prb_utilization_pct", before_row.get("prb_before_pct"), after_rec.get("projected_prb_after_pct")),
        ("rrc_utilization_pct", before_row.get("rrc_before_pct"), after_rec.get("projected_rrc_after_pct")),
        ("rrc_connected_users", before_row.get("rrc_users_before"), after_rec.get("projected_rrc_users_after")),
        ("existing_carrier_count", before_row.get("existing_carrier_count"), before_row.get("existing_carrier_count")),
        ("max_supported_carriers", before_row.get("max_supported_carriers"), before_row.get("max_supported_carriers")),
        ("sector_has_alternate_carrier", before_row.get("sector_has_alternate_carrier"), before_row.get("sector_has_alternate_carrier")),
        ("carrier_addition_possible", before_row.get("carrier_addition_possible"), before_row.get("carrier_addition_possible")),
        ("recommended_band_to_add", before_row.get("recommended_band_to_add"), after_rec.get("recommended_band_to_add", before_row.get("recommended_band_to_add"))),
        ("selected_peer_band", "", after_rec.get("selected_peer_band")),
        ("selected_peer_node_cell_id", "", after_rec.get("selected_peer_node_cell_id")),
        ("resimulation_flow", "", after_rec.get("resimulation_flow")),
    ]
    for feature_name, before_value, after_value in comparisons:
        rows.append(
            {
                "scenario_name": scenario_name,
                "scenario_kind": scenario_kind,
                "feature_name": feature_name,
                "before_value": model3_rules._excel_safe_value(before_value),
                "after_value": model3_rules._excel_safe_value(after_value),
                "actual_state_change": actual_state_change,
                "notes": notes,
            }
        )
    return rows


def _select_scenarios(cell_df: pd.DataFrame, config: model3_rules.Model3RecommendationConfig, logger) -> dict[str, dict[str, Any]]:
    scenarios: dict[str, dict[str, Any]] = {}
    logger_stub = type("LogStub", (), {"info": lambda *args, **kwargs: None})()

    candidates: list[dict[str, Any]] = []
    for sector_id, sector_cells in cell_df.groupby("sector_id", dropna=False):
        if not bool(sector_cells["congested"].any()):
            continue
        lead = _lead_row(sector_cells)
        rec = model3_rules._simulate_recommendation(sector_cells, config, logger_stub, context=None)
        candidates.append(
            {
                "sector_id": sector_id,
                "sector_cells": sector_cells.copy(),
                "lead_row": lead,
                "rec": rec,
            }
        )

    def _pick(kind: str, predicate) -> dict[str, Any]:
        filtered = [c for c in candidates if predicate(c)]
        if not filtered:
            raise RuntimeError(f"Could not find a scenario for {kind}.")
        filtered = sorted(
            filtered,
            key=lambda item: (
                -float(item["lead_row"].get("prb_rrc_pressure") or 0.0),
                -float(item["lead_row"].get("grid_count") or 0.0),
            ),
        )
        return filtered[0]

    scenarios["load_balance"] = _pick("load_balance", lambda c: str(c["rec"]["action"]).startswith("Load Balance"))
    scenarios["carrier_addition"] = _pick("carrier_addition", lambda c: str(c["rec"]["action"]).startswith("Add Carrier"))
    scenarios["sector_split"] = _pick("sector_split", lambda c: str(c["rec"]["action"]) == "Sector Split")

    split_candidates = sorted(
        [c for c in candidates if str(c["rec"]["action"]) == "Sector Split"],
        key=lambda item: (
            -float(item["lead_row"].get("prb_rrc_pressure") or 0.0),
            -float(item["lead_row"].get("grid_count") or 0.0),
        ),
    )
    scenarios["new_site_follow_on"] = split_candidates[1] if len(split_candidates) > 1 else split_candidates[0]

    logger.info(
        "selected_scenarios load_balance=%s carrier_addition=%s sector_split=%s new_site_follow_on=%s",
        scenarios["load_balance"]["sector_id"],
        scenarios["carrier_addition"]["sector_id"],
        scenarios["sector_split"]["sector_id"],
        scenarios["new_site_follow_on"]["sector_id"],
    )
    return scenarios


def _run_debug(
    *,
    config: model3_rules.Model3RecommendationConfig,
    output_dir: Path,
) -> dict[str, Any]:
    logger = _make_logger(output_dir)
    logger.info("start dataset=%s summary=%s", config.dataset_path, config.summary_path)

    dataset_df = pd.read_csv(config.dataset_path)
    summary_payload = json.loads(config.summary_path.read_text(encoding="utf-8"))
    cell_df, inventory_summary = model3_rules._build_cell_inventory(dataset_df, config)
    scenarios = _select_scenarios(cell_df, config, logger)

    logger.info(
        "inventory cell_count=%d sector_count=%d congested_cells=%d",
        inventory_summary["cell_count"],
        inventory_summary["sector_count"],
        inventory_summary["congested_cell_count"],
    )

    context = model3_rules._load_resimulation_context(config, logger)

    summary_rows: list[dict[str, Any]] = []
    feature_rows: list[dict[str, Any]] = []

    for scenario_name, scenario in scenarios.items():
        sector_id = scenario["sector_id"]
        sector_cells = scenario["sector_cells"].copy()
        lead_row = scenario["lead_row"]
        base_rec = scenario["rec"]

        logger.info(
            "scenario_start name=%s sector=%s source_cell=%s action=%s",
            scenario_name,
            sector_id,
            lead_row["Node_Cell_ID"],
            base_rec["action"],
        )

        actual_rec = dict(base_rec)
        actual_state_change = "none"
        notes = ""
        uses_real_resim = False

        if scenario_name == "sector_split":
            actual_rec = model3_rules._simulate_recommendation(sector_cells, config, logger, context=context)
            actual_state_change = "topology_change_local_part3_rerun"
            uses_real_resim = True
            notes = "This branch does a real local PART_3 rerun through baseline, Model 1, Model 2, then reevaluates Model 3."
        elif scenario_name == "load_balance":
            actual_rec = model3_rules._simulate_recommendation(sector_cells, config, logger, context=context)
            actual_state_change = "grid_reassignment_plus_model2_rerun"
            uses_real_resim = True
            notes = (
                "This branch now reassigns a movable subset of overlapping local grids from the congested carrier to the peer carrier "
                "and reruns Model 2 aggregation before reevaluating Model 3."
            )
        elif scenario_name == "carrier_addition":
            actual_rec = model3_rules._simulate_recommendation(sector_cells, config, logger, context=context)
            actual_state_change = "carrier_topology_change_local_part3_rerun"
            uses_real_resim = True
            notes = (
                "This branch now adds a real synthetic carrier to the affected sector in the local PART_3 topology, "
                "reruns baseline, then reruns Model 1 and Model 2 before reevaluating Model 3."
            )
        elif scenario_name == "new_site_follow_on":
            split_rec = model3_rules._simulate_recommendation(sector_cells, config, logger, context=context)
            actual_rec = dict(split_rec)
            actual_rec["action"] = "New Site (Follow-on Recommendation)"
            actual_rec["status"] = "Recommended"
            actual_rec["action_reason"] = (
                "Sector split did not clear congestion enough, so the current engine escalates to new site. "
                "This branch still does not perform a real new-site topology rerun."
            )
            actual_rec["next_step"] = ""
            actual_state_change = "no_new_site_state_change_in_current_engine"
            uses_real_resim = True
            notes = (
                "This scenario proves the follow-on path: the engine can identify that sector split is insufficient, "
                "but it does not yet create a new site, reassign grids, or rerun the models on a changed new-site topology."
            )

        summary_rows.append(
            {
                "scenario_name": scenario_name,
                "scenario_kind": scenario_name,
                "sector_id": sector_id,
                "site_id": lead_row["site_id"],
                "source_node_cell_id": lead_row["Node_Cell_ID"],
                "source_band": lead_row["band"],
                "before_prb_pct": round(float(lead_row["prb_before_pct"]), 3) if pd.notna(lead_row["prb_before_pct"]) else None,
                "before_rrc_pct": round(float(lead_row["rrc_before_pct"]), 3) if pd.notna(lead_row["rrc_before_pct"]) else None,
                "before_rrc_users": round(float(lead_row["rrc_users_before"]), 3) if pd.notna(lead_row["rrc_users_before"]) else None,
                "engine_action": actual_rec.get("action"),
                "engine_status": actual_rec.get("status"),
                "after_prb_pct": actual_rec.get("projected_prb_after_pct"),
                "after_rrc_pct": actual_rec.get("projected_rrc_after_pct"),
                "after_rrc_users": actual_rec.get("projected_rrc_users_after"),
                "selected_peer_band": actual_rec.get("selected_peer_band"),
                "selected_peer_node_cell_id": actual_rec.get("selected_peer_node_cell_id"),
                "recommended_band_to_add": lead_row.get("recommended_band_to_add"),
                "available_bands_to_add": lead_row.get("available_bands_to_add"),
                "existing_carriers": lead_row.get("existing_carriers"),
                "existing_carrier_count": int(lead_row.get("existing_carrier_count") or 0),
                "max_supported_carriers": int(lead_row.get("max_supported_carriers") or 0),
                "uses_real_resimulation": uses_real_resim,
                "actual_state_change": actual_state_change,
                "action_reason": actual_rec.get("action_reason"),
                "next_step": actual_rec.get("next_step"),
                "resimulation_flow": actual_rec.get("resimulation_flow"),
                "notes": notes,
            }
        )
        feature_rows.extend(
            _scenario_feature_rows(
                scenario_name=scenario_name,
                scenario_kind=scenario_name,
                before_row=lead_row,
                after_rec=actual_rec,
                actual_state_change=actual_state_change,
                notes=notes,
            )
        )

        _save_json(
            output_dir / f"{scenario_name}.json",
            {
                "scenario_name": scenario_name,
                "sector_id": sector_id,
                "source_node_cell_id": str(lead_row["Node_Cell_ID"]),
                "before": {
                    "prb_before_pct": float(lead_row["prb_before_pct"]) if pd.notna(lead_row["prb_before_pct"]) else None,
                    "rrc_before_pct": float(lead_row["rrc_before_pct"]) if pd.notna(lead_row["rrc_before_pct"]) else None,
                    "rrc_users_before": float(lead_row["rrc_users_before"]) if pd.notna(lead_row["rrc_users_before"]) else None,
                    "existing_carriers": lead_row.get("existing_carriers"),
                    "existing_carrier_count": int(lead_row.get("existing_carrier_count") or 0),
                    "max_supported_carriers": int(lead_row.get("max_supported_carriers") or 0),
                    "available_bands_to_add": lead_row.get("available_bands_to_add"),
                },
                "engine_result": actual_rec,
                "actual_state_change": actual_state_change,
                "notes": notes,
            },
        )
        logger.info(
            "scenario_done name=%s sector=%s action=%s status=%s after_prb=%s after_rrc=%s",
            scenario_name,
            sector_id,
            actual_rec.get("action"),
            actual_rec.get("status"),
            actual_rec.get("projected_prb_after_pct"),
            actual_rec.get("projected_rrc_after_pct"),
        )

    summary_df = pd.DataFrame(summary_rows)
    feature_df = pd.DataFrame(feature_rows)
    summary_csv = output_dir / "scenario_summary.csv"
    feature_csv = output_dir / "scenario_feature_debug.csv"
    workbook_path = output_dir / "model3_debug_scenarios.xlsx"
    summary_df.to_csv(summary_csv, index=False)
    feature_df.to_csv(feature_csv, index=False)
    _write_workbook(workbook_path, summary_df, feature_df)

    run_summary = {
        "dataset_path": str(config.dataset_path),
        "summary_path": str(config.summary_path),
        "rows_in_dataset": int(len(dataset_df)),
        "inventory_summary": inventory_summary,
        "scenario_output_dir": str(output_dir),
        "files": {
            "log": str(output_dir / "log.txt"),
            "scenario_summary_csv": str(summary_csv),
            "scenario_feature_debug_csv": str(feature_csv),
            "scenario_workbook": str(workbook_path),
        },
        "scenario_names": summary_df["scenario_name"].tolist(),
    }
    _save_json(output_dir / "run_summary.json", run_summary)
    logger.info("finished output_dir=%s", output_dir)
    return run_summary


def main() -> None:
    parser = argparse.ArgumentParser(description="Run focused Model 3 debug scenarios.")
    parser.add_argument("--dataset", type=Path, default=model3_rules.DEFAULT_MODEL3_DATASET)
    parser.add_argument("--summary", type=Path, default=model3_rules.DEFAULT_MODEL3_SUMMARY)
    parser.add_argument("--output-root", type=Path, default=DEFAULT_OUTPUT_ROOT)
    parser.add_argument("--stable-output-dir", type=Path, default=model3_rules.DEFAULT_STABLE_OUTPUT_DIR)
    parser.add_argument("--model1-variant", type=str, default="physical_no_teacher_summary")
    args = parser.parse_args()

    config = model3_rules.Model3RecommendationConfig(
        dataset_path=args.dataset,
        summary_path=args.summary,
        output_root=args.output_root,
        stable_output_dir=args.stable_output_dir,
        model1_variant=args.model1_variant,
    )
    output_dir = _ensure_dir(args.output_root / f"model3_debug_{_timestamp()}")
    result = _run_debug(config=config, output_dir=output_dir)
    print(json.dumps(result, indent=2))


if __name__ == "__main__":
    main()
