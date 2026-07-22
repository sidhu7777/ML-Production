from __future__ import annotations

import argparse
import json
import os
import sys
from pathlib import Path
from typing import Any

import numpy as np
import pandas as pd
from dotenv import load_dotenv
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from sqlalchemy import create_engine, text


ML_ROOT = Path(__file__).resolve().parents[2]
os.chdir(ML_ROOT)
if str(ML_ROOT) not in sys.path:
    sys.path.insert(0, str(ML_ROOT))


OUTPUT_ROOT = ML_ROOT / "models" / "model3_project196_input"
DEFAULT_PROJECT_ID = 196
DEFAULT_REGION = "india"
DEFAULT_GRID_SIZE_M = 25.0
DEFAULT_RRC_SECTOR_CAPACITY = 400.0
KNOWN_BAND_POOL_MHZ = [700.0, 850.0, 900.0, 1800.0, 2100.0, 2300.0]
DEFAULT_MAX_SUPPORTED_CARRIERS = len(KNOWN_BAND_POOL_MHZ)
DEMO_CONGESTION_THRESHOLD = 70.0


def _clean_text(value: object) -> str:
    if pd.isna(value):
        return ""
    text_value = str(value).strip()
    if text_value.lower() in {"", "none", "nan", "null", "<na>"}:
        return ""
    if text_value.endswith(".0"):
        text_value = text_value[:-2]
    return text_value


def _clean_cell_id(value: object) -> str:
    return _clean_text(value).replace("|", "_").strip("_")


def _canonical_site_cell_id(value: object) -> str:
    cell_id = _clean_cell_id(value)
    parts = [part for part in cell_id.split("_") if part]
    if len(parts) >= 3 and parts[0] == parts[1]:
        return "_".join([parts[0], *parts[2:]])
    return cell_id


def _first_non_empty_series(df: pd.DataFrame, candidates: list[str]) -> pd.Series:
    out = pd.Series("", index=df.index, dtype="object")
    for col in candidates:
        if col not in df.columns:
            continue
        values = df[col].map(_clean_cell_id)
        out = out.where(out.astype(str).str.strip().ne(""), values)
    return out


def _to_num(series: pd.Series, default: float = np.nan) -> pd.Series:
    out = pd.to_numeric(series, errors="coerce")
    return out.fillna(default) if not pd.isna(default) else out


def _minmax(series: pd.Series) -> pd.Series:
    values = pd.to_numeric(series, errors="coerce")
    valid = values.dropna()
    if valid.empty:
        return pd.Series(0.0, index=series.index, dtype="float64")
    lo = float(valid.min())
    hi = float(valid.max())
    if not np.isfinite(lo) or not np.isfinite(hi) or hi <= lo:
        return pd.Series(0.0, index=series.index, dtype="float64")
    return ((values.fillna(lo) - lo) / (hi - lo)).clip(0.0, 1.0)


def _format_band(value: object) -> str:
    num = pd.to_numeric(pd.Series([value]), errors="coerce").iloc[0]
    if pd.isna(num):
        return _clean_text(value)
    return str(int(num)) if float(num).is_integer() else f"{float(num):g}"


def _format_list(values: list[object]) -> str:
    cleaned = []
    for value in values:
        num = pd.to_numeric(pd.Series([value]), errors="coerce").iloc[0]
        if pd.notna(num):
            cleaned.append(float(num))
    if not cleaned:
        return ""
    return ",".join(str(int(v)) if float(v).is_integer() else f"{v:g}" for v in sorted(set(cleaned)))


def _set_cell_load(
    cell: pd.DataFrame,
    row_index: object,
    prb_pct: float,
    rrc_pct: float,
    rrc_sector_capacity: float,
) -> None:
    capacity = pd.to_numeric(pd.Series([cell.at[row_index, "estimated_dl_capacity_mbps"]]), errors="coerce").iloc[0]
    capacity = max(0.1, float(capacity) if pd.notna(capacity) else 0.1)
    cell.at[row_index, "proxy_prb_utilization_pct"] = round(float(prb_pct), 3)
    cell.at[row_index, "proxy_rrc_utilization_pct"] = round(float(rrc_pct), 3)
    cell.at[row_index, "proxy_rrc_connected_users"] = round((float(rrc_pct) / 100.0) * float(rrc_sector_capacity), 3)
    cell.at[row_index, "estimated_offered_traffic_mbps"] = round(capacity * (float(prb_pct) / 100.0), 3)


def _apply_model3_demo_scenarios(cell: pd.DataFrame, rrc_sector_capacity: float) -> pd.DataFrame:
    out = cell.copy()
    out["demo_scenario"] = "proxy_load"
    out["demo_scenario_reason"] = "RF/geo proxy load; editable planner input."

    sector_stats = (
        out.groupby("sector_id", dropna=False)
        .agg(
            cell_count=("Node_Cell_ID", "nunique"),
            max_pressure=("proxy_prb_utilization_pct", "max"),
            sector_grid_count=("grid_count", "sum"),
        )
        .reset_index()
        .sort_values(["cell_count", "max_pressure", "sector_grid_count"], ascending=[False, False, False])
    )
    multi_sectors = sector_stats.loc[sector_stats["cell_count"] >= 2, "sector_id"].astype(str).tolist()
    scenario_cycle = [
        "load_balance_success",
        "carrier_addition_required",
        "carrier_blocked_new_site",
        "sector_split_candidate",
    ]
    assignments = {
        sector_id: scenario_cycle[pos % len(scenario_cycle)]
        for pos, sector_id in enumerate(multi_sectors[: min(len(multi_sectors), 12)])
    }

    for sector_id, scenario in assignments.items():
        group = out.loc[out["sector_id"].astype(str).eq(str(sector_id))].copy()
        if group.empty:
            continue
        group = group.sort_values(["grid_count", "proxy_prb_utilization_pct", "proxy_rrc_utilization_pct"], ascending=[False, False, False])
        source_idx = group.index[0]
        peer_indices = list(group.index[1:])

        if scenario == "load_balance_success":
            _set_cell_load(out, source_idx, 88.0, 86.0, rrc_sector_capacity)
            for idx in peer_indices:
                _set_cell_load(out, idx, 42.0, 38.0, rrc_sector_capacity)
            out.loc[group.index, "demo_scenario_reason"] = "One congested carrier, same-sector peers have safe PRB/RRC headroom."

        elif scenario == "carrier_addition_required":
            for rank, idx in enumerate(group.index):
                _set_cell_load(out, idx, 84.0 + min(rank, 2) * 3.0, 82.0 + min(rank, 2) * 2.0, rrc_sector_capacity)
            existing_count = pd.to_numeric(out.loc[group.index, "existing_carrier_count"], errors="coerce").fillna(0).astype(int)
            out.loc[group.index, "input_existing_carrier_count"] = existing_count
            out.loc[group.index, "input_max_supported_carriers"] = np.maximum(existing_count + 1, DEFAULT_MAX_SUPPORTED_CARRIERS)
            add_options = out.loc[group.index, "available_bands_to_add"].fillna("").astype(str).str.strip()
            out.loc[group.index, "input_available_bands_to_add"] = add_options.where(add_options.ne(""), "700")
            out.loc[group.index, "recommended_band_to_add"] = out.loc[group.index, "input_available_bands_to_add"].astype(str).str.split(",").str[0]
            out.loc[group.index, "carrier_addition_possible"] = True
            out.loc[group.index, "demo_scenario_reason"] = "All same-sector carriers are congested; no load-balance headroom, add-band is allowed."

        elif scenario == "sector_split_candidate":
            for rank, idx in enumerate(group.index):
                _set_cell_load(out, idx, 91.0 + min(rank, 2) * 2.0, 88.0 + min(rank, 2) * 2.0, rrc_sector_capacity)
            existing_count = pd.to_numeric(out.loc[group.index, "existing_carrier_count"], errors="coerce").fillna(0).astype(int)
            out.loc[group.index, "input_existing_carrier_count"] = existing_count
            out.loc[group.index, "input_max_supported_carriers"] = existing_count
            out.loc[group.index, "input_available_bands_to_add"] = ""
            out.loc[group.index, "carrier_addition_possible"] = False
            out.loc[group.index, "demo_scenario_reason"] = "Multiple congested carriers with carrier limit reached; full RF runner should branch to sector split."

        else:
            for rank, idx in enumerate(group.index):
                _set_cell_load(out, idx, 94.0 + min(rank, 1), 94.0 + min(rank, 1), rrc_sector_capacity)
            existing_count = pd.to_numeric(out.loc[group.index, "existing_carrier_count"], errors="coerce").fillna(0).astype(int)
            out.loc[group.index, "input_existing_carrier_count"] = existing_count
            out.loc[group.index, "input_max_supported_carriers"] = existing_count
            out.loc[group.index, "input_available_bands_to_add"] = ""
            out.loc[group.index, "carrier_addition_possible"] = False
            out.loc[group.index, "demo_scenario_reason"] = "All same-sector carriers are heavily congested and add-carrier is blocked."

        out.loc[group.index, "demo_scenario"] = scenario

    out["proxy_prb_utilization_pct"] = pd.to_numeric(out["proxy_prb_utilization_pct"], errors="coerce").clip(5.0, 99.0).round(3)
    out["proxy_rrc_utilization_pct"] = pd.to_numeric(out["proxy_rrc_utilization_pct"], errors="coerce").clip(5.0, 99.0).round(3)
    out["proxy_rrc_connected_users"] = ((out["proxy_rrc_utilization_pct"] / 100.0) * float(rrc_sector_capacity)).round(3)
    out["estimated_offered_traffic_mbps"] = (
        pd.to_numeric(out["estimated_dl_capacity_mbps"], errors="coerce").fillna(0.1).clip(lower=0.1)
        * (out["proxy_prb_utilization_pct"] / 100.0)
    ).round(3)
    return out


def _first_existing_column(df: pd.DataFrame, candidates: list[str], default: object = "") -> pd.Series:
    for col in candidates:
        if col in df.columns:
            return df[col]
    return pd.Series(default, index=df.index)


def _strict_identity_key(
    site: pd.Series,
    cell_id: pd.Series,
    sector: pd.Series,
    band: pd.Series,
    operator: pd.Series,
) -> pd.Series:
    parts = [
        site.map(_clean_text),
        cell_id.map(_clean_cell_id),
        sector.map(_clean_text),
        band.map(_format_band).map(_clean_text),
        operator.map(_clean_text),
    ]
    return parts[0] + "|" + parts[1] + "|" + parts[2] + "|" + parts[3] + "|" + parts[4]


def _latest_baseline_job(conn, project_id: int) -> str:
    row = conn.execute(
        text(
            """
            SELECT job_id
            FROM lte_prediction_baseline_results
            WHERE project_id = :project_id
            GROUP BY job_id
            ORDER BY MAX(created_at) DESC
            LIMIT 1
            """
        ),
        {"project_id": int(project_id)},
    ).fetchone()
    if not row or not row[0]:
        raise FileNotFoundError(f"No baseline rows found for project_id={project_id}")
    return str(row[0])


def _read_sql(conn, query: str, params: dict[str, Any]) -> pd.DataFrame:
    return pd.read_sql(text(query), conn, params=params)


def fetch_project_inputs(project_id: int, region: str, baseline_job_id: str | None) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame, str]:
    load_dotenv(ML_ROOT / ".env")
    database_url = os.getenv("DATABASE_URL")
    if not database_url:
        raise RuntimeError("DATABASE_URL is missing from ML/.env")

    engine = create_engine(database_url, pool_pre_ping=True)
    with engine.connect() as conn:
        job_id = baseline_job_id or _latest_baseline_job(conn, project_id)
        baseline_df = _read_sql(
            conn,
            """
            SELECT *
            FROM lte_prediction_baseline_results
            WHERE project_id = :project_id AND job_id = :job_id
            """,
            {"project_id": int(project_id), "job_id": job_id},
        )
        geo_df = _read_sql(
            conn,
            """
            SELECT *
            FROM lte_prediction_geo_features
            WHERE project_id = :project_id AND baseline_job_id = :job_id
              AND LOWER(TRIM(region)) = :region
            """,
            {"project_id": int(project_id), "job_id": job_id, "region": str(region).lower()},
        )
        if geo_df.empty:
            geo_df = _read_sql(
                conn,
                """
                SELECT *
                FROM lte_prediction_geo_features
                WHERE project_id = :project_id AND baseline_job_id = :job_id
                """,
                {"project_id": int(project_id), "job_id": job_id},
            )
        site_df = _read_sql(
            conn,
            """
            SELECT *
            FROM site_prediction
            WHERE tbl_project_id = :project_id
            """,
            {"project_id": int(project_id)},
        )
    return baseline_df, geo_df, site_df, job_id


def _prepare_baseline_geo(baseline_df: pd.DataFrame, geo_df: pd.DataFrame) -> pd.DataFrame:
    baseline = baseline_df.copy()
    baseline["Node_Cell_ID"] = _first_non_empty_series(
        baseline,
        ["rf_identity_key", "site_sector_band_key", "nodeb_id_cell_id", "legacy_nodeb_id_cell_id", "cell_id"],
    )
    if "band" in baseline.columns:
        baseline["baseline_band"] = baseline["band"].map(_format_band)
    else:
        baseline["baseline_band"] = ""
    baseline["topology_match_id"] = baseline["Node_Cell_ID"].map(_canonical_site_cell_id)
    baseline["legacy_nodeb_id_cell_id"] = _first_non_empty_series(
        baseline,
        ["legacy_nodeb_id_cell_id", "nodeb_id_cell_id", "cell_id"],
    )
    baseline["lat_6dp_join"] = pd.to_numeric(baseline.get("lat_6dp", baseline["lat"]), errors="coerce").round(6)
    baseline["lon_6dp_join"] = pd.to_numeric(baseline.get("lon_6dp", baseline["lon"]), errors="coerce").round(6)

    geo = geo_df.copy()
    geo["Node_Cell_ID"] = geo.get("nodeb_id_cell_id", "").map(_clean_cell_id)
    geo["lat_6dp_join"] = pd.to_numeric(geo["lat"], errors="coerce").round(6)
    geo["lon_6dp_join"] = pd.to_numeric(geo["lon"], errors="coerce").round(6)
    geo_keep = [
        col
        for col in [
            "Node_Cell_ID",
            "lat_6dp_join",
            "lon_6dp_join",
            "grid_id",
            "clutter_class",
            "morphology_cluster",
            "building_count",
            "building_area_ratio",
            "road_length_m",
            "green_ratio",
            "water_ratio",
            "nlos_flag",
            "terrain_elevation_m",
            "terrain_slope_deg",
            "site_count_250m",
            "site_count_500m",
            "serving_distance_m",
            "nearest_site_distance_m",
            "mean_nearest3_site_distance_m",
            "azimuth_delta_deg",
        ]
        if col in geo.columns
    ]
    geo = geo[geo_keep].drop_duplicates(subset=["Node_Cell_ID", "lat_6dp_join", "lon_6dp_join"], keep="last")
    merged = baseline.merge(geo, on=["Node_Cell_ID", "lat_6dp_join", "lon_6dp_join"], how="left", validate="many_to_one")
    if "grid_id" not in merged.columns or merged["grid_id"].isna().all():
        merged["grid_id"] = (
            merged["lat_6dp_join"].astype(str)
            + "_"
            + merged["lon_6dp_join"].astype(str)
        )
    return merged


def _prepare_site(site_df: pd.DataFrame) -> pd.DataFrame:
    site = site_df.copy()
    site["site"] = _first_existing_column(site, ["site", "nodeb_id", "site_id"]).map(_clean_text)
    site["cell_id"] = _first_existing_column(site, ["cell_id", "cell"]).map(_clean_cell_id)
    site["sector"] = _first_existing_column(site, ["sector", "sector_id"]).map(_clean_text)
    site["operator"] = _first_existing_column(site, ["operator", "provider", "operator_name", "cluster"]).map(_clean_text)
    site["band_num"] = pd.to_numeric(_first_existing_column(site, ["band", "frequency"]), errors="coerce")
    site["band"] = site["band_num"].map(_format_band).map(_clean_text)
    complete_mask = (
        site["site"].ne("")
        & site["cell_id"].ne("")
        & site["sector"].ne("")
        & site["band"].ne("")
        & site["operator"].ne("")
    )
    site = site.loc[complete_mask].copy()
    if "site_cell_sector_band_operator_key" in site.columns:
        strict_from_db = site["site_cell_sector_band_operator_key"].map(_clean_text)
    elif "site_prediction_key" in site.columns:
        strict_from_db = site["site_prediction_key"].map(_clean_text)
    else:
        strict_from_db = pd.Series("", index=site.index, dtype="object")
    constructed_key = _strict_identity_key(
        site["site"],
        site["cell_id"],
        site["sector"],
        site["band"],
        site["operator"],
    )
    site["Node_Cell_ID"] = strict_from_db.where(strict_from_db.ne(""), constructed_key)
    site["canonical_physical_cell_id"] = site["cell_id"].map(_canonical_site_cell_id)
    site["topology_match_id"] = site["canonical_physical_cell_id"]
    site["sector_id"] = site["site"].astype(str) + "|" + site["sector"].astype(str)
    site["earfcn_num"] = pd.to_numeric(site.get("earfcn", np.nan), errors="coerce")
    site["bw_mhz"] = pd.to_numeric(site.get("bw", np.nan), errors="coerce").fillna(10.0).clip(lower=1.4)
    return site.drop_duplicates(subset=["Node_Cell_ID"], keep="first").reset_index(drop=True)


def _build_cell_input(
    baseline_geo: pd.DataFrame,
    site_df: pd.DataFrame,
    project_id: int,
    baseline_job_id: str,
    grid_size_m: float,
    rrc_sector_capacity: float,
) -> pd.DataFrame:
    site = _prepare_site(site_df)
    baseline = baseline_geo.copy()
    baseline["pred_rsrp"] = pd.to_numeric(baseline["pred_rsrp"], errors="coerce")
    baseline["pred_rsrq"] = pd.to_numeric(baseline["pred_rsrq"], errors="coerce")
    baseline["pred_sinr"] = pd.to_numeric(baseline["pred_sinr"], errors="coerce")
    baseline["interference_gap_db"] = pd.to_numeric(baseline.get("interference_gap_db"), errors="coerce")
    baseline["same_earfcn_interferer_count"] = pd.to_numeric(baseline.get("same_earfcn_interferer_count"), errors="coerce")
    baseline["grid_id"] = baseline["grid_id"].astype(str)

    grouped = baseline.groupby("Node_Cell_ID", dropna=False)
    cell = grouped.agg(
        grid_count=("grid_id", "nunique"),
        point_count=("lat", "count"),
        rsrp_mean=("pred_rsrp", "mean"),
        rsrp_p10=("pred_rsrp", lambda s: s.quantile(0.10)),
        rsrq_mean=("pred_rsrq", "mean"),
        sinr_mean=("pred_sinr", "mean"),
        sinr_p10=("pred_sinr", lambda s: s.quantile(0.10)),
        weak_rsrp_pct=("pred_rsrp", lambda s: float((pd.to_numeric(s, errors="coerce") < -110.0).mean() * 100.0)),
        bad_sinr_pct=("pred_sinr", lambda s: float((pd.to_numeric(s, errors="coerce") < 0.0).mean() * 100.0)),
        interference_gap_p10=("interference_gap_db", lambda s: s.quantile(0.10)),
        same_earfcn_interferer_mean=("same_earfcn_interferer_count", "mean"),
        baseline_band=("baseline_band", lambda s: _format_list(s.dropna().tolist())),
        site_count_500m_mean=("site_count_500m", "mean") if "site_count_500m" in baseline.columns else ("pred_rsrp", "count"),
        serving_distance_p90=("serving_distance_m", lambda s: s.quantile(0.90)) if "serving_distance_m" in baseline.columns else ("pred_rsrp", "count"),
    ).reset_index()
    cell["canonical_physical_cell_id"] = cell["Node_Cell_ID"].map(_canonical_site_cell_id)

    site_keep = [
        col
        for col in [
            "topology_match_id",
            "site",
            "site_name",
            "sector_id",
            "sector",
            "operator",
            "band",
            "earfcn_num",
            "bw_mhz",
            "latitude",
            "longitude",
            "azimuth",
            "height",
            "m_tilt",
            "e_tilt",
            "tx_power",
            "Technology",
        ]
        if col in site.columns
    ]
    cell["topology_match_id"] = cell["Node_Cell_ID"].map(_canonical_site_cell_id)
    site_dedup = site[site_keep + ["Node_Cell_ID"] if "Node_Cell_ID" not in site_keep else site_keep].drop_duplicates(subset=["Node_Cell_ID"], keep="first")
    cell = cell.merge(site_dedup, on="Node_Cell_ID", how="left", validate="many_to_one")
    cell["site_id"] = cell["site"].fillna(cell["Node_Cell_ID"].astype(str).str.split("_").str[0])
    cell["sector_id"] = cell["sector_id"].fillna(cell["site_id"].astype(str) + "|" + cell.get("sector", pd.Series("", index=cell.index)).astype(str))
    cell["earfcn"] = cell["earfcn_num"].map(_format_band)

    cell_band = (
        site.groupby("Node_Cell_ID", dropna=False)
        .agg(
            site_bands_for_cell=("band_num", lambda s: _format_list(s.dropna().tolist())),
            site_band_count_for_cell=("band_num", lambda s: int(pd.to_numeric(s, errors="coerce").dropna().nunique())),
        )
        .reset_index()
    )
    cell = cell.merge(cell_band, on="Node_Cell_ID", how="left", validate="many_to_one")
    cell["band"] = cell["baseline_band"].fillna("").astype(str).str.strip()
    fallback_band = cell["site_bands_for_cell"].fillna("").astype(str).str.strip()
    cell["band"] = cell["band"].where(cell["band"].ne(""), fallback_band)

    # Carrier/add-band logic must use the full site topology inventory for the sector.
    # The row's own `band` comes from the saved baseline when present; site inventory
    # remains for existing-carrier and add-band choices.
    sector_band = (
        site.groupby("sector_id", dropna=False)
        .agg(
            existing_carriers=("band_num", lambda s: _format_list(s.dropna().tolist())),
            existing_carrier_count=("band_num", lambda s: int(pd.to_numeric(s, errors="coerce").dropna().nunique())),
        )
        .reset_index()
    )
    existing_map = {
        row["sector_id"]: set(pd.to_numeric(pd.Series(str(row["existing_carriers"]).split(",")), errors="coerce").dropna().astype(float).tolist())
        for _, row in sector_band.iterrows()
    }
    sector_band["available_bands_to_add"] = [
        _format_list([band for band in KNOWN_BAND_POOL_MHZ if band not in existing_map.get(row["sector_id"], set())])
        for _, row in sector_band.iterrows()
    ]
    sector_band["max_supported_carriers"] = DEFAULT_MAX_SUPPORTED_CARRIERS
    sector_band["carrier_addition_possible"] = sector_band["existing_carrier_count"] < sector_band["max_supported_carriers"]
    cell = cell.merge(sector_band, on="sector_id", how="left", validate="many_to_one")

    burden_score = (
        0.30 * _minmax(cell["grid_count"])
        + 0.20 * _minmax(cell["weak_rsrp_pct"])
        + 0.20 * _minmax(cell["bad_sinr_pct"])
        + 0.15 * (1.0 - _minmax(cell["interference_gap_p10"]))
        + 0.15 * _minmax(cell["same_earfcn_interferer_mean"])
    ).clip(0.0, 1.0)
    cell["proxy_prb_utilization_pct"] = (35.0 + 62.0 * burden_score).clip(5.0, 97.0).round(3)
    cell["proxy_rrc_utilization_pct"] = (25.0 + 75.0 * _minmax(cell["grid_count"])).clip(5.0, 100.0).round(3)
    cell["proxy_rrc_connected_users"] = ((cell["proxy_rrc_utilization_pct"] / 100.0) * float(rrc_sector_capacity)).round(3)
    sinr_linear = np.power(10.0, pd.to_numeric(cell["sinr_mean"], errors="coerce").fillna(0.0) / 10.0)
    spectral_efficiency = (0.65 * np.log2(1.0 + sinr_linear)).clip(0.15, 6.0)
    cell["estimated_dl_capacity_mbps"] = (
        pd.to_numeric(cell.get("bw_mhz", 10.0), errors="coerce").fillna(10.0).clip(lower=1.4)
        * spectral_efficiency
        * 2.0
        * 0.75
    ).clip(lower=0.1).round(3)
    cell["estimated_offered_traffic_mbps"] = (
        cell["estimated_dl_capacity_mbps"] * (cell["proxy_prb_utilization_pct"] / 100.0)
    ).round(3)

    # Editable override columns. Model 3 should prefer these when the planner fills them.
    cell["input_prb_utilization_pct"] = cell["proxy_prb_utilization_pct"]
    cell["input_rrc_utilization_pct"] = cell["proxy_rrc_utilization_pct"]
    cell["input_rrc_connected_users"] = cell["proxy_rrc_connected_users"]
    cell["input_estimated_dl_capacity_mbps"] = cell["estimated_dl_capacity_mbps"]
    cell["input_estimated_offered_traffic_mbps"] = cell["estimated_offered_traffic_mbps"]
    cell["input_available_bands_to_add"] = cell["available_bands_to_add"]
    cell["recommended_band_to_add"] = cell["input_available_bands_to_add"].fillna("").astype(str).str.split(",").str[0].replace("nan", "")
    cell["input_max_supported_carriers"] = cell["max_supported_carriers"]
    cell["input_existing_carrier_count"] = cell["existing_carrier_count"]
    cell["input_sector_rrc_capacity"] = float(rrc_sector_capacity)
    cell = _apply_model3_demo_scenarios(cell, rrc_sector_capacity)
    cell["input_prb_utilization_pct"] = cell["proxy_prb_utilization_pct"]
    cell["input_rrc_utilization_pct"] = cell["proxy_rrc_utilization_pct"]
    cell["input_rrc_connected_users"] = cell["proxy_rrc_connected_users"]
    cell["input_estimated_offered_traffic_mbps"] = cell["estimated_offered_traffic_mbps"]
    cell["input_notes"] = ""
    cell["metric_source"] = "proxy_from_project196_baseline_geo_editable_in_excel"
    cell["project_id"] = int(project_id)
    cell["baseline_job_id"] = baseline_job_id
    cell["grid_size_m"] = float(grid_size_m)

    ordered = [
        "project_id",
        "baseline_job_id",
        "grid_size_m",
        "site_id",
        "sector_id",
        "Node_Cell_ID",
        "canonical_physical_cell_id",
        "band",
        "earfcn",
        "grid_count",
        "point_count",
        "input_prb_utilization_pct",
        "input_rrc_utilization_pct",
        "input_rrc_connected_users",
        "input_estimated_dl_capacity_mbps",
        "input_estimated_offered_traffic_mbps",
        "input_available_bands_to_add",
        "recommended_band_to_add",
        "input_max_supported_carriers",
        "input_existing_carrier_count",
        "input_sector_rrc_capacity",
        "demo_scenario",
        "demo_scenario_reason",
        "proxy_prb_utilization_pct",
        "proxy_rrc_utilization_pct",
        "proxy_rrc_connected_users",
        "estimated_dl_capacity_mbps",
        "estimated_offered_traffic_mbps",
        "existing_carriers",
        "existing_carrier_count",
        "available_bands_to_add",
        "baseline_band",
        "site_bands_for_cell",
        "site_band_count_for_cell",
        "carrier_addition_possible",
        "rsrp_mean",
        "rsrp_p10",
        "rsrq_mean",
        "sinr_mean",
        "sinr_p10",
        "weak_rsrp_pct",
        "bad_sinr_pct",
        "interference_gap_p10",
        "same_earfcn_interferer_mean",
        "site_count_500m_mean",
        "serving_distance_p90",
        "latitude",
        "longitude",
        "azimuth",
        "height",
        "m_tilt",
        "e_tilt",
        "tx_power",
        "metric_source",
        "input_notes",
    ]
    ordered = [col for col in ordered if col in cell.columns]
    return cell[ordered].sort_values(["site_id", "sector_id", "Node_Cell_ID"]).reset_index(drop=True)


def _write_excel(path: Path, sheets: dict[str, pd.DataFrame], summary: dict[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        pd.DataFrame(
            [
                {"field": key, "value": value}
                for key, value in summary.items()
            ]
        ).to_excel(writer, sheet_name="README", index=False)
        for sheet_name, df in sheets.items():
            df.to_excel(writer, sheet_name=sheet_name[:31], index=False)

    wb = load_workbook(path)
    header_fill = PatternFill("solid", fgColor="1F4E78")
    editable_fill = PatternFill("solid", fgColor="FFF2CC")
    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = header_fill
        for idx, cell in enumerate(ws[1], start=1):
            if str(cell.value).startswith("input_"):
                for row in ws.iter_rows(min_col=idx, max_col=idx, min_row=2, max_row=ws.max_row):
                    row[0].fill = editable_fill
        ws.auto_filter.ref = ws.dimensions
        for col in ws.columns:
            max_len = max(len(str(c.value)) if c.value is not None else 0 for c in col[:200])
            ws.column_dimensions[col[0].column_letter].width = min(max(12, max_len + 2), 45)
    wb.save(path)


def build_project196_excel(
    project_id: int = DEFAULT_PROJECT_ID,
    region: str = DEFAULT_REGION,
    baseline_job_id: str | None = None,
    grid_size_m: float = DEFAULT_GRID_SIZE_M,
    rrc_sector_capacity: float = DEFAULT_RRC_SECTOR_CAPACITY,
) -> dict[str, Any]:
    baseline_df, geo_df, site_df, job_id = fetch_project_inputs(project_id, region, baseline_job_id)
    baseline_geo = _prepare_baseline_geo(baseline_df, geo_df)
    cell_input = _build_cell_input(
        baseline_geo,
        site_df,
        project_id=project_id,
        baseline_job_id=job_id,
        grid_size_m=grid_size_m,
        rrc_sector_capacity=rrc_sector_capacity,
    )
    sector_input = (
        cell_input.groupby(["site_id", "sector_id"], dropna=False, as_index=False)
        .agg(
            cell_count=("Node_Cell_ID", "nunique"),
            sector_grid_count=("grid_count", "sum"),
            max_prb_pct=("input_prb_utilization_pct", "max"),
            max_rrc_pct=("input_rrc_utilization_pct", "max"),
            total_rrc_users=("input_rrc_connected_users", "sum"),
            existing_carriers=("existing_carriers", "first"),
            available_bands_to_add=("input_available_bands_to_add", "first"),
            carrier_addition_possible=("carrier_addition_possible", "any"),
        )
        .sort_values(["max_prb_pct", "max_rrc_pct"], ascending=[False, False])
    )
    baseline_grid_input = baseline_geo.copy()
    baseline_grid_cols = [
        "project_id",
        "job_id",
        "Node_Cell_ID",
        "legacy_nodeb_id_cell_id",
        "topology_match_id",
        "baseline_band",
        "band",
        "sector",
        "site_id",
        "node_b_id",
        "cell_id",
        "operator",
        "Technology",
        "grid_id",
        "lat",
        "lat_6dp",
        "lon",
        "lon_6dp",
        "pred_rsrp",
        "pred_rsrq",
        "pred_sinr",
        "pred_rsrp_smoothed",
        "pred_rsrq_smoothed",
        "pred_sinr_smoothed",
        "clutter_class",
        "morphology_cluster",
        "building_count",
        "building_area_ratio",
        "road_length_m",
        "green_ratio",
        "water_ratio",
        "nlos_flag",
        "terrain_elevation_m",
        "terrain_slope_deg",
        "site_count_250m",
        "site_count_500m",
        "serving_distance_m",
        "nearest_site_distance_m",
        "mean_nearest3_site_distance_m",
        "azimuth_delta_deg",
    ]
    baseline_grid_cols = [col for col in baseline_grid_cols if col in baseline_grid_input.columns]
    baseline_grid_input = baseline_grid_input[baseline_grid_cols].sort_values(["Node_Cell_ID", "grid_id", "lat", "lon"]).reset_index(drop=True)

    site_input_cols = [
        "project_id",
        "baseline_job_id",
        "grid_size_m",
        "site_id",
        "sector_id",
        "Node_Cell_ID",
        "canonical_physical_cell_id",
        "band",
        "earfcn",
        "latitude",
        "longitude",
        "azimuth",
        "height",
        "m_tilt",
        "e_tilt",
        "tx_power",
        "grid_count",
        "point_count",
        "existing_carriers",
        "existing_carrier_count",
        "available_bands_to_add",
        "carrier_addition_possible",
    ]
    site_input_cols = [col for col in site_input_cols if col in cell_input.columns]
    site_input = cell_input[site_input_cols].copy().sort_values(["site_id", "sector_id", "Node_Cell_ID"]).reset_index(drop=True)

    geo_features_input = geo_df.copy()
    geo_sort_cols = [col for col in ["nodeb_id_cell_id", "grid_id", "lat", "lon"] if col in geo_features_input.columns]
    if geo_sort_cols:
        geo_features_input = geo_features_input.sort_values(geo_sort_cols).reset_index(drop=True)

    baseline_band_null_rows = int(baseline_df["band"].isna().sum()) if "band" in baseline_df.columns else None
    site_band_values = (
        pd.to_numeric(site_df.get("band", site_df.get("frequency", pd.Series(dtype=float))), errors="coerce")
        .dropna()
        .map(_format_band)
        .drop_duplicates()
        .sort_values()
        .tolist()
    )
    summary = {
        "project_id": int(project_id),
        "region": region,
        "baseline_job_id": job_id,
        "grid_size_m": float(grid_size_m),
        "baseline_rows": int(len(baseline_df)),
        "geo_rows": int(len(geo_df)),
        "site_rows": int(len(site_df)),
        "baseline_grid_rows": int(len(baseline_grid_input)),
        "geo_features_input_rows": int(len(geo_features_input)),
        "site_input_rows": int(len(site_input)),
        "model3_cell_rows": int(len(cell_input)),
        "model3_sector_rows": int(len(sector_input)),
        "baseline_band_null_rows": baseline_band_null_rows,
        "site_band_values": site_band_values,
        "rrc_sector_capacity_default": float(rrc_sector_capacity),
        "important_note": (
            "Project 196 DB baseline/geo tables do not contain real PRB/RRC counters. "
            "input_* columns are editable and default to clearly marked planning proxies."
        ),
    }
    OUTPUT_ROOT.mkdir(parents=True, exist_ok=True)
    excel_path = OUTPUT_ROOT / f"project_{project_id}_model3_input.xlsx"
    csv_path = OUTPUT_ROOT / f"project_{project_id}_model3_cell_input.csv"
    sector_csv_path = OUTPUT_ROOT / f"project_{project_id}_model3_sector_input.csv"
    baseline_grid_csv_path = OUTPUT_ROOT / f"project_{project_id}_model3_baseline_grid_input.csv"
    geo_features_csv_path = OUTPUT_ROOT / f"project_{project_id}_model3_geo_features_input.csv"
    site_input_csv_path = OUTPUT_ROOT / f"project_{project_id}_model3_site_input.csv"
    summary_path = OUTPUT_ROOT / f"project_{project_id}_model3_input_summary.json"
    cell_input.to_csv(csv_path, index=False)
    sector_input.to_csv(sector_csv_path, index=False)
    baseline_grid_input.to_csv(baseline_grid_csv_path, index=False)
    geo_features_input.to_csv(geo_features_csv_path, index=False)
    site_input.to_csv(site_input_csv_path, index=False)
    summary["excel_path"] = str(excel_path)
    summary["cell_input_csv"] = str(csv_path)
    summary["sector_input_csv"] = str(sector_csv_path)
    summary["baseline_grid_input_csv"] = str(baseline_grid_csv_path)
    summary["geo_features_input_csv"] = str(geo_features_csv_path)
    summary["site_input_csv"] = str(site_input_csv_path)
    summary["summary_json"] = str(summary_path)
    sheets = {
        "Model3_Cell_Input": cell_input,
        "Model3_Sector_Summary": sector_input,
        "Baseline_Grid_Input": baseline_grid_input,
        "Geo_Features_Input": geo_features_input,
        "Model3_Site_Input": site_input,
    }
    try:
        _write_excel(excel_path, sheets, summary)
    except PermissionError:
        excel_path = OUTPUT_ROOT / f"project_{project_id}_model3_input_{str(job_id)[:8]}_full_grid.xlsx"
        summary["excel_path"] = str(excel_path)
        _write_excel(excel_path, sheets, summary)
    summary_path.write_text(json.dumps(summary, indent=2, default=str), encoding="utf-8")
    return summary


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Build project 196 Model 3 Excel input from DB baseline/geo/site rows.")
    parser.add_argument("--project-id", type=int, default=DEFAULT_PROJECT_ID)
    parser.add_argument("--region", type=str, default=DEFAULT_REGION)
    parser.add_argument("--baseline-job-id", type=str, default=None)
    parser.add_argument("--grid-size-m", type=float, default=DEFAULT_GRID_SIZE_M)
    parser.add_argument("--rrc-sector-capacity", type=float, default=DEFAULT_RRC_SECTOR_CAPACITY)
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    summary = build_project196_excel(
        project_id=args.project_id,
        region=args.region,
        baseline_job_id=args.baseline_job_id,
        grid_size_m=args.grid_size_m,
        rrc_sector_capacity=args.rrc_sector_capacity,
    )
    print(json.dumps(summary, indent=2, default=str))


if __name__ == "__main__":
    main()
