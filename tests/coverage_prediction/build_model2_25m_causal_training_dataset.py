from __future__ import annotations

import argparse
import csv
import json
import math
from pathlib import Path
from typing import Dict, Iterable, List

import numpy as np
import pandas as pd
from openpyxl import Workbook


BUCKET_ORDER = {"PART_1": 1, "PART_2": 2, "PART_3": 3} 
CLUTTER_DEMAND_WEIGHT = {
    "Dense Urban": 1.0,
    "Urban": 0.78,
    "Suburban": 0.46,
    "Rural/Open": 0.20,
    "Vegetation": 0.18,
    "Water": 0.04,
}
CLUTTER_TRANSITION_LEVEL = {
    "Water": 0,
    "Vegetation": 0,
    "Rural/Open": 0,
    "Suburban": 1,
    "Urban": 2,
    "Dense Urban": 3,
}
BAND_CLASS_LABELS = ("LOW_BAND", "MID_BAND", "HIGH_BAND")
MAX_EXCEL_DATA_ROWS = 1_000_000


def _safe_numeric(df: pd.DataFrame, col: str, default: float = 0.0) -> pd.Series:
    if col not in df.columns:
        return pd.Series(default, index=df.index, dtype="float64")
    return pd.to_numeric(df[col], errors="coerce").fillna(default).astype("float64")


def _robust_norm(series: pd.Series) -> pd.Series:
    values = pd.to_numeric(series, errors="coerce").replace([np.inf, -np.inf], np.nan)
    valid = values.dropna()
    if valid.empty:
        return pd.Series(0.0, index=series.index, dtype="float64")
    lo = float(valid.quantile(0.05))
    hi = float(valid.quantile(0.95))
    if not math.isfinite(lo) or not math.isfinite(hi) or hi <= lo:
        lo = float(valid.min())
        hi = float(valid.max())
    if hi <= lo:
        return pd.Series(0.0, index=series.index, dtype="float64")
    return ((values.fillna(lo) - lo) / (hi - lo)).clip(0.0, 1.0)


def _ratio_vs_first_bucket(df: pd.DataFrame, value_col: str) -> pd.Series:
    work = df.sort_values(["grid_id", "bucket_seq"]).copy()
    current = pd.to_numeric(work[value_col], errors="coerce")
    baseline = pd.to_numeric(work.groupby("grid_id", sort=False)[value_col].transform("first"), errors="coerce")
    denom = baseline.abs().replace(0, np.nan)
    ratio = (current - baseline) / denom
    cold_start = baseline.fillna(0.0).eq(0.0) & current.fillna(0.0).gt(0.0)
    ratio = ratio.where(~cold_start, 1.0)
    return ratio.replace([np.inf, -np.inf], np.nan).fillna(0.0).clip(-1.0, 5.0)


def _classify_band(freq_value: object) -> str:
    try:
        freq = float(freq_value)
    except (TypeError, ValueError):
        return "UNKNOWN"
    if not math.isfinite(freq) or freq <= 0:
        return "UNKNOWN"
    if freq <= 1000:
        return "LOW_BAND"
    if freq <= 2100:
        return "MID_BAND"
    return "HIGH_BAND"


def _spectral_efficiency_from_sinr(sinr: pd.Series) -> pd.Series:
    sinr_linear = np.power(10.0, pd.to_numeric(sinr, errors="coerce").fillna(0.0).clip(-10, 30) / 10.0)
    shannon = np.log2(1.0 + sinr_linear)
    return (0.62 * shannon).clip(0.35, 4.8)


def _aggregate_baseline_rf(baseline_csv: Path, chunksize: int = 300_000) -> pd.DataFrame:
    usecols = [
        "time_bucket",
        "grid_id",
        "Node_Cell_ID",
        "rf_identity_key",
        "site_sector_band_key",
        "node_cell_sector_key",
        "sector_identity_key",
        "pred_rsrp",
        "pred_rsrq",
        "pred_sinr",
        "band",
        "earfcn",
        "Site ID",
        "sector_identity",
        "canonical_sector_id",
        "carrier_load_share",
    ]
    grouped_parts: List[pd.DataFrame] = []
    for chunk in pd.read_csv(baseline_csv, usecols=usecols, chunksize=chunksize):
        chunk["time_bucket"] = chunk["time_bucket"].astype(str)
        chunk["grid_id"] = pd.to_numeric(chunk["grid_id"], errors="coerce").astype("Int64")
        chunk["Node_Cell_ID"] = chunk["Node_Cell_ID"].astype(str)
        for col in ["pred_rsrp", "pred_rsrq", "pred_sinr", "band", "earfcn", "carrier_load_share", "Site ID"]:
            chunk[col] = pd.to_numeric(chunk[col], errors="coerce")
        chunk["serving_cell_key"] = (
            chunk["Site ID"].round(0).astype("Int64").astype(str).replace({"<NA>": ""})
            + "|"
            + chunk["sector_identity"].astype(str).replace({"nan": "", "None": ""})
            + "|"
            + chunk["Node_Cell_ID"].astype(str).replace({"nan": "", "None": ""})
            + "|"
            + chunk["band"].round(0).astype("Int64").astype(str).replace({"<NA>": ""})
        )
        grouped_parts.append(
            chunk.groupby(["time_bucket", "grid_id", "serving_cell_key"], dropna=False).agg(
                rsrp_mean=("pred_rsrp", "mean"),
                rsrp_max=("pred_rsrp", "max"),
                rsrq_mean=("pred_rsrq", "mean"),
                sinr_mean=("pred_sinr", "mean"),
                sample_count=("pred_rsrp", "size"),
                Node_Cell_ID=("Node_Cell_ID", "first"),
                rf_identity_key=("rf_identity_key", "first"),
                site_sector_band_key=("site_sector_band_key", "first"),
                node_cell_sector_key=("node_cell_sector_key", "first"),
                sector_identity_key=("sector_identity_key", "first"),
                band=("band", "first"),
                earfcn=("earfcn", "first"),
                site_id=("Site ID", "first"),
                sector_identity=("sector_identity", "first"),
                canonical_sector_id=("canonical_sector_id", "first"),
                carrier_load_share=("carrier_load_share", "mean"),
            ).reset_index()
        )
    combined = pd.concat(grouped_parts, ignore_index=True)
    return (
        combined.groupby(["time_bucket", "grid_id", "serving_cell_key"], dropna=False)
        .agg(
            rsrp_mean=("rsrp_mean", "mean"),
            rsrp_max=("rsrp_max", "max"),
            rsrq_mean=("rsrq_mean", "mean"),
            sinr_mean=("sinr_mean", "mean"),
            sample_count=("sample_count", "sum"),
            Node_Cell_ID=("Node_Cell_ID", "first"),
            rf_identity_key=("rf_identity_key", "first"),
            site_sector_band_key=("site_sector_band_key", "first"),
            node_cell_sector_key=("node_cell_sector_key", "first"),
            sector_identity_key=("sector_identity_key", "first"),
            band=("band", "first"),
            earfcn=("earfcn", "first"),
            site_id=("site_id", "first"),
            sector_identity=("sector_identity", "first"),
            canonical_sector_id=("canonical_sector_id", "first"),
            carrier_load_share=("carrier_load_share", "mean"),
        )
        .reset_index()
    )


def _grid_rf_features(cell_grid_df: pd.DataFrame) -> pd.DataFrame:
    work = cell_grid_df.copy()
    work["band_class"] = work["band"].map(_classify_band)
    work = work.sort_values(["time_bucket", "grid_id", "rsrp_mean"], ascending=[True, True, False])
    work["rank"] = work.groupby(["time_bucket", "grid_id"]).cumcount() + 1
    serving = work.loc[work["rank"] == 1].copy()
    second = work.loc[work["rank"] == 2, ["time_bucket", "grid_id", "rsrp_mean", "sinr_mean", "band", "Node_Cell_ID", "serving_cell_key"]].rename(
        columns={
            "rsrp_mean": "neighbor1_rsrp",
            "sinr_mean": "neighbor1_sinr",
            "band": "neighbor1_band",
            "Node_Cell_ID": "neighbor1_cell_id",
            "serving_cell_key": "neighbor1_cell_key",
        }
    )
    third = work.loc[work["rank"] == 3, ["time_bucket", "grid_id", "rsrp_mean", "Node_Cell_ID", "serving_cell_key"]].rename(
        columns={"rsrp_mean": "neighbor2_rsrp", "Node_Cell_ID": "neighbor2_cell_id", "serving_cell_key": "neighbor2_cell_key"}
    )
    grid_summary = work.groupby(["time_bucket", "grid_id"], as_index=False).agg(
        candidate_cell_count=("serving_cell_key", "nunique"),
        total_prediction_samples=("sample_count", "sum"),
        mean_candidate_rsrp=("rsrp_mean", "mean"),
        mean_candidate_sinr=("sinr_mean", "mean"),
    )
    band_counts = (
        work.groupby(["time_bucket", "grid_id", "band_class"])["sample_count"]
        .sum()
        .unstack(fill_value=0.0)
        .reset_index()
    )
    for label in BAND_CLASS_LABELS:
        if label not in band_counts.columns:
            band_counts[label] = 0.0
    band_total = band_counts[list(BAND_CLASS_LABELS)].sum(axis=1).replace(0, np.nan)
    band_counts["low_band_ratio"] = (band_counts["LOW_BAND"] / band_total).fillna(0.0)
    band_counts["mid_band_ratio"] = (band_counts["MID_BAND"] / band_total).fillna(0.0)
    band_counts["high_band_ratio"] = (band_counts["HIGH_BAND"] / band_total).fillna(0.0)
    band_counts["carrier_count"] = (
        (band_counts["LOW_BAND"] > 0).astype(int)
        + (band_counts["MID_BAND"] > 0).astype(int)
        + (band_counts["HIGH_BAND"] > 0).astype(int)
    )
    band_counts["dominant_band_class"] = band_counts[["low_band_ratio", "mid_band_ratio", "high_band_ratio"]].idxmax(axis=1)
    band_counts["dominant_band_class"] = band_counts["dominant_band_class"].map(
        {"low_band_ratio": "LOW_BAND", "mid_band_ratio": "MID_BAND", "high_band_ratio": "HIGH_BAND"}
    )

    keep_serving = serving[
        [
            "time_bucket",
            "grid_id",
            "serving_cell_key",
            "Node_Cell_ID",
            "rf_identity_key",
            "site_sector_band_key",
            "node_cell_sector_key",
            "sector_identity_key",
            "site_id",
            "sector_identity",
            "canonical_sector_id",
            "band",
            "earfcn",
            "rsrp_mean",
            "rsrp_max",
            "rsrq_mean",
            "sinr_mean",
            "sample_count",
        ]
    ].rename(
        columns={
            "serving_cell_key": "serving_cell_key",
            "Node_Cell_ID": "serving_cell_id",
            "rf_identity_key": "serving_rf_identity_key",
            "site_sector_band_key": "serving_site_sector_band_key",
            "node_cell_sector_key": "serving_node_cell_sector_key",
            "sector_identity_key": "serving_sector_identity_key",
            "site_id": "serving_site_id",
            "sector_identity": "serving_sector",
            "canonical_sector_id": "serving_canonical_sector",
            "band": "serving_band",
            "earfcn": "serving_earfcn",
            "rsrp_mean": "serving_rsrp_mean",
            "rsrp_max": "serving_rsrp_max",
            "rsrq_mean": "serving_rsrq_mean",
            "sinr_mean": "serving_sinr_mean",
            "sample_count": "serving_sample_count",
        }
    )
    out = keep_serving.merge(second, on=["time_bucket", "grid_id"], how="left")
    out = out.merge(third, on=["time_bucket", "grid_id"], how="left")
    out = out.merge(grid_summary, on=["time_bucket", "grid_id"], how="left")
    out = out.merge(
        band_counts[
            [
                "time_bucket",
                "grid_id",
                "low_band_ratio",
                "mid_band_ratio",
                "high_band_ratio",
                "dominant_band_class",
                "carrier_count",
            ]
        ],
        on=["time_bucket", "grid_id"],
        how="left",
    )
    out["rsrp_gap_to_neighbor1"] = out["serving_rsrp_mean"] - out["neighbor1_rsrp"]
    out["neighbor_interference_index"] = np.power(10.0, (out["neighbor1_rsrp"].fillna(-140.0) - out["serving_rsrp_mean"]) / 10.0)
    return out


def _add_demand_and_cell_prb(grid_df: pd.DataFrame) -> pd.DataFrame:
    out = grid_df.copy()
    out["bucket_seq"] = out["time_bucket"].astype(str).map(BUCKET_ORDER).fillna(0).astype(int)
    for col in [
        "building_count",
        "building_area_ratio",
        "road_density",
        "road_length_m",
        "park_open_area",
        "open_area_ratio",
        "mall_presence",
        "metro_presence",
        "serving_sinr_mean",
        "serving_rsrp_mean",
        "candidate_cell_count",
        "carrier_count",
    ]:
        out[col] = _safe_numeric(out, col, 0.0)
    out["clutter_class"] = out["clutter_class"].fillna("UNKNOWN").astype(str)
    clutter_weight = out["clutter_class"].map(CLUTTER_DEMAND_WEIGHT).fillna(0.35).astype(float)
    building_count_norm = _robust_norm(out["building_count"])
    building_area_norm = _robust_norm(out["building_area_ratio"])
    road_norm = _robust_norm(out["road_density"].combine_first(out["road_length_m"]))
    open_space_discount = (0.45 * _robust_norm(out["park_open_area"])) + (0.30 * _robust_norm(out["open_area_ratio"]))
    activity_anchor = (
        ((out["mall_presence"] > 0).astype(float) * 0.45)
        + ((out["metro_presence"] > 0).astype(float) * 0.35)
        + ((road_norm >= 0.60).astype(float) * 0.20)
    ).clip(0.0, 1.0)
    rf_pressure = (
        ((-92.0 - out["serving_rsrp_mean"]) / 18.0).clip(0.0, 1.0) * 0.55
        + ((8.0 - out["serving_sinr_mean"]) / 18.0).clip(0.0, 1.0) * 0.45
    ).clip(0.0, 1.0)
    out["geo_demand_score"] = (
        (
            0.34 * clutter_weight
            + 0.24 * building_area_norm
            + 0.18 * building_count_norm
            + 0.18 * road_norm
            + 0.16 * activity_anchor
            - open_space_discount
        ).clip(0.0, 1.0)
        * 100.0
    ).round(3)
    out["kpi_demand_score"] = (rf_pressure * 100.0).round(3)

    clutter_level = out["clutter_class"].map(CLUTTER_TRANSITION_LEVEL).fillna(0).astype(int)
    baseline_level = (
        out.assign(_level=clutter_level)
        .sort_values(["grid_id", "bucket_seq"])
        .groupby("grid_id", sort=False)["_level"]
        .transform("first")
        .reindex(out.index)
        .fillna(0)
        .astype(int)
    )
    out["clutter_transition_flag"] = clutter_level.ne(baseline_level).astype(int)
    out["clutter_upgrade_score"] = np.where(clutter_level > baseline_level, clutter_level - baseline_level, 0).astype(float)
    out["building_growth_ratio"] = _ratio_vs_first_bucket(out, "building_area_ratio").round(6)
    out["road_growth_ratio"] = _ratio_vs_first_bucket(out, "road_density").round(6)
    positive_building_growth = np.clip(out["building_growth_ratio"], 0.0, 3.0) / 3.0
    positive_road_growth = np.clip(out["road_growth_ratio"], 0.0, 3.0) / 3.0
    out["development_pressure_score"] = (
        (
            0.24 * building_count_norm
            + 0.26 * building_area_norm
            + 0.16 * road_norm
            + 0.14 * activity_anchor
            + 0.12 * positive_building_growth
            + 0.08 * positive_road_growth
        ).clip(0.0, 1.0)
        * 100.0
    ).round(3)
    out["activity_anchor_score"] = (activity_anchor * 100.0).round(3)

    # Heterogeneous growth: stable grids grow mildly, high-pressure/activity grids grow much faster.
    pressure = (out["development_pressure_score"] / 100.0).clip(0.0, 1.0)
    geo_score = (out["geo_demand_score"] / 100.0).clip(0.0, 1.0)
    bucket_step = (out["bucket_seq"] - 1).clip(lower=0)
    hotspot_intensity = (
        0.34 * np.power(pressure, 1.65)
        + 0.26 * np.power(geo_score, 1.45)
        + 0.18 * activity_anchor
        + 0.12 * positive_building_growth
        + 0.10 * positive_road_growth
        + 0.08 * (out["clutter_upgrade_score"] / 3.0).clip(0.0, 1.0)
    ).clip(0.0, 1.0)
    base_demand = (0.54 * geo_score + 0.24 * rf_pressure + 0.22 * pressure).clip(0.0, 1.0)
    growth_multiplier = 1.0 + (0.035 * bucket_step) + (0.36 * bucket_step * hotspot_intensity) + (0.18 * bucket_step * positive_building_growth)
    out["growth_rate"] = ((growth_multiplier - 1.0) / (bucket_step.replace(0, np.nan))).replace([np.inf, -np.inf], np.nan).fillna(0.0).round(6)
    out["growth_zone_score"] = ((0.42 * geo_score + 0.36 * pressure + 0.22 * hotspot_intensity) * 100.0).round(3)
    out["demand_index"] = (base_demand * growth_multiplier * 100.0).clip(0.0, 140.0).round(3)

    area_factor = _robust_norm(out.get("target_grid_area_m2", pd.Series(625.0, index=out.index))).replace(0, 0.35)
    out["active_users_est"] = (
        (2.0 + (out["demand_index"] / 100.0) * 36.0)
        * (0.72 + 0.50 * area_factor)
        * (1.0 + 0.18 * activity_anchor)
    ).clip(lower=0.0).round(3)
    spectral_eff = _spectral_efficiency_from_sinr(out["serving_sinr_mean"])
    out["spectral_efficiency_bpshz"] = spectral_eff.round(4)
    out["serving_bandwidth_mhz_est"] = 10.0
    out["traffic_demand_est"] = (
        (
            0.050
            + 0.32 * np.power((out["demand_index"] / 100.0).clip(0.0, 1.6), 1.25)
            + 0.18 * hotspot_intensity
            + 0.020 * np.log1p(out["active_users_est"])
        )
        * (0.78 + 0.34 * area_factor)
        * (1.0 + 0.10 * activity_anchor)
    ).clip(lower=0.005).round(6)

    # First pass: compute cell load/capacity from traffic demand. Then use previous bucket
    # PRB as temporal context only for later buckets and recompute once.
    for iteration in range(2):
        cell_capacity = (
            out.groupby(["time_bucket", "serving_cell_key"], dropna=False)
            .agg(
                serving_cell_id=("serving_cell_id", "first"),
                serving_cell_mean_sinr=("serving_sinr_mean", "mean"),
                serving_cell_band=("serving_band", "first"),
                serving_cell_grid_count=("grid_id", "nunique"),
                serving_cell_candidate_count=("candidate_cell_count", "mean"),
                serving_cell_traffic_mbps=("traffic_demand_est", "sum"),
                serving_cell_users=("active_users_est", "sum"),
            )
            .reset_index()
        )
        eff = _spectral_efficiency_from_sinr(cell_capacity["serving_cell_mean_sinr"])
        band = pd.to_numeric(cell_capacity["serving_cell_band"], errors="coerce").fillna(1800.0)
        band_capacity_factor = np.select(
            [band <= 900, (band > 900) & (band <= 2100), band > 2100],
            [0.86, 1.0, 1.16],
            default=1.0,
        )
        cell_capacity["cell_capacity_mbps"] = (10.0 * eff * band_capacity_factor * 3.20).clip(8.0, 180.0)
        cell_capacity["computed_cell_prb_pct"] = (
            100.0 * cell_capacity["serving_cell_traffic_mbps"] / cell_capacity["cell_capacity_mbps"].replace(0, np.nan)
        ).replace([np.inf, -np.inf], np.nan).fillna(0.0).clip(0.0, 140.0)
        cell_capacity["computed_cell_rrc_pct"] = (
            100.0 * cell_capacity["serving_cell_users"] / (260.0 + 55.0 * eff + 3.20 * cell_capacity["cell_capacity_mbps"])
        ).replace([np.inf, -np.inf], np.nan).fillna(0.0).clip(0.0, 140.0)
        cell_capacity["cell_congested_flag"] = (
            (cell_capacity["computed_cell_prb_pct"] >= 70.0) | (cell_capacity["computed_cell_rrc_pct"] >= 70.0)
        ).astype(int)
        if iteration == 0:
            tmp = out.merge(
                cell_capacity[["time_bucket", "serving_cell_key", "computed_cell_prb_pct"]],
                on=["time_bucket", "serving_cell_key"],
                how="left",
            ).sort_values(["grid_id", "bucket_seq"])
            tmp["prev_bucket_prb_context"] = tmp.groupby("grid_id")["computed_cell_prb_pct"].shift(1).fillna(0.0)
            congestion_memory = ((tmp["prev_bucket_prb_context"] - 65.0) / 35.0).clip(0.0, 1.0)
            out = tmp.drop(columns=["computed_cell_prb_pct"])
            out["prev_bucket_prb_context"] = tmp["prev_bucket_prb_context"].round(6)
            out["traffic_demand_est"] = (out["traffic_demand_est"] * (1.0 + 0.10 * congestion_memory)).round(6)

    out = out.merge(cell_capacity.drop(columns=["serving_cell_id"], errors="ignore"), on=["time_bucket", "serving_cell_key"], how="left")
    out["prb_pressure_est"] = out["computed_cell_prb_pct"].round(6)
    out["rrc_pressure_est"] = out["computed_cell_rrc_pct"].round(6)
    out["prb_outlier_flag"] = (out["prb_pressure_est"] > 100.0).astype(int)
    out["capacity_gap_score"] = (np.maximum(out["prb_pressure_est"], out["rrc_pressure_est"]) - 70.0).clip(lower=0.0).round(6)
    out["capacity_context_score"] = (
        (100.0 - out["prb_pressure_est"]).clip(0.0, 100.0) * 0.45
        + out["spectral_efficiency_bpshz"].clip(0.0, 5.0) * 11.0
        + out["carrier_count"].clip(0.0, 4.0) * 4.0
    ).clip(0.0, 100.0).round(3)
    out["demand_feature_source"] = "geo_rf_heterogeneous_growth_cell_traffic_capacity_prb"
    return out


def build_dataset(geo_csv: Path, baseline_csv: Path, output_csv: Path, summary_json: Path) -> tuple[pd.DataFrame, Dict[str, object]]:
    geo = pd.read_csv(geo_csv)
    geo["grid_id"] = pd.to_numeric(geo["grid_id"], errors="coerce").astype("Int64")
    geo["time_bucket"] = geo["time_bucket"].astype(str)
    cell_grid = _aggregate_baseline_rf(baseline_csv)
    rf = _grid_rf_features(cell_grid)
    rf["grid_id"] = pd.to_numeric(rf["grid_id"], errors="coerce").astype("Int64")
    merged = geo.merge(rf, on=["time_bucket", "grid_id"], how="left", validate="one_to_one")
    dataset = _add_demand_and_cell_prb(merged)
    preferred = [
        "time_bucket",
        "bucket_seq",
        "grid_id",
        "grid_row",
        "grid_col",
        "centroid_lat",
        "centroid_lon",
        "clutter_class",
        "building_count",
        "building_area_ratio",
        "road_density",
        "road_length_m",
        "park_open_area",
        "open_area_ratio",
        "mall_presence",
        "metro_presence",
        "serving_cell_key",
        "serving_cell_id",
        "serving_rf_identity_key",
        "serving_site_sector_band_key",
        "serving_site_id",
        "serving_sector",
        "serving_canonical_sector",
        "serving_band",
        "serving_earfcn",
        "serving_rsrp_mean",
        "serving_rsrq_mean",
        "serving_sinr_mean",
        "neighbor1_cell_id",
        "neighbor1_cell_key",
        "neighbor1_rsrp",
        "neighbor2_cell_id",
        "neighbor2_cell_key",
        "neighbor2_rsrp",
        "rsrp_gap_to_neighbor1",
        "neighbor_interference_index",
        "candidate_cell_count",
        "low_band_ratio",
        "mid_band_ratio",
        "high_band_ratio",
        "dominant_band_class",
        "carrier_count",
        "geo_demand_score",
        "kpi_demand_score",
        "growth_rate",
        "development_pressure_score",
        "growth_zone_score",
        "clutter_transition_flag",
        "clutter_upgrade_score",
        "building_growth_ratio",
        "road_growth_ratio",
        "activity_anchor_score",
        "demand_index",
        "active_users_est",
        "traffic_demand_est",
        "prev_bucket_prb_context",
        "serving_cell_grid_count",
        "serving_cell_traffic_mbps",
        "serving_cell_users",
        "cell_capacity_mbps",
        "spectral_efficiency_bpshz",
        "prb_pressure_est",
        "rrc_pressure_est",
        "cell_congested_flag",
        "capacity_context_score",
        "capacity_gap_score",
        "prb_outlier_flag",
        "demand_feature_source",
    ]
    cols = [col for col in preferred if col in dataset.columns] + [col for col in dataset.columns if col not in preferred]
    dataset = dataset[cols].sort_values(["grid_id", "bucket_seq"]).reset_index(drop=True)
    output_csv.parent.mkdir(parents=True, exist_ok=True)
    dataset.to_csv(output_csv, index=False)

    p1 = dataset.loc[dataset["time_bucket"] == "PART_1"].set_index("grid_id")
    p3 = dataset.loc[dataset["time_bucket"] == "PART_3"].set_index("grid_id")
    common = p1.index.intersection(p3.index)
    summary: Dict[str, object] = {
        "output_csv": str(output_csv),
        "rows": int(len(dataset)),
        "unique_grids": int(dataset["grid_id"].nunique(dropna=True)),
        "bucket_counts": {str(k): int(v) for k, v in dataset["time_bucket"].value_counts().sort_index().items()},
        "serving_cells_by_bucket": {
            str(k): int(v) for k, v in dataset.groupby("time_bucket")["serving_cell_key"].nunique(dropna=True).sort_index().items()
        },
        "congested_cells_by_bucket": {
            str(k): int(v)
            for k, v in dataset.loc[dataset["cell_congested_flag"] == 1].groupby("time_bucket")["serving_cell_key"].nunique(dropna=True).sort_index().items()
        },
        "part1_to_part3": {},
    }
    for col in ["traffic_demand_est", "prb_pressure_est", "rrc_pressure_est", "demand_index", "active_users_est"]:
        delta = pd.to_numeric(p3.loc[common, col], errors="coerce") - pd.to_numeric(p1.loc[common, col], errors="coerce")
        summary["part1_to_part3"][col] = {
            "mean_delta": float(delta.mean()),
            "median_delta": float(delta.median()),
            "p95_delta": float(delta.quantile(0.95)),
            "max_delta": float(delta.max()),
            "changed_pct": float((delta.abs() > 1e-9).mean() * 100.0),
        }
    summary_json.write_text(json.dumps(summary, indent=2), encoding="utf-8")
    return dataset, summary


def rewrite_workbook(
    workbook: Path,
    geo_csv: Path,
    model2_csv: Path,
    baseline_csv: Path,
    summary: Dict[str, object],
) -> None:
    wb = Workbook(write_only=True)
    ws = wb.create_sheet("README")
    for row in [
        ["item", "value"],
        ["purpose", "Model 1 / Model 2 training workbook with causal Model 2 traffic-to-cell-PRB sheet"],
        ["model2_logic", "grid demand -> grid traffic -> serving-cell aggregation -> cell capacity -> PRB/RRC -> grid inherits serving-cell KPI"],
        ["model2_rows", summary.get("rows")],
        ["model2_unique_grids", summary.get("unique_grids")],
        ["baseline_rows", sum(1 for _ in open(baseline_csv, "r", encoding="utf-8", errors="ignore")) - 1],
        ["geo_csv", str(geo_csv)],
        ["model2_csv", str(model2_csv)],
        ["baseline_csv", str(baseline_csv)],
        ["summary_json", str(model2_csv.with_suffix(".summary.json"))],
    ]:
        ws.append(row)

    for sheet_name, csv_path in [("Bucket_Geo_25m", geo_csv), ("Model2_Training_25m", model2_csv)]:
        ws = wb.create_sheet(sheet_name)
        with open(csv_path, "r", encoding="utf-8", newline="") as handle:
            for row in csv.reader(handle):
                ws.append(row)

    with open(baseline_csv, "r", encoding="utf-8", newline="") as handle:
        reader = csv.reader(handle)
        header = next(reader)
        sheet_idx = 1
        data_rows = 0
        ws = wb.create_sheet(f"Baseline_25m_RF_{sheet_idx:02d}")
        ws.append(header)
        for row in reader:
            if data_rows >= MAX_EXCEL_DATA_ROWS:
                sheet_idx += 1
                data_rows = 0
                ws = wb.create_sheet(f"Baseline_25m_RF_{sheet_idx:02d}")
                ws.append(header)
            ws.append(row)
            data_rows += 1
    wb.save(workbook)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Build 25m Model 2 causal traffic/capacity training sheet.")
    parser.add_argument("--geo-csv", default="tests/coverage_prediction/data/model1_model2_training_geo_25m.csv")
    parser.add_argument("--baseline-csv", default="tests/coverage_prediction/data/model1_model2_training_baseline_25m_polygon.csv")
    parser.add_argument("--output-csv", default="tests/coverage_prediction/data/model2_training_25m_causal.csv")
    parser.add_argument("--workbook", default="tests/coverage_prediction/data/model1_model2_training_dataset.xlsx")
    parser.add_argument("--skip-workbook", action="store_true")
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    output_csv = Path(args.output_csv)
    summary_json = output_csv.with_suffix(".summary.json")
    _, summary = build_dataset(Path(args.geo_csv), Path(args.baseline_csv), output_csv, summary_json)
    if not args.skip_workbook:
        rewrite_workbook(Path(args.workbook), Path(args.geo_csv), output_csv, Path(args.baseline_csv), summary)
    print(json.dumps(summary, indent=2))


if __name__ == "__main__":
    main()
