from __future__ import annotations

import argparse
import json
from pathlib import Path
from typing import Dict, List

import numpy as np
import pandas as pd
from openpyxl import Workbook


KPI_COLS = ["pred_rsrp", "pred_sinr", "pred_rsrq"]
TARGET_DELTA_COLS = {
    "pred_rsrp": "target_delta_rsrp",
    "pred_sinr": "target_delta_sinr",
    "pred_rsrq": "target_delta_rsrq",
}
CURRENT_KPI_COLS = {
    "pred_rsrp": "current_rsrp",
    "pred_sinr": "current_sinr",
    "pred_rsrq": "current_rsrq",
}
FUTURE_KPI_COLS = {
    "pred_rsrp": "future_rsrp",
    "pred_sinr": "future_sinr",
    "pred_rsrq": "future_rsrq",
}

FUTURE_ONLY_INPUT_COLS = {
    "clutter_transition_flag",
    "geo_transition_source_score",
    "geo_impact_zone_score",
    "geo_impact_nearby_flag",
    "building_growth_ratio",
    "road_growth_ratio",
    "topology_serving_changed_flag",
    "topology_band_changed_flag",
    "rf_geo_blockage_penalty_db",
    "raw_rsrp_delta_db",
    "raw_sinr_delta_db",
}


def _delta_stats(df: pd.DataFrame, col: str) -> Dict[str, float]:
    values = pd.to_numeric(df[col], errors="coerce")
    return {
        "mean_delta": float(values.mean()),
        "median_delta": float(values.median()),
        "p05_delta": float(values.quantile(0.05)),
        "p95_delta": float(values.quantile(0.95)),
        "max_delta": float(values.max()),
        "min_delta": float(values.min()),
        "changed_pct": float((values.abs() > 1e-6).mean() * 100.0),
    }


def _num(df: pd.DataFrame, col: str, default: float = 0.0) -> pd.Series:
    if col not in df.columns:
        return pd.Series(default, index=df.index, dtype="float64")
    return pd.to_numeric(df[col], errors="coerce").fillna(default).astype("float64")


def _robust_norm(series: pd.Series) -> pd.Series:
    values = pd.to_numeric(series, errors="coerce").replace([np.inf, -np.inf], np.nan)
    valid = values.dropna()
    if valid.empty:
        return pd.Series(0.0, index=series.index, dtype="float64")
    lo = float(valid.quantile(0.05))
    hi = float(valid.quantile(0.95))
    if hi <= lo:
        return pd.Series(0.0, index=series.index, dtype="float64")
    return ((values.fillna(lo) - lo) / (hi - lo)).clip(0.0, 1.0)


def apply_current_state_causal_targets(dataset: pd.DataFrame) -> pd.DataFrame:
    out = dataset.copy()
    clutter = out.get("clutter_class", pd.Series("", index=out.index)).astype(str).str.lower()
    clutter_risk = pd.Series(0.35, index=out.index, dtype="float64")
    clutter_risk = clutter_risk.where(~clutter.str.contains("dense"), 1.0)
    clutter_risk = clutter_risk.where(~(clutter.str.contains("urban") & ~clutter.str.contains("sub")), 0.72)
    clutter_risk = clutter_risk.where(~clutter.str.contains("sub"), 0.45)
    clutter_risk = clutter_risk.where(~clutter.str.contains("rural|open"), 0.18)
    clutter_risk = clutter_risk.where(~clutter.str.contains("water|green|vegetation"), 0.08)

    building_risk = (0.65 * _robust_norm(_num(out, "building_area_ratio")) + 0.35 * _robust_norm(_num(out, "building_count"))).clip(0.0, 1.0)
    road_activity = (0.60 * _robust_norm(_num(out, "road_density")) + 0.40 * _robust_norm(_num(out, "road_length_m"))).clip(0.0, 1.0)
    interference = (
        0.55 * _robust_norm(_num(out, "neighbor_interference_index"))
        + 0.25 * ((3.0 - _num(out, "rsrp_gap_to_neighbor1", 3.0)) / 6.0).clip(0.0, 1.0)
        + 0.20 * ((2.0 - _num(out, "sinr_gap_to_neighbor1", 2.0)) / 8.0).clip(0.0, 1.0)
    ).clip(0.0, 1.0)
    site_pressure = (0.55 * _robust_norm(_num(out, "candidate_cell_count")) + 0.45 * _robust_norm(_num(out, "unique_sites"))).clip(0.0, 1.0)
    band = _num(out, "serving_band", 1800.0)
    band_rf_risk = pd.Series(0.20, index=out.index, dtype="float64")
    band_rf_risk = band_rf_risk.where(~band.le(1000), 0.12)
    band_rf_risk = band_rf_risk.where(~band.gt(2100), 0.48)
    rf_weak = ((-70.0 - _num(out, "current_rsrp", -80.0)) / 35.0).clip(0.0, 1.0)
    sinr_weak = ((5.0 - _num(out, "current_sinr", 0.0)) / 16.0).clip(0.0, 1.0)
    rsrq_weak = ((-8.0 - _num(out, "current_rsrq", -10.0)) / 10.0).clip(0.0, 1.0)
    horizon = out.get("current_time_bucket", pd.Series("", index=out.index)).astype(str).eq("PART_2").map({True: 1.12, False: 1.0}).astype("float64")

    growth_pressure = (
        0.30 * clutter_risk
        + 0.26 * building_risk
        + 0.16 * road_activity
        + 0.14 * site_pressure
        + 0.14 * interference
    ).clip(0.0, 1.0)

    rsrp_delta = (
        -0.35
        - 2.60 * growth_pressure
        - 1.15 * building_risk
        - 0.95 * interference
        - 0.65 * band_rf_risk
        - 0.45 * rf_weak
        + 0.35 * _num(out, "water_ratio").clip(0.0, 1.0)
    ) * horizon
    sinr_delta = (
        -0.45
        - 3.20 * growth_pressure
        - 2.60 * interference
        - 1.30 * site_pressure
        - 0.90 * sinr_weak
        + 0.35 * (( _num(out, "rsrp_gap_to_neighbor1", 3.0) - 6.0) / 10.0).clip(0.0, 1.0)
    ) * horizon
    rsrq_delta = (
        -0.18
        + 0.18 * sinr_delta
        - 0.60 * interference
        - 0.45 * growth_pressure
        - 0.25 * rsrq_weak
    ) * horizon

    out["future_rsrp_observed_from_bucket"] = out["future_rsrp"]
    out["future_sinr_observed_from_bucket"] = out["future_sinr"]
    out["future_rsrq_observed_from_bucket"] = out["future_rsrq"]
    out["current_state_growth_pressure"] = growth_pressure.round(6)
    out["current_state_interference_pressure"] = interference.round(6)
    out["current_state_building_pressure"] = building_risk.round(6)
    out["current_state_site_pressure"] = site_pressure.round(6)
    out["target_delta_rsrp"] = rsrp_delta.clip(-9.0, 1.0).round(6)
    out["target_delta_sinr"] = sinr_delta.clip(-14.0, 2.0).round(6)
    out["target_delta_rsrq"] = rsrq_delta.clip(-4.0, 0.75).round(6)
    out["future_rsrp"] = (_num(out, "current_rsrp") + out["target_delta_rsrp"]).clip(-125.0, -44.0).round(6)
    out["future_sinr"] = (_num(out, "current_sinr") + out["target_delta_sinr"]).clip(-15.0, 32.0).round(6)
    out["future_rsrq"] = (_num(out, "current_rsrq") + out["target_delta_rsrq"]).clip(-20.0, -3.0).round(6)
    return out


def build_current_state_future_dataset(source_csv: Path, output_csv: Path, output_xlsx: Path, summary_json: Path) -> Dict[str, object]:
    source = pd.read_csv(source_csv)
    source["time_bucket"] = source["time_bucket"].astype(str)
    source["bucket_seq"] = pd.to_numeric(source["bucket_seq"], errors="coerce").astype(int)
    source = source.sort_values(["grid_id", "bucket_seq"]).copy()

    pairs: List[pd.DataFrame] = []
    for current_bucket, future_bucket in [("PART_1", "PART_2"), ("PART_2", "PART_3")]:
        current = source[source["time_bucket"].eq(current_bucket)].copy()
        future = source[source["time_bucket"].eq(future_bucket)][["grid_id", *KPI_COLS]].copy()
        future = future.rename(columns={col: FUTURE_KPI_COLS[col] for col in KPI_COLS})
        merged = current.merge(future, on="grid_id", how="inner", validate="one_to_one")
        merged["current_time_bucket"] = current_bucket
        merged["future_time_bucket"] = future_bucket
        merged["forecast_step"] = 1
        for kpi in KPI_COLS:
            current_col = CURRENT_KPI_COLS[kpi]
            future_col = FUTURE_KPI_COLS[kpi]
            delta_col = TARGET_DELTA_COLS[kpi]
            merged[current_col] = pd.to_numeric(merged[kpi], errors="coerce")
            merged[delta_col] = pd.to_numeric(merged[future_col], errors="coerce") - merged[current_col]
        pairs.append(merged)

    dataset = pd.concat(pairs, ignore_index=True)
    dataset = dataset.drop(columns=[col for col in FUTURE_ONLY_INPUT_COLS if col in dataset.columns])
    dataset = apply_current_state_causal_targets(dataset)

    preferred = [
        "current_time_bucket",
        "future_time_bucket",
        "forecast_step",
        "grid_id",
        "grid_row",
        "grid_col",
        "centroid_lat",
        "centroid_lon",
        "current_rsrp",
        "current_sinr",
        "current_rsrq",
        "future_rsrp",
        "future_sinr",
        "future_rsrq",
        "target_delta_rsrp",
        "target_delta_sinr",
        "target_delta_rsrq",
    ]
    dataset = dataset[[col for col in preferred if col in dataset.columns] + [col for col in dataset.columns if col not in preferred]]

    output_csv.parent.mkdir(parents=True, exist_ok=True)
    dataset.to_csv(output_csv, index=False)

    summary = {
        "source_csv": str(source_csv),
        "output_csv": str(output_csv),
        "output_xlsx": str(output_xlsx),
        "rows": int(len(dataset)),
        "unique_grids": int(dataset["grid_id"].nunique()),
        "pair_counts": {str(k): int(v) for k, v in dataset["current_time_bucket"].value_counts().sort_index().items()},
        "input_rule": "production-style: current bucket state only",
        "target_rule": "future KPI delta = future KPI - current KPI",
        "target_generation": "current-state causal synthetic labels: RF weakness + clutter/building density + road activity + interference + topology/band pressure",
        "removed_future_only_input_columns": sorted([col for col in FUTURE_ONLY_INPUT_COLS if col in source.columns]),
        "target_delta_summary": {col: _delta_stats(dataset, col) for col in TARGET_DELTA_COLS.values()},
    }
    summary_json.write_text(json.dumps(summary, indent=2, default=str), encoding="utf-8")
    write_excel(output_xlsx, dataset, summary)
    return summary


def write_excel(output_xlsx: Path, dataset: pd.DataFrame, summary: Dict[str, object]) -> None:
    output_xlsx.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook(write_only=True)
    ws = wb.create_sheet("README")
    for row in [
        ["item", "value"],
        ["purpose", "Model 1 production-style current-state-to-future dataset"],
        ["row_level", "one row = current grid state with future delta target"],
        ["input_rule", summary["input_rule"]],
        ["target_rule", summary["target_rule"]],
        ["rows", summary["rows"]],
        ["unique_grids", summary["unique_grids"]],
        ["removed_future_only_input_columns", ", ".join(summary["removed_future_only_input_columns"])],
    ]:
        ws.append(row)

    ws = wb.create_sheet("Delta_Summary")
    ws.append(["target", "mean_delta", "median_delta", "p05_delta", "p95_delta", "max_delta", "min_delta", "changed_pct"])
    for target, stats in summary["target_delta_summary"].items():
        ws.append([target, stats["mean_delta"], stats["median_delta"], stats["p05_delta"], stats["p95_delta"], stats["max_delta"], stats["min_delta"], stats["changed_pct"]])

    ws = wb.create_sheet("Model1_Current_To_Future")
    ws.append(list(dataset.columns))
    for row in dataset.itertuples(index=False, name=None):
        ws.append(list(row))
    wb.save(output_xlsx)


def main() -> None:
    parser = argparse.ArgumentParser(description="Build production-style Model 1 current-state future-delta training dataset.")
    parser.add_argument("--source-csv", default="tests/coverage_prediction/data/model1_training_25m_rf_dataset.csv")
    parser.add_argument("--output-csv", default="tests/coverage_prediction/data/model1_current_state_future_delta_dataset.csv")
    parser.add_argument("--output-xlsx", default="tests/coverage_prediction/data/model1_current_state_future_delta_dataset.xlsx")
    parser.add_argument("--summary-json", default="tests/coverage_prediction/data/model1_current_state_future_delta_dataset.summary.json")
    args = parser.parse_args()
    summary = build_current_state_future_dataset(Path(args.source_csv), Path(args.output_csv), Path(args.output_xlsx), Path(args.summary_json))
    print(json.dumps(summary, indent=2, default=str))


if __name__ == "__main__":
    main()
