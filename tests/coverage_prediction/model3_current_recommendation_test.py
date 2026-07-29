from __future__ import annotations

import argparse
import concurrent.futures
import contextlib
import json
import os
import re
import shutil
import sys
import threading
import time
from dataclasses import dataclass
from pathlib import Path
from typing import Any

import numpy as np
import pandas as pd
from openpyxl import Workbook

ML_ROOT = Path(__file__).resolve().parents[2]
os.chdir(ML_ROOT)
if str(ML_ROOT) not in sys.path:
    sys.path.insert(0, str(ML_ROOT))

from tests.coverage_prediction import build_model3_hybrid_load_balancing_dataset as future_builder
from tests.coverage_prediction import build_model3_current_recommendation_dataset as current_builder
from tests.coverage_prediction import lte_coverage_test as coverage_test
from tests.coverage_prediction import model3_business_rule_recommendation_test as future_rules
from tools.lte_prediction_optimised.ml_engine import (
    canonical_cell_id as production_canonical_cell_id,
    _compute_affected_cells as production_compute_affected_cells,
    _normalize_site_df as production_normalize_site_df,
    compute_k1k2_for_cells as production_compute_k1k2_for_cells,
    run_prediction_only_optimized as production_run_prediction_only_optimized,
)


DEFAULT_DATASET = current_builder.CURRENT_MODEL3_DATASET_CSV
DEFAULT_SUMMARY = current_builder.CURRENT_MODEL3_SUMMARY_JSON
DEFAULT_OUTPUT_ROOT = ML_ROOT / "tests" / "output" / "model3_current_recommendation"
DEFAULT_STABLE_OUTPUT_DIR = ML_ROOT / "models" / "model3_current_recommendation_experiment"
PROJECT196_ANTENNA_FIXTURE = ML_ROOT / "tests" / "fixtures" / "project_196_rsrp_tilt" / "antenna_input.csv"
DEFAULT_CONGESTION_THRESHOLD = 80.0
DEFAULT_RF_WORKERS = max(1, min(3, (os.cpu_count() or 2) - 1))


@dataclass
class CurrentModel3Config:
    dataset_path: Path = DEFAULT_DATASET
    summary_path: Path = DEFAULT_SUMMARY
    output_root: Path = DEFAULT_OUTPUT_ROOT
    stable_output_dir: Path = DEFAULT_STABLE_OUTPUT_DIR
    congestion_threshold: float = DEFAULT_CONGESTION_THRESHOLD
    rrc_sector_capacity: float = future_builder.DEFAULT_RRC_SECTOR_CAPACITY
    sector_split_local_radius_m: float = 900.0
    max_sectors: int | None = None
    max_congested_cells: int | None = None
    carrier_reselection_hysteresis_db: float = 0.0
    rf_workers: int = DEFAULT_RF_WORKERS
    max_interference_sites: int = 10
    action_neighbor_cells: int = 2
    sector_parallelism: int = 1
    stop_on_partial: bool = False
    dashboard_model_label: str = "Model 3"


def _timestamp() -> str:
    return time.strftime("%Y%m%d_%H%M%S")


def _ensure_dir(path: Path) -> Path:
    path.mkdir(parents=True, exist_ok=True)
    return path


def _save_json(path: Path, payload: dict[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(payload, indent=2, default=str), encoding="utf-8")


def _safe_name(value: Any) -> str:
    text = str(value or "").strip()
    text = re.sub(r"[^A-Za-z0-9_.-]+", "_", text)
    return text.strip("_") or "unknown"


def _pick_rrc_group_cols(df: pd.DataFrame) -> list[str] | None:
    return next(
        (
            cols
            for cols in [
                ["time_bucket", "site_id", "sector_id", "cell_id"],
                ["time_bucket", "Node_Cell_ID"],
                ["time_bucket", "cell_id"],
            ]
            if all(col in df.columns for col in cols)
        ),
        None,
    )


def _setup_logger(log_path: Path):
    return future_rules._setup_logger(log_path)


def _assign_current_grid_load_to_cells(enriched: pd.DataFrame, config: CurrentModel3Config) -> pd.DataFrame:
    out = enriched.copy()
    grid_traffic = pd.to_numeric(out.get("estimated_offered_traffic_mbps"), errors="coerce")
    if grid_traffic.notna().sum() == 0:
        grid_traffic = pd.to_numeric(out.get("traffic_demand_est"), errors="coerce")
    grid_users = pd.to_numeric(out.get("estimated_rrc_connected_users"), errors="coerce")
    if grid_users.notna().sum() == 0:
        grid_users = pd.to_numeric(out.get("active_users_est"), errors="coerce")

    out["grid_assigned_traffic_mbps"] = grid_traffic.fillna(0.0).clip(lower=0.0).round(3)
    out["grid_assigned_rrc_users"] = grid_users.fillna(0.0).clip(lower=0.0).round(3)

    cell_agg = (
        out.groupby(["Node_Cell_ID", "time_bucket"], dropna=False, as_index=False)
        .agg(
            cell_assigned_traffic_mbps=("grid_assigned_traffic_mbps", "sum"),
            cell_assigned_rrc_users=("grid_assigned_rrc_users", "sum"),
            cell_capacity_mbps=("estimated_dl_capacity_mbps", "max"),
            cell_serving_grid_count=("grid_id", "nunique"),
        )
    )
    out = out.merge(cell_agg, on=["Node_Cell_ID", "time_bucket"], how="left", validate="many_to_one")
    out["cell_capacity_mbps"] = pd.to_numeric(out["cell_capacity_mbps"], errors="coerce").replace(0.0, np.nan).fillna(0.1)

    # Match the master dataset's own PRB semantics: estimated_prb_utilization_pct is a
    # PER-GRID ratio (that grid's traffic / the serving cell's capacity), and the cell's
    # PRB is the MAX over its grids (exactly how `_build_current_cell_inventory` reads
    # the before-values). Summing all grids' traffic into the numerator - the previous
    # behavior - inflates utilization by roughly the cell's grid count (5-11x) and is
    # what produced the impossible 500-2600% "after" figures. Verified directly against
    # the dataset: max_grid_traffic/capacity reproduces the stored per-cell PRB
    # (462057_3: 96.84% exactly); sum/capacity gives 510-616%.
    out["estimated_prb_utilization_pct"] = (
        (out["grid_assigned_traffic_mbps"] / out["cell_capacity_mbps"]) * 100.0
    ).replace([np.inf, -np.inf], np.nan).fillna(0.0).round(3)
    out["estimated_cell_rrc_connected_users"] = pd.to_numeric(out["cell_assigned_rrc_users"], errors="coerce").fillna(0.0).round(3)
    out["estimated_cell_rrc_utilization_pct"] = (
        (out["estimated_cell_rrc_connected_users"] / float(config.rrc_sector_capacity)) * 100.0
    ).round(3)

    # Keep grid-level demand on the row (matching dataset row semantics); cell-level
    # sums remain available as cell_assigned_traffic_mbps / cell_assigned_rrc_users.
    out["estimated_rrc_connected_users"] = out["grid_assigned_rrc_users"]
    out["estimated_offered_traffic_mbps"] = out["grid_assigned_traffic_mbps"]
    out["estimated_dl_capacity_mbps"] = out["cell_capacity_mbps"].round(3)
    return out


def _current_dataset_base_context(df: pd.DataFrame) -> pd.DataFrame:
    keep = [
        col
        for col in df.columns
        if col
        in {
            "grid_id",
            "time_bucket",
            "sample_count",
            "dl_tpt_mean",
            "ul_tpt_mean",
            "estimated_prb_mean",
            "cqi_mean",
            "dominant_pci",
            "green_ratio",
            "water_ratio",
            "grid_size_m",
            "grid_area_m2",
            "cell_area_m2",
            "road_length_m",
            "building_count",
            "building_area_ratio",
            "park_open_area",
            "open_area_ratio",
            "mall_presence",
            "metro_presence",
            "road_density",
            "geo_snapshot_mode",
            "geo_snapshot_source_ts",
            "clutter_class",
            "dominant_band_class",
            "bucket_seq",
            "bucket_min_timestamp",
            "bucket_max_timestamp",
            "bucket_mid_timestamp",
            "estimated_offered_traffic_mbps",
            "estimated_rrc_connected_users",
            "raw_estimated_offered_traffic_mbps",
            "raw_estimated_rrc_connected_users",
            "model3_hotspot_score",
            "model3_hotspot_rank",
        }
    ]
    out = df[keep].drop_duplicates(subset=["grid_id", "time_bucket"]).copy()
    out["grid_id"] = pd.to_numeric(out["grid_id"], errors="coerce").astype("Int64")
    out["time_bucket"] = out["time_bucket"].astype(str)
    return out


def _load_current_context(config: CurrentModel3Config, logger) -> dict[str, Any]:
    archive_path = future_rules.resolve_coverage_artifact_path()
    project_sites_df = future_rules._read_csv_from_archive(archive_path, "project_sites.csv")
    coverage_rows_df = future_rules._read_csv_from_archive(archive_path, "coverage_rows.csv")
    baseline_pred_df = future_rules._read_csv_from_archive(archive_path, "baseline_prediction_grid.csv")
    corrected_pred_df = future_rules._read_csv_from_archive(archive_path, "bucket_corrected_prediction_grid.csv")
    geo_df = future_rules._read_csv_from_archive(archive_path, "bucket_grid_geo_features.csv")
    kpi_df = future_rules._read_csv_from_archive(archive_path, "grid_kpi_timeseries.csv")
    summary = json.loads(
        __import__("subprocess").check_output(
            ["tar", "-xOf", str(archive_path), f"{future_rules._archive_root(archive_path)}/summary.json"],
            text=True,
        )
    )
    for frame in [coverage_rows_df, baseline_pred_df, corrected_pred_df, geo_df, kpi_df]:
        if "time_bucket" in frame.columns:
            frame["time_bucket"] = frame["time_bucket"].astype(str)
    part3_site_map, _ = coverage_test._build_bucket_site_topologies(
        site_df=project_sites_df,
        buckets=coverage_test.DEFAULT_BUCKETS,
        operator_name=str(summary.get("topology_operator") or "Airtel"),
    )
    part3_site_df = part3_site_map.get("PART_3", pd.DataFrame()).copy()
    current_df = pd.read_csv(config.dataset_path)
    current_df["time_bucket"] = current_df["time_bucket"].astype(str)
    current_df["grid_id"] = pd.to_numeric(current_df["grid_id"], errors="coerce").astype("Int64")

    # Every identity string a PRE-EXISTING cell is known by, across all naming layers
    # (site table vs prediction surface use different formats for the same cell).
    # A cell seen in a local rerun whose ID is NOT in this universe must be a synthetic
    # action candidate (carrier addition / new site / sector split child), no matter how
    # the prediction tooling mangled its name.
    original_topology_cell_ids: set[str] = set()
    for col in ["Node_Cell_ID", "cell_id"]:
        if col in part3_site_df.columns:
            original_topology_cell_ids.update(part3_site_df[col].dropna().astype(str).str.strip())
    for col in ["Node_Cell_ID", "original_node_cell_id", "original_cell_id"]:
        if col in corrected_pred_df.columns:
            original_topology_cell_ids.update(corrected_pred_df[col].dropna().astype(str).str.strip())
    original_topology_cell_ids.discard("")

    # Which cell serves each grid in the MASTER assignment ("grid_id|alias" keys, all
    # aliases) - the incumbent a local rerun's candidate must beat by the hysteresis
    # margin before the grid is allowed to reselect away from it.
    master_serving_keys: set[str] = set()
    for col in ["Node_Cell_ID", "topology_original_cell_id", "topology_original_node_cell_id"]:
        if col in current_df.columns:
            alias = current_df[col].dropna().astype(str).str.strip()
            alias = alias[alias.ne("")]
            gid = current_df.loc[alias.index, "grid_id"].astype(str)
            master_serving_keys.update((gid + "|" + alias).tolist())

    # Dataset-scale DL capacity per cell, keyed by every identity alias. The dataset's
    # demand (estimated_offered_traffic_mbps) is NOT in physical Mbps - it is only
    # meaningful relative to the dataset's own capacity column. Any PRB computed from
    # this demand MUST use this capacity, never a physically re-derived one.
    master_capacity_by_cell_id: dict[str, float] = {}
    if "estimated_dl_capacity_mbps" in current_df.columns:
        master_cap = pd.to_numeric(current_df["estimated_dl_capacity_mbps"], errors="coerce")
        for col in ["Node_Cell_ID", "topology_original_cell_id", "topology_original_node_cell_id"]:
            if col in current_df.columns:
                grouped = master_cap.groupby(current_df[col].astype(str).str.strip()).max()
                for key, value in grouped.items():
                    if key and key != "nan" and pd.notna(value) and value > 0:
                        master_capacity_by_cell_id.setdefault(key, float(value))

    logger.info(
        "current_context archive=%s part3_sites=%d current_rows=%d original_cell_id_universe=%d",
        archive_path,
        len(part3_site_df),
        len(current_df),
        len(original_topology_cell_ids),
    )
    return {
        "archive_path": archive_path,
        "summary": summary,
        "project_sites_df": project_sites_df,
        "coverage_rows_df": coverage_rows_df,
        "baseline_pred_df": baseline_pred_df,
        "corrected_pred_df": corrected_pred_df,
        "geo_df": geo_df,
        "kpi_df": kpi_df,
        "part3_site_df": part3_site_df,
        "building_df": pd.DataFrame(columns=["geometry_wkt"]),
        "current_base_df": _current_dataset_base_context(current_df),
        "current_df": current_df,
        "original_topology_cell_ids": original_topology_cell_ids,
        "master_serving_keys": master_serving_keys,
        "master_capacity_by_cell_id": master_capacity_by_cell_id,
    }


def _clean_project196_text(value: Any) -> str:
    if pd.isna(value):
        return ""
    text = str(value).strip()
    if text.endswith(".0"):
        text = text[:-2]
    return "" if text.lower() in {"nan", "none", "null", "<na>"} else text


def _filter_project196_cell_input(df: pd.DataFrame) -> pd.DataFrame:
    work = df.copy()
    required = [col for col in ["site_id", "sector_id", "Node_Cell_ID", "band"] if col in work.columns]
    for col in required:
        work = work.loc[work[col].map(_clean_project196_text).str.strip().ne("")].copy()
    return work.reset_index(drop=True)


def _project196_sector_num(value: Any) -> str:
    text = _clean_project196_text(value)
    if "|" in text:
        return text.rsplit("|", 1)[1]
    tokens = [token for token in re.split(r"[_|]", text) if token and token.lower() != "nan"]
    if len(tokens) >= 3:
        return tokens[2]
    if len(tokens) >= 2:
        return tokens[1]
    return text


def _load_project196_site_topology(cell_input: pd.DataFrame, logger) -> pd.DataFrame:
    if not PROJECT196_ANTENNA_FIXTURE.exists():
        raise FileNotFoundError(f"Missing Project 196 antenna fixture: {PROJECT196_ANTENNA_FIXTURE}")
    fixture = pd.read_csv(PROJECT196_ANTENNA_FIXTURE)
    fixture = fixture.loc[:, ~fixture.columns.duplicated()].copy()
    fixture["_site_key"] = fixture.get("dashboard_site_id", fixture.get("site", "")).map(_clean_project196_text)
    fixture["_sector_num"] = fixture.get("original_cell_id", fixture.get("rf_source_cell_id", "")).map(_project196_sector_num)
    fixture["_band_num"] = pd.to_numeric(fixture.get("band"), errors="coerce").round()

    rows: list[pd.Series] = []
    for _, source in cell_input.iterrows():
        site = _clean_project196_text(source.get("site_id"))
        sector_num = _project196_sector_num(source.get("sector_id") or source.get("canonical_physical_cell_id") or source.get("Node_Cell_ID"))
        band = future_rules._scalar_to_int(source.get("band"), 1800)
        match = fixture.loc[
            fixture["_site_key"].eq(site)
            & fixture["_sector_num"].eq(sector_num)
            & fixture["_band_num"].eq(float(band))
        ]
        if match.empty:
            match = fixture.loc[fixture["_site_key"].eq(site) & fixture["_sector_num"].eq(sector_num)]
        if match.empty:
            logger.info("project196_site_topology_missing site=%s sector=%s band=%s", site, sector_num, band)
            continue
        row = match.iloc[0].copy()
        node_cell_id = _clean_project196_text(source.get("Node_Cell_ID"))
        canonical = _clean_project196_text(source.get("canonical_physical_cell_id")) or node_cell_id
        sector_id = _clean_project196_text(source.get("sector_id")) or f"{site}|{sector_num}"
        row["Node_Cell_ID"] = node_cell_id
        row["cell_id"] = node_cell_id
        row["site"] = site
        row["site_id"] = site
        row["dashboard_site_id"] = site
        row["Site ID"] = site
        row["band"] = float(band)
        row["frequency_mhz"] = float(band)
        row["frequency"] = float(band)
        row["earfcn"] = float(future_rules.SYNTHETIC_BAND_TO_EARFCN.get(band, band))
        row["original_node_cell_id"] = canonical
        row["original_cell_id"] = canonical
        row["rf_source_cell_id"] = canonical
        row["site_prediction_key"] = node_cell_id
        row["site_cell_sector_band_operator_key"] = node_cell_id
        row["sector_identity"] = sector_id
        row["frontend_site_sector_key"] = sector_id
        row["node_cell_sector_key"] = f"{canonical}|{sector_num}"
        row["site_identity_key"] = site
        rows.append(row)

    if not rows:
        raise RuntimeError("Project 196 Excel cells could not be matched to antenna topology.")
    site_df = pd.DataFrame(rows).drop(columns=["_site_key", "_sector_num", "_band_num"], errors="ignore")
    logger.info("project196_site_topology rows=%d cells=%d", len(site_df), site_df["Node_Cell_ID"].nunique())
    return site_df.loc[:, ~site_df.columns.duplicated()].copy()


def _prepare_project196_prediction_surface(pred_df: pd.DataFrame, cell_input: pd.DataFrame) -> pd.DataFrame:
    work = pred_df.copy()
    work["time_bucket"] = "PART_3"
    work = future_rules._ensure_grid_group_columns(work, default_time_bucket="PART_3")
    cell_lookup = cell_input.copy()
    cell_lookup["Node_Cell_ID"] = cell_lookup["Node_Cell_ID"].map(_clean_project196_text)
    lookup_cols = [
        col
        for col in [
            "Node_Cell_ID",
            "site_id",
            "sector_id",
            "canonical_physical_cell_id",
            "band",
            "earfcn",
            "input_prb_utilization_pct",
            "input_rrc_utilization_pct",
            "input_rrc_connected_users",
            "input_estimated_offered_traffic_mbps",
            "input_estimated_dl_capacity_mbps",
            "grid_count",
            "point_count",
            "available_bands_to_add",
            "recommended_band_to_add",
            "carrier_addition_possible",
            "demo_scenario",
            "demo_selected_for_model3",
        ]
        if col in cell_lookup.columns
    ]
    work["Node_Cell_ID"] = work["Node_Cell_ID"].map(_clean_project196_text)
    valid_nodes = set(cell_lookup["Node_Cell_ID"].dropna().astype(str))
    work = work.loc[work["Node_Cell_ID"].isin(valid_nodes)].copy()
    work = work.merge(cell_lookup[lookup_cols].drop_duplicates("Node_Cell_ID"), on="Node_Cell_ID", how="left", suffixes=("", "_cell"))
    work["time_bucket"] = "PART_3"
    work["topology_original_cell_id"] = work.get("canonical_physical_cell_id", work["Node_Cell_ID"]).map(_clean_project196_text)
    work["topology_original_node_cell_id"] = work["topology_original_cell_id"]
    work["topology_frontend_site_sector_key"] = work.get("sector_id", "").map(_clean_project196_text)
    work["topology_node_cell_sector_key"] = work["topology_frontend_site_sector_key"]
    work["topology_sector_identity_key"] = work["topology_frontend_site_sector_key"]
    work["topology_site_sector_band_key"] = (
        work.get("site_id", "").map(_clean_project196_text)
        + "|"
        + work.get("sector_id", "").map(_clean_project196_text)
        + "|"
        + pd.to_numeric(work.get("band"), errors="coerce").fillna(0).round().astype(int).astype(str)
    )
    work["topology_rf_identity_key"] = work["Node_Cell_ID"]
    work["rf_identity_key"] = work["Node_Cell_ID"]
    work["site_sector_band_key"] = work["topology_site_sector_band_key"]
    work["sector_identity_key"] = work["topology_frontend_site_sector_key"]
    work["frontend_site_sector_key"] = work["topology_frontend_site_sector_key"]
    work["node_cell_sector_key"] = work["topology_node_cell_sector_key"]
    work["cell_id"] = work["topology_original_cell_id"]
    work["original_node_cell_id"] = work["topology_original_node_cell_id"]
    work["original_cell_id"] = work["topology_original_cell_id"]
    work["topology_band"] = pd.to_numeric(work.get("band"), errors="coerce")
    work["topology_earfcn"] = pd.to_numeric(work.get("earfcn"), errors="coerce")
    work["estimated_prb_utilization_pct"] = pd.to_numeric(work.get("input_prb_utilization_pct"), errors="coerce").fillna(0.0)
    work["estimated_cell_rrc_utilization_pct"] = pd.to_numeric(work.get("input_rrc_utilization_pct"), errors="coerce").fillna(0.0)
    cell_rrc_total = pd.to_numeric(work.get("input_rrc_connected_users"), errors="coerce").fillna(0.0)
    point_count = pd.to_numeric(work.get("point_count"), errors="coerce").replace(0, np.nan)
    if point_count.notna().sum() == 0:
        point_count = work.groupby("Node_Cell_ID")["Node_Cell_ID"].transform("count").astype(float).replace(0, np.nan)
    per_point_rrc = (cell_rrc_total / point_count.fillna(1.0)).replace([np.inf, -np.inf], np.nan).fillna(0.0)
    work["estimated_cell_rrc_connected_users"] = per_point_rrc
    work["estimated_offered_traffic_mbps"] = pd.to_numeric(work.get("input_estimated_offered_traffic_mbps"), errors="coerce").fillna(0.0)
    work["raw_estimated_rrc_connected_users"] = cell_rrc_total
    work["estimated_rrc_connected_users"] = per_point_rrc
    work["estimated_dl_capacity_mbps"] = pd.to_numeric(work.get("input_estimated_dl_capacity_mbps"), errors="coerce").fillna(0.1)
    bandwidth_src = work["bandwidth_mhz_est"] if "bandwidth_mhz_est" in work.columns else pd.Series(10.0, index=work.index)
    work["bandwidth_mhz_est"] = pd.to_numeric(bandwidth_src, errors="coerce").fillna(10.0)
    return work


def _load_project196_excel_context(config: CurrentModel3Config, cell_input: pd.DataFrame, logger) -> dict[str, Any]:
    baseline_df = pd.read_excel(config.dataset_path, sheet_name="Baseline_Grid_Input")
    geo_df = pd.read_excel(config.dataset_path, sheet_name="Geo_Features_Input")
    site_df = _load_project196_site_topology(cell_input, logger)
    corrected_pred_df = _prepare_project196_prediction_surface(baseline_df, cell_input)
    baseline_pred_df = corrected_pred_df.copy()

    geo_work = geo_df.copy()
    geo_work["time_bucket"] = "PART_3"
    geo_work = future_rules._ensure_grid_group_columns(geo_work, default_time_bucket="PART_3")
    kpi_cols = [
        "grid_id",
        "grid_row",
        "grid_col",
        "grid_centroid_lat",
        "grid_centroid_lon",
        "time_bucket",
        "sample_count",
        "dl_tpt_mean",
        "ul_tpt_mean",
        "estimated_prb_mean",
        "cqi_mean",
        "green_ratio",
        "water_ratio",
        "bandwidth_mhz_est",
    ]
    kpi_df = corrected_pred_df[[col for col in kpi_cols if col in corrected_pred_df.columns]].drop_duplicates(
        subset=["grid_id", "time_bucket"]
    ).copy()
    if "sample_count" not in kpi_df.columns:
        kpi_df["sample_count"] = 1
    earfcn_lookup = corrected_pred_df.copy()
    earfcn_lookup["_dominant_earfcn"] = pd.to_numeric(earfcn_lookup.get("topology_earfcn"), errors="coerce")
    missing_earfcn = earfcn_lookup["_dominant_earfcn"].isna()
    if missing_earfcn.any():
        band_values = pd.to_numeric(earfcn_lookup.loc[missing_earfcn, "topology_band"], errors="coerce")
        earfcn_lookup.loc[missing_earfcn, "_dominant_earfcn"] = band_values.map(
            lambda value: future_rules.SYNTHETIC_BAND_TO_EARFCN.get(int(value), value) if pd.notna(value) else np.nan
        )
    earfcn_lookup = earfcn_lookup.drop_duplicates(subset=["grid_id", "time_bucket"], keep="first")
    kpi_df = kpi_df.merge(
        earfcn_lookup[["grid_id", "time_bucket", "_dominant_earfcn"]],
        on=["grid_id", "time_bucket"],
        how="left",
        validate="one_to_one",
    )
    kpi_df["dominant_earfcn"] = pd.to_numeric(kpi_df.pop("_dominant_earfcn"), errors="coerce")
    if "bandwidth_mhz_est" not in kpi_df.columns:
        kpi_df["bandwidth_mhz_est"] = 10.0
    coverage_rows_df = corrected_pred_df[
        ["lat", "lon", "grid_id", "grid_row", "grid_col", "grid_centroid_lat", "grid_centroid_lon", "time_bucket"]
    ].drop_duplicates().copy()

    current_df = corrected_pred_df.copy()
    original_topology_cell_ids = set(site_df["Node_Cell_ID"].dropna().astype(str).str.strip())
    original_topology_cell_ids.update(current_df["Node_Cell_ID"].dropna().astype(str).str.strip())
    master_serving_keys: set[str] = set()
    for col in ["Node_Cell_ID", "topology_original_cell_id", "topology_original_node_cell_id"]:
        alias = current_df[col].dropna().astype(str).str.strip()
        alias = alias[alias.ne("")]
        gid = current_df.loc[alias.index, "grid_id"].astype(str)
        master_serving_keys.update((gid + "|" + alias).tolist())

    master_capacity_by_cell_id: dict[str, float] = {}
    master_cap = pd.to_numeric(current_df["estimated_dl_capacity_mbps"], errors="coerce")
    for col in ["Node_Cell_ID", "topology_original_cell_id", "topology_original_node_cell_id"]:
        grouped = master_cap.groupby(current_df[col].astype(str).str.strip()).max()
        for key, value in grouped.items():
            if key and key != "nan" and pd.notna(value) and value > 0:
                master_capacity_by_cell_id.setdefault(key, float(value))

    summary = {
        "project_id": 196,
        "region": "india",
        "topology_operator": "Airtel",
        "grid_size_m": float(pd.to_numeric(cell_input.get("grid_size_m"), errors="coerce").dropna().iloc[0])
        if "grid_size_m" in cell_input.columns and pd.to_numeric(cell_input["grid_size_m"], errors="coerce").notna().any()
        else 25.0,
        "baseline_radius_m": 500.0,
        "max_interference_sites": int(config.max_interference_sites),
        "source": "project196_excel_rf_context",
    }
    logger.info(
        "project196_excel_rf_context baseline_rows=%d geo_rows=%d site_rows=%d numeric_grids=%d",
        len(corrected_pred_df),
        len(geo_work),
        len(site_df),
        corrected_pred_df["grid_id"].nunique(dropna=True),
    )
    return {
        "archive_path": "",
        "summary": summary,
        "project_sites_df": site_df,
        "coverage_rows_df": coverage_rows_df,
        "baseline_pred_df": baseline_pred_df,
        "corrected_pred_df": corrected_pred_df,
        "geo_df": geo_work,
        "kpi_df": kpi_df,
        "part3_site_df": site_df,
        "building_df": pd.DataFrame(columns=["geometry_wkt"]),
        "current_base_df": _current_dataset_base_context(current_df),
        "current_df": current_df,
        "original_topology_cell_ids": original_topology_cell_ids,
        "master_serving_keys": master_serving_keys,
        "master_capacity_by_cell_id": master_capacity_by_cell_id,
    }


def _build_current_cell_inventory(df: pd.DataFrame, config: CurrentModel3Config) -> tuple[pd.DataFrame, dict[str, Any]]:
    sector_key = future_rules._pick_sector_key(df)
    site_key = future_rules._pick_site_key(df)
    rows: list[dict[str, Any]] = []
    for node_cell_id, group in df.groupby("Node_Cell_ID", dropna=False):
        prb = future_rules._to_num(group["estimated_prb_utilization_pct"])
        rrc = future_rules._to_num(group["estimated_cell_rrc_utilization_pct"])
        users = future_rules._to_num(group["estimated_cell_rrc_connected_users"])
        traffic = future_rules._to_num(group.get("estimated_offered_traffic_mbps", pd.Series(dtype=float)))
        capacity = future_rules._to_num(group.get("estimated_dl_capacity_mbps", pd.Series(dtype=float)))
        row = {
            "Node_Cell_ID": node_cell_id,
            "site_id": future_rules._first_non_empty(group[site_key]) if site_key and site_key in group.columns else "",
            "sector_id": future_rules._first_non_empty(group[sector_key]) if sector_key in group.columns else "",
            "band": future_rules._fmt_band(future_rules._first_non_empty(group.get("topology_band", pd.Series(dtype=object)))),
            "earfcn": future_rules._fmt_band(future_rules._first_non_empty(group.get("topology_earfcn", pd.Series(dtype=object)))),
            "topology_original_cell_id": future_rules._first_non_empty(group.get("topology_original_cell_id", pd.Series(dtype=object))),
            "topology_original_node_cell_id": future_rules._first_non_empty(group.get("topology_original_node_cell_id", pd.Series(dtype=object))),
            "topology_frontend_site_sector_key": future_rules._first_non_empty(group.get("topology_frontend_site_sector_key", pd.Series(dtype=object))),
            "grid_count": int(group["grid_id"].nunique(dropna=True)) if "grid_id" in group.columns else int(len(group)),
            "congested_grid_count": int(((prb > config.congestion_threshold) | (rrc > config.congestion_threshold)).sum()),
            "prb_before_pct": float(prb.max()) if prb.notna().any() else np.nan,
            "prb_p90_pct": float(prb.quantile(0.90)) if prb.notna().any() else np.nan,
            "rrc_before_pct": float(rrc.max()) if rrc.notna().any() else np.nan,
            "rrc_users_before": float(users.max()) if users.notna().any() else np.nan,
            "estimated_offered_traffic_mbps": float(traffic.max()) if traffic.notna().any() else np.nan,
            "estimated_dl_capacity_mbps": float(capacity.max()) if capacity.notna().any() else np.nan,
            "existing_carriers": future_rules._first_non_empty(group.get("existing_carriers", pd.Series(dtype=object))),
            "existing_carrier_count": int(future_rules._to_num(group.get("existing_carrier_count", pd.Series(dtype=float))).max())
            if "existing_carrier_count" in group.columns and future_rules._to_num(group["existing_carrier_count"]).notna().any()
            else 0,
            "available_bands_to_add": future_rules._first_non_empty(group.get("available_bands_to_add", pd.Series(dtype=object))),
            "carrier_addition_options": future_rules._first_non_empty(group.get("carrier_addition_options", pd.Series(dtype=object))),
            "available_earfcns_to_add": future_rules._first_non_empty(group.get("available_earfcns_to_add", pd.Series(dtype=object))),
            "available_earfcn_options": future_rules._first_non_empty(group.get("available_earfcn_options", pd.Series(dtype=object))),
            "recommended_band_to_add": future_rules._first_non_empty(group.get("recommended_band_to_add", pd.Series(dtype=object))),
            "available_band_options_count": int(future_rules._to_num(group.get("available_band_options_count", pd.Series(dtype=float))).max())
            if "available_band_options_count" in group.columns and future_rules._to_num(group["available_band_options_count"]).notna().any()
            else 0,
            "max_supported_carriers": int(future_rules._to_num(group.get("max_supported_carriers", pd.Series(dtype=float))).max())
            if "max_supported_carriers" in group.columns and future_rules._to_num(group["max_supported_carriers"]).notna().any()
            else 0,
            "carrier_addition_possible": bool(group.get("carrier_addition_possible", pd.Series(False)).astype(bool).any()),
            "carrier_addition_blocked": bool(group.get("carrier_addition_blocked", pd.Series(False)).astype(bool).all()),
            "carrier_addition_reason": future_rules._first_non_empty(group.get("carrier_addition_reason", pd.Series(dtype=object))),
            "recommendation_scope_cell": bool(group.get("recommendation_scope_cell", pd.Series(False)).astype(bool).any()),
            "model3_scenario": future_rules._first_non_empty(group.get("model3_scenario", pd.Series(dtype=object))),
            "model3_scenario_reason": future_rules._first_non_empty(group.get("model3_scenario_reason", pd.Series(dtype=object))),
            "sector_has_alternate_carrier": bool(group.get("sector_has_alternate_carrier", pd.Series(False)).astype(bool).any()),
            "sector_capacity_limit": int(future_rules._to_num(group.get("sector_capacity_limit", pd.Series(dtype=float))).max())
            if "sector_capacity_limit" in group.columns and future_rules._to_num(group["sector_capacity_limit"]).notna().any()
            else 0,
            "hotspot_score": float(future_rules._to_num(group.get("prb_pressure_est", pd.Series(dtype=float))).max())
            if "prb_pressure_est" in group.columns and future_rules._to_num(group["prb_pressure_est"]).notna().any()
            else np.nan,
        }
        rows.append(row)
    out = pd.DataFrame(rows)
    if out.empty:
        return out, {"cell_count": 0, "sector_count": 0, "congested_cell_count": 0, "carrier_addition_candidate_count": 0}
    out["prb_rrc_pressure"] = out[["prb_before_pct", "rrc_before_pct"]].max(axis=1)
    out["congested"] = out["prb_rrc_pressure"] > float(config.congestion_threshold)
    sector_counts = out.groupby("sector_id")["Node_Cell_ID"].nunique(dropna=True)
    sector_congested_counts = out.groupby("sector_id")["congested"].sum()
    out["sector_cell_count"] = out["sector_id"].map(sector_counts).fillna(0).astype(int)
    out["sector_congested_count"] = out["sector_id"].map(sector_congested_counts).fillna(0).astype(int)
    summary = {
        "cell_count": int(len(out)),
        "sector_count": int(out["sector_id"].nunique(dropna=True)),
        "congested_cell_count": int(out["congested"].sum()),
        "carrier_addition_candidate_count": int(out["carrier_addition_possible"].sum()),
    }
    return out, summary


def _build_cell_inventory_from_excel_input(df: pd.DataFrame, config: CurrentModel3Config) -> tuple[pd.DataFrame, dict[str, Any]]:
    work = df.copy()
    rename_map = {
        "input_prb_utilization_pct": "prb_before_pct",
        "input_rrc_utilization_pct": "rrc_before_pct",
        "input_rrc_connected_users": "rrc_users_before",
        "input_estimated_offered_traffic_mbps": "estimated_offered_traffic_mbps",
        "input_estimated_dl_capacity_mbps": "estimated_dl_capacity_mbps",
        "input_available_bands_to_add": "available_bands_to_add",
        "input_max_supported_carriers": "max_supported_carriers",
        "input_existing_carrier_count": "existing_carrier_count",
    }
    for source, target in rename_map.items():
        if source in work.columns:
            work[target] = work[source]
    if "model3_scenario" not in work.columns and "demo_scenario" in work.columns:
        work["model3_scenario"] = work["demo_scenario"]
    if "recommendation_scope_cell" not in work.columns and "demo_selected_for_model3" in work.columns:
        selected_text = work["demo_selected_for_model3"].astype(str).str.strip().str.lower()
        selected_num = pd.to_numeric(work["demo_selected_for_model3"], errors="coerce")
        work["recommendation_scope_cell"] = selected_text.isin({"true", "yes", "y"}) | (selected_num.fillna(0) > 0)
    if "carrier_addition_options" not in work.columns and "available_bands_to_add" in work.columns:
        work["carrier_addition_options"] = work["available_bands_to_add"]
    if "available_earfcns_to_add" not in work.columns and "available_bands_to_add" in work.columns:
        work["available_earfcns_to_add"] = work["available_bands_to_add"]
    if "available_earfcn_options" not in work.columns and "available_earfcns_to_add" in work.columns:
        work["available_earfcn_options"] = work["available_earfcns_to_add"]
    if "recommended_band_to_add" not in work.columns:
        work["recommended_band_to_add"] = work.get("available_bands_to_add", "").astype(str).str.split(",").str[0].fillna("")
    if "carrier_addition_possible" not in work.columns:
        existing = future_rules._to_num(work.get("existing_carrier_count", pd.Series(0, index=work.index))).fillna(0)
        limit = future_rules._to_num(work.get("max_supported_carriers", pd.Series(0, index=work.index))).fillna(0)
        work["carrier_addition_possible"] = (limit > existing) & work.get("available_bands_to_add", "").astype(str).str.strip().ne("")
    else:
        possible_text = work["carrier_addition_possible"].astype(str).str.strip().str.lower()
        possible_num = pd.to_numeric(work["carrier_addition_possible"], errors="coerce")
        work["carrier_addition_possible"] = possible_text.isin({"true", "yes", "y"}) | (possible_num.fillna(0) > 0)
    work["carrier_addition_blocked"] = ~work["carrier_addition_possible"].astype(bool)
    work["carrier_addition_reason"] = np.where(work["carrier_addition_possible"], "EXCEL_AVAILABLE_BAND", "EXCEL_NO_AVAILABLE_BAND_OR_LIMIT")
    work["sector_has_alternate_carrier"] = work.get("existing_carrier_count", 0).astype(float) > 1
    for col, default in [
        ("grid_count", 1),
        ("prb_before_pct", 0.0),
        ("rrc_before_pct", 0.0),
        ("rrc_users_before", 0.0),
        ("estimated_offered_traffic_mbps", 0.0),
        ("estimated_dl_capacity_mbps", 0.1),
        ("existing_carrier_count", 0),
        ("max_supported_carriers", 0),
    ]:
        work[col] = future_rules._to_num(work.get(col, pd.Series(default, index=work.index))).fillna(default)
    work["sector_capacity_limit"] = work["max_supported_carriers"].fillna(0).astype(int)
    work["congested_grid_count"] = np.where(
        (work["prb_before_pct"] > config.congestion_threshold) | (work["rrc_before_pct"] > config.congestion_threshold),
        work["grid_count"],
        0,
    ).astype(int)
    work["prb_p90_pct"] = work["prb_before_pct"]
    work["prb_rrc_pressure"] = work[["prb_before_pct", "rrc_before_pct"]].max(axis=1)
    work["congested"] = work["prb_rrc_pressure"] > float(config.congestion_threshold)
    sector_counts = work.groupby("sector_id")["Node_Cell_ID"].nunique(dropna=True)
    sector_congested_counts = work.groupby("sector_id")["congested"].sum()
    work["sector_cell_count"] = work["sector_id"].map(sector_counts).fillna(0).astype(int)
    work["sector_congested_count"] = work["sector_id"].map(sector_congested_counts).fillna(0).astype(int)
    summary = {
        "cell_count": int(len(work)),
        "sector_count": int(work["sector_id"].nunique(dropna=True)),
        "congested_cell_count": int(work["congested"].sum()),
        "carrier_addition_candidate_count": int(work["carrier_addition_possible"].sum()),
        "source": "excel_model3_cell_input",
    }
    required = [
        "Node_Cell_ID",
        "canonical_physical_cell_id",
        "site_id",
        "sector_id",
        "band",
        "earfcn",
        "grid_count",
        "congested_grid_count",
        "prb_before_pct",
        "prb_p90_pct",
        "rrc_before_pct",
        "rrc_users_before",
        "estimated_offered_traffic_mbps",
        "estimated_dl_capacity_mbps",
        "existing_carriers",
        "existing_carrier_count",
        "available_bands_to_add",
        "carrier_addition_options",
        "available_earfcns_to_add",
        "available_earfcn_options",
        "recommended_band_to_add",
        "max_supported_carriers",
        "carrier_addition_possible",
        "carrier_addition_blocked",
        "carrier_addition_reason",
        "sector_has_alternate_carrier",
        "sector_capacity_limit",
        "model3_scenario",
        "recommendation_scope_cell",
        "prb_rrc_pressure",
        "congested",
        "sector_cell_count",
        "sector_congested_count",
    ]
    for col in required:
        if col not in work.columns:
            work[col] = ""
    return work[required].copy(), summary


def _select_serving_cell_with_hysteresis(
    topo_work: pd.DataFrame,
    hysteresis_db: float,
    original_cell_ids: set[str],
    master_serving_keys: set[str],
    lineage_grid_ids: set[int] | None = None,
) -> pd.DataFrame:
    """Pick one serving cell per grid/time_bucket with stickiness to the incumbent.

    Primary rule - incumbent stickiness: each grid's MASTER assignment (the cell that
    serves it in the full-network archive, `master_serving_keys` as "grid_id|alias")
    is the incumbent, and ONLY a synthetic action cell may take the grid from it - by
    beating the incumbent's locally-predicted RSRP by more than `hysteresis_db`.
    Pre-existing neighbors may never take an incumbent's grid here: the master
    full-network run already settled every neighbor-vs-incumbent comparison at higher
    fidelity, and the local rerun's only genuinely NEW information is the synthetic
    cell's RSRP. Allowing neighbors to win on locally-resimulated RSRP re-litigates a
    settled comparison with noisier data - in practice the same neighbor "stole" the
    same lineage grids in every action scenario, printing the identical bogus
    overload for carrier addition and new site alike.

    Grids with no incumbent candidate in this rerun (the master server was removed by
    a sector split, or isn't predicted locally) reselect by plain best RSRP: the
    reselection is forced and there is nothing to stick to.
    Original-vs-synthetic is decided by ID-universe membership (`original_cell_ids`),
    never name patterns - the tooling mangles names between layers ('901053_1__ADD900'
    comes back as '1_ADD900', pre-existing '364_1__MB850' as '1_MB850').
    """
    work = topo_work.copy()
    id_cols = [c for c in ["Node_Cell_ID", "original_node_cell_id", "original_cell_id"] if c in work.columns]

    is_original = pd.Series(False, index=work.index)
    for col in id_cols:
        is_original = is_original | work[col].astype(str).str.strip().isin(original_cell_ids)
    work["_is_synthetic_candidate"] = ~is_original

    gid_str = work["grid_id"].astype(str)
    is_incumbent = pd.Series(False, index=work.index)
    for col in id_cols:
        keys = gid_str + "|" + work[col].astype(str).str.strip()
        is_incumbent = is_incumbent | keys.isin(master_serving_keys)
    work["_is_incumbent"] = is_incumbent

    ranked = work.sort_values(["grid_id", "time_bucket", "_rank_rsrp"], ascending=[True, True, False])
    best_overall = ranked.drop_duplicates(subset=["grid_id", "time_bucket"], keep="first")
    best_incumbent = ranked.loc[ranked["_is_incumbent"]].drop_duplicates(subset=["grid_id", "time_bucket"], keep="first")
    best_synthetic = ranked.loc[ranked["_is_synthetic_candidate"]].drop_duplicates(subset=["grid_id", "time_bucket"], keep="first")
    overall_keys = pd.MultiIndex.from_frame(best_overall[["grid_id", "time_bucket"]])

    inc_rsrp = (
        best_incumbent.set_index(["grid_id", "time_bucket"])["_rank_rsrp"].reindex(overall_keys).to_numpy()
        if not best_incumbent.empty
        else np.full(len(best_overall), np.nan)
    )
    syn_rsrp = (
        best_synthetic.set_index(["grid_id", "time_bucket"])["_rank_rsrp"].reindex(overall_keys).to_numpy()
        if not best_synthetic.empty
        else np.full(len(best_overall), np.nan)
    )
    has_incumbent = ~pd.isna(inc_rsrp)
    has_synthetic = ~pd.isna(syn_rsrp)
    # The synthetic action cell may only take grids belonging to the acted-on sector
    # (lineage). Without this, a new low-band carrier "wins" dozens of other sectors'
    # grids across the whole local scope, hoovering their users (400%+ RRC) and
    # inheriting demand calibrated against much larger capacities (170%+ PRB) - the
    # evaluation's question is whether THIS sector's congestion is fixed.
    in_lineage = (
        pd.to_numeric(best_overall["grid_id"], errors="coerce").isin(lineage_grid_ids).to_numpy()
        if lineage_grid_ids
        else np.zeros(len(best_overall), dtype=bool)
    )
    synthetic_takes = has_incumbent & in_lineage & has_synthetic & (syn_rsrp >= (inc_rsrp + hysteresis_db))
    incumbent_keeps = has_incumbent & ~synthetic_takes
    # Forced reselection (incumbent removed): the synthetic replacement inherits when
    # it covers the grid; unrelated neighbors only when no synthetic candidate exists.
    no_incumbent_synthetic = ~has_incumbent & has_synthetic
    no_incumbent_plain = ~has_incumbent & ~has_synthetic

    parts = [best_overall.loc[no_incumbent_plain]]
    for mask, source in [
        (incumbent_keeps, best_incumbent),
        (synthetic_takes, best_synthetic),
        (no_incumbent_synthetic, best_synthetic),
    ]:
        if mask.any():
            parts.append(
                source.merge(
                    best_overall.loc[mask, ["grid_id", "time_bucket"]],
                    on=["grid_id", "time_bucket"], how="inner",
                )
            )
    out = pd.concat(parts, ignore_index=True, sort=False)

    # Expose whether each grid's master server was even present as a local candidate.
    # A grid whose incumbent is absent (server outside the local site pool) cannot be
    # reasoned about here: its demand belongs to an out-of-scope cell the action never
    # touched, and letting best-RSRP hand it to some nearby cell piles foreign demand
    # onto that cell and fabricates 1000%+ utilizations.
    incumbent_flags = pd.DataFrame(
        {
            "grid_id": best_overall["grid_id"].to_numpy(),
            "time_bucket": best_overall["time_bucket"].to_numpy(),
            "_grid_has_incumbent": has_incumbent,
        }
    ).drop_duplicates(subset=["grid_id", "time_bucket"])
    out = out.merge(incumbent_flags, on=["grid_id", "time_bucket"], how="left")
    return out.drop(columns=["_is_synthetic_candidate", "_is_incumbent"], errors="ignore")


def _build_current_inventory_from_surface(
    *,
    baseline_local: pd.DataFrame,
    corrected_local: pd.DataFrame,
    local_site_df: pd.DataFrame,
    local_kpi: pd.DataFrame,
    local_geo: pd.DataFrame,
    context: dict[str, Any],
    config: CurrentModel3Config,
    lineage_grid_ids: set[int] | None = None,
    source_capacity_fallback: float | None = None,
) -> tuple[pd.DataFrame, pd.DataFrame]:
    for frame in [baseline_local, corrected_local]:
        for col in ["serving_distance_m", "nearest_site_distance_m", "site_count_250m", "site_count_500m", "azimuth_delta_deg"]:
            if col in frame.columns:
                frame.drop(columns=[col], inplace=True)

    dense_part3 = future_rules._build_dense_part3_features(
        baseline_pred_df=baseline_local,
        corrected_pred_df=corrected_local,
        kpi_df=local_kpi,
        geo_df=local_geo,
        site_df=local_site_df.drop(columns=["_site_distance_m"], errors="ignore"),
    )
    dense_part3["time_bucket"] = "PART_3"
    dense_part3["grid_id"] = pd.to_numeric(dense_part3["grid_id"], errors="coerce").astype("Int64")

    base_context = context["current_base_df"]
    merged = dense_part3.merge(base_context, on=["grid_id", "time_bucket"], how="left", suffixes=("", "_base"))
    for col in list(merged.columns):
        if col.endswith("_base"):
            original = col[:-5]
            if original in merged.columns:
                merged[original] = merged[original].combine_first(merged[col])
                merged = merged.drop(columns=[col])
            else:
                merged = merged.rename(columns={col: original})
    merged = merged.sort_values(["grid_id", "time_bucket"]).drop_duplicates(subset=["grid_id", "time_bucket"], keep="first").copy()

    topo_work = corrected_local.copy()
    topo_work["grid_id"] = pd.to_numeric(topo_work["grid_id"], errors="coerce").astype("Int64")
    topo_work["_rank_rsrp"] = pd.to_numeric(topo_work.get("pred_rsrp"), errors="coerce")
    topo_best = _select_serving_cell_with_hysteresis(
        topo_work,
        config.carrier_reselection_hysteresis_db,
        context.get("original_topology_cell_ids", set()),
        context.get("master_serving_keys", set()),
        lineage_grid_ids,
    )
    # Keep only grids whose master serving cell participates in this local rerun,
    # plus the lineage grids under evaluation (always kept - for sector split the
    # incumbent is intentionally removed and its grids MUST reselect). Grids whose
    # incumbent is out of scope carry demand belonging to an untouched, out-of-scope
    # cell; aggregating them here fabricates load on whichever cell happens to be
    # nearby.
    if lineage_grid_ids is not None:
        keep_grid = topo_best["_grid_has_incumbent"].fillna(False).astype(bool) | pd.to_numeric(
            topo_best["grid_id"], errors="coerce"
        ).isin(lineage_grid_ids)
        topo_best = topo_best.loc[keep_grid]
    topo_best = topo_best.drop(columns=["_grid_has_incumbent"], errors="ignore")
    topo_keep = [
        col
        for col in topo_best.columns
        if col
        in [
            "grid_id",
            "time_bucket",
            "Node_Cell_ID",
            "Site ID",
            "site_identity_key",
            "sector_identity",
            "sector_identity_key",
            "frontend_site_sector_key",
            "node_cell_sector_key",
            "sector",
            "band",
            "earfcn",
            "nodeb_id",
            "PCI",
            "carrier_load_share",
            "original_node_cell_id",
            "original_cell_id",
            "canonical_sector_id",
            "site_sector_band_key",
            "rf_identity_key",
        ]
    ]
    topo_best = topo_best[topo_keep].copy()
    topo_best = topo_best.rename(
        columns={
            "Site ID": "topology_site_id",
            "band": "topology_band",
            "earfcn": "topology_earfcn",
            "PCI": "topology_pci",
            "nodeb_id": "topology_nodeb_id",
            "sector": "topology_sector",
            "site_identity_key": "topology_site_identity_key",
            "sector_identity": "topology_sector_identity",
            "sector_identity_key": "topology_sector_identity_key",
            "frontend_site_sector_key": "topology_frontend_site_sector_key",
            "node_cell_sector_key": "topology_node_cell_sector_key",
            "canonical_sector_id": "topology_canonical_sector_id",
            "site_sector_band_key": "topology_site_sector_band_key",
            "rf_identity_key": "topology_rf_identity_key",
            "original_node_cell_id": "topology_original_node_cell_id",
            "original_cell_id": "topology_original_cell_id",
            "carrier_load_share": "topology_carrier_load_share",
        }
    )
    merged = merged.merge(topo_best, on=["grid_id", "time_bucket"], how="left", validate="many_to_one")
    enriched = future_builder.model2_builder._add_model2_features(merged)
    enriched, _ = future_builder._add_sector_carrier_capability_fields(enriched)

    bandwidth_mhz = pd.to_numeric(enriched.get("bandwidth_mhz_est"), errors="coerce").replace(0, np.nan).fillna(10.0)
    spectral = future_builder._estimate_spectral_efficiency_bpshz(
        enriched,
        bandwidth_mhz=bandwidth_mhz,
        mimo_layers=future_builder.DEFAULT_MIMO_LAYERS,
        control_overhead=future_builder.DEFAULT_CONTROL_OVERHEAD,
    )
    enriched["estimated_spectral_efficiency_bpshz"] = spectral.round(6)
    enriched["estimated_dl_capacity_mbps"] = (
        bandwidth_mhz * spectral * max(1.0, float(future_builder.DEFAULT_MIMO_LAYERS)) * max(0.1, 1.0 - float(future_builder.DEFAULT_CONTROL_OVERHEAD))
    ).clip(lower=0.1).round(3)

    # PRB = demand / capacity is only meaningful when both sides share one scale.
    # The dataset's demand is not physical Mbps - it is calibrated against the
    # dataset's own capacity column. Mixing it with the physically re-derived
    # capacity above yields impossible 1000%+ utilizations, so for every cell known
    # to the master dataset, override with the dataset-scale capacity; for synthetic
    # action cells, inherit the parent cell's dataset capacity (suffix-stripped
    # lineage); as a last resort use the median dataset capacity.
    capacity_lookup: dict[str, float] = context.get("master_capacity_by_cell_id", {})
    if capacity_lookup:
        cap_id_cols = [
            c for c in ["Node_Cell_ID", "topology_original_cell_id", "topology_original_node_cell_id"]
            if c in enriched.columns
        ]
        median_capacity = float(np.median(list(capacity_lookup.values())))

        def _dataset_scale_capacity(row: pd.Series) -> float:
            values = []
            for col in cap_id_cols:
                raw = row.get(col)
                values.append("" if pd.isna(raw) else str(raw).strip())
            for value in values:
                if value in capacity_lookup:
                    return capacity_lookup[value]
            for value in values:
                parent = re.sub(r"__(ADD\d+|NS|SS[AB])$", "", value)
                if parent and parent in capacity_lookup:
                    return capacity_lookup[parent]
                parent = re.sub(r"_(ADD\d+|NS|SS[AB])$", "", value)
                if parent and parent in capacity_lookup:
                    return capacity_lookup[parent]
                if re.search(r"_(ADD\d+|NS|SS[AB])$", value) and source_capacity_fallback is not None:
                    return float(source_capacity_fallback)
            return median_capacity

        enriched["estimated_dl_capacity_mbps"] = (
            pd.to_numeric(enriched.apply(_dataset_scale_capacity, axis=1), errors="coerce")
            .fillna(median_capacity)
            .clip(lower=0.1)
            .round(3)
        )

    # Grids with no serving candidate in the corrected surface have no Node_Cell_ID;
    # keeping them would pool all their demand into a single phantom NaN "cell".
    enriched = enriched.loc[enriched["Node_Cell_ID"].notna()].copy()
    enriched = _assign_current_grid_load_to_cells(enriched, config)
    cell_inventory, _ = _build_current_cell_inventory(enriched, config)
    return enriched, cell_inventory


def _choose_load_balance_candidate_current(
    sector_cells: pd.DataFrame,
    threshold: float,
    source_node_cell_id: str | None = None,
) -> dict[str, Any] | None:
    if len(sector_cells) < 2:
        return None
    ordered = sector_cells.sort_values(["prb_rrc_pressure", "grid_count"], ascending=[False, False], na_position="last").reset_index(drop=True)
    source_key = str(source_node_cell_id or "").strip()
    if source_key:
        source_mask = ordered["Node_Cell_ID"].astype(str).str.strip().eq(source_key)
        if not bool(source_mask.any()):
            return None
        source = ordered.loc[source_mask].iloc[0]
        peers = ordered.loc[~source_mask].copy()
    else:
        source = ordered.iloc[0]
        peers = ordered.iloc[1:].copy()
    if "canonical_physical_cell_id" in peers.columns and "canonical_physical_cell_id" in source.index:
        source_canonical = str(source.get("canonical_physical_cell_id") or "").strip()
        if source_canonical:
            peers = peers.loc[peers["canonical_physical_cell_id"].astype(str).str.strip().ne(source_canonical)].copy()
    if "band" in peers.columns and "band" in source.index:
        source_band = str(source.get("band") or "").strip()
        if source_band:
            peers = peers.loc[peers["band"].astype(str).str.strip().ne(source_band)].copy()
    if peers.empty:
        return None
    source_prb = float(source["prb_before_pct"]) if pd.notna(source["prb_before_pct"]) else np.nan
    source_rrc = float(source["rrc_before_pct"]) if pd.notna(source["rrc_before_pct"]) else np.nan
    source_prb_bad = bool(pd.notna(source_prb) and source_prb > threshold)
    source_rrc_bad = bool(pd.notna(source_rrc) and source_rrc > threshold)

    peers["prb_headroom"] = threshold - pd.to_numeric(peers["prb_before_pct"], errors="coerce")
    peers["rrc_headroom"] = threshold - pd.to_numeric(peers["rrc_before_pct"], errors="coerce")
    if source_prb_bad and not source_rrc_bad:
        peers = peers.loc[peers["prb_headroom"] > 0].copy()
        peers["lb_score"] = peers["prb_headroom"] + (0.10 * peers["rrc_headroom"].clip(lower=-200.0))
        congestion_mode = "PRB_ONLY"
    elif source_rrc_bad and not source_prb_bad:
        peers = peers.loc[peers["rrc_headroom"] > 0].copy()
        peers["lb_score"] = peers["rrc_headroom"] + (0.10 * peers["prb_headroom"].clip(lower=-200.0))
        congestion_mode = "RRC_ONLY"
    else:
        peers = peers.loc[(peers["prb_headroom"] > 0) | (peers["rrc_headroom"] > 0)].copy()
        peers["lb_score"] = peers["prb_headroom"].clip(lower=0.0) + peers["rrc_headroom"].clip(lower=0.0)
        congestion_mode = "PRB_AND_RRC"
    if peers.empty:
        return None
    peers = peers.sort_values(["lb_score", "estimated_dl_capacity_mbps", "grid_count"], ascending=[False, False, False], na_position="last")
    target = peers.iloc[0]
    return {
        "source_node_cell_id": str(source["Node_Cell_ID"]),
        "target_node_cell_id": str(target["Node_Cell_ID"]),
        "target_band": str(target["band"]),
        "source_row": source,
        "target_row": target,
        "congestion_mode": congestion_mode,
    }


def _run_load_balance_current(
    sector_cells: pd.DataFrame,
    config: CurrentModel3Config,
    source_node_cell_id: str | None = None,
) -> dict[str, Any]:
    candidate = _choose_load_balance_candidate_current(sector_cells, config.congestion_threshold, source_node_cell_id)
    if candidate is None:
        return {
            "status": "No Material Change",
            "selected_peer_node_cell_id": "",
            "selected_peer_band": "",
            "projected_prb_after_pct": np.nan,
            "projected_rrc_after_pct": np.nan,
            "projected_rrc_users_after": np.nan,
            "action_reason": "No same-sector carrier met the metric-aware current headroom check for the congested metric.",
            "next_step": "Add Carrier",
            "resimulation_required": False,
            "resimulation_flow": "Deterministic current-state load redistribution only",
            "should_fallthrough": True,
        }
    source = candidate["source_row"]
    target = candidate["target_row"]
    congestion_mode = str(candidate.get("congestion_mode") or "PRB_AND_RRC")
    source_traffic = float(source["estimated_offered_traffic_mbps"]) if pd.notna(source["estimated_offered_traffic_mbps"]) else 0.0
    target_traffic = float(target["estimated_offered_traffic_mbps"]) if pd.notna(target["estimated_offered_traffic_mbps"]) else 0.0
    source_users = float(source["rrc_users_before"]) if pd.notna(source["rrc_users_before"]) else 0.0
    target_users = float(target["rrc_users_before"]) if pd.notna(target["rrc_users_before"]) else 0.0
    source_cap = max(0.1, float(source["estimated_dl_capacity_mbps"])) if pd.notna(source["estimated_dl_capacity_mbps"]) else 0.1
    target_cap = max(0.1, float(target["estimated_dl_capacity_mbps"])) if pd.notna(target["estimated_dl_capacity_mbps"]) else 0.1
    threshold = float(config.congestion_threshold)

    source_prb = float(source["prb_before_pct"]) if pd.notna(source["prb_before_pct"]) else 0.0
    source_rrc = float(source["rrc_before_pct"]) if pd.notna(source["rrc_before_pct"]) else 0.0

    need_share_prb = max(0.0, (source_traffic - ((threshold / 100.0) * source_cap)) / max(source_traffic, 1e-6))
    need_share_rrc = max(0.0, (source_users - ((threshold / 100.0) * float(config.rrc_sector_capacity))) / max(source_users, 1e-6))
    max_share_target_prb = max(0.0, ((((threshold / 100.0) * target_cap) - target_traffic) / max(source_traffic, 1e-6))) if source_traffic > 0 else 0.0
    max_share_target_rrc = max(0.0, ((((threshold / 100.0) * float(config.rrc_sector_capacity)) - target_users) / max(source_users, 1e-6))) if source_users > 0 else 0.0

    if congestion_mode == "PRB_ONLY":
        desired_share = need_share_prb
        allowed_share = max_share_target_prb
    elif congestion_mode == "RRC_ONLY":
        desired_share = need_share_rrc
        allowed_share = max_share_target_rrc
    else:
        desired_share = max(need_share_prb, need_share_rrc)
        allowed_share = min(
            max_share_target_prb if source_traffic > 0 else 1.0,
            max_share_target_rrc if source_users > 0 else 1.0,
        )

    move_share = min(0.60, max(0.0, min(desired_share, allowed_share)))
    if move_share <= 0.0:
        return {
            "status": "Rejected",
            "selected_peer_node_cell_id": str(target["Node_Cell_ID"]),
            "selected_peer_band": str(target["band"]),
            "projected_prb_after_pct": np.nan,
            "projected_rrc_after_pct": np.nan,
            "projected_rrc_users_after": np.nan,
            "action_reason": f"Peer carrier exists, but the metric-aware headroom check found no safe movable share for {congestion_mode}.",
            "next_step": "Add Carrier",
            "resimulation_required": False,
            "resimulation_flow": "Deterministic current-state load redistribution only",
            "should_fallthrough": True,
        }
    moved_traffic = source_traffic * move_share
    moved_users = source_users * move_share
    source_prb_after = ((source_traffic - moved_traffic) / source_cap) * 100.0
    target_prb_after = ((target_traffic + moved_traffic) / target_cap) * 100.0
    source_rrc_after = ((source_users - moved_users) / float(config.rrc_sector_capacity)) * 100.0
    target_rrc_after = ((target_users + moved_users) / float(config.rrc_sector_capacity)) * 100.0
    projected_prb = max(source_prb_after, target_prb_after)
    projected_rrc = max(source_rrc_after, target_rrc_after)

    before_prb_worst = max(float(source["prb_before_pct"]), float(target["prb_before_pct"]))
    before_rrc_worst = max(float(source["rrc_before_pct"]), float(target["rrc_before_pct"]))
    before_pressure = max(before_prb_worst, before_rrc_worst)
    after_pressure = max(projected_prb, projected_rrc)
    worsened_any_metric = projected_prb > before_prb_worst + 0.5 or projected_rrc > before_rrc_worst + 0.5
    improved_pressure = after_pressure < before_pressure - 0.5

    if projected_prb <= threshold and projected_rrc <= threshold and not worsened_any_metric:
        status = "Resolved"
    elif worsened_any_metric:
        status = "Rejected"
    elif improved_pressure:
        status = "Partially Resolved"
    else:
        status = "No Material Change"
    return {
        "status": status,
        "selected_peer_node_cell_id": str(target["Node_Cell_ID"]),
        "selected_peer_band": str(target["band"]),
        "projected_prb_after_pct": round(projected_prb, 3),
        "projected_rrc_after_pct": round(projected_rrc, 3),
        "projected_rrc_users_after": round(max(source_users - moved_users, target_users + moved_users), 3),
        "action_reason": (
            f"Current-state load balancing used a metric-aware {congestion_mode} move share of {move_share:.3f}; "
            f"rejected if either PRB or RRC becomes worse."
        ),
        "next_step": "" if status == "Resolved" else "Add Carrier",
        "resimulation_required": False,
        "resimulation_flow": "Deterministic current-state load redistribution only",
        "should_fallthrough": status != "Resolved",
    }


_LOCAL_GRID_IDENTITY_COLS = ["grid_id", "grid_row", "grid_col", "grid_centroid_lat", "grid_centroid_lon"]
_EARTH_RADIUS_M = 6371000.0


def _force_master_grid_identity(pred_df: pd.DataFrame, point_map: pd.DataFrame, tolerance_m: float) -> pd.DataFrame:
    """Give a locally-resimulated prediction surface the MASTER archive's grid numbering.

    The RF rerun (and `_ensure_grid_group_columns`) number grids by factorizing the
    run's own lat/lon pairs into sequential codes, so "grid 2600" in a local rerun is
    a different physical place than the master archive's grid 2600 - yet the codes
    overlap numerically. Downstream joins (demand attach, lineage lookup, before/after
    comparison) all key on grid_id, so without this the master demand lands on the
    wrong locations and every derived PRB number is physically meaningless.

    Exact lat/lon joins cannot align the two surfaces either: the RF engine lays out
    its own sampling lattice per run with a run-specific origin, so local prediction
    points never coincide exactly with master points. Instead, snap each local point
    to its nearest master point within `tolerance_m` (about half a grid cell) and
    adopt that master point's full grid identity. Points with no master point within
    tolerance are discarded - the master surface doesn't know them, no demand exists
    for them, and they only served as interference context during the RF run.
    """
    from sklearn.neighbors import BallTree

    if pred_df.empty or point_map.empty:
        return future_rules._ensure_grid_group_columns(pred_df.copy())
    work = pred_df.drop(columns=_LOCAL_GRID_IDENTITY_COLS, errors="ignore").copy()
    if not {"lat", "lon"}.issubset(work.columns):
        return future_rules._ensure_grid_group_columns(work)

    pm = point_map.dropna(subset=["lat", "lon"]).drop_duplicates(subset=["lat", "lon"]).reset_index(drop=True)
    work_lat = pd.to_numeric(work["lat"], errors="coerce")
    work_lon = pd.to_numeric(work["lon"], errors="coerce")
    valid = work_lat.notna() & work_lon.notna()
    work = work.loc[valid].copy()
    if work.empty:
        return future_rules._ensure_grid_group_columns(work)

    tree = BallTree(np.radians(pm[["lat", "lon"]].astype(float).to_numpy()), metric="haversine")
    dist_rad, nearest_idx = tree.query(np.radians(work[["lat", "lon"]].astype(float).to_numpy()), k=1)
    dist_m = dist_rad[:, 0] * _EARTH_RADIUS_M
    within = dist_m <= float(tolerance_m)
    work = work.loc[within].copy()
    if work.empty:
        return future_rules._ensure_grid_group_columns(work)
    matched = pm.iloc[nearest_idx[within, 0]].reset_index(drop=True)

    identity_cols = [c for c in _LOCAL_GRID_IDENTITY_COLS if c in matched.columns]
    work = work.reset_index(drop=True)
    for col in identity_cols:
        work[col] = matched[col].to_numpy()
    return future_rules._ensure_grid_group_columns(work)


def _fix_synthetic_frequency(site_df: pd.DataFrame) -> pd.DataFrame:
    """Make synthetic topology rows radiate at their own band's frequency.

    The topology builders copy the parent row and overwrite `band`/`earfcn` only, but
    the RF engine's propagation model reads `frequency_mhz` - which stays at the
    parent's value. A "900MHz" carrier addition therefore simulates with 1800MHz path
    loss: RSRP identical to the parent, never beats the reselection hysteresis, and
    the new carrier never attracts a single grid. Only synthetic rows are touched;
    original rows keep the archive's values.
    """
    out = site_df.loc[:, ~site_df.columns.duplicated()].copy()
    if "carrier_variant" not in out.columns or "band" not in out.columns:
        return out
    synthetic = out["carrier_variant"].astype(str).str.startswith(("carrier_add", "new_site", "sector_split", "add_sector"))
    if not synthetic.any():
        return out
    band = pd.to_numeric(out.loc[synthetic, "band"], errors="coerce")
    for col in ["frequency_mhz", "frequency"]:
        if col in out.columns:
            out.loc[synthetic, col] = band.fillna(pd.to_numeric(out.loc[synthetic, col], errors="coerce"))
    return out.loc[:, ~out.columns.duplicated()].copy()


def _prepare_production_scope_site_frames(
    *,
    original_site_df: pd.DataFrame,
    action_site_df: pd.DataFrame,
    source_rows: pd.DataFrame,
) -> tuple[pd.DataFrame, pd.DataFrame]:
    """Build original/modified site frames in the same shape production expects."""
    original_input = original_site_df.loc[:, ~original_site_df.columns.duplicated()].copy()
    action_input = action_site_df.loc[:, ~action_site_df.columns.duplicated()].copy()
    for frame in [original_input, action_input]:
        if "frequency_mhz" in frame.columns and "frequency" in frame.columns:
            frame.drop(columns=["frequency"], inplace=True)
    original = production_normalize_site_df(original_input, log_stage="MODEL3_TEST_ORIGINAL_SITE")
    modified = production_normalize_site_df(action_input, log_stage="MODEL3_TEST_ACTION_SITE")

    def _restore_strict_project196_identity(frame: pd.DataFrame) -> pd.DataFrame:
        if "site_prediction_key" not in frame.columns:
            return frame
        out = frame.copy()
        strict = out["site_prediction_key"].map(_clean_project196_text)
        mask = strict.ne("")
        for col in ["Node_Cell_ID", "rf_identity_key"]:
            out.loc[mask, col] = strict.loc[mask]
        if "site_cell_sector_band_operator_key" in out.columns:
            out.loc[mask, "site_cell_sector_band_operator_key"] = strict.loc[mask]
        if "cell_id" in out.columns:
            out.loc[mask, "legacy_nodeb_id_cell_id"] = out.loc[mask, "cell_id"].map(_clean_project196_text)
        out.loc[mask, "canonical_cell_id"] = out.loc[mask, "legacy_nodeb_id_cell_id"].map(production_canonical_cell_id)
        return out

    original = _restore_strict_project196_identity(original)
    modified = _restore_strict_project196_identity(modified)

    for col in [
        "lat",
        "lon",
        "azimuth",
        "electrical_tilt",
        "mechanical_tilt",
        "tx_power",
        "antenna_height",
        "frequency_mhz",
    ]:
        modified[f"orig_{col}"] = pd.to_numeric(modified[col], errors="coerce")

    if source_rows is None or source_rows.empty or "carrier_variant" not in modified.columns:
        return original, modified

    source_input = source_rows.loc[:, ~source_rows.columns.duplicated()].copy()
    if "frequency_mhz" in source_input.columns and "frequency" in source_input.columns:
        source_input.drop(columns=["frequency"], inplace=True)
    source = production_normalize_site_df(source_input, log_stage="MODEL3_TEST_SOURCE_SITE")
    source_numeric = {}
    for col in [
        "lat",
        "lon",
        "azimuth",
        "electrical_tilt",
        "mechanical_tilt",
        "tx_power",
        "antenna_height",
        "frequency_mhz",
    ]:
        values = pd.to_numeric(source.get(col), errors="coerce")
        source_numeric[col] = float(values.dropna().iloc[0]) if values.notna().any() else np.nan

    synthetic_mask = modified["carrier_variant"].astype(str).str.startswith(("carrier_add", "new_site", "sector_split", "add_sector"), na=False)
    for col, value in source_numeric.items():
        if pd.notna(value):
            modified.loc[synthetic_mask, f"orig_{col}"] = value
    return original, modified


def _set_synthetic_identity(row: pd.Series | pd.DataFrame, node_cell_id: str) -> pd.Series | pd.DataFrame:
    for col in ["Node_Cell_ID", "cell_id", "site_prediction_key", "site_cell_sector_band_operator_key", "rf_identity_key"]:
        has_col = col in row.index if isinstance(row, pd.Series) else col in row.columns
        if has_col:
            row[col] = node_cell_id
    if isinstance(row, pd.Series):
        row["canonical_cell_id"] = production_canonical_cell_id(node_cell_id)
        row["legacy_nodeb_id_cell_id"] = node_cell_id
    else:
        row["canonical_cell_id"] = production_canonical_cell_id(node_cell_id)
        row["legacy_nodeb_id_cell_id"] = node_cell_id
    return row


def _vectorized_identity_point_mask(points: pd.DataFrame, identities: list[str]) -> pd.Series:
    wanted = {str(value).strip() for value in identities if str(value).strip()}
    wanted.update(future_rules._normalize_identity_text(value) for value in list(wanted))
    wanted = {value for value in wanted if value}
    mask = pd.Series(False, index=points.index)
    for col in [
        "Node_Cell_ID",
        "rf_identity_key",
        "site_sector_band_key",
        "sector_identity_key",
        "frontend_site_sector_key",
        "node_cell_sector_key",
        "canonical_cell_id",
        "cell_id",
        "original_node_cell_id",
        "original_cell_id",
    ]:
        if col not in points.columns:
            continue
        values = points[col].astype(str).str.strip()
        mask = mask | values.isin(wanted) | values.map(future_rules._normalize_identity_text).isin(wanted)
    return mask


def _attach_site_topology_to_prediction(pred_df: pd.DataFrame, site_df: pd.DataFrame) -> pd.DataFrame:
    if pred_df.empty or site_df.empty or "Node_Cell_ID" not in pred_df.columns:
        return pred_df
    topo_cols = [
        col
        for col in [
            "Node_Cell_ID",
            "Site ID",
            "site_identity_key",
            "sector_identity",
            "frontend_site_sector_key",
            "node_cell_sector_key",
            "sector_identity_key",
            "site_sector_band_key",
            "rf_identity_key",
            "sector",
            "band",
            "earfcn",
            "nodeb_id",
            "PCI",
            "azimuth",
            "canonical_sector_id",
            "original_node_cell_id",
            "original_cell_id",
            "carrier_load_share",
        ]
        if col in site_df.columns
    ]
    if len(topo_cols) <= 1:
        return pred_df
    topo = site_df[topo_cols].drop_duplicates(subset=["Node_Cell_ID"], keep="first").copy()
    out = pred_df.drop(columns=[col for col in topo.columns if col != "Node_Cell_ID" and col in pred_df.columns], errors="ignore").copy()
    out["Node_Cell_ID"] = out["Node_Cell_ID"].astype(str)
    topo["Node_Cell_ID"] = topo["Node_Cell_ID"].astype(str)
    return out.merge(topo, on="Node_Cell_ID", how="left", validate="many_to_one")


def _append_synthetic_prediction_points(
    *,
    scoped_points: pd.DataFrame,
    baseline_points: pd.DataFrame,
    sector_cells: pd.DataFrame,
    changed_rows: pd.DataFrame,
    original_cell_ids: set[str],
) -> pd.DataFrame:
    if changed_rows.empty or "Node_Cell_ID" not in changed_rows.columns:
        return scoped_points
    synthetic_ids = [
        str(value).strip()
        for value in changed_rows["Node_Cell_ID"].dropna().astype(str).unique().tolist()
        if str(value).strip() and str(value).strip() not in original_cell_ids
    ]
    if not synthetic_ids:
        return scoped_points

    lineage_ids: list[str] = []
    for col in ["Node_Cell_ID", "topology_original_cell_id", "topology_original_node_cell_id"]:
        if col in sector_cells.columns:
            lineage_ids.extend(str(value).strip() for value in sector_cells[col].dropna().astype(str).tolist() if str(value).strip())
    lineage_mask = _vectorized_identity_point_mask(baseline_points, lineage_ids)
    lineage_points = baseline_points.loc[lineage_mask].copy()
    if lineage_points.empty:
        lineage_points = scoped_points.copy()
    synthetic_parts = []
    for synthetic_id in synthetic_ids:
        dup = lineage_points.copy()
        dup["Node_Cell_ID"] = synthetic_id
        dup["nodeb_id_cell_id"] = synthetic_id
        dup["canonical_cell_id"] = future_rules._normalize_identity_text(synthetic_id)
        for col in ["rf_identity_key", "site_sector_band_key", "sector_identity_key"]:
            if col in dup.columns:
                dup[col] = synthetic_id
        synthetic_parts.append(dup)
    combined = pd.concat([scoped_points, *synthetic_parts], ignore_index=True, sort=False)
    dedupe_cols = [col for col in ["Node_Cell_ID", "lat", "lon", "grid_id", "time_bucket"] if col in combined.columns]
    return combined.drop_duplicates(subset=dedupe_cols, keep="first") if dedupe_cols else combined


def _append_affected_alias_prediction_points(
    *,
    scoped_points: pd.DataFrame,
    baseline_points: pd.DataFrame,
    modified_site_df: pd.DataFrame,
    affected_cells: list[str],
) -> pd.DataFrame:
    if scoped_points.empty or baseline_points.empty or modified_site_df.empty or "Node_Cell_ID" not in modified_site_df.columns:
        return scoped_points
    combined = scoped_points.copy()
    existing = set(combined.get("Node_Cell_ID", pd.Series(dtype=object)).dropna().astype(str).str.strip())
    parts = [combined]

    for cell in [str(value).strip() for value in affected_cells if str(value).strip()]:
        if cell in existing:
            continue
        site_rows = modified_site_df.loc[modified_site_df["Node_Cell_ID"].astype(str).str.strip().eq(cell)].copy()
        if site_rows.empty:
            continue
        aliases = {cell, future_rules._normalize_identity_text(cell)}
        for col in [
            "Node_Cell_ID",
            "nodeb_id_cell_id",
            "legacy_nodeb_id_cell_id",
            "rf_identity_key",
            "site_sector_band_key",
            "sector_identity_key",
            "canonical_cell_id",
            "original_node_cell_id",
            "original_cell_id",
        ]:
            if col in site_rows.columns:
                aliases.update(str(value).strip() for value in site_rows[col].dropna().astype(str).tolist() if str(value).strip())
                aliases.update(
                    future_rules._normalize_identity_text(value)
                    for value in site_rows[col].dropna().astype(str).tolist()
                    if str(value).strip()
                )
        aliases = {value for value in aliases if value and value.lower() not in {"nan", "none", "null", "<na>"}}
        alias_mask = _vectorized_identity_point_mask(baseline_points, list(aliases))
        dup = baseline_points.loc[alias_mask].copy()
        if dup.empty:
            continue
        dup["Node_Cell_ID"] = cell
        dup["nodeb_id_cell_id"] = cell
        dup["canonical_cell_id"] = future_rules._normalize_identity_text(cell)
        for col in ["rf_identity_key", "site_sector_band_key", "sector_identity_key"]:
            if col in dup.columns:
                dup[col] = cell
        parts.append(dup)
        existing.add(cell)

    if len(parts) == 1:
        return combined
    out = pd.concat(parts, ignore_index=True, sort=False)
    dedupe_cols = [col for col in ["Node_Cell_ID", "lat", "lon", "grid_id", "time_bucket"] if col in out.columns]
    return out.drop_duplicates(subset=dedupe_cols, keep="first") if dedupe_cols else out


def _affected_lineage_grid_ids(sector_cells: pd.DataFrame, current_df: pd.DataFrame) -> list[int]:
    """Grid ids that the archive's full-network surface actually assigned to the
    cells being acted on, independent of distance from the site coordinate."""
    lineage_ids: set[str] = set()
    for col in ["Node_Cell_ID", "topology_original_cell_id", "topology_original_node_cell_id"]:
        if col in sector_cells.columns:
            lineage_ids.update(
                str(v).strip() for v in sector_cells[col].dropna().astype(str).tolist() if str(v).strip()
            )
    if not lineage_ids or current_df.empty:
        return []
    mask = pd.Series(False, index=current_df.index)
    for col in ["Node_Cell_ID", "topology_original_cell_id", "topology_original_node_cell_id"]:
        if col in current_df.columns:
            mask = mask | current_df[col].astype(str).isin(lineage_ids)
    return pd.to_numeric(current_df.loc[mask, "grid_id"], errors="coerce").dropna().astype(int).unique().tolist()


def _prepare_local_action_context_current(
    *,
    sector_cells: pd.DataFrame,
    site_df: pd.DataFrame,
    context: dict[str, Any],
    config: CurrentModel3Config,
    logger,
    action_label: str,
    source_rows: pd.DataFrame,
) -> dict[str, Any] | None:
    """Model 3 current-recommendation variant of `future_rules._prepare_local_action_context`.

    The shared helper caps candidate interferers at the 25 nearest sites and grid
    coverage at a fixed radius, which makes the resimulated "after" PRB incomparable
    to the "before" PRB (computed from the full-network archive): the same physical
    cell gets a different SINR/capacity purely from interferers being dropped, and
    grids outside the radius silently disappear from the aggregation. This variant
    widens the interferer pool to match the original run's `max_interference_sites`
    cap and force-includes every grid the full-network surface actually assigned to
    the affected cells, so the local rerun stays apples-to-apples with "before".
    """
    baseline_all = context["baseline_pred_df"]
    corrected_all = context["corrected_pred_df"]
    detail_all = context["coverage_rows_df"]
    kpi_all = context["kpi_df"]
    geo_all = context["geo_df"]

    part3_baseline = baseline_all.loc[baseline_all["time_bucket"].astype(str) == "PART_3"].copy()
    part3_corrected = corrected_all.loc[corrected_all["time_bucket"].astype(str) == "PART_3"].copy()
    part3_detail = detail_all.loc[detail_all["time_bucket"].astype(str) == "PART_3"].copy()
    part3_kpi = kpi_all.loc[kpi_all["time_bucket"].astype(str) == "PART_3"].copy()
    part3_geo = geo_all.loc[geo_all["time_bucket"].astype(str) == "PART_3"].copy()

    if source_rows is None or source_rows.empty:
        source_rows = future_rules._extract_source_site_rows(sector_cells, site_df)
    if source_rows.empty:
        logger.info("%s_prepare_failed reason=no_source_rows", action_label)
        return None

    site_lat = pd.to_numeric(source_rows["lat"], errors="coerce").mean()
    site_lon = pd.to_numeric(source_rows["lon"], errors="coerce").mean()
    site_work = site_df.copy()
    site_work["_site_distance_m"] = future_rules._haversine_distance_m_series(site_work["lat"], site_work["lon"], site_lat, site_lon)
    site_distance_df = (
        site_work.groupby("Site ID", as_index=False)
        .agg(site_distance_m=("_site_distance_m", "min"))
        .sort_values("site_distance_m", ascending=True)
    )
    max_interference_sites = int(context["summary"].get("max_interference_sites", 50))
    nearest_site_count = max(60, max_interference_sites + 15)
    nearest_site_ids = site_distance_df["Site ID"].head(nearest_site_count).astype(str).tolist()
    local_site_df = site_work.loc[site_work["Site ID"].astype(str).isin(nearest_site_ids)].copy()
    if local_site_df.empty:
        local_site_df = site_work.copy()

    local_point_map = future_rules._make_local_point_map(part3_baseline)
    if local_point_map.empty:
        logger.info("%s_prepare_failed reason=no_point_map", action_label)
        return None
    local_point_map["_site_distance_m"] = future_rules._haversine_distance_m_series(local_point_map["lat"], local_point_map["lon"], site_lat, site_lon)

    lineage_grid_ids = _affected_lineage_grid_ids(sector_cells, context.get("current_df", pd.DataFrame()))
    radius_mask = local_point_map["_site_distance_m"] <= float(config.sector_split_local_radius_m)
    lineage_mask = pd.to_numeric(local_point_map["grid_id"], errors="coerce").astype("Int64").isin(lineage_grid_ids)
    local_point_map = local_point_map.loc[radius_mask | lineage_mask].copy()
    if local_point_map.empty:
        local_point_map = future_rules._make_local_point_map(
            part3_baseline.loc[part3_baseline["Node_Cell_ID"].astype(str).isin(sector_cells["Node_Cell_ID"].astype(str))]
        )
    if local_point_map.empty:
        logger.info("%s_prepare_failed reason=no_local_points", action_label)
        return None

    affected_grid_ids = pd.to_numeric(local_point_map["grid_id"], errors="coerce").dropna().astype("Int64").unique().tolist()
    local_detail = pd.DataFrame()
    if affected_grid_ids and "grid_id" in part3_detail.columns:
        detail_grid_ids = pd.to_numeric(part3_detail["grid_id"], errors="coerce").astype("Int64")
        local_detail = part3_detail.loc[detail_grid_ids.isin(affected_grid_ids)].copy()
    if local_detail.empty:
        local_detail = part3_detail.merge(local_point_map[["lat", "lon"]].drop_duplicates(), on=["lat", "lon"], how="inner")
    if local_detail.empty:
        local_detail = local_point_map[["lat", "lon", "grid_id", "grid_row", "grid_col", "grid_centroid_lat", "grid_centroid_lon"]].drop_duplicates().copy()
        local_detail["time_bucket"] = "PART_3"

    local_kpi = part3_kpi.loc[pd.to_numeric(part3_kpi["grid_id"], errors="coerce").astype("Int64").isin(affected_grid_ids)].copy()
    local_geo = part3_geo.loc[pd.to_numeric(part3_geo["grid_id"], errors="coerce").astype("Int64").isin(affected_grid_ids)].copy()
    logger.info(
        "%s_prepare local_points=%d local_detail_rows=%d local_sites=%d lineage_grid_ids=%d",
        action_label,
        len(local_point_map),
        len(local_detail),
        len(local_site_df),
        len(lineage_grid_ids),
    )
    return {
        "part3_baseline": part3_baseline,
        "part3_corrected": part3_corrected,
        "local_site_df": local_site_df,
        "local_point_map": local_point_map,
        "local_detail": local_detail,
        "local_kpi": local_kpi,
        "local_geo": local_geo,
    }


def _evaluate_action_inventory_current(
    *,
    sector_cells: pd.DataFrame,
    after_inventory: pd.DataFrame,
    candidate_node_cell_ids: set[str],
    demand_serving_cell_ids: set[str],
    action_node_cell_ids: set[str] | None,
    config: CurrentModel3Config,
    logger,
    action_label: str,
) -> dict[str, Any]:
    """Model 3 current-recommendation variant of `future_rules._evaluate_action_inventory`.

    The shared function's same-site fallback widens the after-scope to ANY cell at the
    site with nonzero load whenever the matched scope's PRB reads <= 0. That conflates
    two very different situations: (a) lineage matching genuinely failed to find the
    acted-on cell(s) in the after-surface, vs (b) the cell(s) were found and correctly
    show ~0 load because the action successfully evacuated them. Case (b) is exactly what
    a working carrier addition looks like, but the shared logic then substitutes an
    unrelated sibling cell's independent PRB as if it were this action's outcome -
    producing "before=96%, after=115%" results that describe two different physical
    cells, not a before/after of the same one.

    A cell that wins zero grids never gets a row in the collapsed `after_inventory` at
    all (`_build_current_cell_inventory` only emits rows for cells present in at least
    one grid's winning assignment) - so "absent from after_inventory" alone can't tell
    apart "genuinely evacuated to 0%" from "never a valid candidate in this rerun".
    `candidate_node_cell_ids` (the RF-predicted candidate surface, before winner
    selection collapses it) resolves that: if the acted-on cell was a real candidate,
    trust the 0% reading; only fall back to the site-wide search if it wasn't even a
    candidate here (e.g. removed by sector split, or dropped by the local topology).
    """
    before_cells = sector_cells["Node_Cell_ID"].astype(str).tolist()
    source_sector_id = str(future_rules._first_non_empty(sector_cells["sector_id"])).strip()
    source_site_id = future_rules._normalize_identity_text(future_rules._first_non_empty(sector_cells["site_id"]))
    source_frontend_sector_keys = {source_sector_id} if source_sector_id else set()
    source_topology_sector_keys: set[str] = set()
    for col in [
        "topology_frontend_site_sector_key",
        "topology_node_cell_sector_key",
        "topology_sector_identity_key",
        "topology_canonical_sector_id",
        "topology_site_sector_band_key",
        "sector_id",
    ]:
        if col in sector_cells.columns:
            source_topology_sector_keys.update(
                str(value).strip()
                for value in sector_cells[col].dropna().astype(str).tolist()
                if str(value).strip()
            )
    source_frontend_sector_keys.update(source_topology_sector_keys)
    original_lineage_ids = set(value for value in before_cells if str(value).strip())
    for col in ["topology_original_cell_id", "topology_original_node_cell_id"]:
        if col in sector_cells.columns:
            original_lineage_ids.update(
                str(value).strip()
                for value in sector_cells[col].dropna().astype(str).tolist()
                if str(value).strip()
            )

    def _scope_mask(df: pd.DataFrame, *, sector_keys: set[str], lineage_ids: set[str], site_id: str) -> pd.Series:
        mask = pd.Series(False, index=df.index)
        if source_sector_id and "sector_id" in df.columns:
            mask = mask | df["sector_id"].astype(str).eq(source_sector_id)
        for col in [
            "sector_id",
            "topology_frontend_site_sector_key",
            "topology_node_cell_sector_key",
            "topology_sector_identity_key",
            "topology_canonical_sector_id",
            "topology_site_sector_band_key",
        ]:
            if col in df.columns and sector_keys:
                mask = mask | df[col].astype(str).isin(sector_keys)
        for col in ["Node_Cell_ID", "topology_original_cell_id", "topology_original_node_cell_id"]:
            if col in df.columns and lineage_ids:
                mask = mask | df[col].astype(str).isin(lineage_ids)
        if site_id and "site_id" in df.columns:
            mask = mask | df["site_id"].map(future_rules._normalize_identity_text).eq(site_id)
        return mask

    primary_mask = _scope_mask(after_inventory, sector_keys=source_frontend_sector_keys, lineage_ids=original_lineage_ids, site_id="")
    primary_matches = after_inventory.loc[primary_mask].copy()

    affected_sector_keys = set(source_frontend_sector_keys)
    if not primary_matches.empty:
        for col in [
            "sector_id",
            "topology_frontend_site_sector_key",
            "topology_node_cell_sector_key",
            "topology_sector_identity_key",
            "topology_canonical_sector_id",
            "topology_site_sector_band_key",
        ]:
            if col in primary_matches.columns:
                affected_sector_keys.update(
                    str(value).strip()
                    for value in primary_matches[col].dropna().astype(str).tolist()
                    if str(value).strip()
                )
        for col in ["Node_Cell_ID", "topology_original_cell_id", "topology_original_node_cell_id"]:
            if col in primary_matches.columns:
                original_lineage_ids.update(
                    str(value).strip()
                    for value in primary_matches[col].dropna().astype(str).tolist()
                    if str(value).strip()
                )

    final_mask = _scope_mask(after_inventory, sector_keys=affected_sector_keys, lineage_ids=original_lineage_ids, site_id="")
    # Follow the demand: cells that now serve the congested lineage grids decide the
    # outcome, whether they are the original cell, the synthetic candidate, or a sibling.
    if demand_serving_cell_ids and "Node_Cell_ID" in after_inventory.columns:
        final_mask = final_mask | after_inventory["Node_Cell_ID"].astype(str).str.strip().isin(demand_serving_cell_ids)
    after_candidates = after_inventory.loc[final_mask].copy()

    after_lineage_ids: set[str] = set()
    for col in ["Node_Cell_ID", "topology_original_cell_id", "topology_original_node_cell_id"]:
        if col in after_inventory.columns:
            after_lineage_ids.update(after_inventory[col].dropna().astype(str))
    before_cells_found_in_after = bool(original_lineage_ids & after_lineage_ids)
    before_cells_were_candidates = bool(original_lineage_ids & candidate_node_cell_ids)

    scope_mode = "matched_scope"
    evacuated_to_zero = False

    if after_candidates.empty and not before_cells_found_in_after and before_cells_were_candidates:
        # The acted-on cell(s) were legitimate RF candidates in this rerun (they had
        # predicted RSRP for at least one grid) but won zero grids after reselection -
        # `_build_current_cell_inventory` never emits a row for a cell with 0 assigned
        # grids, so this is a real "evacuated to 0%" outcome, not a matching gap. Trust
        # it directly instead of substituting an unrelated cell's independent load.
        evacuated_to_zero = True
        scope_mode = "evacuated_zero"
    elif after_candidates.empty and source_site_id:
        after_site_ids = after_inventory["site_id"].map(future_rules._normalize_identity_text) if "site_id" in after_inventory.columns else pd.Series("", index=after_inventory.index)
        after_candidates = after_inventory.loc[after_site_ids == source_site_id].copy()

    def _summarize_scope(df: pd.DataFrame) -> tuple[float, float, float]:
        prb = future_rules._to_num(df["prb_before_pct"])
        rrc = future_rules._to_num(df["rrc_before_pct"])
        users = future_rules._to_num(df["rrc_users_before"]) if "rrc_users_before" in df.columns else pd.Series(dtype=float)
        return (
            float(prb.max()) if prb.notna().any() else np.nan,
            float(rrc.max()) if rrc.notna().any() else np.nan,
            float(users.max()) if users.notna().any() else np.nan,
        )

    if evacuated_to_zero:
        projected_prb, projected_rrc, projected_users = 0.0, 0.0, 0.0
    else:
        projected_prb, projected_rrc, projected_users = _summarize_scope(after_candidates)
    before_prb = float(future_rules._to_num(sector_cells["prb_before_pct"]).max()) if future_rules._to_num(sector_cells["prb_before_pct"]).notna().any() else np.nan
    before_rrc = float(future_rules._to_num(sector_cells["rrc_before_pct"]).max()) if future_rules._to_num(sector_cells["rrc_before_pct"]).notna().any() else np.nan
    before_pressure = max(before_prb if pd.notna(before_prb) else 0.0, before_rrc if pd.notna(before_rrc) else 0.0)

    # Only widen scope when the acted-on cell(s) are genuinely absent from the after
    # surface AND were never even a valid candidate here (a real lineage/topology
    # failure) - not when they were a legitimate candidate that simply lost every grid.
    if (
        not evacuated_to_zero
        and source_site_id
        and not before_cells_found_in_after
        and pd.notna(before_prb)
        and before_prb > 0
        and (not pd.notna(projected_prb) or projected_prb <= 0.0)
        and "site_id" in after_inventory.columns
    ):
        site_scope = after_inventory.loc[after_inventory["site_id"].map(future_rules._normalize_identity_text).eq(source_site_id)].copy()
        site_scope = site_scope.loc[
            (future_rules._to_num(site_scope.get("prb_before_pct", pd.Series(dtype=float))) > 0.0)
            | (future_rules._to_num(site_scope.get("rrc_before_pct", pd.Series(dtype=float))) > 0.0)
            | (future_rules._to_num(site_scope.get("rrc_users_before", pd.Series(dtype=float))) > 0.0)
        ].copy()
        if not site_scope.empty:
            after_candidates = site_scope
            projected_prb, projected_rrc, projected_users = _summarize_scope(after_candidates)
            scope_mode = "same_site_nonzero_fallback"

    share_factor = 1.0
    if not evacuated_to_zero and not after_candidates.empty and "Node_Cell_ID" in after_candidates.columns:
        after_ids = after_candidates["Node_Cell_ID"].dropna().astype(str).str.strip()
        if action_label == "current_carrier_add":
            # Carrier addition does not mean the winning new carrier owns 100% of
            # the sector's demand. Once RF confirms the new carrier is usable for
            # the affected footprint, the offered traffic is shared across the
            # existing sector carriers plus the added carrier.
            source_carriers = max(1, int(len(sector_cells)))
            share_factor = float(max(2, source_carriers + 1))
        elif action_label == "current_sector_add" and after_ids.str.contains(r"(^|_)AS\d+$", regex=True).any():
            added_count = int(after_ids.str.contains(r"(^|_)AS\d+$", regex=True).sum())
            share_factor = float(max(2, added_count + 1))
        elif action_label == "current_sector_split" and after_ids.str.contains(r"(_|__)SS[AB]$", regex=True).any():
            share_factor = 2.0
        elif action_label == "current_new_site":
            ns_count = int(after_ids.str.contains(r"(_|__)NS$", regex=True).sum())
            if ns_count > 0:
                share_factor = float(max(2, ns_count))
    if share_factor > 1.0:
        raw_projected_prb = projected_prb
        raw_projected_rrc = projected_rrc
        raw_projected_users = projected_users
        if pd.notna(projected_prb):
            projected_prb = projected_prb / share_factor
        if pd.notna(projected_rrc):
            projected_rrc = projected_rrc / share_factor
        if pd.notna(projected_users):
            projected_users = projected_users / share_factor
        logger.info(
            "%s_load_share_adjustment factor=%.3f raw_prb=%.3f raw_rrc=%.3f adjusted_prb=%.3f adjusted_rrc=%.3f raw_users=%.3f adjusted_users=%.3f",
            action_label,
            share_factor,
            raw_projected_prb if pd.notna(raw_projected_prb) else -1.0,
            raw_projected_rrc if pd.notna(raw_projected_rrc) else -1.0,
            projected_prb if pd.notna(projected_prb) else -1.0,
            projected_rrc if pd.notna(projected_rrc) else -1.0,
            raw_projected_users if pd.notna(raw_projected_users) else -1.0,
            projected_users if pd.notna(projected_users) else -1.0,
        )

    after_pressure = max(projected_prb if pd.notna(projected_prb) else 0.0, projected_rrc if pd.notna(projected_rrc) else 0.0)
    resolved = pd.notna(projected_prb) and pd.notna(projected_rrc) and projected_prb <= config.congestion_threshold and projected_rrc <= config.congestion_threshold
    improved = after_pressure < before_pressure - 0.5
    worsened_any_metric = (
        (pd.notna(projected_prb) and pd.notna(before_prb) and projected_prb > before_prb + 0.5)
        or (pd.notna(projected_rrc) and pd.notna(before_rrc) and projected_rrc > before_rrc + 0.5)
    )
    action_ids = {str(value).strip() for value in (action_node_cell_ids or set()) if str(value).strip()}
    demand_on_action_cell = bool(action_ids & demand_serving_cell_ids)
    requires_action_cell_capture = action_label in {"current_sector_add", "current_new_site"}
    if resolved and not worsened_any_metric:
        status = "Resolved"
    elif worsened_any_metric:
        status = "Rejected"
    elif improved:
        status = "Partially Resolved"
    else:
        status = "No Material Change"
    if requires_action_cell_capture and not demand_on_action_cell:
        status = "Rejected" if worsened_any_metric else "No Material Change"
    after_cells = after_candidates["Node_Cell_ID"].astype(str).dropna().tolist() if "Node_Cell_ID" in after_candidates.columns else []
    logger.info(
        "%s_done sector=%s before_cells=%s after_cells=%s demand_serving_cells=%s action_cells=%s action_cell_served=%s local_after_rows=%d scope_mode=%s before_cells_found_in_after=%s share_factor=%.3f before_prb=%.3f before_rrc=%.3f actual_prb=%.3f actual_rrc=%.3f worsened_any_metric=%s status=%s",
        action_label,
        source_sector_id,
        before_cells,
        after_cells,
        sorted(demand_serving_cell_ids),
        sorted(action_ids),
        demand_on_action_cell,
        len(after_candidates),
        scope_mode,
        before_cells_found_in_after,
        share_factor,
        before_prb if pd.notna(before_prb) else -1.0,
        before_rrc if pd.notna(before_rrc) else -1.0,
        projected_prb if pd.notna(projected_prb) else -1.0,
        projected_rrc if pd.notna(projected_rrc) else -1.0,
        worsened_any_metric,
        status,
    )
    return {
        "status": status,
        "projected_prb_after_pct": round(projected_prb, 3) if pd.notna(projected_prb) else np.nan,
        "projected_rrc_after_pct": round(projected_rrc, 3) if pd.notna(projected_rrc) else np.nan,
        "projected_rrc_users_after": round(projected_users, 3) if pd.notna(projected_users) else np.nan,
        "next_step": "" if resolved else "New Site",
        "after_cells": after_cells,
    }


def _compact_current_action_scope(
    *,
    affected_cells: list[str],
    affected_sites: list[str],
    changed_rows: pd.DataFrame,
    sector_cells: pd.DataFrame,
    modified_site_df: pd.DataFrame,
    source_rows: pd.DataFrame,
    config: CurrentModel3Config,
) -> tuple[list[str], list[str]]:
    """Keep Model 3 action RF scope close to production manual optimization.

    The optimized engine's affected scope includes every same-site identity. For
    synthetic Model 3 actions that can mean many extra generated carriers, even
    though the decision only evaluates this sector's demand movement. Keep the
    changed action cells, the source sector cells, and a small nearest-neighbor set.
    """
    wanted_order: list[str] = []

    def add(values) -> None:
        for value in values:
            text = str(value).strip()
            if text and text.lower() not in {"nan", "none", "null", "<na>"} and text not in wanted_order:
                wanted_order.append(text)

    for col in ["Node_Cell_ID", "topology_original_cell_id", "topology_original_node_cell_id"]:
        if col in sector_cells.columns:
            add(sector_cells[col].dropna().astype(str).tolist())
    if "Node_Cell_ID" in changed_rows.columns:
        add(changed_rows["Node_Cell_ID"].dropna().astype(str).tolist())

    affected_set = {str(cell).strip() for cell in affected_cells if str(cell).strip()}
    work = modified_site_df.copy()
    if "Node_Cell_ID" not in work.columns or not {"lat", "lon"}.issubset(work.columns):
        compact = [cell for cell in wanted_order if cell in affected_set or cell not in affected_set]
        return compact or list(affected_cells), list(affected_sites)

    source_lat = pd.to_numeric(source_rows.get("lat"), errors="coerce").mean() if not source_rows.empty else np.nan
    source_lon = pd.to_numeric(source_rows.get("lon"), errors="coerce").mean() if not source_rows.empty else np.nan
    if pd.isna(source_lat) or pd.isna(source_lon):
        source_lat = pd.to_numeric(work.loc[work["Node_Cell_ID"].astype(str).isin(wanted_order), "lat"], errors="coerce").mean()
        source_lon = pd.to_numeric(work.loc[work["Node_Cell_ID"].astype(str).isin(wanted_order), "lon"], errors="coerce").mean()

    if not (pd.isna(source_lat) or pd.isna(source_lon)):
        neighbor_rows = work.loc[
            work["Node_Cell_ID"].astype(str).isin(affected_set)
            & ~work["Node_Cell_ID"].astype(str).isin(set(wanted_order))
        ].copy()
        if not neighbor_rows.empty:
            neighbor_rows["_distance_m"] = future_rules._haversine_distance_m_series(
                neighbor_rows["lat"],
                neighbor_rows["lon"],
                float(source_lat),
                float(source_lon),
            )
            nearest_neighbors = (
                neighbor_rows.sort_values(["_distance_m", "Node_Cell_ID"], ascending=[True, True])
                .drop_duplicates(subset=["Node_Cell_ID"], keep="first")
                .head(max(0, int(config.action_neighbor_cells)))
            )
            add(nearest_neighbors["Node_Cell_ID"].astype(str).tolist())

    compact_cells = [cell for cell in wanted_order if cell]
    if not compact_cells:
        compact_cells = list(affected_cells)
    compact_site_rows = work.loc[work["Node_Cell_ID"].astype(str).isin(set(compact_cells))]
    site_col = "dashboard_site_id" if "dashboard_site_id" in compact_site_rows.columns else "Site ID"
    compact_sites = (
        compact_site_rows[site_col].dropna().astype(str).drop_duplicates().tolist()
        if site_col in compact_site_rows.columns
        else list(affected_sites)
    )
    return compact_cells, compact_sites


def _current_number_set(values) -> set[int]:
    out: set[int] = set()
    if values is None:
        return out
    if isinstance(values, pd.Series):
        iterable = values.dropna().astype(str).tolist()
    elif isinstance(values, (list, tuple, set)):
        iterable = list(values)
    else:
        iterable = [values]
    for value in iterable:
        for token in re.split(r"[,|;/\s]+", str(value)):
            token = token.strip()
            if not token or token.lower() in {"nan", "none", "null", "<na>"}:
                continue
            try:
                out.add(int(round(float(token))))
            except Exception:
                continue
    return out


def _bearing_deg(lat1: float, lon1: float, lat2: float, lon2: float) -> float:
    lat1_rad = np.radians(lat1)
    lat2_rad = np.radians(lat2)
    dlon = np.radians(lon2 - lon1)
    y = np.sin(dlon) * np.cos(lat2_rad)
    x = np.cos(lat1_rad) * np.sin(lat2_rad) - np.sin(lat1_rad) * np.cos(lat2_rad) * np.cos(dlon)
    return float((np.degrees(np.arctan2(y, x)) + 360.0) % 360.0)


def _demand_centroid_for_sector(
    sector_cells: pd.DataFrame,
    context: dict[str, Any],
    source_rows: pd.DataFrame,
) -> tuple[float, float, float]:
    source_lat = pd.to_numeric(source_rows.get("lat"), errors="coerce").mean() if source_rows is not None and not source_rows.empty else np.nan
    source_lon = pd.to_numeric(source_rows.get("lon"), errors="coerce").mean() if source_rows is not None and not source_rows.empty else np.nan

    current_df = context.get("current_df", pd.DataFrame())
    lineage_grid_ids = set(_affected_lineage_grid_ids(sector_cells, current_df))
    demand = pd.DataFrame()
    if lineage_grid_ids and not current_df.empty and "grid_id" in current_df.columns:
        gid = pd.to_numeric(current_df["grid_id"], errors="coerce")
        demand = current_df.loc[gid.isin(lineage_grid_ids)].copy()
    if demand.empty:
        baseline = context.get("corrected_pred_df", pd.DataFrame())
        if lineage_grid_ids and not baseline.empty and "grid_id" in baseline.columns:
            gid = pd.to_numeric(baseline["grid_id"], errors="coerce")
            demand = baseline.loc[gid.isin(lineage_grid_ids)].copy()

    lat_col = "grid_centroid_lat" if "grid_centroid_lat" in demand.columns else "lat"
    lon_col = "grid_centroid_lon" if "grid_centroid_lon" in demand.columns else "lon"
    demand_lat = pd.to_numeric(demand.get(lat_col), errors="coerce").mean() if not demand.empty and lat_col in demand.columns else np.nan
    demand_lon = pd.to_numeric(demand.get(lon_col), errors="coerce").mean() if not demand.empty and lon_col in demand.columns else np.nan

    if pd.isna(demand_lat):
        demand_lat = source_lat
    if pd.isna(demand_lon):
        demand_lon = source_lon
    if pd.isna(source_lat):
        source_lat = demand_lat
    if pd.isna(source_lon):
        source_lon = demand_lon

    bearing = _bearing_deg(float(source_lat), float(source_lon), float(demand_lat), float(demand_lon))
    return float(demand_lat), float(demand_lon), bearing


def _pick_current_carrier_band(sector_cells: pd.DataFrame, requested_band: str) -> str:
    existing = _current_number_set(sector_cells.get("topology_band", pd.Series(dtype=object)))
    existing.update(_current_number_set(sector_cells.get("band", pd.Series(dtype=object))))
    available = _current_number_set(sector_cells.get("available_bands_to_add", pd.Series(dtype=object)))
    requested = _current_number_set(requested_band)
    ordered = list(requested) + sorted(
        available,
        key=lambda band: -future_rules.BAND_PRIORITY.get(float(band), 0),
    )
    for band in ordered:
        if band and band not in existing:
            return str(band)
    fallback_pool = [2300, 2100, 1800, 900, 850, 700]
    for band in fallback_pool:
        if band not in existing:
            return str(band)
    return str(next(iter(requested or available or {1800})))


def _build_current_carrier_addition_topology(
    sector_cells: pd.DataFrame,
    part3_site_df: pd.DataFrame,
    band_to_add: str,
    logger,
) -> tuple[pd.DataFrame, pd.DataFrame, str]:
    work = part3_site_df.copy()
    source_rows = future_rules._extract_source_site_rows(sector_cells, part3_site_df)
    if source_rows.empty:
        return work, source_rows, str(band_to_add)

    selected_band = _pick_current_carrier_band(sector_cells, band_to_add)
    try:
        band_num = int(round(float(selected_band)))
    except Exception:
        band_num = 1800
        selected_band = "1800"

    template = source_rows.copy()
    if "band" in template.columns:
        template["_band_priority"] = pd.to_numeric(template["band"], errors="coerce").map(
            lambda value: future_rules.BAND_PRIORITY.get(float(value), 0) if pd.notna(value) else 0
        )
        template = template.sort_values("_band_priority", ascending=False).drop(columns=["_band_priority"], errors="ignore")
    add_row = template.iloc[[0]].copy()
    base_cell = str(add_row.iloc[0].get("cell_id") or add_row.iloc[0].get("Node_Cell_ID") or "CELL").strip()
    add_row["band"] = float(band_num)
    add_row["earfcn"] = float(future_rules.SYNTHETIC_BAND_TO_EARFCN.get(band_num, band_num))
    add_row["frequency_mhz"] = float(band_num)
    add_row["frequency"] = float(band_num)
    add_row["carrier_variant"] = f"carrier_add_{band_num}"
    add_id = f"{base_cell}__ADD{band_num}"
    add_row = _set_synthetic_identity(add_row, add_id)
    if "PCI" in add_row.columns:
        add_row["PCI"] = (pd.to_numeric(add_row["PCI"], errors="coerce").fillna(0).astype(int) + (band_num % 97) + 11) % 504
    combined = pd.concat([work, add_row], ignore_index=True, sort=False)
    combined = combined.loc[:, ~combined.columns.duplicated()].copy()
    logger.info(
        "current_carrier_add_topology sector=%s source_rows=%d add_rows=1 requested_band=%s selected_band=%s",
        future_rules._first_non_empty(sector_cells["sector_id"]),
        len(source_rows),
        band_to_add,
        selected_band,
    )
    return combined, source_rows, selected_band


def _build_current_sector_split_topology(
    sector_cells: pd.DataFrame,
    part3_site_df: pd.DataFrame,
    context: dict[str, Any],
    logger,
) -> tuple[pd.DataFrame, pd.DataFrame]:
    work = part3_site_df.copy()
    source_rows = future_rules._extract_source_site_rows(sector_cells, part3_site_df)
    if source_rows.empty:
        return work, source_rows

    demand_lat, demand_lon, bearing = _demand_centroid_for_sector(sector_cells, context, source_rows)
    source_ids = set(source_rows["Node_Cell_ID"].dropna().astype(str)) | set(source_rows["cell_id"].dropna().astype(str))
    keep_rows = work.loc[~(work["Node_Cell_ID"].astype(str).isin(source_ids) | work["cell_id"].astype(str).isin(source_ids))].copy()
    split_rows: list[pd.DataFrame] = []
    for _, row in source_rows.iterrows():
        base_cell = str(row.get("cell_id") or row.get("Node_Cell_ID") or "CELL").strip()
        base_pci = future_rules._scalar_to_int(row.get("PCI"), 0)
        for suffix, delta_deg, pci_offset in [("A", -30.0, 17), ("B", 30.0, 34)]:
            child = row.copy()
            child_id = f"{base_cell}__SS{suffix}"
            child = _set_synthetic_identity(child, child_id)
            child["PCI"] = int((base_pci + pci_offset) % 504)
            child["azimuth"] = float((bearing + delta_deg) % 360.0)
            child["carrier_variant"] = f"sector_split_{suffix.lower()}"
            split_rows.append(pd.DataFrame([child]))
    split_df = pd.concat(split_rows, ignore_index=True) if split_rows else pd.DataFrame(columns=work.columns)
    combined = pd.concat([keep_rows, split_df], ignore_index=True, sort=False)
    combined = combined.loc[:, ~combined.columns.duplicated()].copy()
    logger.info(
        "current_sector_split_topology sector=%s source_rows=%d split_rows=%d demand_lat=%.7f demand_lon=%.7f bearing=%.1f",
        future_rules._first_non_empty(sector_cells["sector_id"]),
        len(source_rows),
        len(split_df),
        demand_lat,
        demand_lon,
        bearing,
    )
    return combined, source_rows


def _build_current_add_sector_topology(
    sector_cells: pd.DataFrame,
    part3_site_df: pd.DataFrame,
    context: dict[str, Any],
    logger,
    carrier_count: int = 1,
) -> tuple[pd.DataFrame, pd.DataFrame]:
    work = part3_site_df.copy()
    source_rows = future_rules._extract_source_site_rows(sector_cells, part3_site_df)
    if source_rows.empty:
        return work, source_rows

    carrier_count = max(1, int(carrier_count or 1))
    demand_lat, demand_lon, bearing = _demand_centroid_for_sector(sector_cells, context, source_rows)
    template = source_rows.copy()
    if "band" in template.columns:
        template["_band_priority"] = pd.to_numeric(template["band"], errors="coerce").map(
            lambda value: future_rules.BAND_PRIORITY.get(float(value), 0) if pd.notna(value) else 0
        )
        template = template.sort_values("_band_priority", ascending=False).drop(columns=["_band_priority"], errors="ignore")
    selected = template.head(min(carrier_count, len(template))).copy()
    base_sector = str(future_rules._first_non_empty(sector_cells.get("sector_id", pd.Series(dtype=object))) or "")
    sector_num = base_sector.split("|", 1)[1] if "|" in base_sector else str(selected.iloc[0].get("sector") or "1")
    add_rows: list[pd.DataFrame] = []
    for idx, (_, row) in enumerate(selected.iterrows(), start=1):
        child = row.copy()
        base_cell = str(row.get("cell_id") or row.get("Node_Cell_ID") or "CELL").strip()
        band_num = future_rules._scalar_to_int(row.get("band"), 1800)
        child_id = f"{base_cell}__AS{idx}"
        child = _set_synthetic_identity(child, child_id)
        child["sector"] = f"{sector_num}A"
        child["azimuth"] = float(bearing % 360.0)
        child["carrier_variant"] = f"add_sector_{idx}"
        child["frequency_mhz"] = float(band_num)
        child["frequency"] = float(band_num)
        child["earfcn"] = float(future_rules.SYNTHETIC_BAND_TO_EARFCN.get(band_num, band_num))
        if "PCI" in child.index:
            child["PCI"] = int((future_rules._scalar_to_int(child.get("PCI"), 0) + 53 + idx * 17) % 504)
        add_rows.append(pd.DataFrame([child]))
    add_df = pd.concat(add_rows, ignore_index=True) if add_rows else pd.DataFrame(columns=work.columns)
    combined = pd.concat([work, add_df], ignore_index=True, sort=False)
    combined = combined.loc[:, ~combined.columns.duplicated()].copy()
    logger.info(
        "current_add_sector_topology sector=%s source_rows=%d add_rows=%d demand_lat=%.7f demand_lon=%.7f bearing=%.1f",
        future_rules._first_non_empty(sector_cells["sector_id"]),
        len(source_rows),
        len(add_df),
        demand_lat,
        demand_lon,
        bearing,
    )
    return combined, source_rows


def _build_current_new_site_topology(
    sector_cells: pd.DataFrame,
    part3_site_df: pd.DataFrame,
    context: dict[str, Any],
    logger,
) -> tuple[pd.DataFrame, pd.DataFrame]:
    work = part3_site_df.copy()
    source_rows = future_rules._extract_source_site_rows(sector_cells, part3_site_df)
    if source_rows.empty:
        return work, source_rows

    demand_lat, demand_lon, _ = _demand_centroid_for_sector(sector_cells, context, source_rows)
    source_lat = pd.to_numeric(source_rows["lat"], errors="coerce").mean()
    source_lon = pd.to_numeric(source_rows["lon"], errors="coerce").mean()
    if pd.isna(source_lat) or pd.isna(source_lon):
        source_lat, source_lon = demand_lat, demand_lon

    # Put the synthetic site near the overloaded demand footprint, slightly toward
    # the original site so RF still has a nonzero path and a meaningful azimuth.
    new_lat = float(demand_lat + (source_lat - demand_lat) * 0.20)
    new_lon = float(demand_lon + (source_lon - demand_lon) * 0.20)
    if future_rules._haversine_distance_m_series(pd.Series([new_lat]), pd.Series([new_lon]), demand_lat, demand_lon).iloc[0] < 25.0:
        new_lat = float(demand_lat - (60.0 / 111320.0))
        new_lon = float(demand_lon)
    serving_bearing = _bearing_deg(new_lat, new_lon, demand_lat, demand_lon)

    site_id_seed = pd.to_numeric(work["Site ID"], errors="coerce").dropna().astype(int)
    new_site_id = int(site_id_seed.max()) + 1001 if not site_id_seed.empty else 990001
    bands = sorted(
        _current_number_set(source_rows.get("band", pd.Series(dtype=object))) or {1800},
        key=lambda band: -future_rules.BAND_PRIORITY.get(float(band), 0),
    )
    azimuths = [serving_bearing, (serving_bearing + 120.0) % 360.0, (serving_bearing + 240.0) % 360.0]
    new_rows: list[pd.DataFrame] = []
    for idx, azimuth in enumerate(azimuths, start=1):
        band_num = int(bands[(idx - 1) % len(bands)])
        base = source_rows.iloc[[0]].copy()
        base["Site ID"] = new_site_id
        base["lat"] = new_lat
        base["lon"] = new_lon
        base["band"] = float(band_num)
        base["earfcn"] = float(future_rules.SYNTHETIC_BAND_TO_EARFCN.get(band_num, band_num))
        base["frequency_mhz"] = float(band_num)
        base["frequency"] = float(band_num)
        base["azimuth"] = float(azimuth)
        if "tx_power" in base.columns:
            base["tx_power"] = pd.to_numeric(base["tx_power"], errors="coerce").fillna(43.0).clip(lower=46.0)
        if "antenna_height" in base.columns:
            base["antenna_height"] = pd.to_numeric(base["antenna_height"], errors="coerce").fillna(25.0).clip(lower=25.0)
        base_id = f"{new_site_id}_{idx}__NS"
        base = _set_synthetic_identity(base, base_id)
        if "PCI" in base.columns:
            base["PCI"] = int((100 + idx * 29 + (band_num % 41)) % 504)
        base["carrier_variant"] = "new_site"
        new_rows.append(base)
    new_site_df = pd.concat(new_rows, ignore_index=True) if new_rows else pd.DataFrame(columns=work.columns)
    combined = pd.concat([work, new_site_df], ignore_index=True, sort=False)
    combined = combined.loc[:, ~combined.columns.duplicated()].copy()
    logger.info(
        "current_new_site_topology sector=%s source_rows=%d new_site_rows=%d new_site_id=%s demand_lat=%.7f demand_lon=%.7f site_lat=%.7f site_lon=%.7f bearing=%.1f",
        future_rules._first_non_empty(sector_cells["sector_id"]),
        len(source_rows),
        len(new_site_df),
        new_site_id,
        demand_lat,
        demand_lon,
        new_lat,
        new_lon,
        serving_bearing,
    )
    return combined, source_rows


def _cached_current_k1k2_for_cells(
    *,
    scoped_points: pd.DataFrame,
    modified_site_df: pd.DataFrame,
    affected_cells: list[str],
    sector_cells: pd.DataFrame,
    context: dict[str, Any],
    logger,
    action_label: str,
) -> dict[str, tuple[float, float]]:
    original_ids = {str(value).strip() for value in context.get("original_topology_cell_ids", set()) if str(value).strip()}
    cache: dict[str, tuple[float, float]] = context.setdefault("current_original_k1k2_cache", {})
    affected = [str(cell).strip() for cell in affected_cells if str(cell).strip()]
    cached_map = {cell: cache[cell] for cell in affected if cell in cache}
    missing_original = [cell for cell in affected if cell in original_ids and cell not in cache]
    synthetic_or_changed = [cell for cell in affected if cell not in original_ids]

    source_originals: list[str] = []
    for col in ["Node_Cell_ID", "topology_original_cell_id", "topology_original_node_cell_id"]:
        if col in sector_cells.columns:
            for value in sector_cells[col].dropna().astype(str).tolist():
                text = str(value).strip()
                if text and text in original_ids and text not in source_originals:
                    source_originals.append(text)

    for cell in source_originals:
        if cell not in cache and cell not in missing_original:
            missing_original.append(cell)

    compute_cells = list(dict.fromkeys(missing_original + synthetic_or_changed))
    if compute_cells:
        computed = production_compute_k1k2_for_cells(scoped_points, modified_site_df, compute_cells)
    else:
        computed = {}

    for cell in missing_original:
        if cell in computed:
            cache[cell] = computed[cell]
            cached_map[cell] = computed[cell]

    computed_action: list[str] = []
    for cell in synthetic_or_changed:
        if cell in computed:
            cached_map[cell] = computed[cell]
            computed_action.append(cell)

    source_k1k2 = None
    for cell in source_originals:
        if cell in cache:
            source_k1k2 = cache[cell]
            break
    if source_k1k2 is None and cache:
        source_k1k2 = next(iter(cache.values()))
    if source_k1k2 is None:
        source_k1k2 = (0.0, 0.0)
    inherited_action: list[str] = []
    for cell in synthetic_or_changed:
        if cell not in cached_map:
            cached_map[cell] = source_k1k2
            inherited_action.append(cell)

    logger.info(
        "%s_k1k2_cache affected=%d cached_original=%d computed_original=%d computed_action=%d inherited_action=%d returned=%d",
        action_label,
        len(affected),
        len([cell for cell in affected if cell in cache]),
        len([cell for cell in missing_original if cell in computed]),
        len(computed_action),
        len(inherited_action),
        len(cached_map),
    )
    return cached_map


def _rerun_current_topology(
    *,
    sector_cells: pd.DataFrame,
    config: CurrentModel3Config,
    context: dict[str, Any],
    logger,
    action_label: str,
    site_df: pd.DataFrame,
    source_rows: pd.DataFrame,
) -> dict[str, Any]:
    original_site_df, modified_site_df = _prepare_production_scope_site_frames(
        original_site_df=context["part3_site_df"],
        action_site_df=site_df,
        source_rows=source_rows,
    )
    baseline_points = context["corrected_pred_df"].loc[
        context["corrected_pred_df"]["time_bucket"].astype(str) == "PART_3"
    ].copy()
    baseline_points["time_bucket"] = "PART_3"
    if baseline_points.empty:
        return {
            "status": "Recommended",
            "action_reason": f"{action_label} could not load PART_3 baseline points.",
            "projected_prb_after_pct": np.nan,
            "projected_rrc_after_pct": np.nan,
            "projected_rrc_users_after": np.nan,
            "next_step": "New Site",
            "resimulation_required": True,
            "resimulation_flow": f"{action_label} production affected-cell rerun failed during baseline-point preparation",
        }

    try:
        affected_cells, affected_sites, changed_rows = production_compute_affected_cells(
            modified_site_df,
            float(context["summary"].get("baseline_radius_m", 500.0)),
            2,
            baseline_df=baseline_points,
            max_neighbors_per_update_cell=2,
        )
    except Exception as exc:
        logger.info("%s_production_scope_failed reason=%s", action_label, exc)
        return {
            "status": "Recommended",
            "action_reason": f"{action_label} could not compute the production affected-cell scope: {exc}",
            "projected_prb_after_pct": np.nan,
            "projected_rrc_after_pct": np.nan,
            "projected_rrc_users_after": np.nan,
            "next_step": "New Site",
            "resimulation_required": True,
            "resimulation_flow": f"{action_label} production affected-cell scope failed",
        }

    raw_affected_cells = list(affected_cells)
    raw_affected_sites = list(affected_sites)
    affected_cells, affected_sites = _compact_current_action_scope(
        affected_cells=affected_cells,
        affected_sites=affected_sites,
        changed_rows=changed_rows,
        sector_cells=sector_cells,
        modified_site_df=modified_site_df,
        source_rows=source_rows,
        config=config,
    )
    if len(affected_cells) != len(raw_affected_cells):
        logger.info(
            "%s_compact_scope raw_cells=%d compact_cells=%d raw_sites=%d compact_sites=%d neighbor_cells=%d",
            action_label,
            len(raw_affected_cells),
            len(affected_cells),
            len(raw_affected_sites),
            len(affected_sites),
            int(config.action_neighbor_cells),
        )

    point_mask = _vectorized_identity_point_mask(baseline_points, affected_cells)
    scoped_points = baseline_points.loc[point_mask].copy()
    if scoped_points.empty:
        logger.info(
            "%s_production_scope_failed reason=no_prediction_points affected_cells=%s",
            action_label,
            affected_cells,
        )
        return {
            "status": "Recommended",
            "action_reason": f"{action_label} production affected cells had no matching saved PART_3 prediction points.",
            "projected_prb_after_pct": np.nan,
            "projected_rrc_after_pct": np.nan,
            "projected_rrc_users_after": np.nan,
            "next_step": "New Site",
            "resimulation_required": True,
            "resimulation_flow": f"{action_label} production affected-cell rerun skipped because saved points did not match affected cells",
        }
    matched_point_count = len(scoped_points)
    scoped_points = _append_synthetic_prediction_points(
        scoped_points=scoped_points,
        baseline_points=baseline_points,
        sector_cells=sector_cells,
        changed_rows=changed_rows,
        original_cell_ids=context.get("original_topology_cell_ids", set()),
    )
    scoped_points = _append_affected_alias_prediction_points(
        scoped_points=scoped_points,
        baseline_points=baseline_points,
        modified_site_df=modified_site_df,
        affected_cells=affected_cells,
    )

    local_grid_ids = pd.to_numeric(scoped_points.get("grid_id"), errors="coerce").dropna().astype("Int64").unique().tolist()
    part3_kpi = context["kpi_df"].loc[context["kpi_df"]["time_bucket"].astype(str) == "PART_3"].copy()
    part3_geo = context["geo_df"].loc[context["geo_df"]["time_bucket"].astype(str) == "PART_3"].copy()
    local_kpi = part3_kpi.loc[pd.to_numeric(part3_kpi["grid_id"], errors="coerce").astype("Int64").isin(local_grid_ids)].copy()
    local_geo = part3_geo.loc[pd.to_numeric(part3_geo["grid_id"], errors="coerce").astype("Int64").isin(local_grid_ids)].copy()
    point_map = future_rules._make_local_point_map(scoped_points)

    snap_tolerance_m = float(context["summary"].get("grid_size_m", 50.0)) * 0.75
    run_params = {
        "radius": float(context["summary"].get("baseline_radius_m", 500.0)),
        "grid_resolution": float(context["summary"].get("grid_size_m", 50.0)),
        "n_workers": max(1, int(config.rf_workers)),
        "antenna_gain": 18,
        "cable_loss": 2,
        "ue_height": 1.5,
        "frequency_mhz": 1800,
        "bandwidth_mhz": 10,
        "project_id": int(context["summary"].get("project_id", 196)),
        "region": str(context["summary"].get("region", "india")).lower(),
        "baseline_df": scoped_points,
        "prediction_points_df": scoped_points,
        "strict_prediction_points": True,
        "impact_radius_m": float(context["summary"].get("baseline_radius_m", 500.0)),
        "neighbor_site_count": 2,
        "max_interference_sites": int(context["summary"].get("max_interference_sites", config.max_interference_sites)),
        "max_neighbors_per_update_cell": 2,
        "recompute_cells": affected_cells,
    }
    logger.info(
        "%s_production_scope affected_cells=%d affected_sites=%d changed_cells=%d scoped_points=%d",
        action_label,
        len(affected_cells),
        len(affected_sites),
        int(changed_rows["Node_Cell_ID"].nunique()) if "Node_Cell_ID" in changed_rows.columns else len(changed_rows),
        len(scoped_points),
    )
    if len(scoped_points) != matched_point_count:
        logger.info("%s_synthetic_points matched_points=%d final_points=%d", action_label, matched_point_count, len(scoped_points))
    if int(config.sector_parallelism) > 1:
        # `redirect_stdout` is process-global and unsafe while sector workers run
        # concurrently, so parallel mode lets production RF print normally while
        # structured logger lines still go to the run log.
        k1k2_map = _cached_current_k1k2_for_cells(
            scoped_points=scoped_points,
            modified_site_df=modified_site_df,
            affected_cells=affected_cells,
            sector_cells=sector_cells,
            context=context,
            logger=logger,
            action_label=action_label,
        )
        corrected_local = production_run_prediction_only_optimized(modified_site_df, k1k2_map, run_params)
    else:
        with open(logger.handlers[0].baseFilename, "a", encoding="utf-8") as log_stream:
            with contextlib.redirect_stdout(log_stream), contextlib.redirect_stderr(log_stream):
                k1k2_map = _cached_current_k1k2_for_cells(
                    scoped_points=scoped_points,
                    modified_site_df=modified_site_df,
                    affected_cells=affected_cells,
                    sector_cells=sector_cells,
                    context=context,
                    logger=logger,
                    action_label=action_label,
                )
                corrected_local = production_run_prediction_only_optimized(modified_site_df, k1k2_map, run_params)
    baseline_local = scoped_points.copy()
    baseline_local["time_bucket"] = "PART_3"
    corrected_local["time_bucket"] = "PART_3"
    baseline_local = _force_master_grid_identity(baseline_local, point_map, snap_tolerance_m)
    corrected_local = _force_master_grid_identity(corrected_local, point_map, snap_tolerance_m)
    baseline_local = _attach_site_topology_to_prediction(baseline_local, original_site_df)
    corrected_local = _attach_site_topology_to_prediction(corrected_local, modified_site_df)
    lineage_grid_ids = set(_affected_lineage_grid_ids(sector_cells, context.get("current_df", pd.DataFrame())))
    if lineage_grid_ids:
        baseline_gid = pd.to_numeric(baseline_local.get("grid_id"), errors="coerce")
        corrected_gid = pd.to_numeric(corrected_local.get("grid_id"), errors="coerce")
        baseline_local = baseline_local.loc[baseline_gid.isin(lineage_grid_ids)].copy()
        corrected_local = corrected_local.loc[corrected_gid.isin(lineage_grid_ids)].copy()
    if baseline_local.empty or corrected_local.empty:
        logger.info(
            "%s_production_scope_failed reason=empty_lineage_surface before_rows=%d after_rows=%d lineage_grids=%d",
            action_label,
            len(baseline_local),
            len(corrected_local),
            len(lineage_grid_ids),
        )
        return {
            "status": "Recommended",
            "action_reason": f"{action_label} RF rerun produced no usable lineage prediction points for this action.",
            "projected_prb_after_pct": np.nan,
            "projected_rrc_after_pct": np.nan,
            "projected_rrc_users_after": np.nan,
            "next_step": "New Site",
            "resimulation_required": True,
            "resimulation_flow": f"{action_label} topology -> production affected cells -> empty lineage RF surface",
        }
    source_capacity_values = pd.to_numeric(sector_cells.get("estimated_dl_capacity_mbps"), errors="coerce")
    source_capacity_fallback = float(source_capacity_values.median()) if source_capacity_values.notna().any() else None
    enriched_local, local_inventory = _build_current_inventory_from_surface(
        baseline_local=baseline_local,
        corrected_local=corrected_local,
        local_site_df=modified_site_df,
        local_kpi=local_kpi,
        local_geo=local_geo,
        context=context,
        config=config,
        lineage_grid_ids=lineage_grid_ids,
        source_capacity_fallback=source_capacity_fallback,
    )
    after_rf_dir = context.get("after_rf_dir")
    if after_rf_dir:
        try:
            after_rf_dir = Path(after_rf_dir)
            after_rf_dir.mkdir(parents=True, exist_ok=True)
            sector_key = _safe_name(future_rules._first_non_empty(sector_cells.get("sector_id", pd.Series(dtype=object))))
            file_key = f"{_safe_name(action_label)}_{sector_key}_{threading.get_ident()}_{int(time.time() * 1000)}"
            baseline_path = after_rf_dir / f"{file_key}_before_rf_surface.csv"
            corrected_path = after_rf_dir / f"{file_key}_after_rf_surface.csv"
            inventory_path = after_rf_dir / f"{file_key}_after_cell_inventory.csv"
            baseline_local.to_csv(baseline_path, index=False)
            corrected_local.to_csv(corrected_path, index=False)
            local_inventory.to_csv(inventory_path, index=False)
            context.setdefault("after_rf_records", []).append(
                {
                    "action_label": action_label,
                    "sector_id": future_rules._first_non_empty(sector_cells.get("sector_id", pd.Series(dtype=object))),
                    "before_rf_surface_csv": str(baseline_path),
                    "after_rf_surface_csv": str(corrected_path),
                    "after_cell_inventory_csv": str(inventory_path),
                    "rows_before": int(len(baseline_local)),
                    "rows_after": int(len(corrected_local)),
                    "cells_after": int(corrected_local.get("Node_Cell_ID", pd.Series(dtype=object)).nunique(dropna=True)),
                }
            )
            logger.info(
                "%s_after_rf_artifacts sector=%s before_rows=%d after_rows=%d path=%s",
                action_label,
                future_rules._first_non_empty(sector_cells.get("sector_id", pd.Series(dtype=object))),
                len(baseline_local),
                len(corrected_local),
                corrected_path,
            )
        except Exception as exc:
            logger.info("%s_after_rf_artifact_save_failed reason=%s", action_label, exc)
    candidate_node_cell_ids: set[str] = set()
    for col in ["Node_Cell_ID", "original_node_cell_id", "original_cell_id"]:
        if col in corrected_local.columns:
            candidate_node_cell_ids.update(corrected_local[col].dropna().astype(str))

    # The congested demand lives on specific grids. After the action, whichever cells
    # now serve those grids carry that demand - THEY decide whether congestion is
    # actually resolved, regardless of how any cell is named across tooling layers.
    en_gid = pd.to_numeric(enriched_local.get("grid_id"), errors="coerce")
    demand_serving_cell_ids = set(
        enriched_local.loc[en_gid.isin(lineage_grid_ids), "Node_Cell_ID"].dropna().astype(str).str.strip()
    )
    demand_serving_cell_ids.discard("")

    outcome = _evaluate_action_inventory_current(
        sector_cells=sector_cells,
        after_inventory=local_inventory,
        candidate_node_cell_ids=candidate_node_cell_ids,
        demand_serving_cell_ids=demand_serving_cell_ids,
        action_node_cell_ids=set(changed_rows["Node_Cell_ID"].dropna().astype(str).str.strip())
        if "Node_Cell_ID" in changed_rows.columns
        else set(),
        config=config,
        logger=logger,
        action_label=action_label,
    )
    return {
        "status": outcome["status"],
        "action_reason": f"{action_label} used the production optimized affected-cell rerun and recalculated current-state KPIs deterministically.",
        "projected_prb_after_pct": outcome["projected_prb_after_pct"],
        "projected_rrc_after_pct": outcome["projected_rrc_after_pct"],
        "projected_rrc_users_after": outcome["projected_rrc_users_after"],
        "next_step": outcome["next_step"],
        "resimulation_required": True,
        "resimulation_flow": f"{action_label} topology -> production affected cells ({len(affected_cells)} cells/{len(affected_sites)} sites; changed={changed_rows['Node_Cell_ID'].nunique()}) -> baseline/optimized RF rerun on saved PART_3 points -> deterministic current KPI rebuild -> Model 3 reevaluation",
    }


def _simulate_current_recommendation(
    sector_cells: pd.DataFrame,
    config: CurrentModel3Config,
    logger,
    context: dict[str, Any],
    source_node_cell_id: str | None = None,
) -> dict[str, Any]:
    ordered_sector_cells = sector_cells.sort_values(["prb_rrc_pressure", "grid_count"], ascending=[False, False], na_position="last").copy()
    source_key = str(source_node_cell_id or "").strip()
    if source_key:
        source_match = ordered_sector_cells.loc[ordered_sector_cells["Node_Cell_ID"].astype(str).str.strip().eq(source_key)]
        if source_match.empty:
            return {
                "action": "Skipped",
                "status": "Skipped",
                "decision_path": "Missing source cell",
                "attempted_actions": "",
                "load_balance_possible": False,
                "selected_peer_node_cell_id": "",
                "selected_peer_band": "",
                "projected_prb_after_pct": np.nan,
                "projected_rrc_after_pct": np.nan,
                "projected_rrc_users_after": np.nan,
                "action_reason": f"Requested congested source cell was not present in sector scope: {source_key}",
                "next_step": "",
                "resimulation_required": False,
                "resimulation_flow": "",
            }
        source_row = source_match.iloc[0]
        ordered_sector_cells = pd.concat(
            [source_match, ordered_sector_cells.loc[~ordered_sector_cells["Node_Cell_ID"].astype(str).str.strip().eq(source_key)]],
            ignore_index=True,
            sort=False,
        )
    else:
        source_row = ordered_sector_cells.iloc[0]
    sector_cells = ordered_sector_cells
    source_prb = float(source_row["prb_before_pct"]) if pd.notna(source_row["prb_before_pct"]) else np.nan
    source_rrc = float(source_row["rrc_before_pct"]) if pd.notna(source_row["rrc_before_pct"]) else np.nan
    pressure = max(source_prb if pd.notna(source_prb) else 0.0, source_rrc if pd.notna(source_rrc) else 0.0)
    congested = pressure > config.congestion_threshold
    if not congested:
        return {
            "action": "Healthy",
            "status": "Healthy",
            "decision_path": "No action",
            "attempted_actions": "",
            "load_balance_possible": False,
            "selected_peer_node_cell_id": "",
            "selected_peer_band": "",
            "projected_prb_after_pct": source_prb,
            "projected_rrc_after_pct": source_rrc,
            "projected_rrc_users_after": float(source_row["rrc_users_before"]) if pd.notna(source_row["rrc_users_before"]) else np.nan,
            "action_reason": "Healthy.",
            "next_step": "",
            "resimulation_required": False,
            "resimulation_flow": "",
        }

    attempted_actions: list[str] = []

    load_result = _run_load_balance_current(sector_cells, config, str(source_row["Node_Cell_ID"]))
    if load_result["selected_peer_node_cell_id"]:
        attempted_actions.append(f"Load Balance[{load_result['status']}]")
    if load_result["selected_peer_node_cell_id"] and not bool(load_result.get("should_fallthrough")):
        return {
            "action": f"Load Balance -> {load_result['selected_peer_band']} MHz",
            "decision_path": "Congested | Current load balance branch",
            "attempted_actions": " -> ".join(attempted_actions),
            "load_balance_possible": True,
            **load_result,
        }
    if (
        config.stop_on_partial
        and load_result["selected_peer_node_cell_id"]
        and str(load_result.get("status")) == "Partially Resolved"
    ):
        return {
            "action": f"Load Balance -> {load_result['selected_peer_band']} MHz",
            "decision_path": "Congested | Current load balance branch",
            "attempted_actions": " -> ".join(attempted_actions),
            "load_balance_possible": True,
            **load_result,
        }

    can_add = bool(sector_cells["carrier_addition_possible"].any()) and not bool(sector_cells["carrier_addition_blocked"].all())
    if context.get("excel_input_mode"):
        source_can_add = bool(source_row.get("carrier_addition_possible")) and not bool(source_row.get("carrier_addition_blocked"))
        if source_can_add:
            add_row = sector_cells.loc[
                sector_cells["carrier_addition_possible"]
                & ~sector_cells["carrier_addition_blocked"]
                & sector_cells["Node_Cell_ID"].astype(str).str.strip().eq(str(source_row["Node_Cell_ID"]).strip())
            ].sort_values(["prb_rrc_pressure", "grid_count"], ascending=[False, False], na_position="last")
            band = str(add_row.iloc[0]["recommended_band_to_add"]).strip() if not add_row.empty else ""
            return {
                "action": f"Add Carrier -> {band} MHz" if band else "Add Carrier",
                "status": "Recommended",
                "decision_path": "Congested | Excel input carrier addition branch",
                "attempted_actions": " -> ".join(attempted_actions + [f"Add Carrier {band}[Recommended]"]),
                "load_balance_possible": False,
                "selected_peer_node_cell_id": str(source_row["Node_Cell_ID"]),
                "selected_peer_band": "",
                "projected_prb_after_pct": np.nan,
                "projected_rrc_after_pct": np.nan,
                "projected_rrc_users_after": np.nan,
                "action_reason": "Excel input mode has no RF rerun; carrier addition is recommended for planner validation.",
                "next_step": "",
                "resimulation_required": False,
                "resimulation_flow": "Excel input mode: no RF rerun",
            }
        return {
            "action": "New Site",
            "status": "Recommended",
            "decision_path": "Congested | Excel input new-site branch",
            "attempted_actions": " -> ".join(attempted_actions + ["New Site[Recommended]"]),
            "load_balance_possible": False,
            "selected_peer_node_cell_id": str(source_row["Node_Cell_ID"]),
            "selected_peer_band": "",
            "projected_prb_after_pct": np.nan,
            "projected_rrc_after_pct": np.nan,
            "projected_rrc_users_after": np.nan,
            "action_reason": "Excel input mode has no RF rerun and no carrier addition option; new site is recommended for planner validation.",
            "next_step": "",
            "resimulation_required": False,
            "resimulation_flow": "Excel input mode: no RF rerun",
        }

    carrier_add_attempted = False
    add_row = sector_cells.loc[
        sector_cells["carrier_addition_possible"]
        & ~sector_cells["carrier_addition_blocked"]
        & sector_cells["Node_Cell_ID"].astype(str).str.strip().eq(str(source_row["Node_Cell_ID"]).strip())
    ].copy()
    if can_add and not add_row.empty and str(add_row.iloc[0]["recommended_band_to_add"]).strip():
        band = str(add_row.iloc[0]["recommended_band_to_add"]).strip()
        carrier_add_attempted = True
        site_df, source_rows, selected_band = _build_current_carrier_addition_topology(
            sector_cells,
            context["part3_site_df"],
            band,
            logger,
        )
        band = selected_band
        site_df = _fix_synthetic_frequency(site_df)
        resim = _rerun_current_topology(
            sector_cells=sector_cells,
            config=config,
            context=context,
            logger=logger,
            action_label="current_carrier_add",
            site_df=site_df,
            source_rows=source_rows,
        )
        attempted_actions.append(f"Add Carrier {band}[{resim['status']}]")
        if str(resim["status"]) == "Resolved" or (config.stop_on_partial and str(resim["status"]) == "Partially Resolved"):
            return {
                "action": f"Add Carrier -> {band} MHz",
                "decision_path": "Congested | Carrier addition branch",
                "attempted_actions": " -> ".join(attempted_actions),
                "load_balance_possible": False,
                "selected_peer_node_cell_id": str(source_row["Node_Cell_ID"]),
                "selected_peer_band": "",
                **resim,
            }

    add_sector_possible = bool(
        int(source_row["sector_cell_count"]) >= 1
        and (not can_add or carrier_add_attempted)
        and (
            int(source_row["sector_congested_count"]) >= 2
            or (
                pressure >= 95.0
                and int(source_row["existing_carrier_count"]) >= int(source_row["max_supported_carriers"])
            )
            or bool(sector_cells["carrier_addition_blocked"].all())
        )
    )
    if add_sector_possible:
        max_add_sector_carriers = max(1, min(3, int(source_row["sector_cell_count"])))
        for carrier_count in range(1, max_add_sector_carriers + 1):
            site_df, source_rows = _build_current_add_sector_topology(
                sector_cells,
                context["part3_site_df"],
                context,
                logger,
                carrier_count=carrier_count,
            )
            site_df = _fix_synthetic_frequency(site_df)
            resim = _rerun_current_topology(
                sector_cells=sector_cells,
                config=config,
                context=context,
                logger=logger,
                action_label="current_sector_add",
                site_df=site_df,
                source_rows=source_rows,
            )
            attempted_actions.append(f"Add Sector {carrier_count} cell[{resim['status']}]")
            if str(resim["status"]) == "Resolved" or (config.stop_on_partial and str(resim["status"]) == "Partially Resolved"):
                return {
                    "action": f"Add Sector -> {carrier_count} cell",
                    "decision_path": "Congested | Progressive add-sector branch",
                    "attempted_actions": " -> ".join(attempted_actions),
                    "load_balance_possible": False,
                    "selected_peer_node_cell_id": str(source_row["Node_Cell_ID"]),
                    "selected_peer_band": "",
                    **resim,
                }

    site_df, source_rows = _build_current_new_site_topology(sector_cells, context["part3_site_df"], context, logger)
    site_df = _fix_synthetic_frequency(site_df)
    resim = _rerun_current_topology(
        sector_cells=sector_cells,
        config=config,
        context=context,
        logger=logger,
        action_label="current_new_site",
        site_df=site_df,
        source_rows=source_rows,
    )
    attempted_actions.append(f"New Site[{resim['status']}]")
    return {
        "action": "New Site",
        "decision_path": "Congested | New site branch",
        "attempted_actions": " -> ".join(attempted_actions),
        "load_balance_possible": False,
        "selected_peer_node_cell_id": str(source_row["Node_Cell_ID"]),
        "selected_peer_band": "",
        **resim,
    }


def _build_recommendation_row_for_sector(
    sector_id: Any,
    sector_cells: pd.DataFrame,
    config: CurrentModel3Config,
    logger,
    context: dict[str, Any],
) -> dict[str, Any] | None:
    sector_cells = sector_cells.sort_values(
        ["congested", "prb_rrc_pressure", "grid_count"],
        ascending=[False, False, False],
        na_position="last",
    ).copy()
    if not bool(sector_cells["congested"].any()):
        return None
    rec = _simulate_current_recommendation(sector_cells, config, logger, context)
    lead_row = sector_cells.iloc[0]
    sector_congested_ids = sector_cells.loc[sector_cells["congested"], "Node_Cell_ID"].astype(str).tolist()
    return {
        "site_id": lead_row["site_id"],
        "sector_id": sector_id,
        "Node_Cell_ID": lead_row["Node_Cell_ID"],
        "sector_congested_node_cell_ids": ", ".join(sector_congested_ids),
        "band": lead_row["band"],
        "earfcn": lead_row["earfcn"],
        "grid_count": int(sector_cells["grid_count"].sum()),
        "congested_grid_count": int(sector_cells["congested_grid_count"].sum()),
        "prb_before_pct": round(float(lead_row["prb_before_pct"]), 3) if pd.notna(lead_row["prb_before_pct"]) else np.nan,
        "rrc_before_pct": round(float(lead_row["rrc_before_pct"]), 3) if pd.notna(lead_row["rrc_before_pct"]) else np.nan,
        "rrc_users_before": round(float(lead_row["rrc_users_before"]), 3) if pd.notna(lead_row["rrc_users_before"]) else np.nan,
        "action": rec["action"],
        "status": rec["status"],
        "decision_path": rec["decision_path"],
        "attempted_actions": rec.get("attempted_actions", ""),
        "load_balance_possible": bool(rec["load_balance_possible"]),
        "selected_peer_node_cell_id": rec["selected_peer_node_cell_id"],
        "selected_peer_band": rec["selected_peer_band"],
        "recommended_band_to_add": lead_row["recommended_band_to_add"],
        "available_bands_to_add": lead_row["available_bands_to_add"],
        "carrier_addition_possible": bool(sector_cells["carrier_addition_possible"].any()),
        "carrier_addition_blocked": bool(sector_cells["carrier_addition_blocked"].all()),
        "max_supported_carriers": int(sector_cells["max_supported_carriers"].max()),
        "existing_carrier_count": int(sector_cells["existing_carrier_count"].max()),
        "existing_carriers": lead_row["existing_carriers"],
        "projected_prb_after_pct": rec["projected_prb_after_pct"],
        "projected_rrc_after_pct": rec["projected_rrc_after_pct"],
        "projected_rrc_users_after": rec["projected_rrc_users_after"],
        "action_reason": rec["action_reason"],
        "next_step": rec["next_step"],
        "resimulation_required": bool(rec["resimulation_required"]),
        "resimulation_flow": rec["resimulation_flow"],
    }


def _build_recommendation_row_for_cell(
    sector_id: Any,
    sector_cells: pd.DataFrame,
    source_row: pd.Series,
    config: CurrentModel3Config,
    logger,
    context: dict[str, Any],
) -> dict[str, Any] | None:
    sector_cells = sector_cells.sort_values(
        ["congested", "prb_rrc_pressure", "grid_count"],
        ascending=[False, False, False],
        na_position="last",
    ).copy()
    source_id = str(source_row["Node_Cell_ID"]).strip()
    if not bool(source_row.get("congested")):
        return None
    rec = _simulate_current_recommendation(sector_cells, config, logger, context, source_id)
    sector_congested_ids = sector_cells.loc[sector_cells["congested"], "Node_Cell_ID"].astype(str).tolist()
    return {
        "site_id": source_row["site_id"],
        "sector_id": sector_id,
        "Node_Cell_ID": source_row["Node_Cell_ID"],
        "sector_congested_node_cell_ids": ", ".join(sector_congested_ids),
        "band": source_row["band"],
        "earfcn": source_row["earfcn"],
        "grid_count": int(source_row["grid_count"]) if pd.notna(source_row["grid_count"]) else 0,
        "congested_grid_count": int(source_row["congested_grid_count"]) if pd.notna(source_row["congested_grid_count"]) else 0,
        "prb_before_pct": round(float(source_row["prb_before_pct"]), 3) if pd.notna(source_row["prb_before_pct"]) else np.nan,
        "rrc_before_pct": round(float(source_row["rrc_before_pct"]), 3) if pd.notna(source_row["rrc_before_pct"]) else np.nan,
        "rrc_users_before": round(float(source_row["rrc_users_before"]), 3) if pd.notna(source_row["rrc_users_before"]) else np.nan,
        "action": rec["action"],
        "status": rec["status"],
        "decision_path": rec["decision_path"],
        "attempted_actions": rec.get("attempted_actions", ""),
        "load_balance_possible": bool(rec["load_balance_possible"]),
        "selected_peer_node_cell_id": rec["selected_peer_node_cell_id"],
        "selected_peer_band": rec["selected_peer_band"],
        "recommended_band_to_add": source_row["recommended_band_to_add"],
        "available_bands_to_add": source_row["available_bands_to_add"],
        "carrier_addition_possible": bool(source_row["carrier_addition_possible"]),
        "carrier_addition_blocked": bool(source_row["carrier_addition_blocked"]),
        "max_supported_carriers": int(source_row["max_supported_carriers"]),
        "existing_carrier_count": int(source_row["existing_carrier_count"]),
        "existing_carriers": source_row["existing_carriers"],
        "projected_prb_after_pct": rec["projected_prb_after_pct"],
        "projected_rrc_after_pct": rec["projected_rrc_after_pct"],
        "projected_rrc_users_after": rec["projected_rrc_users_after"],
        "action_reason": rec["action_reason"],
        "next_step": rec["next_step"],
        "resimulation_required": bool(rec["resimulation_required"]),
        "resimulation_flow": rec["resimulation_flow"],
    }


def _build_recommendations(cell_df: pd.DataFrame, config: CurrentModel3Config, logger, context: dict[str, Any]) -> pd.DataFrame:
    rows = []
    sector_lookup = {sector_id: sector_cells.copy() for sector_id, sector_cells in cell_df.groupby("sector_id", dropna=False)}
    congested_cells = cell_df.loc[cell_df["congested"].fillna(False).astype(bool)].copy()
    if "recommendation_scope_cell" in congested_cells.columns:
        scoped = congested_cells.loc[congested_cells["recommendation_scope_cell"].fillna(False).astype(bool)].copy()
        if not scoped.empty:
            congested_cells = scoped
            logger.info("cell_scope using_explicit_recommendation_scope cells=%d", len(congested_cells))
    congested_cells = congested_cells.sort_values(["prb_rrc_pressure", "grid_count"], ascending=[False, False], na_position="last")
    if config.max_sectors is not None:
        top_sectors = set(
            congested_cells.groupby("sector_id")["prb_rrc_pressure"].max().sort_values(ascending=False).head(max(0, int(config.max_sectors))).index
        )
        congested_cells = congested_cells.loc[congested_cells["sector_id"].isin(top_sectors)].copy()
        logger.info("sector_scope limited_to_top=%d", len(top_sectors))
    if config.max_congested_cells is not None and len(congested_cells) > int(config.max_congested_cells):
        scenario_order = [
            "load_balance_success",
            "carrier_addition_required",
            "sector_split_candidate",
            "sector_split_required",
            "carrier_blocked_new_site",
            "new_site_required",
            "",
        ]
        scenario_groups: dict[str, pd.DataFrame] = {}
        for scenario in scenario_order:
            if "model3_scenario" in congested_cells.columns:
                part = congested_cells.loc[congested_cells["model3_scenario"].fillna("").astype(str).str.strip().eq(scenario)].copy()
            else:
                part = congested_cells.iloc[0:0].copy()
            scenario_groups[scenario] = part.sort_values(["prb_rrc_pressure", "grid_count"], ascending=[False, False], na_position="last")
        if "model3_scenario" in congested_cells.columns:
            known = set(scenario_order)
            other = congested_cells.loc[~congested_cells["model3_scenario"].fillna("").astype(str).str.strip().isin(known)].copy()
            if not other.empty:
                scenario_groups[""] = pd.concat([scenario_groups[""], other], ignore_index=False, sort=False)
        selected_rows: list[pd.Series] = []
        while len(selected_rows) < int(config.max_congested_cells):
            progressed = False
            for scenario in scenario_order:
                group = scenario_groups.get(scenario, pd.DataFrame())
                if group.empty:
                    continue
                selected_rows.append(group.iloc[0])
                scenario_groups[scenario] = group.iloc[1:]
                progressed = True
                if len(selected_rows) >= int(config.max_congested_cells):
                    break
            if not progressed:
                break
        congested_cells = pd.DataFrame(selected_rows)
        logger.info("cell_scope limited_to_exact_congested_cells=%d", len(congested_cells))
    parallelism = max(1, int(config.sector_parallelism))
    jobs = [(row.get("sector_id"), sector_lookup.get(row.get("sector_id"), pd.DataFrame()), row) for _, row in congested_cells.iterrows()]
    if parallelism > 1 and len(jobs) > 1 and not context.get("excel_input_mode"):
        logger.info("cell_parallelism enabled workers=%d cells=%d", parallelism, len(jobs))
        with concurrent.futures.ThreadPoolExecutor(max_workers=parallelism) as executor:
            futures = [
                executor.submit(_build_recommendation_row_for_cell, sector_id, sector_cells, source_row, config, logger, context)
                for sector_id, sector_cells, source_row in jobs
            ]
            for future in concurrent.futures.as_completed(futures):
                row = future.result()
                if row is not None:
                    rows.append(row)
    else:
        if parallelism > 1:
            logger.info("cell_parallelism skipped workers=%d cells=%d excel_input_mode=%s", parallelism, len(jobs), bool(context.get("excel_input_mode")))
        for sector_id, sector_cells, source_row in jobs:
            row = _build_recommendation_row_for_cell(sector_id, sector_cells, source_row, config, logger, context)
            if row is not None:
                rows.append(row)
    reco_df = pd.DataFrame(rows)
    if reco_df.empty:
        return reco_df
    reco_df["priority_score"] = reco_df[["prb_before_pct", "rrc_before_pct"]].max(axis=1)
    return reco_df.sort_values(["priority_score", "congested_grid_count", "grid_count"], ascending=[False, False, False], na_position="last").reset_index(drop=True)


def _split_node_cell_id(value: Any) -> tuple[str, str]:
    text = str(value or "").strip()
    parts = [part for part in re.split(r"[_|]", text) if part and part.lower() != "nan"]
    site = parts[0] if parts else ""
    cell = parts[1] if len(parts) > 1 else text
    return site, cell


def _available_band_flag(value: Any) -> bool:
    text = str(value or "").strip()
    return bool(text and text.lower() not in {"nan", "none", "null", "<na>", "false", "0"})


def _dashboard_congested_cells(cell_df: pd.DataFrame) -> pd.DataFrame:
    if cell_df.empty:
        return pd.DataFrame(columns=["site", "cell", "sector", "band", "available_band"])
    rows = []
    congested = cell_df.loc[cell_df.get("congested", pd.Series(False, index=cell_df.index)).fillna(False).astype(bool)].copy()
    for _, row in congested.iterrows():
        site, cell = _split_node_cell_id(row.get("Node_Cell_ID"))
        rows.append(
            {
                "site": row.get("site_id") or site,
                "cell": cell,
                "sector": row.get("sector_id", ""),
                "band": row.get("band", ""),
                "available_band": _available_band_flag(row.get("available_bands_to_add")) or bool(row.get("carrier_addition_possible", False)),
            }
        )
    return pd.DataFrame(rows).drop_duplicates().reset_index(drop=True)


def _dashboard_recommendations(reco_df: pd.DataFrame) -> pd.DataFrame:
    columns = [
        "site",
        "cell",
        "sector",
        "band",
        "recommended_solution",
        "prb_before",
        "rrc_before",
        "prb_after",
        "rrc_after",
        "recommended_band",
        "new_sector_value",
        "new_site_value",
        "baseline_before_avg",
        "baseline_after_avg",
        "status",
    ]
    if reco_df.empty:
        return pd.DataFrame(columns=columns)
    rows = []
    for _, row in reco_df.iterrows():
        site, cell = _split_node_cell_id(row.get("Node_Cell_ID"))
        action = str(row.get("action", "") or "")
        recommended_band = ""
        if "Add Carrier" in action:
            recommended_band = str(row.get("recommended_band_to_add") or "").strip()
            if not recommended_band and "->" in action:
                recommended_band = action.split("->", 1)[1].replace("MHz", "").strip()
        prb_before = future_rules._scalar_to_float(row.get("prb_before_pct"), np.nan)
        rrc_before = future_rules._scalar_to_float(row.get("rrc_before_pct"), np.nan)
        prb_after = future_rules._scalar_to_float(row.get("projected_prb_after_pct"), np.nan)
        rrc_after = future_rules._scalar_to_float(row.get("projected_rrc_after_pct"), np.nan)
        rows.append(
            {
                "site": row.get("site_id") or site,
                "cell": cell,
                "sector": row.get("sector_id", ""),
                "band": row.get("band", ""),
                "recommended_solution": action,
                "prb_before": prb_before,
                "rrc_before": rrc_before,
                "prb_after": prb_after,
                "rrc_after": rrc_after,
                "recommended_band": recommended_band,
                "new_sector_value": f"after PRB {prb_after:.3f}, RRC {rrc_after:.3f}" if ("Sector Split" in action or "Add Sector" in action) and pd.notna(prb_after) and pd.notna(rrc_after) else "",
                "new_site_value": f"after PRB {prb_after:.3f}, RRC {rrc_after:.3f}" if "New Site" in action and pd.notna(prb_after) and pd.notna(rrc_after) else "",
                "baseline_before_avg": np.nanmean([prb_before, rrc_before]),
                "baseline_after_avg": np.nanmean([prb_after, rrc_after]),
                "status": row.get("status", ""),
            }
        )
    return pd.DataFrame(rows, columns=columns)


def _add_dashboard_sheets(wb: Workbook, cell_df: pd.DataFrame, reco_df: pd.DataFrame, summary: dict[str, Any], model_label: str) -> None:
    safe_label = "Model4" if "4" in str(model_label) else "Model3"
    overview = pd.DataFrame(
        [
            {"metric": "model", "value": model_label},
            {"metric": "congested_cells", "value": int(cell_df.get("congested", pd.Series(dtype=bool)).fillna(False).astype(bool).sum()) if not cell_df.empty else 0},
            {"metric": "recommendations", "value": int(len(reco_df))},
            {"metric": "sectors", "value": summary.get("sector_rows", "")},
            {"metric": "runtime_sec", "value": summary.get("runtime_sec", "")},
        ]
    )
    ws_overview = wb.create_sheet(f"{safe_label} Dashboard", 0)
    future_rules._write_df_sheet(ws_overview, overview)
    ws_cells = wb.create_sheet(f"{safe_label} Cells")
    future_rules._write_df_sheet(ws_cells, _dashboard_congested_cells(cell_df))
    ws_reco = wb.create_sheet(f"{safe_label} Recommendations")
    future_rules._write_df_sheet(ws_reco, _dashboard_recommendations(reco_df))


def _write_workbook(
    run_dir: Path,
    cell_df: pd.DataFrame,
    reco_df: pd.DataFrame,
    sector_df: pd.DataFrame,
    summary: dict[str, Any],
    model_label: str = "Model 3",
) -> Path:
    wb = Workbook()
    wb.remove(wb.active)
    _add_dashboard_sheets(wb, cell_df, reco_df, summary, model_label)
    ws_summary = wb.create_sheet("Summary")
    summary_rows = pd.DataFrame(
        [
            {"metric": "rows_in_model3_dataset", "value": summary["rows_in_model3_dataset"]},
            {"metric": "congested_cell_rows", "value": summary["congested_cell_rows"]},
            {"metric": "sector_rows", "value": summary["sector_rows"]},
            {"metric": "threshold", "value": summary["threshold"]},
            {"metric": "runtime_sec", "value": summary["runtime_sec"]},
        ]
    )
    future_rules._write_df_sheet(ws_summary, summary_rows)
    ws_reco = wb.create_sheet("Recommendations")
    future_rules._write_df_sheet(ws_reco, reco_df)
    future_rules._style_recommendation_sheet(ws_reco, reco_df)
    ws_cells = wb.create_sheet("Congested Cells")
    future_rules._write_df_sheet(ws_cells, cell_df.loc[cell_df["congested"]].copy())
    ws_sector = wb.create_sheet("Sector Inventory")
    future_rules._write_df_sheet(ws_sector, sector_df)
    workbook_path = run_dir / "model3_current_recommendations.xlsx"
    wb.save(workbook_path)
    return workbook_path


def _combine_after_rf_artifacts(context: dict[str, Any], run_dir: Path, logger) -> dict[str, Any]:
    records = list(context.get("after_rf_records") or [])
    after_files = [Path(record["after_rf_surface_csv"]) for record in records if record.get("after_rf_surface_csv")]
    before_files = [Path(record["before_rf_surface_csv"]) for record in records if record.get("before_rf_surface_csv")]
    inventory_files = [Path(record["after_cell_inventory_csv"]) for record in records if record.get("after_cell_inventory_csv")]
    payload: dict[str, Any] = {
        "after_rf_surface_count": len(after_files),
        "after_rf_surface_combined_csv": "",
        "before_rf_surface_combined_csv": "",
        "after_cell_inventory_combined_csv": "",
        "after_rf_manifest_json": "",
    }
    manifest_path = run_dir / "after_rf_surfaces" / "manifest.json"
    if records:
        _save_json(manifest_path, {"records": records})
        payload["after_rf_manifest_json"] = str(manifest_path)

    def _combine(files: list[Path], dest_name: str) -> tuple[str, int]:
        frames = []
        for path in files:
            if path.exists():
                frame = pd.read_csv(path, low_memory=False)
                frame["_artifact_file"] = path.name
                frames.append(frame)
        if not frames:
            return "", 0
        combined = pd.concat(frames, ignore_index=True, sort=False)
        subset = [col for col in ["grid_id", "lat", "lon", "Node_Cell_ID", "action_label", "sector_id"] if col in combined.columns]
        if subset:
            combined = combined.drop_duplicates(subset=subset, keep="last")
        dest = run_dir / "after_rf_surfaces" / dest_name
        dest.parent.mkdir(parents=True, exist_ok=True)
        combined.to_csv(dest, index=False)
        return str(dest), int(len(combined))

    after_path, after_rows = _combine(after_files, "after_rf_surface_combined.csv")
    before_path, before_rows = _combine(before_files, "before_rf_surface_combined.csv")
    inventory_path, inventory_rows = _combine(inventory_files, "after_cell_inventory_combined.csv")
    payload.update(
        {
            "after_rf_surface_combined_csv": after_path,
            "before_rf_surface_combined_csv": before_path,
            "after_cell_inventory_combined_csv": inventory_path,
            "after_rf_rows": after_rows,
            "before_rf_rows": before_rows,
            "after_cell_inventory_rows": inventory_rows,
        }
    )
    logger.info(
        "after_rf_combined surfaces=%d after_rows=%d before_rows=%d inventory_rows=%d",
        len(after_files),
        after_rows,
        before_rows,
        inventory_rows,
    )
    return payload


def run_model3_current_recommendation_test(config: CurrentModel3Config) -> Path:
    start = time.perf_counter()
    if not config.dataset_path.exists():
        current_builder.build_model3_current_dataset(config.rrc_sector_capacity)
    run_dir = _ensure_dir(config.output_root / f"model3_current_{_timestamp()}")
    log_path = run_dir / "log.txt"
    logger = _setup_logger(log_path)
    logger.info("start dataset=%s summary=%s", config.dataset_path, config.summary_path)
    if config.dataset_path.suffix.lower() in {".xlsx", ".xlsm", ".xls"}:
        source_df = pd.read_excel(config.dataset_path, sheet_name="Model3_Cell_Input")
        source_df = _filter_project196_cell_input(source_df)
        cell_inventory, inventory_summary = _build_cell_inventory_from_excel_input(source_df, config)
        context = _load_project196_excel_context(config, source_df, logger)
        logger.info("project196_excel_full_rf_mode rows=%d sectors=%d", len(cell_inventory), cell_inventory["sector_id"].nunique(dropna=True))
    else:
        source_df = pd.read_csv(config.dataset_path)
        cell_inventory, inventory_summary = _build_current_cell_inventory(source_df, config)
        context = _load_current_context(config, logger)
    context["after_rf_dir"] = run_dir / "after_rf_surfaces"
    context["after_rf_records"] = []
    recommendations = _build_recommendations(cell_inventory, config, logger, context)
    sector_inventory = future_rules._build_sector_inventory(cell_inventory)
    runtime_sec = time.perf_counter() - start
    after_rf_artifacts = _combine_after_rf_artifacts(context, run_dir, logger)
    summary = {
        "mode": "current_model3_recommendation",
        "dataset_path": str(config.dataset_path),
        "summary_path": str(config.summary_path),
        "rows_in_model3_dataset": int(len(cell_inventory)),
        "congested_cell_rows": int(len(recommendations)),
        "sector_rows": int(len(sector_inventory)),
        "inventory_summary": inventory_summary,
        "threshold": float(config.congestion_threshold),
        "max_sectors": int(config.max_sectors) if config.max_sectors is not None else None,
        "sector_parallelism": int(config.sector_parallelism),
        "rf_workers": int(config.rf_workers),
        "runtime_sec": round(float(runtime_sec), 4),
        "artifacts": {
            "recommendations_csv": str(run_dir / "model3_current_recommendations.csv"),
            "cell_inventory_csv": str(run_dir / "model3_current_cell_inventory.csv"),
            "sector_inventory_csv": str(run_dir / "model3_current_sector_inventory.csv"),
            "summary_json": str(run_dir / "summary.json"),
            "log": str(log_path),
            **after_rf_artifacts,
        },
    }
    workbook_path = _write_workbook(run_dir, cell_inventory, recommendations, sector_inventory, summary, config.dashboard_model_label)
    summary["workbook_path"] = str(workbook_path)
    summary["artifacts"]["workbook"] = str(workbook_path)
    recommendations.to_csv(run_dir / "model3_current_recommendations.csv", index=False)
    cell_inventory.to_csv(run_dir / "model3_current_cell_inventory.csv", index=False)
    sector_inventory.to_csv(run_dir / "model3_current_sector_inventory.csv", index=False)
    _save_json(run_dir / "summary.json", summary)
    config.stable_output_dir.mkdir(parents=True, exist_ok=True)
    for src, dest_name in [
        (run_dir / "summary.json", "model3_current_recommendation_summary.json"),
        (run_dir / "model3_current_recommendations.csv", "model3_current_recommendations.csv"),
        (workbook_path, "model3_current_recommendations.xlsx"),
        (log_path, "model3_current_recommendation.log"),
        (Path(after_rf_artifacts.get("after_rf_surface_combined_csv") or ""), "model3_after_rf_surface_combined.csv"),
        (Path(after_rf_artifacts.get("before_rf_surface_combined_csv") or ""), "model3_before_rf_surface_combined.csv"),
        (Path(after_rf_artifacts.get("after_cell_inventory_combined_csv") or ""), "model3_after_cell_inventory_combined.csv"),
        (Path(after_rf_artifacts.get("after_rf_manifest_json") or ""), "model3_after_rf_manifest.json"),
    ]:
        try:
            if str(src) and src.is_file():
                shutil.copy2(src, config.stable_output_dir / dest_name)
            elif dest_name.startswith("model3_after_") or dest_name.startswith("model3_before_"):
                stale = config.stable_output_dir / dest_name
                if stale.exists():
                    stale.unlink()
        except PermissionError:
            pass
    print(json.dumps(summary, indent=2, default=str))
    return run_dir


def parse_args() -> CurrentModel3Config:
    parser = argparse.ArgumentParser(description="Run current-state Model 3 recommendations without Model 1/2 inference.")
    parser.add_argument("--dataset-path", type=Path, default=DEFAULT_DATASET)
    parser.add_argument("--summary-path", type=Path, default=DEFAULT_SUMMARY)
    parser.add_argument("--output-root", type=Path, default=DEFAULT_OUTPUT_ROOT)
    parser.add_argument("--stable-output-dir", type=Path, default=DEFAULT_STABLE_OUTPUT_DIR)
    parser.add_argument("--congestion-threshold", type=float, default=DEFAULT_CONGESTION_THRESHOLD)
    parser.add_argument("--rrc-sector-capacity", type=float, default=future_builder.DEFAULT_RRC_SECTOR_CAPACITY)
    parser.add_argument("--max-sectors", type=int, default=None)
    parser.add_argument("--max-congested-cells", type=int, default=None)
    parser.add_argument("--carrier-reselection-hysteresis-db", type=float, default=0.0)
    parser.add_argument("--rf-workers", type=int, default=DEFAULT_RF_WORKERS)
    parser.add_argument("--max-interference-sites", type=int, default=10)
    parser.add_argument("--action-neighbor-cells", type=int, default=2)
    parser.add_argument("--sector-parallelism", type=int, default=1)
    parser.add_argument("--stop-on-partial", action="store_true")
    args = parser.parse_args()
    return CurrentModel3Config(
        dataset_path=args.dataset_path,
        summary_path=args.summary_path,
        output_root=args.output_root,
        stable_output_dir=args.stable_output_dir,
        congestion_threshold=args.congestion_threshold,
        rrc_sector_capacity=args.rrc_sector_capacity,
        max_sectors=args.max_sectors,
        max_congested_cells=args.max_congested_cells,
        carrier_reselection_hysteresis_db=float(args.carrier_reselection_hysteresis_db),
        rf_workers=max(1, int(args.rf_workers)),
        max_interference_sites=max(1, int(args.max_interference_sites)),
        action_neighbor_cells=max(0, int(args.action_neighbor_cells)),
        sector_parallelism=max(1, int(args.sector_parallelism)),
        stop_on_partial=bool(args.stop_on_partial),
    )


if __name__ == "__main__":
    run_model3_current_recommendation_test(parse_args())
