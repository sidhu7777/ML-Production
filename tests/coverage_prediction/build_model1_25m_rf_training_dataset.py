from __future__ import annotations

import argparse
import json
import math
from pathlib import Path
from typing import Dict, List

import numpy as np
import pandas as pd
from openpyxl import Workbook


BUCKET_ORDER = {"PART_1": 1, "PART_2": 2, "PART_3": 3}
BAND_CLASS_LABELS = ("LOW_BAND", "MID_BAND", "HIGH_BAND")


def _classify_band(freq_value: object) -> str:
    try:
        freq = float(freq_value)
    except (TypeError, ValueError):
        return "UNKNOWN"
    if not math.isfinite(freq) or freq <= 0:
        return "UNKNOWN"
    if freq <= 1000:
        return "LOW_BAND"
    if freq <= 2100:
        return "MID_BAND"
    return "HIGH_BAND"


def _safe_numeric(df: pd.DataFrame, col: str, default: float = 0.0) -> pd.Series:
    if col not in df.columns:
        return pd.Series(default, index=df.index, dtype="float64")
    return pd.to_numeric(df[col], errors="coerce").fillna(default).astype("float64")


def _ratio_vs_first_bucket(df: pd.DataFrame, value_col: str) -> pd.Series:
    work = df.sort_values(["grid_id", "bucket_seq"]).copy()
    current = pd.to_numeric(work[value_col], errors="coerce")
    baseline = pd.to_numeric(work.groupby("grid_id", sort=False)[value_col].transform("first"), errors="coerce")
    denom = baseline.abs().replace(0, np.nan)
    ratio = (current - baseline) / denom
    cold_start = baseline.fillna(0.0).eq(0.0) & current.fillna(0.0).gt(0.0)
    ratio = ratio.where(~cold_start, 1.0)
    return ratio.replace([np.inf, -np.inf], np.nan).fillna(0.0).clip(-1.0, 5.0)


def _robust_norm(series: pd.Series) -> pd.Series:
    values = pd.to_numeric(series, errors="coerce").replace([np.inf, -np.inf], np.nan)
    valid = values.dropna()
    if valid.empty:
        return pd.Series(0.0, index=series.index, dtype="float64")
    lo = float(valid.quantile(0.05))
    hi = float(valid.quantile(0.95))
    if not math.isfinite(lo) or not math.isfinite(hi) or hi <= lo:
        lo = float(valid.min())
        hi = float(valid.max())
    if hi <= lo:
        return pd.Series(0.0, index=series.index, dtype="float64")
    return ((values.fillna(lo) - lo) / (hi - lo)).clip(0.0, 1.0)


def _physical_cell_key(chunk: pd.DataFrame) -> pd.Series:
    site = pd.to_numeric(chunk["Site ID"], errors="coerce").round(0).astype("Int64").astype(str).replace({"<NA>": ""})
    sector = chunk["sector_identity"].astype(str).replace({"nan": "", "None": ""})
    cell = chunk["Node_Cell_ID"].astype(str).replace({"nan": "", "None": ""})
    band = pd.to_numeric(chunk["band"], errors="coerce").round(0).astype("Int64").astype(str).replace({"<NA>": ""})
    return site + "|" + sector + "|" + cell + "|" + band


def aggregate_baseline_rf(baseline_csv: Path, chunksize: int = 300_000) -> pd.DataFrame:
    usecols = [
        "time_bucket",
        "grid_id",
        "Node_Cell_ID",
        "pred_rsrp",
        "pred_rsrq",
        "pred_sinr",
        "band",
        "earfcn",
        "Site ID",
        "sector_identity",
        "canonical_sector_id",
        "carrier_load_share",
        "azimuth",
    ]
    grouped_parts: List[pd.DataFrame] = []
    for chunk in pd.read_csv(baseline_csv, usecols=usecols, chunksize=chunksize):
        chunk["time_bucket"] = chunk["time_bucket"].astype(str)
        chunk["grid_id"] = pd.to_numeric(chunk["grid_id"], errors="coerce").astype("Int64")
        chunk["Node_Cell_ID"] = chunk["Node_Cell_ID"].astype(str)
        for col in ["pred_rsrp", "pred_rsrq", "pred_sinr", "band", "earfcn", "carrier_load_share", "Site ID", "azimuth"]:
            chunk[col] = pd.to_numeric(chunk[col], errors="coerce")
        chunk["serving_cell_key"] = _physical_cell_key(chunk)
        grouped_parts.append(
            chunk.groupby(["time_bucket", "grid_id", "serving_cell_key"], dropna=False)
            .agg(
                candidate_rsrp_mean=("pred_rsrp", "mean"),
                candidate_rsrp_max=("pred_rsrp", "max"),
                candidate_rsrq_mean=("pred_rsrq", "mean"),
                candidate_sinr_mean=("pred_sinr", "mean"),
                candidate_sample_count=("pred_rsrp", "size"),
                Node_Cell_ID=("Node_Cell_ID", "first"),
                band=("band", "first"),
                earfcn=("earfcn", "first"),
                site_id=("Site ID", "first"),
                sector_identity=("sector_identity", "first"),
                canonical_sector_id=("canonical_sector_id", "first"),
                carrier_load_share=("carrier_load_share", "mean"),
                azimuth=("azimuth", "first"),
            )
            .reset_index()
        )
    combined = pd.concat(grouped_parts, ignore_index=True)
    return (
        combined.groupby(["time_bucket", "grid_id", "serving_cell_key"], dropna=False)
        .agg(
            candidate_rsrp_mean=("candidate_rsrp_mean", "mean"),
            candidate_rsrp_max=("candidate_rsrp_max", "max"),
            candidate_rsrq_mean=("candidate_rsrq_mean", "mean"),
            candidate_sinr_mean=("candidate_sinr_mean", "mean"),
            candidate_sample_count=("candidate_sample_count", "sum"),
            Node_Cell_ID=("Node_Cell_ID", "first"),
            band=("band", "first"),
            earfcn=("earfcn", "first"),
            site_id=("site_id", "first"),
            sector_identity=("sector_identity", "first"),
            canonical_sector_id=("canonical_sector_id", "first"),
            carrier_load_share=("carrier_load_share", "mean"),
            azimuth=("azimuth", "first"),
        )
        .reset_index()
    )


def grid_rf_dataset(cell_grid: pd.DataFrame) -> pd.DataFrame:
    work = cell_grid.copy()
    work["band_class"] = work["band"].map(_classify_band)
    work = work.sort_values(["time_bucket", "grid_id", "candidate_rsrp_mean"], ascending=[True, True, False])
    work["rank"] = work.groupby(["time_bucket", "grid_id"]).cumcount() + 1
    serving = work.loc[work["rank"] == 1].copy()
    second = work.loc[
        work["rank"] == 2,
        ["time_bucket", "grid_id", "serving_cell_key", "Node_Cell_ID", "candidate_rsrp_mean", "candidate_sinr_mean", "band"],
    ].rename(
        columns={
            "serving_cell_key": "neighbor1_cell_key",
            "Node_Cell_ID": "neighbor1_cell_id",
            "candidate_rsrp_mean": "neighbor1_rsrp",
            "candidate_sinr_mean": "neighbor1_sinr",
            "band": "neighbor1_band",
        }
    )
    third = work.loc[
        work["rank"] == 3,
        ["time_bucket", "grid_id", "serving_cell_key", "Node_Cell_ID", "candidate_rsrp_mean", "candidate_sinr_mean"],
    ].rename(
        columns={
            "serving_cell_key": "neighbor2_cell_key",
            "Node_Cell_ID": "neighbor2_cell_id",
            "candidate_rsrp_mean": "neighbor2_rsrp",
            "candidate_sinr_mean": "neighbor2_sinr",
        }
    )
    grid_summary = work.groupby(["time_bucket", "grid_id"], as_index=False).agg(
        candidate_cell_count=("serving_cell_key", "nunique"),
        total_prediction_samples=("candidate_sample_count", "sum"),
        mean_candidate_rsrp=("candidate_rsrp_mean", "mean"),
        max_candidate_rsrp=("candidate_rsrp_max", "max"),
        mean_candidate_rsrq=("candidate_rsrq_mean", "mean"),
        mean_candidate_sinr=("candidate_sinr_mean", "mean"),
        std_candidate_rsrp=("candidate_rsrp_mean", "std"),
        std_candidate_sinr=("candidate_sinr_mean", "std"),
        unique_sites=("site_id", "nunique"),
    )
    band_counts = (
        work.groupby(["time_bucket", "grid_id", "band_class"])["candidate_sample_count"]
        .sum()
        .unstack(fill_value=0.0)
        .reset_index()
    )
    for label in BAND_CLASS_LABELS:
        if label not in band_counts.columns:
            band_counts[label] = 0.0
    band_total = band_counts[list(BAND_CLASS_LABELS)].sum(axis=1).replace(0, np.nan)
    band_counts["low_band_ratio"] = (band_counts["LOW_BAND"] / band_total).fillna(0.0)
    band_counts["mid_band_ratio"] = (band_counts["MID_BAND"] / band_total).fillna(0.0)
    band_counts["high_band_ratio"] = (band_counts["HIGH_BAND"] / band_total).fillna(0.0)
    band_counts["carrier_count"] = (
        (band_counts["LOW_BAND"] > 0).astype(int)
        + (band_counts["MID_BAND"] > 0).astype(int)
        + (band_counts["HIGH_BAND"] > 0).astype(int)
    )
    band_counts["dominant_band_class"] = band_counts[["low_band_ratio", "mid_band_ratio", "high_band_ratio"]].idxmax(axis=1)
    band_counts["dominant_band_class"] = band_counts["dominant_band_class"].map(
        {"low_band_ratio": "LOW_BAND", "mid_band_ratio": "MID_BAND", "high_band_ratio": "HIGH_BAND"}
    )

    keep = serving[
        [
            "time_bucket",
            "grid_id",
            "serving_cell_key",
            "Node_Cell_ID",
            "site_id",
            "sector_identity",
            "canonical_sector_id",
            "band",
            "earfcn",
            "azimuth",
            "candidate_rsrp_mean",
            "candidate_rsrp_max",
            "candidate_rsrq_mean",
            "candidate_sinr_mean",
            "candidate_sample_count",
        ]
    ].rename(
        columns={
            "Node_Cell_ID": "serving_cell_id",
            "site_id": "serving_site_id",
            "sector_identity": "serving_sector",
            "canonical_sector_id": "serving_canonical_sector",
            "band": "serving_band",
            "earfcn": "serving_earfcn",
            "azimuth": "serving_azimuth",
            "candidate_rsrp_mean": "pred_rsrp",
            "candidate_rsrp_max": "pred_rsrp_max",
            "candidate_rsrq_mean": "pred_rsrq",
            "candidate_sinr_mean": "pred_sinr",
            "candidate_sample_count": "serving_sample_count",
        }
    )
    out = keep.merge(second, on=["time_bucket", "grid_id"], how="left")
    out = out.merge(third, on=["time_bucket", "grid_id"], how="left")
    out = out.merge(grid_summary, on=["time_bucket", "grid_id"], how="left")
    out = out.merge(
        band_counts[
            [
                "time_bucket",
                "grid_id",
                "low_band_ratio",
                "mid_band_ratio",
                "high_band_ratio",
                "carrier_count",
                "dominant_band_class",
            ]
        ],
        on=["time_bucket", "grid_id"],
        how="left",
    )
    out["rsrp_gap_to_neighbor1"] = out["pred_rsrp"] - out["neighbor1_rsrp"]
    out["sinr_gap_to_neighbor1"] = out["pred_sinr"] - out["neighbor1_sinr"]
    out["neighbor_interference_index"] = np.power(10.0, (out["neighbor1_rsrp"].fillna(-140.0) - out["pred_rsrp"]) / 10.0)
    return out


def add_geo_impact_zones(dataset: pd.DataFrame, radius_cells: int = 4) -> pd.DataFrame:
    out = dataset.copy()
    row_col_ready = {"grid_row", "grid_col"}.issubset(out.columns)
    building_growth = np.clip(pd.to_numeric(out["building_growth_ratio"], errors="coerce").fillna(0.0), 0.0, 2.0) / 2.0
    road_growth = np.clip(pd.to_numeric(out["road_growth_ratio"], errors="coerce").fillna(0.0), 0.0, 2.0) / 2.0
    clutter_transition = pd.to_numeric(out["clutter_transition_flag"], errors="coerce").fillna(0.0).clip(0.0, 1.0)
    activity = (
        (pd.to_numeric(out["mall_presence"], errors="coerce").fillna(0.0) > 0).astype(float) * 0.35
        + (pd.to_numeric(out["metro_presence"], errors="coerce").fillna(0.0) > 0).astype(float) * 0.35
    ).clip(0.0, 1.0)
    out["geo_transition_source_score"] = (
        0.48 * clutter_transition
        + 0.28 * building_growth
        + 0.18 * road_growth
        + 0.06 * activity
    ).clip(0.0, 1.0).round(6)
    out["geo_impact_zone_score"] = out["geo_transition_source_score"]
    out["geo_impact_nearby_flag"] = 0

    if not row_col_ready:
        return out

    impacted_parts: List[pd.DataFrame] = []
    base_cols = ["time_bucket", "grid_id", "grid_row", "grid_col", "geo_transition_source_score"]
    for bucket, bucket_df in out[base_cols].groupby("time_bucket", sort=False):
        bucket_df = bucket_df.copy()
        bucket_df["grid_row"] = pd.to_numeric(bucket_df["grid_row"], errors="coerce").astype("Int64")
        bucket_df["grid_col"] = pd.to_numeric(bucket_df["grid_col"], errors="coerce").astype("Int64")
        sources = bucket_df.loc[bucket_df["geo_transition_source_score"] > 0.01, ["grid_row", "grid_col", "geo_transition_source_score"]].dropna()
        if sources.empty:
            impacted_parts.append(bucket_df[["time_bucket", "grid_id"]].assign(_zone_score=0.0, _nearby_flag=0))
            continue
        contributions: List[pd.DataFrame] = []
        for dr in range(-radius_cells, radius_cells + 1):
            for dc in range(-radius_cells, radius_cells + 1):
                distance_cells = math.sqrt(float(dr * dr + dc * dc))
                if distance_cells > radius_cells:
                    continue
                weight = max(0.0, 1.0 - (distance_cells / (radius_cells + 0.001)))
                if weight <= 0:
                    continue
                shifted = sources.copy()
                shifted["grid_row"] = shifted["grid_row"] + dr
                shifted["grid_col"] = shifted["grid_col"] + dc
                shifted["_weighted_score"] = shifted["geo_transition_source_score"] * weight
                shifted["_nearby_flag"] = 0 if dr == 0 and dc == 0 else 1
                contributions.append(shifted[["grid_row", "grid_col", "_weighted_score", "_nearby_flag"]])
        contrib = pd.concat(contributions, ignore_index=True)
        contrib = (
            contrib.groupby(["grid_row", "grid_col"], as_index=False)
            .agg(_zone_score=("_weighted_score", "max"), _nearby_flag=("_nearby_flag", "max"))
        )
        merged = bucket_df[["time_bucket", "grid_id", "grid_row", "grid_col"]].merge(contrib, on=["grid_row", "grid_col"], how="left")
        merged["_zone_score"] = merged["_zone_score"].fillna(0.0).clip(0.0, 1.0)
        merged["_nearby_flag"] = merged["_nearby_flag"].fillna(0).astype(int)
        impacted_parts.append(merged[["time_bucket", "grid_id", "_zone_score", "_nearby_flag"]])

    impact = pd.concat(impacted_parts, ignore_index=True)
    out = out.merge(impact, on=["time_bucket", "grid_id"], how="left")
    out["geo_impact_zone_score"] = np.maximum(out["geo_impact_zone_score"], out["_zone_score"].fillna(0.0)).clip(0.0, 1.0).round(6)
    out["geo_impact_nearby_flag"] = np.maximum(out["geo_impact_nearby_flag"], out["_nearby_flag"].fillna(0).astype(int))
    out = out.drop(columns=["_zone_score", "_nearby_flag"], errors="ignore")
    return out


def apply_realistic_rf_evolution(dataset: pd.DataFrame) -> pd.DataFrame:
    out = dataset.sort_values(["grid_id", "bucket_seq"]).copy()
    for col in ["pred_rsrp", "pred_rsrq", "pred_sinr", "neighbor1_rsrp", "neighbor1_sinr", "neighbor_interference_index"]:
        out[f"raw_{col}"] = pd.to_numeric(out[col], errors="coerce")

    grouped = out.groupby("grid_id", sort=False)
    base_rsrp = grouped["raw_pred_rsrp"].transform("first")
    base_sinr = grouped["raw_pred_sinr"].transform("first")
    base_rsrq = grouped["raw_pred_rsrq"].transform("first")
    base_neighbor_rsrp = grouped["raw_neighbor1_rsrp"].transform("first")
    base_neighbor_sinr = grouped["raw_neighbor1_sinr"].transform("first")
    base_interference = grouped["raw_neighbor_interference_index"].transform("first")
    base_cell = grouped["serving_cell_key"].transform("first").astype(str)
    base_band = grouped["serving_band"].transform("first")

    bucket_step = (out["bucket_seq"] - 1).clip(lower=0)
    serving_changed = out["serving_cell_key"].astype(str).ne(base_cell)
    band_changed = pd.to_numeric(out["serving_band"], errors="coerce").ne(pd.to_numeric(base_band, errors="coerce"))

    building_growth = np.clip(pd.to_numeric(out["building_growth_ratio"], errors="coerce").fillna(0.0), 0.0, 2.0) / 2.0
    road_growth = np.clip(pd.to_numeric(out["road_growth_ratio"], errors="coerce").fillna(0.0), 0.0, 2.0) / 2.0
    clutter_penalty = out["clutter_transition_flag"].fillna(0).astype(float).clip(0.0, 1.0)
    density_norm = _robust_norm(out["building_area_ratio"]) * 0.65 + _robust_norm(out["road_density"]) * 0.35
    activity = (
        (pd.to_numeric(out["mall_presence"], errors="coerce").fillna(0.0) > 0).astype(float) * 0.55
        + (pd.to_numeric(out["metro_presence"], errors="coerce").fillna(0.0) > 0).astype(float) * 0.45
    ).clip(0.0, 1.0)
    geo_blockage = (
        1.20 * building_growth
        + 0.55 * road_growth
        + 1.10 * clutter_penalty
        + 0.45 * density_norm
        + 0.35 * activity
        + 1.35 * pd.to_numeric(out["geo_impact_zone_score"], errors="coerce").fillna(0.0).clip(0.0, 1.0)
    ).clip(0.0, 3.2)

    raw_rsrp_delta = out["raw_pred_rsrp"] - base_rsrp
    raw_sinr_delta = out["raw_pred_sinr"] - base_sinr
    raw_rsrq_delta = out["raw_pred_rsrq"] - base_rsrq
    raw_neighbor_rsrp_delta = out["raw_neighbor1_rsrp"] - base_neighbor_rsrp
    raw_neighbor_sinr_delta = out["raw_neighbor1_sinr"] - base_neighbor_sinr
    interference_delta = (out["raw_neighbor_interference_index"] - base_interference).replace([np.inf, -np.inf], np.nan).fillna(0.0)

    stable = ~(serving_changed | band_changed)
    rsrp_lower = np.where(stable, -4.0, -10.0)
    rsrp_upper = np.where(stable, 4.0, 10.0)
    sinr_lower = np.where(stable, -5.0, -13.0)
    sinr_upper = np.where(stable, 5.0, 13.0)

    rsrp_delta = np.clip(raw_rsrp_delta.fillna(0.0), rsrp_lower, rsrp_upper)
    sinr_delta = np.clip(raw_sinr_delta.fillna(0.0), sinr_lower, sinr_upper)

    rsrp_geo_penalty = np.where(bucket_step > 0, geo_blockage * (0.55 + 0.20 * bucket_step), 0.0)
    sinr_interference_penalty = (
        np.clip(interference_delta, 0.0, 2.0) * 1.20
        + geo_blockage * 0.55
        + np.where(band_changed, 0.45, 0.0)
    )
    sinr_interference_relief = np.clip(-interference_delta, 0.0, 2.0) * 0.80

    out["topology_serving_changed_flag"] = serving_changed.astype(int)
    out["topology_band_changed_flag"] = band_changed.astype(int)
    out["rf_geo_blockage_penalty_db"] = rsrp_geo_penalty.round(4)
    out["raw_rsrp_delta_db"] = raw_rsrp_delta.round(4)
    out["raw_sinr_delta_db"] = raw_sinr_delta.round(4)

    out["pred_rsrp"] = (base_rsrp + rsrp_delta - rsrp_geo_penalty).clip(-125.0, -44.0).round(6)
    out["pred_sinr"] = (base_sinr + sinr_delta - sinr_interference_penalty + sinr_interference_relief).clip(-15.0, 32.0).round(6)
    sinr_quality_delta = (out["pred_sinr"] - base_sinr).fillna(0.0)
    load_pressure_proxy = (
        0.45 * pd.to_numeric(out["neighbor_interference_index"], errors="coerce").fillna(0.0).clip(0.0, 3.0)
        + 0.30 * pd.to_numeric(out["geo_impact_zone_score"], errors="coerce").fillna(0.0).clip(0.0, 1.0)
        + 0.20 * (pd.to_numeric(out["candidate_cell_count"], errors="coerce").fillna(1.0).clip(1.0, 12.0) / 12.0)
        + 0.05 * np.where(band_changed, 1.0, 0.0)
    ).clip(0.0, 2.2)
    rsrq_delta = (
        0.16 * sinr_quality_delta
        - 0.85 * load_pressure_proxy
        - 0.20 * geo_blockage
        + np.clip(raw_rsrq_delta.fillna(0.0), -1.0, 1.0) * 0.35
    )
    out["rf_load_pressure_proxy"] = load_pressure_proxy.round(6)
    out["pred_rsrq"] = (base_rsrq + np.clip(rsrq_delta, -4.5, 3.0)).clip(-20.0, -3.0).round(6)
    out["neighbor1_rsrp"] = (base_neighbor_rsrp + np.clip(raw_neighbor_rsrp_delta.fillna(0.0), -9.0, 9.0) - 0.65 * geo_blockage).round(6)
    out["neighbor1_sinr"] = (base_neighbor_sinr + np.clip(raw_neighbor_sinr_delta.fillna(0.0), -11.0, 11.0) - 0.45 * geo_blockage).round(6)
    out["rsrp_gap_to_neighbor1"] = (out["pred_rsrp"] - out["neighbor1_rsrp"]).round(6)
    out["sinr_gap_to_neighbor1"] = (out["pred_sinr"] - out["neighbor1_sinr"]).round(6)
    out["neighbor_interference_index"] = np.power(10.0, (out["neighbor1_rsrp"].fillna(-140.0) - out["pred_rsrp"]) / 10.0).round(8)
    out["mean_candidate_rsrp"] = (out["mean_candidate_rsrp"] + (out["pred_rsrp"] - out["raw_pred_rsrp"]) * 0.45).round(6)
    out["mean_candidate_sinr"] = (out["mean_candidate_sinr"] + (out["pred_sinr"] - out["raw_pred_sinr"]) * 0.45).round(6)
    return out


def build_model1_dataset(geo_csv: Path, baseline_csv: Path, output_xlsx: Path, output_csv: Path, summary_json: Path) -> tuple[pd.DataFrame, Dict[str, object]]:
    geo = pd.read_csv(geo_csv)
    geo["grid_id"] = pd.to_numeric(geo["grid_id"], errors="coerce").astype("Int64")
    geo["time_bucket"] = geo["time_bucket"].astype(str)
    geo["bucket_seq"] = geo["time_bucket"].map(BUCKET_ORDER).fillna(0).astype(int)

    rf = grid_rf_dataset(aggregate_baseline_rf(baseline_csv))
    rf["grid_id"] = pd.to_numeric(rf["grid_id"], errors="coerce").astype("Int64")
    dataset = geo.merge(rf, on=["time_bucket", "grid_id"], how="left", validate="one_to_one")
    dataset["bucket_seq"] = dataset["time_bucket"].map(BUCKET_ORDER).fillna(dataset["bucket_seq"]).astype(int)

    for col in [
        "building_area_ratio",
        "building_count",
        "road_density",
        "road_length_m",
        "mall_presence",
        "metro_presence",
        "park_open_area",
        "open_area_ratio",
    ]:
        dataset[col] = _safe_numeric(dataset, col, 0.0)
    dataset["building_growth_ratio"] = _ratio_vs_first_bucket(dataset, "building_area_ratio").round(6)
    dataset["road_growth_ratio"] = _ratio_vs_first_bucket(dataset, "road_density").round(6)
    dataset["clutter_class"] = dataset["clutter_class"].fillna("UNKNOWN").astype(str)
    first_clutter = dataset.sort_values(["grid_id", "bucket_seq"]).groupby("grid_id")["clutter_class"].transform("first")
    dataset["clutter_transition_flag"] = dataset["clutter_class"].ne(first_clutter).astype(int)
    dataset = add_geo_impact_zones(dataset)
    dataset = apply_realistic_rf_evolution(dataset)

    preferred = [
        "time_bucket",
        "bucket_seq",
        "grid_id",
        "grid_row",
        "grid_col",
        "centroid_lat",
        "centroid_lon",
        "target_grid_area_m2",
        "clutter_class",
        "clutter_transition_flag",
        "geo_transition_source_score",
        "geo_impact_zone_score",
        "geo_impact_nearby_flag",
        "building_count",
        "building_area_ratio",
        "building_growth_ratio",
        "road_density",
        "road_length_m",
        "road_growth_ratio",
        "park_open_area",
        "open_area_ratio",
        "mall_presence",
        "metro_presence",
        "serving_cell_key",
        "serving_cell_id",
        "serving_site_id",
        "serving_sector",
        "serving_canonical_sector",
        "serving_band",
        "serving_earfcn",
        "serving_azimuth",
        "topology_serving_changed_flag",
        "topology_band_changed_flag",
        "neighbor1_cell_key",
        "neighbor1_cell_id",
        "neighbor1_band",
        "neighbor1_rsrp",
        "neighbor1_sinr",
        "neighbor2_cell_key",
        "neighbor2_cell_id",
        "neighbor2_rsrp",
        "neighbor2_sinr",
        "rsrp_gap_to_neighbor1",
        "sinr_gap_to_neighbor1",
        "neighbor_interference_index",
        "candidate_cell_count",
        "unique_sites",
        "total_prediction_samples",
        "low_band_ratio",
        "mid_band_ratio",
        "high_band_ratio",
        "carrier_count",
        "dominant_band_class",
        "mean_candidate_rsrp",
        "max_candidate_rsrp",
        "mean_candidate_rsrq",
        "mean_candidate_sinr",
        "std_candidate_rsrp",
        "std_candidate_sinr",
        "raw_pred_rsrp",
        "raw_pred_rsrq",
        "raw_pred_sinr",
        "raw_rsrp_delta_db",
        "raw_sinr_delta_db",
        "rf_geo_blockage_penalty_db",
        "rf_load_pressure_proxy",
        "pred_rsrp",
        "pred_rsrq",
        "pred_sinr",
        "pred_rsrp_max",
    ]
    cols = [col for col in preferred if col in dataset.columns] + [col for col in dataset.columns if col not in preferred]
    dataset = dataset[cols].sort_values(["grid_id", "bucket_seq"]).reset_index(drop=True)

    output_csv.parent.mkdir(parents=True, exist_ok=True)
    dataset.to_csv(output_csv, index=False)

    summary = make_summary(dataset, output_xlsx, output_csv)
    summary_json.write_text(json.dumps(summary, indent=2, default=str), encoding="utf-8")
    write_excel(output_xlsx, dataset, summary)
    return dataset, summary


def _delta_stats(p1: pd.DataFrame, p3: pd.DataFrame, common: pd.Index, col: str) -> Dict[str, float]:
    delta = pd.to_numeric(p3.loc[common, col], errors="coerce") - pd.to_numeric(p1.loc[common, col], errors="coerce")
    return {
        "mean_delta": float(delta.mean()),
        "median_delta": float(delta.median()),
        "p75_delta": float(delta.quantile(0.75)),
        "p95_delta": float(delta.quantile(0.95)),
        "max_delta": float(delta.max()),
        "min_delta": float(delta.min()),
        "changed_pct": float((delta.abs() > 1e-6).mean() * 100.0),
    }


def make_summary(dataset: pd.DataFrame, output_xlsx: Path, output_csv: Path) -> Dict[str, object]:
    p1 = dataset.loc[dataset["time_bucket"] == "PART_1"].set_index("grid_id")
    p3 = dataset.loc[dataset["time_bucket"] == "PART_3"].set_index("grid_id")
    common = p1.index.intersection(p3.index)
    delta_cols = [
        "pred_rsrp",
        "pred_rsrq",
        "pred_sinr",
        "mean_candidate_rsrp",
        "mean_candidate_sinr",
        "neighbor1_rsrp",
        "rsrp_gap_to_neighbor1",
    ]
    serving_changed = (
        p1.loc[common, "serving_cell_key"].astype(str).ne(p3.loc[common, "serving_cell_key"].astype(str)).mean() * 100.0
    )
    band_changed = (
        pd.to_numeric(p1.loc[common, "serving_band"], errors="coerce")
        .ne(pd.to_numeric(p3.loc[common, "serving_band"], errors="coerce"))
        .mean()
        * 100.0
    )
    summary: Dict[str, object] = {
        "output_xlsx": str(output_xlsx),
        "output_csv": str(output_csv),
        "rows": int(len(dataset)),
        "unique_grids": int(dataset["grid_id"].nunique(dropna=True)),
        "bucket_counts": {str(k): int(v) for k, v in dataset["time_bucket"].value_counts().sort_index().items()},
        "serving_cells_by_bucket": {
            str(k): int(v) for k, v in dataset.groupby("time_bucket")["serving_cell_key"].nunique(dropna=True).sort_index().items()
        },
        "part1_to_part3_serving_cell_changed_pct": float(serving_changed),
        "part1_to_part3_serving_band_changed_pct": float(band_changed),
        "part1_to_part3": {col: _delta_stats(p1, p3, common, col) for col in delta_cols if col in dataset.columns},
    }
    return summary


def write_excel(output_xlsx: Path, dataset: pd.DataFrame, summary: Dict[str, object]) -> None:
    output_xlsx.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook(write_only=True)

    ws = wb.create_sheet("README")
    for row in [
        ["item", "value"],
        ["purpose", "Model 1 25m RF/KPI training dataset"],
        ["row_level", "one row = grid_id + time_bucket"],
        ["serving_identity", "site + sector + cell + band"],
        ["do_not_train_on", "grid_id, time_bucket, bucket_seq"],
        ["train_targets", "pred_rsrp, pred_rsrq, pred_sinr"],
        ["rows", summary["rows"]],
        ["unique_grids", summary["unique_grids"]],
        ["serving_cell_changed_pct_part1_to_part3", summary["part1_to_part3_serving_cell_changed_pct"]],
        ["serving_band_changed_pct_part1_to_part3", summary["part1_to_part3_serving_band_changed_pct"]],
    ]:
        ws.append(row)

    ws = wb.create_sheet("RF_Evolution_Summary")
    ws.append(["metric", "mean_delta", "median_delta", "p75_delta", "p95_delta", "max_delta", "min_delta", "changed_pct"])
    for metric, stats in summary["part1_to_part3"].items():
        ws.append(
            [
                metric,
                stats["mean_delta"],
                stats["median_delta"],
                stats["p75_delta"],
                stats["p95_delta"],
                stats["max_delta"],
                stats["min_delta"],
                stats["changed_pct"],
            ]
        )

    ws = wb.create_sheet("Model1_Training_25m")
    ws.append(list(dataset.columns))
    for row in dataset.itertuples(index=False, name=None):
        ws.append(list(row))

    wb.save(output_xlsx)


def main() -> None:
    parser = argparse.ArgumentParser(description="Build separate 25m Model 1 RF training Excel.")
    parser.add_argument("--geo-csv", default="tests/coverage_prediction/data/model1_model2_training_geo_25m.csv")
    parser.add_argument("--baseline-csv", default="tests/coverage_prediction/data/model1_model2_training_baseline_25m_polygon.csv")
    parser.add_argument("--output-xlsx", default="tests/coverage_prediction/data/model1_training_25m_rf_dataset.xlsx")
    parser.add_argument("--output-csv", default="tests/coverage_prediction/data/model1_training_25m_rf_dataset.csv")
    parser.add_argument("--summary-json", default="tests/coverage_prediction/data/model1_training_25m_rf_dataset.summary.json")
    args = parser.parse_args()

    _, summary = build_model1_dataset(
        Path(args.geo_csv),
        Path(args.baseline_csv),
        Path(args.output_xlsx),
        Path(args.output_csv),
        Path(args.summary_json),
    )
    print(json.dumps(summary, indent=2, default=str))


if __name__ == "__main__":
    main()
