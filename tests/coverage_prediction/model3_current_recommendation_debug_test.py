from __future__ import annotations

import argparse
import json
import os
import sys
import time
from dataclasses import dataclass
from pathlib import Path
from typing import Any

import numpy as np
import pandas as pd
from openpyxl import Workbook

ML_ROOT = Path(__file__).resolve().parents[2]
os.chdir(ML_ROOT)
if str(ML_ROOT) not in sys.path:
    sys.path.insert(0, str(ML_ROOT))

from tests.coverage_prediction import build_model3_current_recommendation_dataset as current_builder
from tests.coverage_prediction import model3_business_rule_recommendation_test as future_rules
from tests.coverage_prediction import model3_current_recommendation_test as current_rules


DEFAULT_OUTPUT_ROOT = ML_ROOT / "tests" / "output" / "model3_current_recommendation_debug"
DEFAULT_STABLE_OUTPUT_DIR = ML_ROOT / "models" / "model3_current_recommendation_debug"


@dataclass
class Model3DebugConfig:
    dataset_path: Path = current_builder.CURRENT_MODEL3_DATASET_CSV
    output_root: Path = DEFAULT_OUTPUT_ROOT
    stable_output_dir: Path = DEFAULT_STABLE_OUTPUT_DIR
    congestion_threshold: float = current_rules.DEFAULT_CONGESTION_THRESHOLD


def _timestamp() -> str:
    return time.strftime("%Y%m%d_%H%M%S")


def _ensure_dir(path: Path) -> Path:
    path.mkdir(parents=True, exist_ok=True)
    return path


def _save_json(path: Path, payload: dict[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(payload, indent=2, default=str), encoding="utf-8")


def _normalize_bool(series: pd.Series) -> pd.Series:
    return series.astype(str).str.strip().str.lower().isin({"true", "1", "yes"})


def _sector_branch_audit(cell_df: pd.DataFrame, threshold: float) -> pd.DataFrame:
    rows: list[dict[str, Any]] = []
    for sector_id, group in cell_df.groupby("sector_id", dropna=False):
        group = group.copy()
        if not bool(group["congested"].any()):
            continue

        group["prb_before_pct"] = pd.to_numeric(group["prb_before_pct"], errors="coerce")
        group["rrc_before_pct"] = pd.to_numeric(group["rrc_before_pct"], errors="coerce")
        group["grid_count"] = pd.to_numeric(group["grid_count"], errors="coerce")
        group["prb_rrc_pressure"] = pd.to_numeric(group["prb_rrc_pressure"], errors="coerce")
        source = group.sort_values(["prb_rrc_pressure", "grid_count"], ascending=[False, False], na_position="last").iloc[0]
        peers = group.loc[group["Node_Cell_ID"].astype(str) != str(source["Node_Cell_ID"])].copy()

        has_alternate = bool(_normalize_bool(group["sector_has_alternate_carrier"]).any())
        can_add = bool(_normalize_bool(group["carrier_addition_possible"]).any()) and not bool(_normalize_bool(group["carrier_addition_blocked"]).all())
        strict_lb = peers.loc[(peers["prb_before_pct"] < threshold) & (peers["rrc_before_pct"] < threshold)].copy()
        prb_only_spare = peers.loc[peers["prb_before_pct"] < threshold].copy()
        rrc_only_spare = peers.loc[peers["rrc_before_pct"] < threshold].copy()

        if not has_alternate:
            lb_reason = "NO_ALTERNATE_CARRIER"
        elif peers.empty:
            lb_reason = "NO_PEER_ROWS"
        elif not strict_lb.empty:
            lb_reason = "STRICT_PEER_AVAILABLE"
        elif prb_only_spare.empty and rrc_only_spare.empty:
            lb_reason = "ALL_PEERS_PRB_AND_RRC_CONGESTED"
        elif prb_only_spare.empty:
            lb_reason = "PEER_RRC_OK_BUT_PRB_CONGESTED"
        elif rrc_only_spare.empty:
            lb_reason = "PEER_PRB_OK_BUT_RRC_CONGESTED"
        else:
            lb_reason = "NO_STRICT_MATCH_AFTER_FILTER"

        dry_run = current_rules._run_load_balance_current(group, current_rules.CurrentModel3Config(congestion_threshold=threshold))
        if dry_run["selected_peer_node_cell_id"]:
            decision_without_resim = f"LOAD_BALANCE::{dry_run['status']}"
        elif can_add and str(source.get("recommended_band_to_add", "")).strip():
            decision_without_resim = f"CARRIER_ADD::{str(source.get('recommended_band_to_add')).strip()}"
        elif int(source.get("sector_cell_count", 0) or 0) >= 1:
            decision_without_resim = "SECTOR_SPLIT"
        else:
            decision_without_resim = "NEW_SITE"

        rows.append(
            {
                "sector_id": sector_id,
                "site_id": source["site_id"],
                "source_node_cell_id": source["Node_Cell_ID"],
                "source_band": source["band"],
                "sector_cell_count": int(group["Node_Cell_ID"].nunique(dropna=True)),
                "sector_congested_count": int(group["congested"].sum()),
                "source_prb_before_pct": round(float(source["prb_before_pct"]), 3) if pd.notna(source["prb_before_pct"]) else np.nan,
                "source_rrc_before_pct": round(float(source["rrc_before_pct"]), 3) if pd.notna(source["rrc_before_pct"]) else np.nan,
                "has_alternate_carrier": has_alternate,
                "peer_count": int(len(peers)),
                "strict_load_balance_peer_count": int(len(strict_lb)),
                "peer_prb_lt_threshold_count": int(len(prb_only_spare)),
                "peer_rrc_lt_threshold_count": int(len(rrc_only_spare)),
                "load_balance_reason": lb_reason,
                "strict_load_balance_peer_ids": ", ".join(strict_lb["Node_Cell_ID"].astype(str).tolist()),
                "peer_ids": ", ".join(peers["Node_Cell_ID"].astype(str).tolist()),
                "carrier_addition_possible": can_add,
                "carrier_addition_blocked": bool(_normalize_bool(group["carrier_addition_blocked"]).all()),
                "recommended_band_to_add": str(source.get("recommended_band_to_add", "")).strip(),
                "existing_carriers": str(source.get("existing_carriers", "")).strip(),
                "decision_without_resim": decision_without_resim,
                "load_balance_dry_run_status": dry_run["status"],
                "load_balance_dry_run_next_step": dry_run["next_step"],
                "load_balance_dry_run_peer": dry_run["selected_peer_node_cell_id"],
                "load_balance_dry_run_peer_band": dry_run["selected_peer_band"],
                "load_balance_dry_run_projected_prb_after_pct": dry_run["projected_prb_after_pct"],
                "load_balance_dry_run_projected_rrc_after_pct": dry_run["projected_rrc_after_pct"],
            }
        )
    audit_df = pd.DataFrame(rows)
    if audit_df.empty:
        return audit_df
    return audit_df.sort_values(["source_prb_before_pct", "source_rrc_before_pct"], ascending=[False, False], na_position="last").reset_index(drop=True)


def _implementation_bug_notes(audit_df: pd.DataFrame) -> pd.DataFrame:
    notes = [
        {
            "bug_key": "ENGINE_CASCADE_IMPLEMENTED",
            "severity": "info",
            "description": "Current Model 3 now cascades through Load Balance -> Carrier Addition -> Sector Split -> New Site within the same sector evaluation. Remaining issues are no longer caused by single-step stopping.",
        },
        {
            "bug_key": "LOAD_BALANCE_FILTER_TOO_STRICT",
            "severity": "high",
            "description": "Load balance only accepts peers where both PRB and RRC are already below threshold before any move. That rejects sectors where a peer has some spare room but fails one strict pre-check.",
        },
        {
            "bug_key": "LOAD_BALANCE_NO_FALLTHROUGH_AFTER_REJECTED",
            "severity": "high",
            "description": "If a load-balance peer exists, _simulate_current_recommendation returns immediately even when the dry-run status is Partially Resolved or Rejected, so it never falls through to carrier addition in that same branch.",
        },
        {
            "bug_key": "SECTOR_SPLIT_GATE_TOO_PERMISSIVE",
            "severity": "medium",
            "description": "Sector split is allowed whenever sector_cell_count >= 1, which means almost every congested sector can jump to split once carrier addition is unavailable.",
        },
        {
            "bug_key": "CURRENT_MODEL3_FULL_RUN_IS_EXPENSIVE",
            "severity": "medium",
            "description": "Each topology-changing action triggers a real local baseline rerun, so an unbounded run across all congested sectors becomes very expensive unless the test is scoped.",
        },
    ]
    if not audit_df.empty:
        strict = int((audit_df["load_balance_reason"] == "STRICT_PEER_AVAILABLE").sum())
        alternates = int(audit_df["has_alternate_carrier"].sum())
        notes.append(
            {
                "bug_key": "DATASET_EXPLAINS_MANY_CARRIER_ADD_BRANCHES",
                "severity": "info",
                "description": f"Out of {len(audit_df)} congested sectors, {alternates} have alternate carriers but only {strict} pass the current strict load-balance peer filter. So many add-carrier branches are caused by current dataset pressure, not only by code order.",
            }
        )
    return pd.DataFrame(notes)


def _write_workbook(run_dir: Path, audit_df: pd.DataFrame, summary_df: pd.DataFrame, notes_df: pd.DataFrame) -> Path:
    wb = Workbook()
    wb.remove(wb.active)
    ws_summary = wb.create_sheet("Summary")
    future_rules._write_df_sheet(ws_summary, summary_df)
    ws_audit = wb.create_sheet("Sector Audit")
    future_rules._write_df_sheet(ws_audit, audit_df)
    ws_notes = wb.create_sheet("Bug Notes")
    future_rules._write_df_sheet(ws_notes, notes_df)
    workbook_path = run_dir / "model3_current_debug.xlsx"
    wb.save(workbook_path)
    return workbook_path


def run_model3_current_debug(config: Model3DebugConfig) -> Path:
    if not config.dataset_path.exists():
        current_builder.build_model3_current_dataset()

    run_dir = _ensure_dir(config.output_root / f"model3_current_debug_{_timestamp()}")
    source_df = pd.read_csv(config.dataset_path)
    cell_df, inventory_summary = current_rules._build_current_cell_inventory(
        source_df,
        current_rules.CurrentModel3Config(dataset_path=config.dataset_path, congestion_threshold=config.congestion_threshold),
    )
    audit_df = _sector_branch_audit(cell_df, config.congestion_threshold)
    notes_df = _implementation_bug_notes(audit_df)

    summary = {
        "mode": "model3_current_debug",
        "dataset_path": str(config.dataset_path),
        "cell_inventory_rows": int(len(cell_df)),
        "sector_audit_rows": int(len(audit_df)),
        "inventory_summary": inventory_summary,
        "threshold": float(config.congestion_threshold),
        "engine_shape": {
            "current_behavior": "cascading_per_sector",
            "expected_behavior": "cascading_load_balance_then_carrier_then_sector_split_then_new_site",
            "workers_used_for_topology_rerun": 1,
            "parallel_sector_execution": False,
        },
        "counts": {
            "strict_load_balance_available": int((audit_df["load_balance_reason"] == "STRICT_PEER_AVAILABLE").sum()) if not audit_df.empty else 0,
            "alternate_carrier_but_no_strict_peer": int(((audit_df["has_alternate_carrier"]) & (audit_df["load_balance_reason"] != "STRICT_PEER_AVAILABLE")).sum()) if not audit_df.empty else 0,
            "all_peers_prb_and_rrc_congested": int((audit_df["load_balance_reason"] == "ALL_PEERS_PRB_AND_RRC_CONGESTED").sum()) if not audit_df.empty else 0,
            "peer_rrc_ok_but_prb_congested": int((audit_df["load_balance_reason"] == "PEER_RRC_OK_BUT_PRB_CONGESTED").sum()) if not audit_df.empty else 0,
            "peer_prb_ok_but_rrc_congested": int((audit_df["load_balance_reason"] == "PEER_PRB_OK_BUT_RRC_CONGESTED").sum()) if not audit_df.empty else 0,
            "decision_load_balance": int(audit_df["decision_without_resim"].astype(str).str.startswith("LOAD_BALANCE::").sum()) if not audit_df.empty else 0,
            "decision_carrier_add": int(audit_df["decision_without_resim"].astype(str).str.startswith("CARRIER_ADD::").sum()) if not audit_df.empty else 0,
            "decision_sector_split": int((audit_df["decision_without_resim"] == "SECTOR_SPLIT").sum()) if not audit_df.empty else 0,
            "decision_new_site": int((audit_df["decision_without_resim"] == "NEW_SITE").sum()) if not audit_df.empty else 0,
        },
    }

    summary_df = pd.DataFrame(
        [{"metric": key, "value": value} for key, value in {
            "cell_inventory_rows": summary["cell_inventory_rows"],
            "sector_audit_rows": summary["sector_audit_rows"],
            "strict_load_balance_available": summary["counts"]["strict_load_balance_available"],
            "alternate_carrier_but_no_strict_peer": summary["counts"]["alternate_carrier_but_no_strict_peer"],
            "all_peers_prb_and_rrc_congested": summary["counts"]["all_peers_prb_and_rrc_congested"],
            "peer_rrc_ok_but_prb_congested": summary["counts"]["peer_rrc_ok_but_prb_congested"],
            "peer_prb_ok_but_rrc_congested": summary["counts"]["peer_prb_ok_but_rrc_congested"],
            "decision_load_balance": summary["counts"]["decision_load_balance"],
            "decision_carrier_add": summary["counts"]["decision_carrier_add"],
            "decision_sector_split": summary["counts"]["decision_sector_split"],
            "decision_new_site": summary["counts"]["decision_new_site"],
            "parallel_sector_execution": False,
            "workers_used_for_topology_rerun": 1,
        }.items()]
    )

    workbook_path = _write_workbook(run_dir, audit_df, summary_df, notes_df)
    audit_csv = run_dir / "model3_current_sector_audit.csv"
    notes_csv = run_dir / "model3_current_bug_notes.csv"
    summary_json = run_dir / "summary.json"
    audit_df.to_csv(audit_csv, index=False)
    notes_df.to_csv(notes_csv, index=False)
    _save_json(summary_json, summary)

    config.stable_output_dir.mkdir(parents=True, exist_ok=True)
    for src, dest_name in [
        (audit_csv, "model3_current_sector_audit.csv"),
        (notes_csv, "model3_current_bug_notes.csv"),
        (summary_json, "model3_current_debug_summary.json"),
        (workbook_path, "model3_current_debug.xlsx"),
    ]:
        if src.exists():
            try:
                src.replace(config.stable_output_dir / dest_name)
            except Exception:
                import shutil
                shutil.copy2(src, config.stable_output_dir / dest_name)

    print(json.dumps(summary, indent=2, default=str))
    return run_dir


def parse_args() -> Model3DebugConfig:
    parser = argparse.ArgumentParser(description="Debug current Model 3 branch selection without expensive reruns.")
    parser.add_argument("--dataset-path", type=Path, default=current_builder.CURRENT_MODEL3_DATASET_CSV)
    parser.add_argument("--output-root", type=Path, default=DEFAULT_OUTPUT_ROOT)
    parser.add_argument("--stable-output-dir", type=Path, default=DEFAULT_STABLE_OUTPUT_DIR)
    parser.add_argument("--congestion-threshold", type=float, default=current_rules.DEFAULT_CONGESTION_THRESHOLD)
    args = parser.parse_args()
    return Model3DebugConfig(
        dataset_path=args.dataset_path,
        output_root=args.output_root,
        stable_output_dir=args.stable_output_dir,
        congestion_threshold=args.congestion_threshold,
    )


if __name__ == "__main__":
    run_model3_current_debug(parse_args())
