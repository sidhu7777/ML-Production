"""Compare two LTE tilt recommendation output folders.

This debug is intentionally artifact-first: it compares saved CSV/JSON/XLSX
outputs from two runs so we can identify whether a difference came from input
grid scores, target/candidate search, validation, or final export.
"""

from __future__ import annotations

import argparse
import hashlib
import json
from pathlib import Path
from typing import Any

import pandas as pd


DEFAULT_GOOD_DIR = Path(
    r"C:\Users\PC\Desktop\ML\outputs\temp_de3e9b7e-7831-40ea-a296-1db936a7d493"
)
DEFAULT_STRACER_DIR = Path(
    r"C:\Users\PC\Desktop\S-Tracer Exe\S-Tracer Exe\ML\outputs\temp_731389d7-3d3c-4c4f-bf42-afeb625c7b27"
)

GOOD_CODE_DIR = Path(r"C:\Users\PC\Desktop\ML\tools\lte_tilt_recommandation")
STRACER_CODE_DIR = Path(
    r"C:\Users\PC\Desktop\S-Tracer Exe\S-Tracer Exe\ML\tools\lte_tilt_recommandation"
)

SUMMARY_KEYS = [
    "candidate_name",
    "update_count",
    "baseline_bad_count",
    "candidate_bad_count",
    "net_bad_reduction",
    "rf_delta_matched_row_count",
    "recompute_cell_count",
    "frontend_rsrp_before_bad_grid_count",
    "frontend_rsrq_before_bad_grid_count",
    "frontend_sinr_before_bad_grid_count",
    "combined_any_before_bad_grid_count",
    "weighted_frontend_before_bad_grid_count",
    "combined_rsrp_weight",
    "combined_rsrq_weight",
    "combined_sinr_weight",
    "changed_cells",
]


def _read_json(path: Path) -> dict[str, Any]:
    if not path.exists():
        return {}
    with path.open("r", encoding="utf-8") as handle:
        return json.load(handle)


def _sha16(path: Path) -> str | None:
    if not path.exists():
        return None
    h = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            h.update(chunk)
    return h.hexdigest()[:16]


def _normalized_sha16(path: Path) -> str | None:
    if not path.exists():
        return None
    text = path.read_text(encoding="utf-8", errors="replace")
    text = text.replace("\r\n", "\n").replace("\r", "\n")
    return hashlib.sha256(text.encode("utf-8")).hexdigest()[:16]


def _safe_float(value: Any) -> float:
    try:
        if pd.isna(value):
            return 0.0
        return float(value)
    except Exception:
        return 0.0


def _grid_report(run_dir: Path) -> tuple[dict[str, Any], pd.DataFrame]:
    path = run_dir / "frontend_grid_scores.csv"
    if not path.exists():
        return {"exists": False}, pd.DataFrame()
    df = pd.read_csv(path)
    report: dict[str, Any] = {
        "exists": True,
        "rows": len(df),
        "sha16": _sha16(path),
    }
    for col in ["scenario_id", "created_at"]:
        if col in df.columns:
            values = df[col].dropna().astype(str)
            report[f"{col}_unique"] = values.nunique()
            report[f"{col}_sample"] = ",".join(values.drop_duplicates().head(5).tolist())
    for col in ["is_bad_rsrp", "is_bad_rsrq", "is_bad_sinr", "is_bad_combined"]:
        if col in df.columns:
            report[col] = int(df[col].fillna(False).astype(bool).sum())
    for col in [
        "rsrp_severity",
        "rsrq_severity",
        "sinr_severity",
        "combined_weighted_severity",
        "baseline_point_count",
    ]:
        if col in df.columns:
            report[f"{col}_sum"] = round(float(pd.to_numeric(df[col], errors="coerce").fillna(0).sum()), 3)
    top_cols = [
        col
        for col in [
            "grid_id",
            "baseline_point_count",
            "baseline_avg_rsrp",
            "baseline_avg_rsrq",
            "baseline_avg_sinr",
            "rsrp_severity",
            "rsrq_severity",
            "sinr_severity",
            "combined_weighted_severity",
        ]
        if col in df.columns
    ]
    top = pd.DataFrame()
    if "combined_weighted_severity" in df.columns and top_cols:
        top = df.sort_values("combined_weighted_severity", ascending=False).head(12)[top_cols]
    return report, top


def _candidate_report(run_dir: Path) -> tuple[dict[str, Any], pd.DataFrame]:
    path = run_dir / "candidate_validation_results.csv"
    if not path.exists():
        return {"exists": False}, pd.DataFrame()
    df = pd.read_csv(path)
    report: dict[str, Any] = {
        "exists": True,
        "rows": len(df),
        "sha16": _sha16(path),
    }
    if "update_count" in df.columns:
        report["max_update_count_evaluated"] = int(pd.to_numeric(df["update_count"], errors="coerce").max())
        report["evaluated_update_counts"] = (
            pd.to_numeric(df["update_count"], errors="coerce")
            .dropna()
            .astype(int)
            .value_counts()
            .sort_index()
            .to_dict()
        )
    for col in [
        "score",
        "net_bad_reduction",
        "rf_delta_matched_row_count",
        "combined_any_net_bad_grid_reduction",
        "weighted_frontend_net_bad_grid_reduction",
    ]:
        if col in df.columns:
            report[f"best_{col}"] = round(float(pd.to_numeric(df[col], errors="coerce").max()), 6)
    sort_col = "score" if "score" in df.columns else None
    top_cols = [
        col
        for col in [
            "candidate_name",
            "update_count",
            "score",
            "net_bad_reduction",
            "rf_delta_matched_row_count",
            "frontend_sinr_net_bad_grid_reduction",
            "combined_any_net_bad_grid_reduction",
            "constraints_passed",
            "changed_cells",
        ]
        if col in df.columns
    ]
    top = pd.DataFrame()
    if sort_col and top_cols:
        top = df.sort_values(sort_col, ascending=False).head(12)[top_cols]
    return report, top


def _report_recommendation_count(run_dir: Path) -> int | None:
    path = run_dir / "RF_Optimization_Report.xlsx"
    if not path.exists():
        return None
    try:
        from openpyxl import load_workbook

        wb = load_workbook(path, read_only=True, data_only=True)
        if "Recommendations" not in wb.sheetnames:
            return None
        ws = wb["Recommendations"]
        return max(ws.max_row - 1, 0)
    except Exception as exc:
        print(f"[WARN] could not read workbook {path}: {exc}")
        return None


def _compare_summaries(good_dir: Path, stracer_dir: Path) -> pd.DataFrame:
    good = _read_json(good_dir / "best_candidate_summary.json")
    stracer = _read_json(stracer_dir / "best_candidate_summary.json")
    rows = []
    for key in SUMMARY_KEYS:
        rows.append({"field": key, "good": good.get(key), "stracer": stracer.get(key)})
    return pd.DataFrame(rows)


def _compare_code() -> pd.DataFrame:
    rows = []
    for name in [
        "candidate_validation.py",
        "cell_identity.py",
        "geo_logic.py",
        "recommendation_engine.py",
        "etilt_optimizer_cd2.py",
        "services.py",
    ]:
        good = GOOD_CODE_DIR / name
        stracer = STRACER_CODE_DIR / name
        good_hash = _normalized_sha16(good)
        stracer_hash = _normalized_sha16(stracer)
        rows.append(
            {
                "file": name,
                "logic_hash_good": good_hash,
                "logic_hash_stracer": stracer_hash,
                "same_normalized": good_hash == stracer_hash,
            }
        )
    return pd.DataFrame(rows)


def _print_frame(title: str, df: pd.DataFrame) -> None:
    print(f"\n[{title}]")
    if df.empty:
        print("(empty)")
    else:
        print(df.to_string(index=False, max_colwidth=160))


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--good-dir", default=str(DEFAULT_GOOD_DIR))
    parser.add_argument("--stracer-dir", default=str(DEFAULT_STRACER_DIR))
    args = parser.parse_args()

    good_dir = Path(args.good_dir)
    stracer_dir = Path(args.stracer_dir)

    print(f"[COMPARE] good_dir={good_dir}")
    print(f"[COMPARE] stracer_dir={stracer_dir}")

    _print_frame("best_candidate_summary", _compare_summaries(good_dir, stracer_dir))

    good_grid, good_grid_top = _grid_report(good_dir)
    stracer_grid, stracer_grid_top = _grid_report(stracer_dir)
    _print_frame("frontend_grid_scores_summary", pd.DataFrame([
        {"run": "good", **good_grid},
        {"run": "stracer", **stracer_grid},
    ]))
    _print_frame("top_grid_severity_good", good_grid_top)
    _print_frame("top_grid_severity_stracer", stracer_grid_top)

    good_cand, good_cand_top = _candidate_report(good_dir)
    stracer_cand, stracer_cand_top = _candidate_report(stracer_dir)
    _print_frame("candidate_validation_summary", pd.DataFrame([
        {"run": "good", **good_cand},
        {"run": "stracer", **stracer_cand},
    ]))
    _print_frame("top_candidates_good", good_cand_top)
    _print_frame("top_candidates_stracer", stracer_cand_top)

    _print_frame("rf_report_recommendation_rows", pd.DataFrame([
        {"run": "good", "recommendation_rows": _report_recommendation_count(good_dir)},
        {"run": "stracer", "recommendation_rows": _report_recommendation_count(stracer_dir)},
    ]))

    _print_frame("logic_file_compare", _compare_code())

    print("\n[DIAGNOSIS_HINT]")
    print(
        "If frontend_grid_scores differ, the issue is before candidate search: "
        "different grid/baseline data entered the same optimizer. If grid scores match "
        "but candidate_validation differs, then optimizer logic/config differs."
    )


if __name__ == "__main__":
    main()
